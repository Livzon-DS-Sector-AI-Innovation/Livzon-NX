import uuid
from pathlib import Path

import pytest
from pypdf import PdfWriter
from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject
from services import dazah_agent_service

from services.dazah_agent_service import (
    AgentBackendSource,
    AgentBackendV2Request,
    AgentTrustedSubject,
    DazahAIAgent,
    _attachment_catalog_instruction,
    _try_basic_command_response,
    _user_message_with_attachments,
)


def _payload(**values) -> AgentBackendV2Request:
    return AgentBackendV2Request(
        protocol_version="2.0",
        run_id=uuid.uuid4(),
        trace_id=uuid.uuid4(),
        session_id="web:session-1",
        subject=AgentTrustedSubject(
            tenant_id="test",
            user_id=str(uuid.uuid4()),
            source="web",
        ),
        source=AgentBackendSource(platform="web"),
        **values,
    )


def test_dazah_proxy_keeps_multimodal_message_parts() -> None:
    agent = object.__new__(DazahAIAgent)
    assert agent._model_supports_vision() is True


@pytest.mark.asyncio
async def test_help_command_returns_without_model_execution() -> None:
    response = await _try_basic_command_response(_payload(message="/help"))

    assert response is not None
    assert "`/new`" in response.message
    assert "`/restart`" in response.message
    assert "`/memory status`" in response.message
    assert "`/memory clear confirm`" in response.message
    assert "群聊不读取或修改个人记忆" in response.message
    assert response.tool_trace == []


@pytest.mark.asyncio
async def test_restrat_typo_is_not_a_command() -> None:
    response = await _try_basic_command_response(_payload(message="/restrat"))

    assert response is None


@pytest.mark.asyncio
async def test_status_command_reports_current_channel() -> None:
    response = await _try_basic_command_response(_payload(message="/status"))

    assert response is not None
    assert "渠道：Web" in response.message


@pytest.mark.asyncio
async def test_tasks_command_returns_recent_progress(monkeypatch) -> None:
    async def fake_execute(operation, *, params):
        assert operation == "agent.list_automation_runs"
        assert params["scope"] == "mine"
        return {
            "data": {
                "items": [
                    {"id": str(uuid.uuid4()), "status": "failed", "error_code": "tool.timeout"}
                ]
            }
        }

    monkeypatch.setattr(dazah_agent_service, "_execute_deterministic_operation", fake_execute)

    response = await _try_basic_command_response(_payload(message="/tasks"))

    assert response is not None
    assert "最近任务进度" in response.message
    assert "tool.timeout" in response.message
    assert "/retry" in response.message


@pytest.mark.asyncio
async def test_retry_command_returns_real_confirmation(monkeypatch) -> None:
    run_id = str(uuid.uuid4())
    confirmation = {
        "id": str(uuid.uuid4()),
        "operation": "agent.retry_automation_run",
        "summary": "重试失败运行",
        "risk_level": "medium",
        "status": "pending",
        "expires_at": "2026-08-06T08:00:00Z",
    }

    async def fake_execute(operation, *, params):
        assert operation == "agent.retry_automation_run"
        assert params == {"run_id": run_id}
        return {"requires_confirmation": True, "confirmation": confirmation}

    monkeypatch.setattr(dazah_agent_service, "_execute_deterministic_operation", fake_execute)

    response = await _try_basic_command_response(_payload(message=f"/retry {run_id}"))

    assert response is not None
    assert response.pending_confirmations == [confirmation]


def test_document_attachment_is_added_as_user_content() -> None:
    payload = _payload(
        message="请总结",
        attachments=[
            {
                "filename": "记录.txt",
                "content_type": "text/plain",
                "size": 12,
                "kind": "document",
                "text": "批次状态正常",
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "记录.txt" in content
    assert "批次状态正常" in content


def test_persistent_attachment_catalog_routes_follow_up_operations() -> None:
    attachment_id = str(uuid.uuid4())
    payload = _payload(
        message="继续修改销售数据.xlsx",
        attachment_catalog=[
            {
                "attachment_id": attachment_id,
                "filename": "销售数据.xlsx",
                "kind": "document",
                "version": 2,
            }
        ],
    )

    instruction = _attachment_catalog_instruction(payload)

    assert attachment_id in instruction
    assert "agent.read_attachment" in instruction
    assert "agent.mutate_tabular_attachment" in instruction
    assert "agent.delete_attachment" in instruction


def test_gateway_cached_pdf_text_is_extracted_without_exposing_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    pdf_path = tmp_path / "cache" / "documents" / "示例文档.pdf"
    pdf_path.parent.mkdir(parents=True)
    writer = PdfWriter()
    page = writer.add_blank_page(width=300, height=300)
    font = DictionaryObject(
        {
            NameObject("/Type"): NameObject("/Font"),
            NameObject("/Subtype"): NameObject("/Type1"),
            NameObject("/BaseFont"): NameObject("/Helvetica"),
        }
    )
    page[NameObject("/Resources")] = DictionaryObject(
        {NameObject("/Font"): DictionaryObject({NameObject("/F1"): writer._add_object(font)})}
    )
    stream = DecodedStreamObject()
    stream.set_data(b"BT /F1 12 Tf 40 200 Td (PDF acceptance text) Tj ET")
    page[NameObject("/Contents")] = writer._add_object(stream)
    with pdf_path.open("wb") as output:
        writer.write(output)
    payload = _payload(
        message="请总结",
        attachments=[
            {
                "filename": "示例文档.pdf",
                "content_type": "application/pdf",
                "kind": "document",
                "local_path": str(pdf_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "PDF acceptance text" in content
    assert str(pdf_path) not in content


def test_gateway_cached_csv_supports_chinese_encoding_without_exposing_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    csv_path = tmp_path / "cache" / "documents" / "批次.csv"
    csv_path.parent.mkdir(parents=True)
    csv_path.write_bytes("批次,状态\nB-001,合格\n".encode("gb18030"))
    payload = _payload(
        message="读取表格",
        attachments=[
            {
                "filename": "批次.csv",
                "content_type": "text/csv",
                "kind": "document",
                "local_path": str(csv_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "[CSV 数据]" in content
    assert "批次\t状态" in content
    assert "B-001\t合格" in content
    assert str(csv_path) not in content


def test_gateway_cached_xlsx_extracts_multiple_sheets_without_exposing_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    from openpyxl import Workbook

    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    xlsx_path = tmp_path / "cache" / "documents" / "批次.xlsx"
    xlsx_path.parent.mkdir(parents=True)
    workbook = Workbook()
    active = workbook.active
    active.title = "批次"
    active.append(["批次号", "状态"])
    active.append(["B-001", "合格"])
    summary = workbook.create_sheet("汇总")
    summary.append(["总数", 1])
    workbook.save(xlsx_path)
    payload = _payload(
        message="读取工作簿",
        attachments=[
            {
                "filename": "批次.xlsx",
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "kind": "document",
                "local_path": str(xlsx_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "[工作表: 批次]" in content
    assert "批次号\t状态" in content
    assert "B-001\t合格" in content
    assert "[工作表: 汇总]" in content
    assert "总数\t1" in content
    assert str(xlsx_path) not in content


def test_gateway_cached_xls_uses_legacy_workbook_reader(
    tmp_path: Path,
    monkeypatch,
) -> None:
    import xlrd

    class FakeSheet:
        nrows = 2
        ncols = 2

        @staticmethod
        def cell_value(row_index: int, column_index: int) -> str:
            return (("批次号", "状态"), ("B-002", "待检"))[row_index][column_index]

    class FakeWorkbook:
        @staticmethod
        def sheet_names() -> list[str]:
            return ["旧版数据"]

        @staticmethod
        def sheet_by_name(_name: str) -> FakeSheet:
            return FakeSheet()

        @staticmethod
        def release_resources() -> None:
            return None

    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    xls_path = tmp_path / "cache" / "documents" / "旧版批次.xls"
    xls_path.parent.mkdir(parents=True)
    xls_path.write_bytes(b"legacy workbook")
    monkeypatch.setattr(xlrd, "open_workbook", lambda *_args, **_kwargs: FakeWorkbook())
    payload = _payload(
        message="读取旧版工作簿",
        attachments=[
            {
                "filename": "旧版批次.xls",
                "content_type": "application/vnd.ms-excel",
                "kind": "document",
                "local_path": str(xls_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "[工作表: 旧版数据]" in content
    assert "批次号\t状态" in content
    assert "B-002\t待检" in content
    assert str(xls_path) not in content


def test_unextractable_cached_document_does_not_expose_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    document_path = tmp_path / "cache" / "documents" / "扫描件.bin"
    document_path.parent.mkdir(parents=True)
    document_path.write_bytes(b"binary")
    payload = _payload(
        message="请总结",
        attachments=[
            {
                "filename": "扫描件.bin",
                "content_type": "application/octet-stream",
                "kind": "document",
                "local_path": str(document_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "无法提取文本内容" in content
    assert str(document_path) not in content


def test_cached_audio_does_not_expose_path(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    audio_path = tmp_path / "cache" / "audio" / "语音.opus"
    audio_path.parent.mkdir(parents=True)
    audio_path.write_bytes(b"audio")
    payload = _payload(
        message="请转写",
        attachments=[
            {
                "filename": "语音.opus",
                "content_type": "audio/opus",
                "kind": "audio",
                "local_path": str(audio_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "当前未启用内容转写" in content
    assert str(audio_path) not in content


def test_image_attachment_builds_multimodal_user_content() -> None:
    payload = _payload(
        message="识别图片",
        attachments=[
            {
                "filename": "现场.png",
                "content_type": "image/png",
                "size": 3,
                "kind": "image",
                "data_base64": "YWJj",
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, list)
    assert content[0]["type"] == "text"
    assert content[1]["image_url"]["url"] == "data:image/png;base64,YWJj"


def test_gateway_cached_image_is_loaded_only_from_hermes_cache(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path))
    image_path = tmp_path / "cache" / "images" / "现场.png"
    image_path.parent.mkdir(parents=True)
    image_path.write_bytes(b"abc")
    payload = _payload(
        message="识别图片",
        attachments=[
            {
                "filename": "现场.png",
                "content_type": "image/png",
                "kind": "image",
                "local_path": str(image_path),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, list)
    assert content[1]["image_url"]["url"] == "data:image/png;base64,YWJj"


def test_gateway_attachment_rejects_path_outside_hermes_cache(
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setenv("HERMES_HOME", str(tmp_path / "hermes"))
    outside = tmp_path / "outside.txt"
    outside.write_text("sensitive", encoding="utf-8")
    payload = _payload(
        message="读取附件",
        attachments=[
            {
                "filename": "outside.txt",
                "content_type": "text/plain",
                "kind": "document",
                "local_path": str(outside),
            }
        ],
    )

    content = _user_message_with_attachments(payload)

    assert isinstance(content, str)
    assert "附件不可读取" in content
    assert "sensitive" not in content
