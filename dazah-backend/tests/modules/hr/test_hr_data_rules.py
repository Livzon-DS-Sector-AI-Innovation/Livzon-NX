from datetime import date, datetime
from types import SimpleNamespace as _SimpleNamespace
from typing import Any
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from app.modules.hr import api
from app.modules.hr.models import Employee
from app.modules.hr.service import (
    DepartmentService,
    EmployeeService,
    _dept_mapping_to_dict,
    _extract_number,
    _extract_text,
    _ms_to_date,
    _parse_contract_date,
    _parse_date,
    _parse_feishu_record,
)

SimpleNamespace: Any = _SimpleNamespace


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ([{"text": "数组文本"}], "数组文本"),
        ({"text": "字典文本"}, "字典文本"),
        ({"value": [{"text": "嵌套文本"}]}, "嵌套文本"),
        (None, ""),
        (123, "123"),
        ([], "[]"),
    ],
)
def test_extract_text(value: object, expected: str) -> None:
    assert _extract_text(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [(1, 1), (2.9, 2), ({"value": [3]}, 3), ({"value": []}, None), ("4", None)],
)
def test_extract_number(value: object, expected: int | None) -> None:
    assert _extract_number(value) == expected


def test_hr_date_parsers_support_feishu_formats() -> None:
    timestamp_ms = datetime(2026, 8, 20).timestamp() * 1000

    assert _ms_to_date(timestamp_ms) == date(2026, 8, 20)
    assert _ms_to_date(0) is None
    assert _parse_date(timestamp_ms) == date(2026, 8, 20)
    assert _parse_date("2026-08-20T10:00:00") == date(2026, 8, 20)
    assert _parse_date("bad") is None
    assert _parse_date(-1) is None
    assert _parse_contract_date([{"text": "2026/08/20"}]) == date(2026, 8, 20)
    assert _parse_contract_date("2026-08-20") == date(2026, 8, 20)
    assert _parse_contract_date("20260820") == date(2026, 8, 20)
    assert _parse_contract_date("") is None


def test_parse_feishu_record_maps_full_employee_payload_and_fallbacks() -> None:
    timestamp_ms = datetime(2026, 8, 20).timestamp() * 1000
    record = {
        "record_id": "rec-1",
        "updated_time": "2026-08-20T08:30:00Z",
        "fields": {
            "序号": {"value": [7]},
            "工号": [{"text": "1001"}],
            "姓名": "张三",
            "部门": "质量部",
            "职位": "检验员",
            "技能证书": ["化验员证"],
            "技能证书复审时间": "2027/08/20",
            "性别": "男",
            "拟转正日期": timestamp_ms,
            "进厂时间": timestamp_ms,
            "首次签订合同日期": timestamp_ms,
            "合同截止日期（2）": "2028-08-20",
            "手机": "13800000000",
            "备注": "重点培养",
            "在职状态": "",
        },
    }

    parsed = _parse_feishu_record(record)

    assert parsed["feishu_record_id"] == "rec-1"
    assert parsed["seq_number"] == 7
    assert parsed["employee_number"] == "1001"
    assert parsed["department"] == "质量部"
    assert parsed["position"] == "检验员"
    assert parsed["qualifications"] == ["化验员证"]
    assert parsed["hire_date"] == date(2026, 8, 20)
    assert parsed["phone"] == "13800000000"
    assert parsed["remarks"] == ["重点培养"]
    assert parsed["status"] == "在职"
    assert parsed["feishu_synced_at"] == date(2026, 8, 20)


def test_parse_feishu_record_uses_today_for_missing_or_invalid_update_time() -> None:
    missing = _parse_feishu_record({"fields": {}})
    invalid = _parse_feishu_record({"updated_time": "invalid", "fields": {}})

    assert missing["feishu_synced_at"] == date.today()
    assert invalid["feishu_synced_at"] == date.today()
    assert missing["status"] == "在职"


def test_department_mapping_serializes_timestamps() -> None:
    now = datetime(2026, 8, 20, 8, 30)
    mapping: Any = SimpleNamespace(
        id=uuid4(),
        source_name="质量管理部",
        target_name="质量部",
        match_level="exact",
        mapping_type="manual",
        priority=10,
        enabled=True,
        remark=None,
        created_at=now,
        updated_at=None,
    )

    result = _dept_mapping_to_dict(mapping)

    assert result["created_at"] == now.isoformat()
    assert result["updated_at"] is None
    assert result["enabled"] is True


def _department(
    name: str,
    *,
    parent_id: object = None,
    sort_order: int = 0,
    headcount: int | None = None,
    current_count: int | None = None,
) -> SimpleNamespace:
    now = datetime(2026, 8, 20)
    return SimpleNamespace(
        id=uuid4(),
        name=name,
        code=name,
        description=f"{name}说明",
        leader_name="负责人",
        parent_id=parent_id,
        feishu_open_department_id=f"fs-{name}",
        sort_order=sort_order,
        headcount=headcount,
        current_count=current_count,
        responsibilities="职责",
        category="职能",
        created_at=now,
        updated_at=now,
    )


def test_build_department_tree_sorts_roots_and_attaches_children() -> None:
    root = _department("总经办", sort_order=5)
    workshop = _department("201车间", sort_order=1)
    finance = _department("财务部", sort_order=9)
    child = _department("财务科", parent_id=finance.id, sort_order=2)

    roots = DepartmentService._build_dept_tree([workshop, child, finance, root])

    assert [item.name for item in roots] == ["总经办", "财务部", "201车间"]
    assert finance._children_list == [child]
    assert DepartmentService._build_dept_tree([]) == []


@pytest.mark.asyncio
async def test_get_department_tree_calculates_vacancy_and_nested_nodes() -> None:
    session: Any = MagicMock()
    service = DepartmentService(session)
    service.repo = AsyncMock()
    root = _department("质量管理部", headcount=10, current_count=7)
    child = _department("QA部", parent_id=root.id)
    service.repo.list_all_departments.return_value = [root, child]

    result = await service.get_department_tree()

    assert result[0]["vacancy"] == 3
    assert result[0]["children"][0]["name"] == "QA部"
    assert result[0]["children"][0]["vacancy"] is None
    service.repo.list_all_departments.return_value = []
    assert await service.get_department_tree() == []


def test_employee_to_bitable_fields_maps_modern_fields_and_filters_empty_values() -> (
    None
):
    employee = Employee(
        id=uuid4(),
        employee_number="1001",
        name="张三",
        department="质量部",
        position="检验员",
        hire_date=date(2026, 8, 20),
        factory_entry_date=date(2026, 8, 19),
        contract_end_2=date(2028, 8, 20),
        qualifications=["化验员证"],
        remarks=["重点培养"],
        status="在职",
    )
    service = EmployeeService(MagicMock())

    fields = service._to_bitable_fields(employee)

    assert fields["工号"] == 1001
    assert fields["姓名"] == "张三"
    assert fields["一级部门"] == "质量部"
    assert isinstance(fields["入职日期"], int)
    assert fields["合同截止日期（2）"] == "2028-08-20"
    assert fields["技能证书"] == ["化验员证"]
    assert fields["备注"] == ["重点培养"]
    assert "联系电话" not in fields


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (None, None),
        ("", None),
        (date(2026, 8, 20), date(2026, 8, 20)),
        (datetime(2026, 8, 20, 10), date(2026, 8, 20)),
        ("2026-08-20", date(2026, 8, 20)),
        ("2026/08/20", date(2026, 8, 20)),
        ("2026.08.20\n09:00~10:30", date(2026, 8, 20)),
        ("2026年08月20日", date(2026, 8, 20)),
        ("bad", None),
    ],
)
def test_parse_excel_date(value: object, expected: date | None) -> None:
    assert api._parse_excel_date(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("09:00~10:30", 1.5),
        ("9:15 至 10:00", 0.75),
        ("23:00-24:00", None),
        ("10:00-09:00", None),
        ("无时间", None),
        (None, None),
    ],
)
def test_calc_duration_from_text(value: object, expected: float | None) -> None:
    assert api._calc_duration_from_text(value) == expected


def test_header_detection_alias_mapping_and_candidate_selection() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["某部门培训台账"])
    ws.append(["日期/时间", "培训内容", "培训师", "受训部门", "备注"])
    ws.append(["2026-08-20", "GMP", "张三", "质量部", "完成"])

    header_row, headers = api._locate_header_row(ws)
    mapped_row, col_map = api._read_excel_header_map(ws)
    candidate_row, candidate_headers = api._find_candidate_header_row(ws)

    assert header_row == 2
    assert headers[:3] == ["日期/时间", "培训内容", "培训师"]
    assert mapped_row == 2
    assert col_map == {
        0: "training_datetime",
        1: "training_content",
        2: "instructor",
        3: "involved_depts",
        4: "remarks",
    }
    assert candidate_row == 2
    assert candidate_headers == headers
    assert api._map_headers_by_alias(["未知", "培训日期"]) == {"1": "training_date"}


def test_header_detection_returns_empty_for_unrecognized_sheet() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["说明", "无表头"])

    assert api._locate_header_row(ws) == (0, [])
    assert api._read_excel_header_map(ws) == (0, {})
    assert api._find_candidate_header_row(ws) == (0, [])


def test_cell_text_and_clip_normalize_excel_values() -> None:
    assert api._cell_text(None) is None
    assert api._cell_text("  ") is None
    assert api._cell_text(" 内容 ") == "内容"
    assert api._cell_text(2.0) == "2"
    assert api._cell_text(2.5) == "2.5"
    assert api._clip("abcdef", 3) == "abc"
    assert api._clip(None, 3) is None


@pytest.mark.asyncio
async def test_import_rows_with_mapping_creates_records_and_skips_blanks() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["培训时间", "培训内容", "培训时长（h）", "授课人", "备注"])
    ws.append(["2026.08.20\n09:00~10:30", "GMP培训", "", "张三", 1.0])
    ws.append([None, None, None, None, None])
    ws.append(["2026-08-21", "安全培训", "2", "李四", "完成"])
    col_map = {
        0: "training_datetime",
        1: "training_content",
        2: "duration_hours",
        3: "instructor",
        4: "remarks",
    }
    record_service: Any = AsyncMock()

    created = await api._import_rows_with_mapping(
        ws, 1, col_map, "质量部", record_service
    )

    assert created == 2
    assert record_service.create_record.await_count == 2
    first = record_service.create_record.await_args_list[0].args[0]
    second = record_service.create_record.await_args_list[1].args[0]
    assert first.training_date == date(2026, 8, 20)
    assert first.duration_hours == 1.5
    assert first.training_datetime == "2026.08.20 09:00~10:30"
    assert first.remarks == "1"
    assert second.duration_hours == 2


def test_count_data_rows_returns_first_three_samples() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["培训日期", "培训内容"])
    for index in range(5):
        ws.append([f"2026-08-{index + 1:02d}", f"课程{index}\n说明"])
    ws.append([None, None])

    count, samples = api._count_data_rows(
        ws, 1, {0: "training_date", 1: "training_content"}
    )

    assert count == 5
    assert len(samples) == 3
    assert samples[0] == ["2026-08-01", "课程0 说明"]


def test_generate_employee_training_ledger_excel_contains_employee_and_records() -> (
    None
):
    buffer = api._generate_training_ledger_excel(
        {
            "name": "张三",
            "gender": "男",
            "employee_number": "1001",
            "department": "质量部",
            "position": "检验员",
            "hire_date": "2026-08-20",
        },
        [
            {
                "training_date": "2026-08-21",
                "training_subject": "GMP",
                "training_method": "面授",
                "duration_hours": 2,
                "trainer": "李四",
                "assessment_result": "合格",
                "remarks": "完成",
            }
        ],
    )
    workbook = load_workbook(buffer)
    ws = workbook["员工培训台账"]

    assert ws["B3"].value == "张三"
    assert ws["F3"].value == "1001"
    assert ws["B8"].value == "GMP"
    assert ws.max_row >= 20


def test_generate_department_ledger_excel_uses_expected_headers_and_dates() -> None:
    buffer = api._generate_dept_ledger_excel(
        "质量部",
        [
            {
                "training_datetime": "2026-08-20 09:00",
                "training_date": date(2026, 8, 20),
                "duration_hours": 1.5,
                "training_content": "GMP",
                "instructor": "张三",
            }
        ],
    )
    workbook = load_workbook(buffer)
    ws = workbook["年度培训统计表"]

    assert ws["A1"].value == "质量部 年度培训统计表"
    assert ws["A2"].value == "培训时间"
    assert ws["B3"].value == "2026-08-20"
    assert ws["D3"].value == "GMP"


def test_generate_annual_plan_excel_pads_template_and_formats_confirmation() -> None:
    buffer = api._generate_annual_plan_excel(
        {"year": 2027, "department": "质量部"},
        [
            {
                "month": "第一季度",
                "duration_hours": 4,
                "content_and_textbook": "GMP法规",
                "target_audience": "全员",
                "position_and_count": "质量部/20人",
                "training_method": "面授",
                "tracking_status": "计划中",
                "confirmer": "负责人",
                "confirm_date": "2026-08-20",
                "remarks": "重点",
            }
        ],
    )
    workbook = load_workbook(buffer)
    ws = workbook["年度培训计划"]

    assert ws["A1"].value == "2027 年培训计划"
    assert ws["A2"].value == "部门：质量部"
    assert ws["B4"].value == "第一季度\n4课时"
    assert ws["H4"].value == "负责人 / 2026.08.20"
    assert ws.max_row >= 17
