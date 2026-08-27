from __future__ import annotations

from copy import copy
from pathlib import Path

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.modules.registration.service import certificate as certificate_service

SHEET_HEADERS: dict[str, list[str]] = {
    "国外注册": [
        "序号",
        "证照名称",
        "证书编号",
        "国家/发证机关",
        "发证日期",
        "有效期/复验期",
        "产品范围",
        "质量标准",
        "页数",
        "备注",
    ],
    "国内注册": [
        "序号",
        "证照名称",
        "受理号",
        "批件号",
        "编号",
        "发证机关",
        "发证日期",
        "有效期/复验期",
        "产品范围",
        "质量标准",
        "页数",
        "备注",
    ],
    "国内GMP": [
        "序号",
        "证照名称",
        "编号",
        "发证机关",
        "发证日期",
        "有效期/复验期",
        "产品范围",
        "质量标准",
        "页数",
        "备注",
    ],
    "国际GMP": [
        "序号",
        "证照名称",
        "编号",
        "国家/发证机关",
        "发证日期",
        "有效期/复验期",
        "产品范围",
        "质量标准",
        "页数",
        "备注",
    ],
}


def build_certificate_workbook_fixture(
    workbook_path: Path,
    rows_by_sheet: dict[str, list[list[str | int | None]]],
    *,
    header_overrides: dict[str, list[str]] | None = None,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, headers in SHEET_HEADERS.items():
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(headers)
        )
        worksheet.cell(row=1, column=1).value = sheet_name
        final_headers = (header_overrides or {}).get(sheet_name, headers)
        for column_index, header in enumerate(final_headers, start=1):
            worksheet.cell(row=2, column=column_index).value = header
        for row_offset, row_values in enumerate(
            rows_by_sheet.get(sheet_name, []), start=3
        ):
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_offset, column=column_index).value = value

    workbook.save(workbook_path)
    workbook.close()


def build_certificate_export_template_fixture(workbook_path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "国外注册"

    for sheet_index, (sheet_name, headers) in enumerate(SHEET_HEADERS.items()):
        worksheet = (
            workbook.worksheets[0]
            if sheet_index == 0
            else workbook.create_sheet(sheet_name)
        )
        worksheet.title = sheet_name
        worksheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(headers)
        )
        worksheet.cell(row=1, column=1).value = sheet_name

        for column_index, header in enumerate(headers, start=1):
            header_cell = worksheet.cell(row=2, column=column_index)
            header_cell.value = header
            header_cell.font = Font(name="宋体", size=11, bold=True)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            data_cell = worksheet.cell(row=3, column=column_index)
            data_cell.value = f"模板值{column_index}"
            data_cell.font = Font(name="宋体", size=10)
            data_cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            data_cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        worksheet.row_dimensions[1].height = 36
        worksheet.row_dimensions[2].height = 26
        worksheet.row_dimensions[3].height = 28

        if sheet_name == "国内GMP":
            worksheet.merge_cells("B15:B16")
            worksheet.merge_cells("D15:D16")

    workbook.save(workbook_path)
    workbook.close()


def patch_certificate_workbook_path(
    monkeypatch: pytest.MonkeyPatch, workbook_path: Path
) -> None:
    monkeypatch.setattr(
        certificate_service, "_get_certificate_workbook_path", lambda: workbook_path
    )


async def reset_certificate_entries(client: AsyncClient) -> None:
    """通过 API 软删除现有证书记录。

    必须使用与后续写操作相同的 client 会话：client 与 db_session fixture 是两个
    独立连接，跨连接批量 UPDATE 会持有行锁导致服务端写入阻塞（死锁）。
    """
    response = await client.get("/api/v1/registration/certificate-management/workbook")
    if response.status_code == 404:
        return
    assert response.status_code == 200
    for sheet in response.json()["data"]["sheets"]:
        for row in sheet["rows"]:
            delete_response = await client.delete(
                f"/api/v1/registration/certificate-management/entries/{row['id']}"
            )
            assert delete_response.status_code == 200


@pytest.mark.asyncio
async def test_get_certificate_workbook_detail_returns_all_named_sheets(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    workbook_path = tmp_path / "certificate-seed.xlsx"
    build_certificate_workbook_fixture(
        workbook_path,
        {
            "国外注册": [
                [
                    1,
                    "国外证书A",
                    "A-001",
                    "FDA",
                    "2026.01.01",
                    "2028.01.01",
                    "多拉菌素",
                    "内控",
                    2,
                    "备注A",
                ]
            ],
            "国内注册": [
                [
                    1,
                    "国内证书B",
                    "SLH-1",
                    "PJ-1",
                    "CN-001",
                    "NMPA",
                    "2026.02.02",
                    "2028.02.02",
                    "洛伐他汀",
                    "内控",
                    3,
                    "备注B",
                ]
            ],
        },
    )
    patch_certificate_workbook_path(monkeypatch, workbook_path)

    response = await client.get("/api/v1/registration/certificate-management/workbook")

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["workbook_name"] == "certificate-seed.xlsx"
    assert [sheet["sheet_key"] for sheet in body["sheets"]] == [
        "international-registration",
        "domestic-registration",
        "domestic-gmp",
        "international-gmp",
    ]
    assert body["sheets"][0]["rows"][0]["values"]["证照名称"] == "国外证书A"
    assert body["sheets"][1]["rows"][0]["values"]["证照名称"] == "国内证书B"


@pytest.mark.asyncio
async def test_import_certificate_workbook_overwrites_all_entries(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    seed_path = tmp_path / "seed.xlsx"
    build_certificate_workbook_fixture(
        seed_path,
        {
            "国外注册": [
                [
                    1,
                    "旧国外证书",
                    "OLD-001",
                    "FDA",
                    "2024.01.01",
                    "2026.01.01",
                    "旧产品",
                    "旧标准",
                    1,
                    "旧备注",
                ]
            ],
        },
    )
    patch_certificate_workbook_path(monkeypatch, seed_path)

    initial_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    assert initial_response.status_code == 200
    assert (
        initial_response.json()["data"]["sheets"][0]["rows"][0]["values"]["证照名称"]
        == "旧国外证书"
    )

    import_path = tmp_path / "import.xlsx"
    build_certificate_workbook_fixture(
        import_path,
        {
            "国外注册": [
                [
                    1,
                    "国外证书A",
                    "A-001",
                    "FDA",
                    "2026.01.01",
                    "2028.01.01",
                    "多拉菌素",
                    "内控",
                    2,
                    "备注A",
                ]
            ],
            "国内注册": [
                [
                    1,
                    "国内证书B",
                    "SLH-1",
                    "PJ-1",
                    "CN-001",
                    "NMPA",
                    "2026.02.02",
                    "2028.02.02",
                    "洛伐他汀",
                    "内控",
                    3,
                    "备注B",
                ]
            ],
        },
    )

    with import_path.open("rb") as handle:
        response = await client.post(
            "/api/v1/registration/certificate-management/workbook/import",
            files={
                "file": (
                    "import.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["imported_sheet_count"] == 4
    assert body["imported_record_count"] == 2
    assert body["replaced_record_count"] == 1

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    workbook_body = workbook_response.json()["data"]
    assert workbook_body["sheets"][0]["rows"][0]["values"]["证照名称"] == "国外证书A"
    assert workbook_body["sheets"][1]["rows"][0]["values"]["证照名称"] == "国内证书B"

    # 旧记录被软删除后不再出现在台账中
    all_names = [
        row["values"]["证照名称"]
        for sheet in workbook_body["sheets"]
        for row in sheet["rows"]
    ]
    assert "旧国外证书" not in all_names
    assert len(all_names) == 2


@pytest.mark.asyncio
async def test_import_certificate_workbook_rolls_back_on_invalid_header(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    seed_path = tmp_path / "seed.xlsx"
    build_certificate_workbook_fixture(
        seed_path,
        {
            "国外注册": [
                [
                    1,
                    "旧国外证书",
                    "OLD-001",
                    "FDA",
                    "2024.01.01",
                    "2026.01.01",
                    "旧产品",
                    "旧标准",
                    1,
                    "旧备注",
                ]
            ],
        },
    )
    patch_certificate_workbook_path(monkeypatch, seed_path)

    initial_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    assert initial_response.status_code == 200

    broken_path = tmp_path / "broken.xlsx"
    build_certificate_workbook_fixture(
        broken_path,
        {
            "国外注册": [
                [
                    1,
                    "国外证书A",
                    "A-001",
                    "FDA",
                    "2026.01.01",
                    "2028.01.01",
                    "多拉菌素",
                    "内控",
                    2,
                    "备注A",
                ]
            ]
        },
        header_overrides={
            "国外注册": [
                "序号",
                "错误列名",
                "证书编号",
                "国家/发证机关",
                "发证日期",
                "有效期/复验期",
                "产品范围",
                "质量标准",
                "页数",
                "备注",
            ]
        },
    )

    with broken_path.open("rb") as handle:
        response = await client.post(
            "/api/v1/registration/certificate-management/workbook/import",
            files={
                "file": (
                    "broken.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 400
    assert "国外注册" in response.json()["message"]

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    workbook_body = workbook_response.json()["data"]
    assert workbook_body["sheets"][0]["rows"][0]["values"]["证照名称"] == "旧国外证书"


@pytest.mark.asyncio
async def test_import_certificate_workbook_ignores_sequence_only_placeholder_rows(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    seed_path = tmp_path / "seed.xlsx"
    build_certificate_workbook_fixture(seed_path, {})
    patch_certificate_workbook_path(monkeypatch, seed_path)

    import_path = tmp_path / "import-with-placeholders.xlsx"
    build_certificate_workbook_fixture(
        import_path,
        {
            "国内注册": [
                [
                    1,
                    "营业执照",
                    "/",
                    "91640221574877733M",
                    "91640221574877733M",
                    "平罗县审批服务管理局",
                    "2024.09.30",
                    "长期",
                    "食品添加剂",
                    "——",
                    "1",
                    "——",
                ],
                [
                    2,
                    "兽药生产许可证",
                    "000018250603001",
                    "/",
                    "兽药生产证字30014号",
                    "宁夏回族自治区农业农村厅",
                    "2025.06.19",
                    "2021.07.08-2026.07.07",
                    "预混剂",
                    "盐酸林可霉素",
                    "1",
                    "——",
                ],
                [40, None, None, None, None, None, None, None, None, None, None, None],
                [41, None, None, None, None, None, None, None, None, None, None, None],
            ],
        },
    )

    with import_path.open("rb") as handle:
        response = await client.post(
            "/api/v1/registration/certificate-management/workbook/import",
            files={
                "file": (
                    "import-with-placeholders.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["imported_record_count"] == 2

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    workbook_body = workbook_response.json()["data"]
    domestic_registration_rows = next(
        sheet["rows"]
        for sheet in workbook_body["sheets"]
        if sheet["sheet_key"] == "domestic-registration"
    )
    assert len(domestic_registration_rows) == 2
    assert domestic_registration_rows[0]["values"]["证照名称"] == "营业执照"
    assert domestic_registration_rows[1]["values"]["证照名称"] == "兽药生产许可证"


@pytest.mark.asyncio
async def test_import_certificate_workbook_preserves_distinct_rows(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    seed_path = tmp_path / "seed.xlsx"
    build_certificate_workbook_fixture(seed_path, {})
    patch_certificate_workbook_path(monkeypatch, seed_path)

    import_path = tmp_path / "import-international-gmp.xlsx"
    build_certificate_workbook_fixture(
        import_path,
        {
            "国际GMP": [
                [
                    1,
                    "GMP compliance inspection result notification",
                    "A-001",
                    "日本/PMDA",
                    "2023.11.30",
                    "——",
                    "美伐他汀",
                    "内控",
                    1,
                    "客户A",
                ],
                [
                    2,
                    "GMP compliance inspection result notification",
                    "A-002",
                    "日本/PMDA",
                    "2023.10.10",
                    "——",
                    "美伐他汀",
                    "内控",
                    2,
                    "客户B",
                ],
                [
                    3,
                    "GMP compliance inspection result notification",
                    "B-001",
                    "日本/PMDA",
                    "2019.06.17",
                    "——",
                    "洛伐他汀",
                    "USP",
                    1,
                    "客户C",
                ],
                [
                    4,
                    "GMP compliance inspection result notification",
                    "B-002",
                    "日本/PMDA",
                    "2023.11.30",
                    "——",
                    "洛伐他汀",
                    "USP",
                    2,
                    "客户D",
                ],
            ]
        },
    )

    with import_path.open("rb") as handle:
        response = await client.post(
            "/api/v1/registration/certificate-management/workbook/import",
            files={
                "file": (
                    "import-international-gmp.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    assert response.json()["data"]["imported_record_count"] == 4

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    workbook_body = workbook_response.json()["data"]
    international_gmp_rows = next(
        sheet["rows"]
        for sheet in workbook_body["sheets"]
        if sheet["sheet_key"] == "international-gmp"
    )
    assert [row["sequence"] for row in international_gmp_rows] == [1, 2, 3, 4]
    assert [row["values"]["编号"] for row in international_gmp_rows] == [
        "A-001",
        "A-002",
        "B-001",
        "B-002",
    ]


@pytest.mark.asyncio
async def test_import_certificate_workbook_preserves_raw_cell_whitespace(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    seed_path = tmp_path / "seed.xlsx"
    build_certificate_workbook_fixture(seed_path, {})
    patch_certificate_workbook_path(monkeypatch, seed_path)

    import_path = tmp_path / "import-raw-whitespace.xlsx"
    build_certificate_workbook_fixture(
        import_path,
        {
            "国内GMP": [
                [
                    1,
                    "HALAL  Certificate ",
                    "NO-001",
                    "Indonesia/LP-POM",
                    "2022.11.16",
                    "2026.11.15",
                    "L-苯丙氨酸 ",
                    "FCC9",
                    2,
                    " ——",
                ],
            ],
            "国际GMP": [
                [
                    1,
                    "GMP通知",
                    "5130508013279\n[Pravastatin \nsodium]",
                    "日本/PMDA\n",
                    "2023.11.30",
                    "——",
                    "美伐他汀",
                    "内控",
                    1,
                    "备注",
                ],
            ],
        },
    )

    with import_path.open("rb") as handle:
        response = await client.post(
            "/api/v1/registration/certificate-management/workbook/import",
            files={
                "file": (
                    "import-raw-whitespace.xlsx",
                    handle,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/workbook"
    )
    workbook_body = workbook_response.json()["data"]
    domestic_gmp_rows = next(
        sheet["rows"]
        for sheet in workbook_body["sheets"]
        if sheet["sheet_key"] == "domestic-gmp"
    )
    international_gmp_rows = next(
        sheet["rows"]
        for sheet in workbook_body["sheets"]
        if sheet["sheet_key"] == "international-gmp"
    )

    assert domestic_gmp_rows[0]["values"]["证照名称"] == "HALAL  Certificate "
    assert domestic_gmp_rows[0]["values"]["产品范围"] == "L-苯丙氨酸 "
    assert domestic_gmp_rows[0]["values"]["备注"] == " ——"
    assert (
        international_gmp_rows[0]["values"]["编号"]
        == "5130508013279\n[Pravastatin \nsodium]"
    )
    assert international_gmp_rows[0]["values"]["国家/发证机关"] == "日本/PMDA\n"


@pytest.mark.asyncio
async def test_export_certificate_workbook_returns_xlsx(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    template_path = tmp_path / "certificate-template.xlsx"
    build_certificate_export_template_fixture(template_path)
    patch_certificate_workbook_path(monkeypatch, template_path)

    response = await client.get(
        "/api/v1/registration/certificate-management/workbook/export"
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=" in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_create_certificate_entry_auto_increments_sequence(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    await reset_certificate_entries(client)
    workbook_path = tmp_path / "certificate-seed.xlsx"
    build_certificate_workbook_fixture(
        workbook_path,
        {
            "国内注册": [
                [
                    1,
                    "国内证书1",
                    "SLH-1",
                    "PJ-1",
                    "CN-001",
                    "NMPA",
                    "2026.01.01",
                    "2028.01.01",
                    "多拉菌素",
                    "内控",
                    1,
                    "备注1",
                ],
                [
                    2,
                    "国内证书2",
                    "SLH-2",
                    "PJ-2",
                    "CN-002",
                    "NMPA",
                    "2026.02.02",
                    "2028.02.02",
                    "洛伐他汀",
                    "USP",
                    2,
                    "备注2",
                ],
            ],
        },
    )
    patch_certificate_workbook_path(monkeypatch, workbook_path)

    response = await client.post(
        "/api/v1/registration/certificate-management/entries",
        json={
            "sheet_key": "domestic-registration",
            "certificate_name": "新增国内证书",
            "acceptance_number": "SLH-3",
            "approval_number": "PJ-3",
            "certificate_number": "CN-003",
            "issuing_authority": "NMPA",
            "issue_date": "2026.03.03",
            "validity_period": "2028.03.03",
            "product_scope": "霉酚酸",
            "quality_standard": "EP",
            "page_count": 3,
            "remarks": "备注3",
        },
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["source_sequence"] == 3

    workbook_response = await client.get(
        "/api/v1/registration/certificate-management/sheets/domestic-registration"
    )
    assert workbook_response.status_code == 200
    rows = workbook_response.json()["data"]["rows"]
    assert rows[-1]["sequence"] == 3
    assert rows[-1]["values"]["证照名称"] == "新增国内证书"


def test_fill_certificate_sheet_copies_template_style_when_expanding(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国外注册"]
    worksheet.cell(row=3, column=1).value = 1

    rows = [
        {
            "序号": index,
            "证照名称": f"国外证书{index}",
            "证书编号": f"A-{index:03d}",
            "国家/发证机关": "FDA",
            "发证日期": "2026.01.01",
            "有效期/复验期": "2028.01.01",
            "产品范围": "多拉菌素",
            "质量标准": "内控",
            "页数": 2,
            "备注": "备注",
        }
        for index in range(1, 6)
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国外注册"],
        start_row=3,
        template_row=3,
    )

    assert worksheet.cell(row=7, column=2).value == "国外证书5"
    assert copy(worksheet.cell(row=7, column=2)._style) == copy(
        worksheet.cell(row=3, column=2)._style
    )
    assert worksheet.row_dimensions[7].height == worksheet.row_dimensions[3].height

    workbook.close()


def test_fill_certificate_sheet_overrides_placeholder_row_style_with_data_style(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template-with-placeholder-style.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国内注册"]

    worksheet.cell(row=3, column=1).value = 1
    worksheet.cell(row=4, column=1).value = 2
    for column_index in range(1, len(SHEET_HEADERS["国内注册"]) + 1):
        placeholder_cell = worksheet.cell(row=4, column=column_index)
        placeholder_cell.font = Font(name="Arial", size=10, color="FF0000")
        placeholder_cell.alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=False
        )
        placeholder_cell.border = Border()

    rows = [
        {
            "序号": 1,
            "证照名称": "国内证书1",
            "受理号": "SLH-001",
            "批件号": "PJ-001",
            "编号": "CN-001",
            "发证机关": "NMPA",
            "发证日期": "2026.01.01",
            "有效期/复验期": "2028.01.01",
            "产品范围": "多拉菌素",
            "质量标准": "内控",
            "页数": 1,
            "备注": "备注1",
        },
        {
            "序号": 2,
            "证照名称": "国内证书2",
            "受理号": "SLH-002",
            "批件号": "PJ-002",
            "编号": "CN-002",
            "发证机关": "NMPA",
            "发证日期": "2026.02.02",
            "有效期/复验期": "2028.02.02",
            "产品范围": "洛伐他汀",
            "质量标准": "USP",
            "页数": 2,
            "备注": "备注2",
        },
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国内注册"],
        start_row=3,
        template_row=3,
    )

    assert copy(worksheet.cell(row=4, column=2)._style) == copy(
        worksheet.cell(row=3, column=2)._style
    )
    assert (
        worksheet.cell(row=4, column=2).font.color
        != worksheet.cell(row=3, column=2).font.color
        or worksheet.cell(row=3, column=2).font.color is None
    )
    assert (
        worksheet.cell(row=4, column=2).alignment.horizontal
        == worksheet.cell(row=3, column=2).alignment.horizontal
    )

    workbook.close()


def test_fill_certificate_sheet_reuses_blank_template_rows_before_appending(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template-reuse-blank-rows.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国内GMP"]

    for row_index in range(3, 13):
        worksheet.cell(row=row_index, column=1).value = row_index - 2
        worksheet.cell(row=row_index, column=2).value = f"原有证书{row_index - 2}"

    rows = [
        {
            "序号": index,
            "证照名称": f"国内GMP证书{index}",
            "编号": f"GMP-{index:03d}",
            "发证机关": "宁夏回族自治区农业农村厅",
            "发证日期": "2026.01.01",
            "有效期/复验期": "2028.01.01",
            "产品范围": "多拉菌素",
            "质量标准": "内控",
            "页数": 1,
            "备注": "备注",
        }
        for index in range(1, 12)
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国内GMP"],
        start_row=3,
        template_row=3,
    )

    assert worksheet.cell(row=13, column=1).value == 11
    assert worksheet.cell(row=13, column=2).value == "国内GMP证书11"
    assert worksheet.cell(row=17, column=1).value is None

    workbook.close()


def test_fill_certificate_sheet_handles_merged_placeholder_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "template-with-merged-placeholder.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国内GMP"]
    for row_index in range(3, 17):
        worksheet.cell(row=row_index, column=1).value = row_index - 2

    rows = [
        {
            "序号": index,
            "证照名称": f"国内GMP证书{index}",
            "编号": f"GMP-{index:03d}",
            "发证机关": "宁夏回族自治区农业农村厅",
            "发证日期": "2026.01.01",
            "有效期/复验期": "2028.01.01",
            "产品范围": "多拉菌素",
            "质量标准": "内控",
            "页数": 1,
            "备注": "备注",
        }
        for index in range(1, 15)
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国内GMP"],
        start_row=3,
        template_row=3,
    )

    assert "B15:B16" not in {str(item) for item in worksheet.merged_cells.ranges}
    assert "D15:D16" not in {str(item) for item in worksheet.merged_cells.ranges}
    assert worksheet.cell(row=15, column=2).value == "国内GMP证书13"
    assert worksheet.cell(row=16, column=2).value == "国内GMP证书14"

    workbook.close()


def test_fill_certificate_sheet_preserves_unused_template_rows_and_merges(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template-with-structure-rows.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国内注册"]
    worksheet.cell(row=3, column=1).value = 1
    worksheet.cell(row=4, column=1).value = 2

    for row_index in range(50, 67):
        worksheet.cell(row=row_index, column=1).value = row_index - 10
    worksheet.cell(row=67, column=1).value = "序号"

    domestic_gmp_sheet = workbook["国内GMP"]
    original_merges = {str(item) for item in domestic_gmp_sheet.merged_cells.ranges}

    rows = [
        {
            "序号": 1,
            "证照名称": "国内证书1",
            "受理号": "SLH-001",
            "批件号": "PJ-001",
            "编号": "CN-001",
            "发证机关": "NMPA",
            "发证日期": "2026.01.01",
            "有效期/复验期": "2028.01.01",
            "产品范围": "多拉菌素",
            "质量标准": "内控",
            "页数": 1,
            "备注": "备注1",
        },
        {
            "序号": 2,
            "证照名称": "国内证书2",
            "受理号": "SLH-002",
            "批件号": "PJ-002",
            "编号": "CN-002",
            "发证机关": "NMPA",
            "发证日期": "2026.02.02",
            "有效期/复验期": "2028.02.02",
            "产品范围": "洛伐他汀",
            "质量标准": "USP",
            "页数": 2,
            "备注": "备注2",
        },
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国内注册"],
        start_row=3,
        template_row=3,
    )

    certificate_service._fill_certificate_sheet(
        domestic_gmp_sheet,
        [
            {
                "序号": 1,
                "证照名称": "国内GMP证书1",
                "编号": "GMP-001",
                "发证机关": "宁夏回族自治区农业农村厅",
                "发证日期": "2026.01.01",
                "有效期/复验期": "2028.01.01",
                "产品范围": "多拉菌素",
                "质量标准": "内控",
                "页数": 1,
                "备注": "备注1",
            },
            {
                "序号": 2,
                "证照名称": "国内GMP证书2",
                "编号": "GMP-002",
                "发证机关": "宁夏回族自治区农业农村厅",
                "发证日期": "2026.02.02",
                "有效期/复验期": "2028.02.02",
                "产品范围": "霉酚酸",
                "质量标准": "USP",
                "页数": 2,
                "备注": "备注2",
            },
        ],
        headers=SHEET_HEADERS["国内GMP"],
        start_row=3,
        template_row=3,
    )

    assert worksheet.cell(row=50, column=1).value == 40
    assert worksheet.cell(row=67, column=1).value == "序号"
    assert {
        str(item) for item in domestic_gmp_sheet.merged_cells.ranges
    } == original_merges

    workbook.close()


def test_fill_certificate_sheet_preserves_existing_placeholder_values(
    tmp_path: Path,
) -> None:
    template_path = tmp_path / "template-with-placeholders.xlsx"
    build_certificate_export_template_fixture(template_path)
    workbook = load_workbook(template_path)
    worksheet = workbook["国际GMP"]

    worksheet.cell(row=3, column=6).value = "——"
    worksheet.cell(row=3, column=10).value = "——"
    worksheet.cell(row=4, column=3).value = "——"
    worksheet.cell(row=4, column=8).value = "——"

    rows = [
        {
            "序号": 1,
            "证照名称": "国际GMP证书1",
            "编号": "GMP-001",
            "国家/发证机关": "日本/PMDA",
            "发证日期": "2026.01.01",
            "有效期/复验期": None,
            "产品范围": "多拉菌素",
            "质量标准": "USP",
            "页数": 1,
            "备注": None,
        },
        {
            "序号": 2,
            "证照名称": "国际GMP证书2",
            "编号": None,
            "国家/发证机关": "美国/FDA",
            "发证日期": "2026.02.02",
            "有效期/复验期": "——",
            "产品范围": "霉酚酸",
            "质量标准": None,
            "页数": 2,
            "备注": "备注2",
        },
    ]

    certificate_service._fill_certificate_sheet(
        worksheet,
        rows,
        headers=SHEET_HEADERS["国际GMP"],
        start_row=3,
        template_row=3,
    )

    assert worksheet.cell(row=3, column=6).value == "——"
    assert worksheet.cell(row=3, column=10).value == "——"
    assert worksheet.cell(row=4, column=3).value == "——"
    assert worksheet.cell(row=4, column=8).value == "——"

    workbook.close()
