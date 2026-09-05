from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace as _SimpleNamespace
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]

from app.core.exceptions import AppException, NotFoundException
from app.modules.registration.models import (
    RegistrationDeclarationProgressWorkbookVersion,
)
from app.modules.registration.service import declaration_progress as progress

SimpleNamespace: Any = _SimpleNamespace


def _definition(
    *, supports_sub_records: bool = True
) -> progress.DeclarationProgressSheetDefinition:
    template = progress.DeclarationProgressSheetTemplate(
        worksheet_title="测试表",
        sheet_key="test-sheet",
        sheet_name="测试表",
        sheet_title="测试标题",
        column_labels=["序号", "项目名称", "产品", "主字段", "子字段"],
        main_end_label="主字段",
        supports_sub_records=supports_sub_records,
    )
    return progress.DeclarationProgressSheetDefinition(
        worksheet_title=template.worksheet_title,
        sheet_key=template.sheet_key,
        sheet_name=template.sheet_name,
        sheet_title=template.sheet_title,
        supports_sub_records=template.supports_sub_records,
        columns=progress._build_columns(template),
    )


def _version(
    *,
    group_id: Any = None,
    version_number: int = 1,
    sequence: int = 1,
    values: dict[str, str | None] | None = None,
    style_marks: dict[str, str | None] | None = None,
) -> RegistrationDeclarationProgressWorkbookVersion:
    definition = _definition()
    keys = [column.key for column in definition.columns]
    item = RegistrationDeclarationProgressWorkbookVersion(
        record_group_id=group_id or uuid4(),
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        sheet_title=definition.sheet_title,
        source_sequence=sequence,
        version_number=version_number,
        source_row_number=version_number + 4,
        values_data=values
        or dict(zip(keys, [str(sequence), "项目A", "产品A", "主值", "子值"])),
        style_marks=style_marks or {},
        project_name="项目A",
        product_name="产品A",
    )
    item.id = uuid4()
    return item


def _save_full_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for template in progress.DECLARATION_PROGRESS_SHEET_TEMPLATES:
        worksheet = workbook.create_sheet(template.worksheet_title)
        for _ in range(3):
            worksheet.append([])
        worksheet.append(template.column_labels)
        row = [None] * len(template.column_labels)
        row[0] = "1"  # type: ignore[call-overload]
        for index, label in enumerate(template.column_labels):
            if label in {"项目名称", "项目名称 "}:
                row[index] = "项目A"  # type: ignore[call-overload]
            elif label in {"产品", "产品名称", "涉及产品"}:
                row[index] = "产品A"  # type: ignore[call-overload]
            elif index > 0 and row[index] is None:
                row[index] = f"值{index}"  # type: ignore[call-overload]
        worksheet.append(row)
        worksheet.cell(5, 2).font = Font(color="0000FF")
        if template.supports_sub_records:
            child = [None] * len(template.column_labels)
            child[-1] = "子记录更新"  # type: ignore[call-overload]
            worksheet.append(child)
            worksheet.cell(6, len(template.column_labels)).font = Font(color="FF0000")
    workbook.create_sheet("Sheet1")
    workbook.save(path)


def test_progress_text_key_and_sequence_helpers() -> None:
    assert progress._normalize_text(None) == ""
    assert progress._normalize_text(" a\r\n b ") == "a\nb"
    assert progress._normalize_optional_text(None) is None
    assert progress._normalize_optional_text("  ") is None
    assert progress._normalize_optional_text(" x ") == "x"
    assert progress._slugify("Hello World") == "hello_world"
    assert len(progress._slugify("项目")) == 12
    used: set[str] = set()
    assert progress._build_column_key("Name", used) == "name"
    assert progress._build_column_key("Name", used) == "name_2"
    assert progress._extract_sequence_number(None) is None
    assert progress._extract_sequence_number("编号 18") == 18
    assert progress._extract_sequence_number("none") is None


@pytest.mark.parametrize(
    ("color", "expected"),
    [
        (None, None),
        (SimpleNamespace(type="rgb", rgb="FF0000FE"), progress.STYLE_MARK_NEW),
        (SimpleNamespace(type="rgb", rgb="FFFF0000"), progress.STYLE_MARK_UPDATED),
        (SimpleNamespace(type="rgb", rgb="FF1010F0"), progress.STYLE_MARK_NEW),
        (SimpleNamespace(type="rgb", rgb="FFF01010"), progress.STYLE_MARK_UPDATED),
        (SimpleNamespace(type="rgb", rgb="FF202020"), None),
        (SimpleNamespace(type="theme", theme=4, tint=0.0), progress.STYLE_MARK_NEW),
        (SimpleNamespace(type="theme", theme=5, tint=0.0), progress.STYLE_MARK_UPDATED),
        (SimpleNamespace(type="theme", theme=1, tint=0.0), None),
        (SimpleNamespace(type="indexed", indexed=8), None),
    ],
)
def test_style_classification(color: Any, expected: str | None) -> None:
    cell: Any = SimpleNamespace(font=SimpleNamespace(color=color))
    assert progress._classify_style_mark(cell) == expected


def test_font_color_parser_handles_unknown_and_broken_colors() -> None:
    assert (
        progress._get_font_color_key(SimpleNamespace(font=SimpleNamespace(color=None)))
        is None
    )
    unknown: Any = SimpleNamespace(
        font=SimpleNamespace(color=SimpleNamespace(type="auto"))
    )
    assert progress._get_font_color_key(unknown) is None

    class BrokenColor:
        @property
        def type(self: Any) -> Any:
            raise ValueError("broken")

    assert (
        progress._get_font_color_key(
            SimpleNamespace(font=SimpleNamespace(color=BrokenColor()))
        )
        is None
    )


def test_column_building_and_header_matching_variants() -> None:
    with_children = _definition()
    assert [column.is_main for column in with_children.columns] == [
        True,
        True,
        True,
        True,
        False,
    ]
    without_children = _definition(supports_sub_records=False)
    assert all(column.is_main for column in without_children.columns)

    template = progress.DECLARATION_PROGRESS_SHEET_TEMPLATES[0]
    assert progress._header_matches(template.column_labels.copy(), template) is True
    assert progress._header_matches(template.column_labels[:-1], template) is False
    wrong = template.column_labels.copy()
    wrong[0] = "错误"
    assert progress._header_matches(wrong, template) is False

    domestic = next(
        item
        for item in progress.DECLARATION_PROGRESS_SHEET_TEMPLATES
        if item.sheet_key == "domestic-completed"
    )
    formula_headers = domestic.column_labels.copy()
    formula_headers[0] = "='[1]5.2019历史公式"
    assert progress._header_matches(formula_headers, domestic) is True


def test_parse_definitions_find_keys_and_missing_file(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "progress.xlsx"
    path.touch()
    monkeypatch.setattr(progress, "_get_workbook_path", lambda: path)
    definitions, updated_at = progress._parse_workbook_definitions()
    assert len(definitions) == len(progress.DECLARATION_PROGRESS_SHEET_TEMPLATES)
    assert updated_at is not None
    first = progress._find_sheet_definition(definitions[0].sheet_key)
    assert progress._get_project_name_key(first)
    assert progress._get_product_name_key(first)
    with pytest.raises(NotFoundException):
        progress._find_sheet_definition("missing")

    monkeypatch.setattr(
        progress, "_get_workbook_path", lambda: tmp_path / "missing.xlsx"
    )
    # 文件缺失时使用静态模板定义 + None 时间，不再抛 404
    definitions, updated_at = progress._parse_workbook_definitions()
    assert len(definitions) == len(progress.DECLARATION_PROGRESS_SHEET_TEMPLATES)
    assert updated_at is None


def test_load_complete_progress_workbook_with_styles(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "progress.xlsx"
    _save_full_workbook(path)
    monkeypatch.setattr(progress, "_get_workbook_path", lambda: path)
    versions = progress._load_seed_versions()
    assert versions
    assert any(version.version_number == 2 for version in versions)
    assert any(version.style_marks for version in versions)
    assert all(version.source_sequence >= 1 for version in versions)


def test_load_workbook_rejects_missing_unexpected_and_bad_headers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = tmp_path / "source.xlsx"
    _save_full_workbook(source)
    monkeypatch.setattr(progress, "_get_workbook_path", lambda: source)

    missing = tmp_path / "missing.xlsx"
    workbook = Workbook()
    workbook.save(missing)
    with pytest.raises(AppException, match="缺少子表"):
        progress._load_versions_from_workbook_path(missing)

    unexpected = tmp_path / "unexpected.xlsx"
    _save_full_workbook(unexpected)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "额外表"
    workbook.save(tmp_path / "only-extra.xlsx")
    with pytest.raises(AppException, match="缺少子表"):
        progress._load_versions_from_workbook_path(tmp_path / "only-extra.xlsx")

    from openpyxl import load_workbook

    workbook = load_workbook(unexpected)
    workbook.create_sheet("额外表")
    workbook.save(unexpected)
    with pytest.raises(AppException, match="未定义子表"):
        progress._load_versions_from_workbook_path(unexpected)

    bad_header = tmp_path / "bad-header.xlsx"
    _save_full_workbook(bad_header)
    workbook = load_workbook(bad_header)
    first_template = progress.DECLARATION_PROGRESS_SHEET_TEMPLATES[0]
    workbook[first_template.worksheet_title].cell(4, 1, "错误")
    workbook.save(bad_header)
    with pytest.raises(AppException, match="表头不匹配"):
        progress._load_versions_from_workbook_path(bad_header)


def test_load_worksheet_without_subrecords_creates_independent_records() -> None:
    definition = _definition(supports_sub_records=False)
    workbook = Workbook()
    worksheet = workbook.active
    for _ in range(4):
        worksheet.append([])
    worksheet.append([None, "项目A", "产品A", "主值", "子值"])
    worksheet.append(["2", None, None, None, None])
    versions = progress._load_versions_from_worksheet(definition, worksheet)
    assert len(versions) == 1
    assert versions[0].source_sequence == 1


def test_record_detail_summary_and_response_builders() -> None:
    definition = _definition()
    group_id = uuid4()
    first = _version(group_id=group_id)
    keys = [column.key for column in definition.columns]
    second = _version(
        group_id=group_id,
        version_number=2,
        values=dict(zip(keys, ["1", "项目A", "产品A", "主值", "更新"])),
        style_marks={keys[-1]: progress.STYLE_MARK_UPDATED},
    )
    single = _version(sequence=2)
    assert progress._build_history_record(first).version == 1
    assert progress._build_record([first, second]).history_count == 2
    assert (
        progress._build_record([single], include_history_records=False).history_records
        == []
    )
    summary = progress._build_sheet_summary(
        definition,
        [progress._build_record([first, second]), progress._build_record([single])],
    )
    assert summary.records_with_history == 1
    assert summary.child_column_count == 1
    detail = progress._build_sheet_detail(definition, [second, single, first])
    assert [item.sequence for item in detail.records] == [1, 2]
    compact = progress._build_sheet_detail(
        definition,
        [first, second],
        include_columns=False,
        include_records=False,
        include_history_records=False,
    )
    assert compact.records == []
    assert compact.summary.total_records == 1
    assert (
        progress._normalize_entry_values(definition, {keys[1]: " x "})[keys[1]] == "x"
    )
    assert progress._build_entry_response(second).style_marks


def test_export_rows_and_fill_sheet_preserve_child_changes_and_styles() -> None:
    definition = _definition()
    keys = [column.key for column in definition.columns]
    group_id = uuid4()
    first = _version(
        group_id=group_id,
        style_marks={keys[1]: progress.STYLE_MARK_NEW},
    )
    second = _version(
        group_id=group_id,
        version_number=2,
        values=dict(zip(keys, ["1", "项目A", "产品A", "主值", "更新"])),
        style_marks={keys[-1]: progress.STYLE_MARK_UPDATED},
    )
    rows = progress._build_export_rows(definition, [first, second])
    assert rows[0].values[keys[0]] == "1"
    assert rows[1].values[keys[1]] is None
    assert rows[1].values[keys[-1]] == "更新"
    assert rows[1].style_marks[keys[-1]] == progress.STYLE_MARK_UPDATED

    workbook = Workbook()
    worksheet = workbook.active
    for _ in range(3):
        worksheet.append([])
    worksheet.append([column.label for column in definition.columns])
    worksheet.append(["old"] * len(definition.columns))
    worksheet.merge_cells("A5:A6")
    progress._fill_declaration_progress_sheet(worksheet, definition, rows)
    assert worksheet.cell(5, 2).value == "项目A"
    assert worksheet.cell(6, len(keys)).value == "更新"
    assert "A5:A6" in {str(item) for item in worksheet.merged_cells.ranges}
    progress._apply_declaration_progress_vertical_merges(
        worksheet, definition, [], start_row=5
    )
    no_children = _definition(supports_sub_records=False)
    progress._apply_declaration_progress_vertical_merges(
        worksheet, no_children, rows, start_row=5
    )


def test_export_rows_without_subrecords_keeps_each_snapshot() -> None:
    definition = _definition(supports_sub_records=False)
    item = _version()
    rows = progress._build_export_rows(definition, [item])
    assert rows[0].values


def test_seed_versions_skipped_when_file_missing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """文件缺失时种子加载返回空列表，页面展示空态而非 404。"""
    monkeypatch.setattr(
        progress, "_get_workbook_path", lambda: tmp_path / "missing.xlsx"
    )
    assert progress._load_seed_versions() == []


def test_load_versions_from_upload_without_config_file(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """配置文件缺失时，导入仍可用静态模板定义解析上传文件并建档。"""
    monkeypatch.setattr(
        progress, "_get_workbook_path", lambda: tmp_path / "missing.xlsx"
    )
    upload_path = tmp_path / "upload.xlsx"
    _save_full_workbook(upload_path)
    versions = progress._load_versions_from_workbook_path(upload_path)
    assert versions
    assert any(version.version_number == 2 for version in versions)
