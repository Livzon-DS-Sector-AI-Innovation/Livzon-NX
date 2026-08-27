from __future__ import annotations

from pathlib import Path

import pytest
from fastapi import UploadFile
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import update
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.registration.models import RegistrationProjectLedgerVersion
from app.modules.registration.schemas.project_ledger import ProjectLedgerEntryInput
from app.modules.registration.service import project_ledger as project_ledger_service

PROJECT_LEDGER_HEADERS = [
    "序号",
    "项目名称",
    "产品",
    "药政活动",
    "",
]


def build_project_ledger_workbook_fixture(
    workbook_path: Path,
    rows_by_sheet: dict[str, list[list[str | int | None]]],
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheets = [
        "1.国际注册（关联审评机制）",
        "2.国际注册（原料药单独审评机制）",
        "3.国内注册（关联审评机制）",
        "4.国内注册（原料药单独审评机制）",
    ]

    for sheet_name in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.cell(row=1, column=1).value = sheet_name
        for column_index, header in enumerate(PROJECT_LEDGER_HEADERS, start=1):
            worksheet.cell(row=2, column=column_index).value = header

        for row_offset, row_values in enumerate(
            rows_by_sheet.get(sheet_name, []), start=3
        ):
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_offset, column=column_index).value = value

    workbook.save(workbook_path)
    workbook.close()


def build_project_ledger_workbook_fixture_with_merged_template_row(
    workbook_path: Path,
    rows_by_sheet: dict[str, list[list[str | int | None]]],
) -> None:
    build_project_ledger_workbook_fixture(workbook_path, rows_by_sheet)
    workbook = load_workbook(workbook_path)
    try:
        worksheet = workbook["1.国际注册（关联审评机制）"]
        worksheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=3)
        workbook.save(workbook_path)
    finally:
        workbook.close()


def patch_project_ledger_workbook_path(
    monkeypatch: pytest.MonkeyPatch,
    workbook_path: Path,
) -> None:
    monkeypatch.setattr(
        project_ledger_service, "_get_workbook_path", lambda: workbook_path
    )


async def reset_project_ledger_entries(db_session: AsyncSession) -> None:
    """仅供 service 级测试（同一 db_session 会话内读写）使用。"""
    await db_session.execute(
        update(RegistrationProjectLedgerVersion).values(is_deleted=True)
    )
    await db_session.commit()


PROJECT_LEDGER_SHEET_KEYS = [
    "international-associated-review",
    "international-standalone-review",
    "domestic-associated-review",
    "domestic-standalone-review",
]


async def reset_project_ledger_entries_via_api(client: AsyncClient) -> None:
    """通过 API 软删除现有台账记录，供 client 路径的测试使用。

    必须与后续写操作共用同一 client 会话：client 与 db_session fixture 是两个
    独立连接，跨连接批量 UPDATE 会持有行锁导致服务端写入阻塞（死锁）。
    """
    record_ids: list[str] = []
    for sheet_key in PROJECT_LEDGER_SHEET_KEYS:
        response = await client.get(
            f"/api/v1/registration/project-ledger/sheets/{sheet_key}"
        )
        if response.status_code == 404:
            continue
        assert response.status_code == 200
        record_ids.extend(
            record["record_id"] for record in response.json()["data"]["records"]
        )

    # Collect all groups before deleting: after the last group is removed the
    # next read may seed the patched workbook again, making the reset partial.
    for record_id in record_ids:
        delete_response = await client.delete(
            f"/api/v1/registration/project-ledger/entries/{record_id}"
        )
        assert delete_response.status_code == 200


def get_column_key_by_label(columns: list[dict[str, str]], label: str) -> str:
    for column in columns:
        if isinstance(column, dict):
            if column["label"] == label:
                return column["key"]
            continue
        if getattr(column, "label", None) == label:
            return getattr(column, "key")
    raise AssertionError(f"未找到列: {label}")


@pytest.mark.asyncio
async def test_get_project_ledger_workbook_returns_named_sheets(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-workbook.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
            ],
            "3.国内注册（关联审评机制）": [
                [2, "国内项目B", "洛伐他汀", "审评", "进行中"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    response = await client.get("/api/v1/registration/project-ledger/workbook")

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["workbook_name"] == "1. 注册台账.xlsx"
    assert len(payload["sheets"]) == 4


@pytest.mark.asyncio
async def test_export_project_ledger_workbook_endpoint_returns_xlsx_file(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-export-endpoint.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    response = await client.get("/api/v1/registration/project-ledger/workbook/export")

    assert response.status_code == 200
    assert (
        response.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        "1.%20%E6%B3%A8%E5%86%8C%E5%8F%B0%E8%B4%A6.xlsx"
        in response.headers["content-disposition"]
    )


@pytest.mark.asyncio
async def test_get_project_ledger_overview_returns_all_named_sheets(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-seed.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
            "3.国内注册（关联审评机制）": [
                [2, "国内项目B", "洛伐他汀", "受理", "已完成"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    response = await client.get("/api/v1/registration/project-ledger/overview")

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["workbook_name"] == "1. 注册台账.xlsx"
    assert [sheet["sheet_key"] for sheet in body["sheets"]] == [
        "international-associated-review",
        "international-standalone-review",
        "domestic-associated-review",
        "domestic-standalone-review",
    ]
    assert body["total_records"] == 2
    assert body["records_with_history"] == 1
    first_sheet = body["sheets"][0]
    assert first_sheet["sheet_name"] == "国际注册（关联审评机制）"
    assert first_sheet["summary"]["total_records"] == 1
    assert first_sheet["summary"]["records_with_history"] == 1

    detail_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )
    assert detail_response.status_code == 200
    detail = detail_response.json()["data"]
    project_name_key = get_column_key_by_label(detail["columns"], "项目名称")
    assert detail["records"][0]["latest_values"][project_name_key] == "国际项目A"


@pytest.mark.asyncio
async def test_get_project_ledger_sheet_detail_uses_new_endpoint(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-detail.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["sheet_key"] == "international-associated-review"
    assert payload["summary"]["total_records"] == 1
    assert payload["records"][0]["history_count"] == 2
    activity_note_key = get_column_key_by_label(payload["columns"], "药政活动说明")
    assert payload["records"][0]["latest_values"][activity_note_key] == "已反馈"


@pytest.mark.asyncio
async def test_create_project_ledger_entry_auto_increments_sequence(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-create.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "3.国内注册（关联审评机制）": [
                [1, "国内项目A", "多拉菌素", "受理", "已提交"],
                [2, "国内项目B", "洛伐他汀", "审评", "进行中"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    before_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/domestic-associated-review"
    )
    assert before_response.status_code == 200
    before_payload = before_response.json()["data"]
    project_name_key = get_column_key_by_label(before_payload["columns"], "项目名称")
    product_key = get_column_key_by_label(before_payload["columns"], "产品")
    activity_key = get_column_key_by_label(before_payload["columns"], "药政活动")
    activity_note_key = get_column_key_by_label(
        before_payload["columns"], "药政活动说明"
    )

    response = await client.post(
        "/api/v1/registration/project-ledger/entries",
        json={
            "sheet_key": "domestic-associated-review",
            "values": {
                project_name_key: "国内项目C",
                product_key: "霉酚酸",
                activity_key: "补件",
                activity_note_key: "等待回执",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["sheet_key"] == "domestic-associated-review"
    assert body["sequence"] == 3

    sheet_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/domestic-associated-review"
    )
    assert sheet_response.status_code == 200
    sheet_payload = sheet_response.json()["data"]
    records = sheet_payload["records"]
    assert records[-1]["sequence"] == 3
    assert records[-1]["latest_values"][project_name_key] == "国内项目C"


@pytest.mark.asyncio
async def test_update_project_ledger_entry_updates_latest_snapshot_in_place(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-update-main.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    before_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )
    assert before_response.status_code == 200
    before_payload = before_response.json()["data"]
    project_name_key = get_column_key_by_label(before_payload["columns"], "项目名称")
    product_key = get_column_key_by_label(before_payload["columns"], "产品")
    activity_key = get_column_key_by_label(before_payload["columns"], "药政活动")
    activity_note_key = get_column_key_by_label(
        before_payload["columns"], "药政活动说明"
    )
    record_id = before_payload["records"][0]["record_id"]

    response = await client.put(
        f"/api/v1/registration/project-ledger/entries/{record_id}",
        json={
            "sheet_key": "international-associated-review",
            "values": {
                project_name_key: "国际项目A-修订",
                product_key: "多拉菌素",
                activity_key: "补件",
                activity_note_key: "主记录已修订",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["version_number"] == 2
    assert body["values"][project_name_key] == "国际项目A-修订"

    after_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )
    assert after_response.status_code == 200
    after_payload = after_response.json()["data"]
    record = after_payload["records"][0]
    assert record["history_count"] == 2
    assert record["latest_values"][project_name_key] == "国际项目A-修订"

    history_response = await client.get(
        f"/api/v1/registration/project-ledger/entries/{record_id}/history"
    )
    assert history_response.status_code == 200
    history_payload = history_response.json()["data"]
    assert (
        history_payload["history_records"][-1]["values"][activity_note_key]
        == "主记录已修订"
    )


@pytest.mark.asyncio
async def test_create_project_ledger_sub_record_appends_history_and_keeps_latest(
    client: AsyncClient,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "project-ledger-add-sub-record.xlsx"
    build_project_ledger_workbook_fixture(
        workbook_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, workbook_path)
    await reset_project_ledger_entries_via_api(client)

    before_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )
    assert before_response.status_code == 200
    before_payload = before_response.json()["data"]
    project_name_key = get_column_key_by_label(before_payload["columns"], "项目名称")
    product_key = get_column_key_by_label(before_payload["columns"], "产品")
    activity_key = get_column_key_by_label(before_payload["columns"], "药政活动")
    activity_note_key = get_column_key_by_label(
        before_payload["columns"], "药政活动说明"
    )
    record_id = before_payload["records"][0]["record_id"]

    response = await client.post(
        f"/api/v1/registration/project-ledger/entries/{record_id}/sub-records",
        json={
            "sheet_key": "international-associated-review",
            "values": {
                project_name_key: "国际项目A",
                product_key: "多拉菌素",
                activity_key: "获批",
                activity_note_key: "子记录新增后以最新一条为准",
            },
        },
    )

    assert response.status_code == 200
    body = response.json()["data"]
    assert body["version_number"] == 3
    assert body["values"][activity_key] == "获批"
    assert body["values"][activity_note_key] == "子记录新增后以最新一条为准"

    after_response = await client.get(
        "/api/v1/registration/project-ledger/sheets/international-associated-review"
    )
    assert after_response.status_code == 200
    after_payload = after_response.json()["data"]
    record = after_payload["records"][0]
    assert record["history_count"] == 3
    assert record["latest_values"][activity_key] == "获批"
    assert record["latest_values"][activity_note_key] == "子记录新增后以最新一条为准"

    history_response = await client.get(
        f"/api/v1/registration/project-ledger/entries/{record_id}/history"
    )
    assert history_response.status_code == 200
    history_payload = history_response.json()["data"]
    assert history_payload["history_records"][-1]["values"][activity_key] == "获批"


@pytest.mark.asyncio
async def test_import_project_ledger_workbook_replaces_all_sheets(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-template.xlsx"
    build_project_ledger_workbook_fixture(
        template_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    import_path = tmp_path / "project-ledger-import.xlsx"
    build_project_ledger_workbook_fixture(
        import_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目C", "霉酚酸", "立项", "已提交"],
            ],
            "4.国内注册（原料药单独审评机制）": [
                [1, "国内项目D", "盐酸林可霉素", "受理", "已完成"],
            ],
        },
    )

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    with import_path.open("rb") as file_obj:
        result = await service.import_workbook(
            UploadFile(
                file=file_obj,
                filename="1. 注册台账.xlsx",
            )
        )

    assert result.workbook_name == "1. 注册台账.xlsx"
    assert result.imported_records == 2
    assert result.sheet_record_counts == {
        "international-associated-review": 1,
        "domestic-standalone-review": 1,
    }

    overview = await service.get_overview()
    assert overview.total_records == 2
    assert overview.records_with_history == 0
    first_sheet = next(
        sheet
        for sheet in overview.sheets
        if sheet.sheet_key == "international-associated-review"
    )
    fourth_sheet = next(
        sheet
        for sheet in overview.sheets
        if sheet.sheet_key == "domestic-standalone-review"
    )
    assert first_sheet.summary.total_records == 1
    assert fourth_sheet.summary.total_records == 1

    first_sheet_detail = await service.get_sheet_detail(
        "international-associated-review"
    )
    project_name_key = get_column_key_by_label(first_sheet_detail.columns, "项目名称")
    assert first_sheet_detail.records[0].latest_values[project_name_key] == "国际项目C"


@pytest.mark.asyncio
async def test_import_project_ledger_workbook_ignores_sequence_only_placeholder_rows(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-import-placeholder-template.xlsx"
    build_project_ledger_workbook_fixture(template_path, {})
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    import_path = tmp_path / "project-ledger-import-placeholder.xlsx"
    build_project_ledger_workbook_fixture(
        import_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [2, None, None, None, None],
            ],
        },
    )

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    with import_path.open("rb") as file_obj:
        result = await service.import_workbook(
            UploadFile(
                file=file_obj,
                filename="1. 注册台账.xlsx",
            )
        )

    assert result.imported_records == 1
    assert result.sheet_record_counts == {"international-associated-review": 1}

    overview = await service.get_overview()
    first_sheet = next(
        sheet
        for sheet in overview.sheets
        if sheet.sheet_key == "international-associated-review"
    )
    assert first_sheet.summary.total_records == 1


@pytest.mark.asyncio
async def test_export_project_ledger_workbook_keeps_named_sheets(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-export-template.xlsx"
    build_project_ledger_workbook_fixture(
        template_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
            "2.国际注册（原料药单独审评机制）": [
                [1, "国际项目B", "盐酸强力霉素", "补件", "进行中"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    export_path, download_name = await service.export_workbook()

    assert download_name == "1. 注册台账.xlsx"

    workbook = load_workbook(export_path, data_only=True)
    try:
        assert workbook.sheetnames == [
            "1.国际注册（关联审评机制）",
            "2.国际注册（原料药单独审评机制）",
            "3.国内注册（关联审评机制）",
            "4.国内注册（原料药单独审评机制）",
        ]

        international_sheet = workbook["1.国际注册（关联审评机制）"]
        standalone_sheet = workbook["2.国际注册（原料药单独审评机制）"]

        assert international_sheet.cell(row=3, column=1).value == "1"
        assert international_sheet.cell(row=3, column=2).value == "国际项目A"
        assert international_sheet.cell(row=4, column=1).value in (None, "")
        assert international_sheet.cell(row=4, column=4).value == "补件"
        assert international_sheet.cell(row=4, column=5).value == "已反馈"
        assert standalone_sheet.cell(row=3, column=2).value == "国际项目B"
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_project_ledger_workbook_handles_merged_template_rows(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-export-merged-template.xlsx"
    build_project_ledger_workbook_fixture_with_merged_template_row(
        template_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "已提交"],
                [None, None, None, "补件", "已反馈"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    export_path, download_name = await service.export_workbook()

    assert download_name == "1. 注册台账.xlsx"
    workbook = load_workbook(export_path, data_only=True)
    try:
        worksheet = workbook["1.国际注册（关联审评机制）"]
        assert worksheet.cell(row=3, column=2).value == "国际项目A"
        assert worksheet.cell(row=4, column=4).value == "补件"
    finally:
        workbook.close()

    workbook_with_merges = load_workbook(export_path)
    try:
        worksheet = workbook_with_merges["1.国际注册（关联审评机制）"]
        merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
        assert "B5:C5" in merged_ranges
    finally:
        workbook_with_merges.close()


@pytest.mark.asyncio
async def test_export_project_ledger_workbook_preserves_existing_placeholder_values(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-export-placeholder-template.xlsx"
    build_project_ledger_workbook_fixture(
        template_path,
        {
            "1.国际注册（关联审评机制）": [
                [1, "国际项目A", "多拉菌素", "立项", "__"],
                [None, None, None, "补件", "__"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    export_path, _ = await service.export_workbook()

    workbook = load_workbook(export_path, data_only=True)
    try:
        worksheet = workbook["1.国际注册（关联审评机制）"]
        assert worksheet.cell(row=4, column=5).value in (None, "")
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_export_project_ledger_workbook_keeps_sub_record_rows_incremental(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    # 全程使用同一 db_session 会话（service 级）：跨连接的 client 写入对导出不可见
    await reset_project_ledger_entries(db_session)
    template_path = tmp_path / "project-ledger-export-sub-record-template.xlsx"
    build_project_ledger_workbook_fixture(
        template_path,
        {
            "4.国内注册（原料药单独审评机制）": [
                [1, "兽药批文", "多拉菌素", "首次递交", "2021.09.09"],
                [2, "饲料添加剂批文", "苯丙氨酸", "首次递交", "2023.12.07"],
            ],
        },
    )
    patch_project_ledger_workbook_path(monkeypatch, template_path)

    service = project_ledger_service.ProjectLedgerWorkbookService(db_session)
    detail = await service.get_sheet_detail("domestic-standalone-review")
    activity_key = get_column_key_by_label(detail.columns, "药政活动")
    note_key = get_column_key_by_label(detail.columns, "药政活动说明")
    record_id = detail.records[0].record_id

    await service.create_sub_record(
        record_id,
        ProjectLedgerEntryInput(
            sheet_key="domestic-standalone-review",
            values={
                activity_key: "换发申请",
                note_key: "2022.10.20",
            },
        ),
    )

    export_path, _ = await service.export_workbook()

    workbook = load_workbook(export_path, data_only=True)
    try:
        worksheet = workbook["4.国内注册（原料药单独审评机制）"]
        assert worksheet.cell(row=3, column=1).value == "1"
        assert worksheet.cell(row=3, column=2).value == "兽药批文"
        assert worksheet.cell(row=3, column=3).value == "多拉菌素"
        assert worksheet.cell(row=3, column=4).value == "首次递交"
        assert worksheet.cell(row=3, column=5).value == "2021.09.09"

        assert worksheet.cell(row=4, column=1).value in (None, "")
        assert worksheet.cell(row=4, column=2).value in (None, "")
        assert worksheet.cell(row=4, column=3).value in (None, "")
        assert worksheet.cell(row=4, column=4).value == "换发申请"
        assert worksheet.cell(row=4, column=5).value == "2022.10.20"

        assert worksheet.cell(row=5, column=1).value == "2"
        assert worksheet.cell(row=5, column=2).value == "饲料添加剂批文"
        assert worksheet.cell(row=5, column=3).value == "苯丙氨酸"
    finally:
        workbook.close()

    workbook_with_merges = load_workbook(export_path)
    try:
        worksheet = workbook_with_merges["4.国内注册（原料药单独审评机制）"]
        merged_ranges = {str(item) for item in worksheet.merged_cells.ranges}
        assert "A3:A4" in merged_ranges
        assert "B3:B4" in merged_ranges
        assert "C3:C4" in merged_ranges
    finally:
        workbook_with_merges.close()
