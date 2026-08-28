from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.modules.registration.models import RegistrationCertificateEntry
from app.modules.registration.service import certificate
from app.modules.registration.service.certificate import CertificateWorkbookService


def _entry(**overrides: object) -> RegistrationCertificateEntry:
    values: dict[str, object] = {
        "id": uuid4(),
        "sheet_key": "international-registration",
        "sheet_name": "国外注册",
        "sheet_title": "国外注册证书",
        "source_sequence": 1,
        "certificate_name": "证书A",
        "acceptance_number": "受理号",
        "approval_number": "批件号",
        "certificate_number": "CERT-1",
        "issuing_authority": "国家药监局",
        "issue_date": "2026.01.01",
        "validity_period": "2026.01.01-2026.12.31",
        "expiry_date": "2026.12.31",
        "product_scope": "产品A\n产品B",
        "quality_standard": "标准",
        "page_count": 3,
        "remarks": "备注",
        "created_at": datetime(2026, 1, 1),
        "updated_at": datetime(2026, 1, 2),
    }
    values.update(overrides)
    return RegistrationCertificateEntry(**values)


def test_certificate_normalization_mapping_and_summary_helpers() -> None:
    assert certificate._normalize_text(" a\r\n b ") == "a\nb"
    assert certificate._normalize_text(date(2026, 1, 2)) == "2026.01.02"
    assert certificate._preserve_text(datetime(2026, 1, 2)) == "2026.01.02"
    assert certificate._slugify("Certificate Number") == "certificate_number"
    assert len(certificate._slugify("证书编号")) == 12
    assert certificate._normalize_optional(" - ") is None
    assert certificate._normalize_optional(" 有效 ") == "有效"
    assert certificate._preserve_optional(" ") == " "
    assert certificate._is_template_placeholder_value("——")
    assert not certificate._is_template_placeholder_value(None)
    assert certificate._extract_date("有效期：2026/8/26") == date(2026, 8, 26)
    assert certificate._extract_date("长期") is None
    assert certificate._extract_date("bad") is None
    assert certificate._format_date(date(2026, 8, 26)) == "2026.08.26"
    assert certificate._format_date(None) is None
    assert certificate._calculate_expiry_status("2000.01.01") == "已过期"
    assert (
        certificate._calculate_expiry_status(
            (date.today() + timedelta(days=30)).strftime("%Y.%m.%d")
        )
        == "90天内到期"
    )
    assert (
        certificate._calculate_expiry_status(
            (date.today() + timedelta(days=120)).strftime("%Y.%m.%d")
        )
        == "有效"
    )
    assert certificate._calculate_expiry_status("长期") == "无明确期限"
    assert certificate._parse_page_count("共 12 页") == 12
    assert certificate._parse_page_count("无") is None
    assert certificate._split_lines("产品A\n-\n产品B") == ["产品A", "产品B"]
    assert certificate._normalize_group_text(" A\n B ") == "a b"
    assert certificate._normalize_department_text(" qa 部 ") == "QA部"
    assert certificate._is_qa_department("QA部")
    assert certificate._is_qa_department("质量保证部")
    assert not certificate._is_qa_department("生产部")
    assert certificate._extract_sequence(" 12 ") == 12
    assert certificate._extract_sequence("bad") is None

    payload = certificate._map_values_to_entry_payload(
        "domestic-registration",
        "国内注册",
        "标题",
        {
            "证照名称": "证书",
            "编号": "NO-1",
            "发证机关": "机关",
            "有效期/复验期": "2026.01.01-2026.12.31",
            "页数": "20页",
        },
        source_sequence=2,
    )
    assert payload["certificate_number"] == "NO-1"
    assert payload["page_count"] == 20
    assert payload["expiry_date"] == "2026.12.31"

    entry = _entry()
    assert certificate._build_row_values(entry)["证书编号"] == "CERT-1"
    for sheet_key in ("domestic-registration", "domestic-gmp", "international-gmp"):
        assert certificate._build_row_values(_entry(sheet_key=sheet_key))
    assert certificate._build_sheet_row(entry).expiry_status
    assert certificate._build_record_summary(entry).certificate_name == "证书A"
    summary = certificate._build_sheet_summary("international-registration", [entry])
    assert summary.total_records == 1
    assert summary.product_count == 2
    assert certificate._build_reminder_content([entry] * 11, 90).count("证书A") == 10
    assert (
        certificate._count_due_entries_for_setting(
            [
                _entry(
                    expiry_date=(date.today() + timedelta(days=1)).strftime("%Y.%m.%d")
                )
            ],
            90,
        )
        == 1
    )


def _certificate_workbook() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for meta in certificate.CERTIFICATE_SHEET_CONFIG:
        sheet = workbook.create_sheet(str(meta["name"]))
        headers = certificate._get_sheet_headers(meta)
        sheet.cell(1, 1).value = str(meta["name"])
        for index, header in enumerate(headers, start=1):
            sheet.cell(2, index).value = header
        for index, header in enumerate(headers, start=1):
            sheet.cell(3, index).value = (
                1 if index == 1 else "证书" if header == "证照名称" else "2026.01.01"
            )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_certificate_workbook_parser_and_sheet_writer() -> None:
    entries = certificate._load_entries_from_workbook_bytes(_certificate_workbook())
    assert len(entries) == len(certificate.CERTIFICATE_SHEET_CONFIG)
    assert entries[0].source_sequence == 1
    incomplete = load_workbook(BytesIO(_certificate_workbook()))
    incomplete.remove(incomplete["国际GMP"])
    incomplete_stream = BytesIO()
    incomplete.save(incomplete_stream)
    with pytest.raises(Exception, match="缺少子表"):
        certificate._load_entries_from_workbook_bytes(incomplete_stream.getvalue())

    workbook = Workbook()
    sheet = workbook.active
    headers = certificate._get_sheet_headers(certificate.CERTIFICATE_SHEET_CONFIG[0])
    for index, header in enumerate(headers, start=1):
        sheet.cell(3, index).value = header
        sheet.cell(4, index).value = 1 if index == 1 else "旧模板"
    certificate._fill_certificate_sheet(
        sheet,
        [{"序号": 1, "证照名称": "新证书"}, {"序号": 2, "证照名称": "第二证书"}],
        headers=headers,
        start_row=3,
        template_row=4,
    )
    assert sheet.cell(4, 2).value == "新证书"
    assert sheet.cell(5, 2).value == "第二证书"


@pytest.mark.asyncio
async def test_certificate_reminder_batches_filter_and_send() -> None:
    due = _entry(expiry_date=(date.today() + timedelta(days=2)).strftime("%Y.%m.%d"))
    expired = _entry(expiry_date="2000.01.01")
    repo = SimpleNamespace(
        count_entries=AsyncMock(return_value=2),
        get_reminder_setting=AsyncMock(
            return_value=SimpleNamespace(
                is_enabled=True,
                reminder_days=90,
                recipient_open_id="open-1",
                recipient_name="QA",
                recipient_department="QA",
            )
        ),
        list_entries=AsyncMock(return_value=[due, expired]),
        reminder_notification_exists=AsyncMock(return_value=False),
        create_reminder_notifications=AsyncMock(),
    )
    service = CertificateWorkbookService(SimpleNamespace(commit=AsyncMock()))
    service.repository = repo
    service._get_qa_reminder_recipient_by_open_id = AsyncMock(
        return_value=SimpleNamespace(
            open_id="open-1",
            name="QA",
            department="QA",
            enterprise_email="qa@example.test",
        )
    )
    batches = await service.find_due_reminder_batch()
    assert batches[0]["recipient_receive_id"] == "qa@example.test"
    assert batches[0]["entries"] == [due]

    from app.modules.registration.service import certificate as certificate_module

    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(
        certificate_module, "send_user_card", AsyncMock(return_value=True)
    )
    try:
        await service.send_due_reminder_batch(batches[0])
    finally:
        monkeypatch.undo()
    repo.create_reminder_notifications.assert_awaited_once()

    repo.get_reminder_setting.return_value = SimpleNamespace(
        is_enabled=False,
        reminder_days=90,
        recipient_open_id=None,
        recipient_name=None,
        recipient_department=None,
    )
    assert await service.find_due_reminder_batch() == []


@pytest.mark.asyncio
async def test_certificate_service_dashboard_crud_import_and_export_paths(
    monkeypatch: pytest.MonkeyPatch, tmp_path: object
) -> None:
    workbook_path = tmp_path / "certificates.xlsx"  # type: ignore[union-attr]
    workbook_path.write_bytes(_certificate_workbook())
    monkeypatch.setattr(
        certificate, "_get_certificate_workbook_path", lambda: workbook_path
    )

    entry = _entry()
    session = SimpleNamespace(
        commit=AsyncMock(), rollback=AsyncMock(), flush=AsyncMock()
    )
    repo = SimpleNamespace(
        count_entries=AsyncMock(return_value=1),
        list_entries=AsyncMock(return_value=[entry]),
        get_next_source_sequence=AsyncMock(return_value=8),
        create_entry=AsyncMock(return_value=entry),
        get_by_id=AsyncMock(return_value=entry),
        update_entry=AsyncMock(return_value=entry),
        soft_delete=AsyncMock(),
        soft_delete_many=AsyncMock(return_value=1),
        create_entries=AsyncMock(return_value=[entry]),
    )
    service = CertificateWorkbookService(session)
    service.repository = repo

    overview = await service.get_overview()
    detail = await service.get_workbook_detail()
    sheet = await service.get_sheet_detail("international-registration")
    assert overview.total_records == 1
    assert len(detail.sheets) == len(certificate.CERTIFICATE_SHEET_CONFIG)
    assert sheet.rows[0].values["证照名称"] == "证书A"

    created = await service.create_entry(
        certificate.CertificateEntryCreate(
            sheet_key="international-registration",
            certificate_name="新增证书",
            validity_period="2026.01.01-2027.01.01",
        )
    )
    assert created.certificate_name == "证书A"
    updated = await service.update_entry(
        entry.id,
        certificate.CertificateEntryUpdate(
            certificate_name="更新证书", validity_period="2027.12.31"
        ),
    )
    assert updated.certificate_name == "证书A"
    await service.delete_entry(entry.id)
    repo.get_by_id.side_effect = [None, None]
    with pytest.raises(Exception, match="证书台账记录"):
        await service.update_entry(entry.id, certificate.CertificateEntryUpdate())
    with pytest.raises(Exception, match="证书台账记录"):
        await service.delete_entry(entry.id)

    upload = SimpleNamespace(filename="import.xlsx")
    monkeypatch.setattr(
        certificate,
        "read_upload_secure",
        AsyncMock(return_value=("import.xlsx", _certificate_workbook())),
    )
    imported = await service.import_workbook(upload)
    assert imported.replaced_record_count == 1
    repo.list_entries.return_value = [entry]
    exported_path, exported_name = await service.export_workbook()
    assert exported_path.exists()
    assert exported_name.endswith(".xlsx")

    monkeypatch.setattr(
        certificate, "_get_certificate_workbook_path", lambda: tmp_path / "missing.xlsx"
    )  # type: ignore[union-attr]
    with pytest.raises(Exception, match="药政证书台账文件"):
        await service.export_workbook()


@pytest.mark.asyncio
async def test_certificate_reminder_settings_and_recipient_fallbacks(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session = SimpleNamespace(
        commit=AsyncMock(), rollback=AsyncMock(), flush=AsyncMock()
    )
    repo = SimpleNamespace(
        count_entries=AsyncMock(return_value=1),
        get_reminder_setting=AsyncMock(return_value=None),
        save_reminder_setting=AsyncMock(),
        list_entries=AsyncMock(return_value=[]),
        reminder_notification_exists=AsyncMock(return_value=True),
    )
    service = CertificateWorkbookService(session)
    service.repository = repo

    from app.modules.quality import public_api

    contacts = AsyncMock(
        return_value={
            "items": [
                {"open_id": "", "name": "空", "department": "QA"},
                {"open_id": "prod", "name": "生产", "department": "生产部"},
                {"open_id": "qa", "name": "QA", "department": "质量保证部"},
                {"open_id": "qa", "name": "QA2", "department": "QA部"},
            ]
        }
    )
    monkeypatch.setattr(public_api, "get_department_contact_list_from_feishu", contacts)
    options = await service.list_reminder_recipient_options()
    assert [item.open_id for item in options] == ["qa"]
    contacts.side_effect = RuntimeError("feishu down")
    assert await service.list_reminder_recipient_options() == []

    with pytest.raises(Exception, match="必须选择通知人"):
        await service.update_reminder_settings(
            certificate.CertificateReminderSettingUpdate(
                is_enabled=True, reminder_days=30
            )
        )
    with pytest.raises(Exception, match="QA 联系人范围"):
        await service.update_reminder_settings(
            certificate.CertificateReminderSettingUpdate(
                is_enabled=True, reminder_days=30, recipient_open_id="unknown"
            )
        )
    service._get_qa_reminder_recipient_by_open_id = AsyncMock(
        return_value=SimpleNamespace(
            open_id="qa", name="QA", department="QA部", enterprise_email=None
        )
    )
    saved = await service.update_reminder_settings(
        certificate.CertificateReminderSettingUpdate(
            is_enabled=False, reminder_days=30, recipient_open_id="qa"
        )
    )
    assert saved.is_enabled is False
    repo.get_reminder_setting.return_value = SimpleNamespace(
        is_enabled=True,
        reminder_days=30,
        recipient_open_id="qa",
        recipient_name="QA",
        recipient_department="QA部",
    )
    assert (await service.get_reminder_settings()).pending_count == 0
