from __future__ import annotations

from datetime import UTC, date, datetime, timedelta
from io import BytesIO
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from app.core.exceptions import AppException, NotFoundException
from app.modules.registration.models import RegistrationCertificateEntry
from app.modules.registration.service import certificate as cert


def _entry(
    sheet_key: str = "international-registration", **overrides: Any
) -> RegistrationCertificateEntry:
    values = {
        "sheet_key": sheet_key,
        "sheet_name": "国外注册",
        "sheet_title": "证书标题",
        "source_sequence": 1,
        "certificate_name": "测试证书",
        "acceptance_number": "受理号",
        "approval_number": "批件号",
        "certificate_number": "CERT-1",
        "issuing_authority": "监管机构",
        "issue_date": "2025.01.01",
        "validity_period": "至 2027.01.01",
        "expiry_date": "2027.01.01",
        "product_scope": "产品A\n产品B",
        "quality_standard": "标准",
        "page_count": 3,
        "remarks": "备注",
    }
    values.update(overrides)
    entry = RegistrationCertificateEntry(**values)
    entry.id = uuid4()
    entry.created_at = datetime.now(UTC)
    entry.updated_at = datetime.now(UTC)
    return entry


def _certificate_workbook_bytes(
    *, bad_header: bool = False, missing_last: bool = False
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    configs = (
        cert.CERTIFICATE_SHEET_CONFIG[:-1]
        if missing_last
        else cert.CERTIFICATE_SHEET_CONFIG
    )
    for config in configs:
        worksheet = workbook.create_sheet(str(config["name"]))
        worksheet.cell(1, 1, f"{config['name']}标题")
        headers = cert._get_sheet_headers(config)
        name_index = headers.index("证照名称")
        if bad_header and config is configs[0]:
            headers[1] = "错误表头"
        worksheet.append(headers)
        row = [1, *[f"值-{label}" for label in config["columns"]]]  # type: ignore[attr-defined]
        row[name_index] = f"{config['name']}证书"
        if "有效期/复验期" in headers:
            row[headers.index("有效期/复验期")] = "2025.01.01-2028.12.31"
        if "页数" in headers:
            row[headers.index("页数")] = "共 12 页"
        worksheet.append(row)
        worksheet.append([2, *([None] * (len(headers) - 1))])
        worksheet.append(["非数字", "忽略记录", *([None] * (len(headers) - 2))])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_certificate_normalizers_cover_date_text_and_placeholders() -> None:
    assert cert._normalize_text(None) == ""
    assert cert._normalize_text(datetime(2026, 8, 20)) == "2026.08.20"
    assert cert._normalize_text(date(2026, 8, 20)) == "2026.08.20"
    assert cert._normalize_text(" a\r\n b ") == "a\nb"
    assert cert._preserve_text(None) == ""
    assert cert._preserve_text(date(2026, 8, 20)) == "2026.08.20"
    assert cert._preserve_text(" a\r\nb ") == " a\nb "
    assert cert._slugify("Hello World") == "hello_world"
    assert len(cert._slugify("证书")) == 12
    assert cert._normalize_optional(None) is None
    assert cert._normalize_optional(" -- ") is None
    assert cert._normalize_optional(" value ") == "value"
    assert cert._preserve_optional(None) is None
    assert cert._preserve_optional(" value ") == " value "
    assert cert._is_template_placeholder_value(None) is False
    assert cert._is_template_placeholder_value("") is False
    assert cert._is_template_placeholder_value("   ") is True
    assert cert._is_template_placeholder_value("__—/") is True
    assert cert._is_template_placeholder_value("value") is False


def test_certificate_date_status_page_and_group_helpers() -> None:
    assert cert._extract_date(None) is None
    assert cert._extract_date("长期") is None
    assert cert._extract_date("无日期") is None
    assert cert._extract_date("2026.02.30") is None
    assert cert._extract_date("签发 2025/1/1 到期 2027.02.03") == date(2027, 2, 3)
    assert cert._format_date(date(2026, 8, 20)) == "2026.08.20"
    assert cert._format_date(None) is None
    assert cert._calculate_expiry_status(None) == "无明确期限"
    assert (
        cert._calculate_expiry_status((date.today() - timedelta(days=1)).isoformat())
        == "已过期"
    )
    assert (
        cert._calculate_expiry_status((date.today() + timedelta(days=30)).isoformat())
        == "90天内到期"
    )
    assert (
        cert._calculate_expiry_status((date.today() + timedelta(days=180)).isoformat())
        == "有效"
    )
    assert cert._parse_page_count(None) is None
    assert cert._parse_page_count("无") is None
    assert cert._parse_page_count("共12页") == 12
    assert cert._split_lines(None) == []
    assert cert._split_lines("A\n-\n B ") == ["A", "B"]
    assert cert._normalize_group_text(None) == ""
    assert cert._normalize_group_text(" A  B ") == "a b"
    assert cert._normalize_department_text(None) == ""
    assert cert._is_qa_department("质量保证部") is True
    assert cert._is_qa_department(" qa ") is True
    assert cert._is_qa_department("注册部") is False


def test_sheet_lookup_and_sequence_helpers() -> None:
    assert cert._get_sheet_meta("international-registration")["name"] == "国外注册"
    assert cert._get_sheet_meta_by_name("国内注册")["key"] == "domestic-registration"
    with pytest.raises(NotFoundException):
        cert._get_sheet_meta("missing")
    with pytest.raises(NotFoundException):
        cert._get_sheet_meta_by_name("missing")
    assert cert._extract_sequence(None) is None
    assert cert._extract_sequence(3) == 3
    assert cert._extract_sequence(" 4 ") == 4
    assert cert._extract_sequence("4a") is None


@pytest.mark.parametrize(
    "sheet_key",
    [
        "international-registration",
        "domestic-registration",
        "domestic-gmp",
        "international-gmp",
    ],
)
def test_payload_and_row_mapping_for_every_sheet(sheet_key: str) -> None:
    meta = cert._get_sheet_meta(sheet_key)
    values = {str(label): f"值-{label}" for label in meta["columns"]}  # type: ignore[attr-defined]
    values["有效期/复验期"] = "到期 2028.08.20"
    values["页数"] = "8页"
    payload = cert._map_values_to_entry_payload(
        sheet_key,
        str(meta["name"]),
        "标题",
        values,  # type: ignore[arg-type]
        source_sequence=2,
    )
    assert payload["expiry_date"] == "2028.08.20"
    assert payload["page_count"] == 8
    entry = _entry(sheet_key, sheet_name=str(meta["name"]))
    row = cert._build_row_values(entry)
    assert row["证照名称"] == "测试证书"
    assert cert._build_sheet_row(entry).sequence == 1
    assert cert._build_record_summary(entry).certificate_name == "测试证书"
    assert cert._build_entry_response(entry).sheet_key == sheet_key


def test_sheet_summary_reminder_content_and_due_count() -> None:
    expired = _entry(expiry_date=(date.today() - timedelta(days=1)).isoformat())
    due = _entry(
        source_sequence=2,
        certificate_name="即将到期",
        expiry_date=(date.today() + timedelta(days=10)).isoformat(),
    )
    undated = _entry(source_sequence=3, certificate_name="长期", expiry_date=None)
    summary = cert._build_sheet_summary(
        "international-registration", [expired, due, undated]
    )
    assert summary.total_records == 3
    assert summary.expired_count == 1
    assert summary.due_90_count == 1
    assert summary.product_count == 2
    assert (
        cert._build_sheet_summary("international-registration", []).total_records == 0
    )
    assert cert._count_due_entries_for_setting([expired, due, undated], 30) == 1

    entries = [
        _entry(source_sequence=i, certificate_name=f"证书{i}") for i in range(12)
    ]
    content = cert._build_reminder_content(entries, 90)
    assert "到期前 90 天" in content
    assert "其余还有 **2** 份" in content


def test_parse_complete_certificate_workbook() -> None:
    entries = cert._load_entries_from_workbook_bytes(_certificate_workbook_bytes())
    assert len(entries) == 4
    assert {entry.sheet_key for entry in entries} == {
        "international-registration",
        "domestic-registration",
        "domestic-gmp",
        "international-gmp",
    }
    assert all(entry.page_count == 12 for entry in entries)


def test_certificate_workbook_rejects_missing_sheet_bad_header_and_empty_name() -> None:
    with pytest.raises(AppException, match="缺少子表"):
        cert._load_entries_from_workbook_bytes(
            _certificate_workbook_bytes(missing_last=True)
        )
    with pytest.raises(AppException, match="表头不匹配"):
        cert._load_entries_from_workbook_bytes(
            _certificate_workbook_bytes(bad_header=True)
        )

    workbook = load_workbook(BytesIO(_certificate_workbook_bytes()))
    first = workbook[str(cert.CERTIFICATE_SHEET_CONFIG[0]["name"])]
    headers = cert._get_sheet_headers(cert.CERTIFICATE_SHEET_CONFIG[0])
    first.cell(3, headers.index("证照名称") + 1, "")
    with pytest.raises(AppException, match="证照名称为空"):
        cert._parse_sheet_entries(first, cert.CERTIFICATE_SHEET_CONFIG[0])


def test_fill_certificate_sheet_reuses_rows_clears_stale_and_adds_rows() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["序号", "证照名称", "备注"]
    worksheet.append(["标题"])
    worksheet.append(headers)
    worksheet.append([1, "旧证书", "旧备注"])
    worksheet.append(headers)
    worksheet.append([2, "待删除", "旧备注"])
    worksheet.merge_cells("B6:C6")
    worksheet.cell(6, 1, 3)

    rows = [
        {"序号": "1", "证照名称": "更新证书", "备注": None},
        {"序号": "4", "证照名称": "新增证书", "备注": "新增"},
        {"序号": "5", "证照名称": "扩展证书", "备注": "扩展"},
    ]
    cert._fill_certificate_sheet(
        worksheet,
        rows,  # type: ignore[arg-type]
        headers=headers,
        start_row=3,
        template_row=3,
    )
    values = [
        worksheet.cell(row=index, column=2).value
        for index in range(3, worksheet.max_row + 1)
    ]
    assert "更新证书" in values
    assert "新增证书" in values
    assert "扩展证书" in values
    assert "待删除" not in values


def test_clear_sheet_data_area_noop_and_merged_cells() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    cert._clear_sheet_data_area(worksheet, row_indexes=set(), column_count=2)
    worksheet.merge_cells("A2:B2")
    worksheet.cell(2, 1, "merged")
    cert._clear_sheet_data_area(worksheet, row_indexes={2}, column_count=2)
    assert worksheet.cell(2, 1).value is None
