from __future__ import annotations

from types import SimpleNamespace
from unittest.mock import AsyncMock
from uuid import uuid4

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font

from app.core.exceptions import AppException, NotFoundException
from app.modules.registration.schemas.declaration_progress import (
    DeclarationProgressEntryInput,
)
from app.modules.registration.service import declaration_progress
from app.modules.registration.service.declaration_progress import (
    DeclarationProgressWorkbookService,
)


def _definition() -> declaration_progress.DeclarationProgressSheetDefinition:
    template = declaration_progress.DECLARATION_PROGRESS_SHEET_TEMPLATES[0]
    return declaration_progress.DeclarationProgressSheetDefinition(
        worksheet_title=template.worksheet_title,
        sheet_key=template.sheet_key,
        sheet_name=template.sheet_name,
        sheet_title=template.sheet_title,
        supports_sub_records=True,
        columns=declaration_progress._build_columns(template),
    )


def test_declaration_progress_styles_grouping_and_export_helpers() -> None:
    template = declaration_progress.DECLARATION_PROGRESS_SHEET_TEMPLATES[0]
    columns = declaration_progress._build_columns(template)
    assert columns[0].key != columns[1].key
    assert declaration_progress._normalize_text(" a\r\n b ") == "a\nb"
    assert declaration_progress._normalize_optional_text(" ") is None
    assert declaration_progress._extract_sequence_number("项目 12") == 12
    assert declaration_progress._extract_sequence_number("无") is None
    assert declaration_progress._header_matches(template.column_labels, template)
    assert not declaration_progress._header_matches(
        template.column_labels[:-1], template
    )
    domestic = next(
        item
        for item in declaration_progress.DECLARATION_PROGRESS_SHEET_TEMPLATES
        if item.sheet_key == "domestic-completed"
    )
    altered = domestic.column_labels.copy()
    altered[1] = "='[1]5.2019项目"
    assert declaration_progress._header_matches(altered, domestic)

    workbook = Workbook()
    sheet = workbook.active
    black = Font(color="FF000000")
    blue = Font(color="FF0000FE")
    red = Font(color="FFFF0000")
    sheet["A1"].font = black
    sheet["B1"].font = blue
    sheet["C1"].font = red
    assert declaration_progress._classify_style_mark(sheet["A1"]) is None
    assert declaration_progress._classify_style_mark(sheet["B1"]) == "new"
    assert declaration_progress._classify_style_mark(sheet["C1"]) == "updated"
    assert (
        declaration_progress._classify_style_mark(
            SimpleNamespace(font=SimpleNamespace(color=None))
        )
        is None
    )

    definition = _definition()
    sequence_key = definition.columns[0].key
    project_key = definition.columns[1].key
    child_key = definition.columns[-1].key
    for row, values in (
        (5, {sequence_key: "1", project_key: "项目A", child_key: "初始"}),
        (6, {sequence_key: None, child_key: "补充"}),
        (7, {sequence_key: "2", project_key: "项目B", child_key: "第二"}),
    ):
        for column_index, column in enumerate(definition.columns, start=1):
            sheet.cell(row=row, column=column_index).value = values.get(column.key)
    versions = declaration_progress._load_versions_from_worksheet(definition, sheet)
    assert len(versions) == 3
    for version in versions:
        version.id = uuid4()
    assert versions[1].values_data[project_key] == "项目A"
    detail = declaration_progress._build_sheet_detail(definition, versions)
    assert detail.summary.total_records == 2
    assert detail.summary.records_with_history == 1
    rows = declaration_progress._build_export_rows(definition, versions)
    assert len(rows) == 3
    declaration_progress._fill_declaration_progress_sheet(sheet, definition, rows)
    assert sheet.cell(5, 1).value == "1"


@pytest.mark.asyncio
async def test_declaration_progress_service_create_update_sub_record_and_delete(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    definition = _definition()
    sequence_key = definition.columns[0].key
    project_key = definition.columns[1].key
    child_key = definition.columns[-1].key
    latest = SimpleNamespace(
        id=uuid4(),
        record_group_id=uuid4(),
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        sheet_title=definition.sheet_title,
        source_sequence=5,
        version_number=1,
        source_row_number=5,
        values_data={sequence_key: "5", project_key: "项目", child_key: "旧"},
        style_marks={project_key: "new"},
        project_name="项目",
        product_name=None,
    )
    repository = SimpleNamespace(
        count_versions=AsyncMock(return_value=1),
        get_max_source_sequence=AsyncMock(return_value=4),
        generate_group_id=uuid4,
        create_version=AsyncMock(side_effect=lambda value: value),
        get_latest_version_by_group=AsyncMock(return_value=latest),
        get_next_version_number=AsyncMock(return_value=2),
        list_versions_by_group=AsyncMock(return_value=[latest]),
        soft_delete_group=AsyncMock(),
    )
    session = SimpleNamespace(
        commit=AsyncMock(),
        flush=AsyncMock(),
        execute=AsyncMock(return_value=SimpleNamespace(scalar_one=lambda: latest)),
    )
    service = DeclarationProgressWorkbookService(session)
    service.repository = repository
    monkeypatch.setattr(
        declaration_progress,
        "_find_sheet_definition",
        lambda _key: definition,
    )

    created = await service.create_entry(
        DeclarationProgressEntryInput(
            sheet_key=definition.sheet_key,
            values={project_key: "新项目"},
        )
    )
    assert created.sequence == 5
    updated = await service.update_entry(
        latest.record_group_id,
        DeclarationProgressEntryInput(
            sheet_key=definition.sheet_key,
            values={project_key: "项目", child_key: "新值"},
        ),
    )
    assert updated.values[child_key] == "新值"
    sub_record = await service.create_sub_record(
        latest.record_group_id,
        DeclarationProgressEntryInput(
            sheet_key=definition.sheet_key,
            values={child_key: "再次更新"},
        ),
    )
    assert sub_record.version_number == 2
    await service.delete_entry(latest.record_group_id)
    repository.soft_delete_group.assert_awaited_once_with(latest.record_group_id)

    with pytest.raises(AppException, match="至少修改"):
        await service.create_sub_record(
            latest.record_group_id,
            DeclarationProgressEntryInput(
                sheet_key=definition.sheet_key,
                values={child_key: "新值"},
            ),
        )
    repository.list_versions_by_group.return_value = []
    with pytest.raises(NotFoundException):
        await service.get_entry_history(uuid4())
