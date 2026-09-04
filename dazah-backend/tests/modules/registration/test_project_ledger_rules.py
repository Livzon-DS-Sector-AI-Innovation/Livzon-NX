from __future__ import annotations

from pathlib import Path
from typing import Any
from uuid import uuid4

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from app.core.exceptions import AppException, NotFoundException
from app.modules.registration.models import RegistrationProjectLedgerVersion
from app.modules.registration.schemas.project_ledger import ProjectLedgerColumn
from app.modules.registration.service import project_ledger as ledger


def _definition() -> ledger.ProjectLedgerSheetDefinition:
    return ledger.ProjectLedgerSheetDefinition(
        sheet_key="test-sheet",
        sheet_name="测试子表",
        sheet_title="测试标题",
        worksheet_title="测试表",
        columns=[
            ProjectLedgerColumn(key="sequence", label="序号"),
            ProjectLedgerColumn(key="project", label="项目名称"),
            ProjectLedgerColumn(key="product", label="产品"),
            ProjectLedgerColumn(key="activity", label="药政活动"),
            ProjectLedgerColumn(key="activity_note", label="药政活动说明"),
        ],
    )


def _version(
    *,
    group_id: Any = None,
    version_number: int = 1,
    sequence: int = 1,
    values: dict[str, str | None] | None = None,
) -> RegistrationProjectLedgerVersion:
    item = RegistrationProjectLedgerVersion(
        record_group_id=group_id or uuid4(),
        sheet_key="test-sheet",
        sheet_name="测试子表",
        sheet_title="测试标题",
        source_sequence=sequence,
        version_number=version_number,
        source_row_number=version_number + 2,
        values_data=values
        or {
            "sequence": str(sequence),
            "project": "项目A",
            "product": "产品A",
            "activity": "申报",
            "activity_note": None,
        },
        project_name="项目A",
        product_name="产品A",
    )
    item.id = uuid4()
    return item


def _save_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title in ledger.PROJECT_LEDGER_SHEET_CONFIG:
        worksheet = workbook.create_sheet(title)
        worksheet.append([f"{title}标题"])
        worksheet.append(["序号", "项目名称", "产品", "药政活动", ""])
        worksheet.append(["1", "项目A", "产品A", "首次递交", "初次"])
        worksheet.append([None, None, None, "补充资料", "回复"])
        worksheet.append(["2", None, None, None, None])
    workbook.save(path)


def test_project_ledger_normalizers_and_keys() -> None:
    assert ledger._normalize_text(None) == ""
    assert ledger._normalize_text(" a\r\n b ") == "a\nb"
    assert ledger._slugify("Hello World") == "hello_world"
    assert len(ledger._slugify("项目")) == 12
    used: set[str] = set()
    assert ledger._build_column_key("Name", used) == "name"
    assert ledger._build_column_key("Name", used) == "name_2"
    assert ledger._normalize_header_label(" 标题 ", None, 0) == "标题"
    assert ledger._normalize_header_label("", "药政活动", 1) == "药政活动说明"
    assert ledger._normalize_header_label("", "产品", 2) == "补充信息3"
    assert ledger._normalize_header_labels(["序号", "药政活动", ""]) == [
        "序号",
        "药政活动",
        "药政活动说明",
    ]
    assert ledger._extract_sequence_number(None) is None
    assert ledger._extract_sequence_number("序号 12") == 12
    assert ledger._extract_sequence_number("none") is None
    assert ledger._normalize_optional_text(None) is None
    assert ledger._normalize_optional_text("  ") is None
    assert ledger._normalize_optional_text(" x ") == "x"
    assert ledger._resolve_sheet_key_and_name("未知表")[1] == "未知表"


def test_parse_workbook_definitions_and_seed_versions(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "ledger.xlsx"
    _save_workbook(path)
    monkeypatch.setattr(ledger, "_get_workbook_path", lambda: path)

    definitions, updated_at = ledger._parse_workbook_definitions()
    assert len(definitions) == 4
    assert updated_at is not None
    assert definitions[0].columns[-1].label == "药政活动说明"
    assert ledger._find_sheet_definition(definitions[0].sheet_key).sheet_name
    with pytest.raises(NotFoundException):
        ledger._find_sheet_definition("missing")

    versions = ledger._load_seed_versions()
    assert len(versions) == 8
    assert versions[0].version_number == 1
    assert versions[1].version_number == 2
    project_key = next(
        column.key for column in definitions[0].columns if column.label == "项目名称"
    )
    assert versions[1].values_data[project_key] == "项目A"


def test_parse_workbook_definitions_handles_missing_file_and_skips_empty_sheets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    missing = tmp_path / "missing.xlsx"
    monkeypatch.setattr(ledger, "_get_workbook_path", lambda: missing)
    # 文件缺失时回退静态列定义（4 个子表），不再抛 404
    definitions, updated_at = ledger._parse_workbook_definitions()
    assert len(definitions) == len(ledger.PROJECT_LEDGER_SHEET_CONFIG)
    assert updated_at is None
    assert all(definition.columns for definition in definitions)

    path = tmp_path / "minimal.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "只有标题"
    worksheet.append(["标题"])
    empty_data = workbook.create_sheet("空数据")
    empty_data.append(["标题"])
    empty_data.append([None])
    workbook.save(path)
    monkeypatch.setattr(ledger, "_get_workbook_path", lambda: path)
    definitions, _ = ledger._parse_workbook_definitions()
    assert definitions == []


def test_load_versions_rejects_header_mismatch() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试表"
    worksheet.append(["标题"])
    worksheet.append(["错误", "项目名称", "产品", "药政活动", ""])
    with pytest.raises(AppException, match="表头不匹配"):
        ledger._load_versions_from_worksheet(_definition(), worksheet)


def test_load_workbook_rejects_missing_and_unexpected_sheets(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    definition = _definition()
    monkeypatch.setattr(
        ledger,
        "_parse_workbook_definitions",
        lambda *args, **kwargs: ([definition], None),
    )
    missing_path = tmp_path / "missing-sheet.xlsx"
    Workbook().save(missing_path)
    with pytest.raises(AppException, match="缺少子表"):
        ledger._load_versions_from_workbook_path(missing_path)

    unexpected_path = tmp_path / "unexpected.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "测试表"
    worksheet.append(["标题"])
    worksheet.append(["序号", "项目名称", "产品", "药政活动", "药政活动说明"])
    extra = workbook.create_sheet("额外表")
    extra.append(["业务数据"])
    workbook.save(unexpected_path)
    with pytest.raises(AppException, match="未定义子表"):
        ledger._load_versions_from_workbook_path(unexpected_path)


def test_record_sheet_and_response_builders_cover_options() -> None:
    definition = _definition()
    group_id = uuid4()
    first = _version(group_id=group_id)
    second = _version(
        group_id=group_id,
        version_number=2,
        values={
            "sequence": "1",
            "project": "项目A",
            "product": "产品A",
            "activity": "补充资料",
            "activity_note": "回复",
        },
    )
    single = _version(sequence=2)
    assert ledger._build_history_record(first).version == 1
    assert ledger._build_record([first, second]).history_count == 2
    assert (
        ledger._build_record([first], include_history_records=False).history_records
        == []
    )
    summary = ledger._build_sheet_summary(
        definition,
        [ledger._build_record([first, second]), ledger._build_record([single])],
    )
    assert summary.records_with_history == 1
    assert summary.total_history_versions == 1

    detail = ledger._build_sheet_detail(definition, [second, single, first])
    assert [record.sequence for record in detail.records] == [1, 2]
    compact = ledger._build_sheet_detail(
        definition,
        [first, second],
        include_columns=False,
        include_records=False,
        include_history_records=False,
    )
    assert compact.columns == []
    assert compact.records == []
    assert compact.summary.total_records == 1
    assert ledger._build_entry_response(second).version_number == 2
    assert (
        ledger._normalize_entry_values(definition, {"project": " x "})["project"] == "x"
    )
    assert ledger._get_project_name_key(definition) == "project"
    assert ledger._get_product_name_key(definition) == "product"


def test_export_rows_include_only_history_changes() -> None:
    definition = _definition()
    group_id = uuid4()
    first = _version(group_id=group_id)
    second = _version(
        group_id=group_id,
        version_number=2,
        values={
            "sequence": "1",
            "project": "项目A",
            "product": "产品A",
            "activity": "补充资料",
            "activity_note": None,
        },
    )
    rows = ledger._build_export_rows(definition, [first, second])
    assert rows[0]["sequence"] == "1"
    assert rows[1]["sequence"] is None
    assert rows[1]["project"] is None
    assert rows[1]["activity"] == "补充资料"


def test_fill_project_ledger_sheet_resets_merges_and_extends_rows() -> None:
    definition = _definition()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["标题"])
    worksheet.append([column.label for column in definition.columns])
    worksheet.append(["old", "old", "old", "old", "old"])
    worksheet.merge_cells("A3:A4")
    rows = [
        {
            "sequence": "1",
            "project": "项目A",
            "product": "产品A",
            "activity": "首次",
            "activity_note": None,
        },
        {
            "sequence": None,
            "project": None,
            "product": None,
            "activity": "补充",
            "activity_note": "回复",
        },
        {
            "sequence": "2",
            "project": "项目B",
            "product": "产品B",
            "activity": "首次",
            "activity_note": None,
        },
    ]
    ledger._fill_project_ledger_sheet(worksheet, definition, rows)
    assert worksheet.cell(3, 2).value == "项目A"
    assert worksheet.cell(4, 4).value == "补充"
    assert worksheet.cell(5, 2).value == "项目B"
    assert "A3:A4" in {str(item) for item in worksheet.merged_cells.ranges}
    ledger._apply_project_ledger_vertical_merges(worksheet, definition, [], start_row=3)


def test_reset_merges_ignores_ranges_outside_target() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:B1")
    worksheet.merge_cells("D3:E3")
    ledger._reset_project_ledger_merges(
        worksheet,
        start_row=3,
        end_row=4,
        column_count=2,
    )
    assert {str(item) for item in worksheet.merged_cells.ranges} == {"A1:B1", "D3:E3"}


def test_load_versions_falls_back_to_upload_definitions_when_config_missing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """配置文件缺失时，导入应回退用上传文件自身解析定义并成功建档。"""
    config_path = tmp_path / "missing.xlsx"
    monkeypatch.setattr(ledger, "_get_workbook_path", lambda: config_path)

    upload_path = tmp_path / "upload.xlsx"
    _save_workbook(upload_path)

    versions = ledger._load_versions_from_workbook_path(upload_path)
    assert versions
    assert any(version.version_number == 2 for version in versions)


@pytest.mark.asyncio
async def test_import_workbook_creates_config_file_when_missing(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """无源文件时导入建档成功，且上传内容落盘到配置路径（create-if-missing）。"""
    from types import SimpleNamespace
    from unittest.mock import AsyncMock

    config_path = tmp_path / "missing-config" / "1. 注册台账.xlsx"
    monkeypatch.setattr(ledger, "_get_workbook_path", lambda: config_path)

    upload_path = tmp_path / "upload.xlsx"
    _save_workbook(upload_path)
    content = upload_path.read_bytes()

    session = SimpleNamespace(commit=AsyncMock(), rollback=AsyncMock())
    service = ledger.ProjectLedgerWorkbookService(session)
    service.repository = SimpleNamespace(
        count_versions=AsyncMock(return_value=0),
        replace_all_versions=AsyncMock(),
    )

    monkeypatch.setattr(
        ledger,
        "read_upload_secure",
        AsyncMock(return_value=("1. 注册台账.xlsx", content)),
    )
    upload = SimpleNamespace(filename="1. 注册台账.xlsx")

    result = await service.import_workbook(upload)
    assert result.imported_records > 0
    assert config_path.exists()
    assert config_path.read_bytes() == content
