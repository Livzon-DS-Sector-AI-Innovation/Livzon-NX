import base64
import hashlib
import uuid
from io import BytesIO
from types import SimpleNamespace

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from app.modules.agent import agent_tools
from app.modules.agent.attachment_service import AgentAttachmentService
from app.modules.agent.service import AgentService
from app.modules.agent.tools import ToolContext


class FakeDb:
    async def flush(self) -> None:
        return None


class FailingDb:
    async def flush(self) -> None:
        raise RuntimeError("database flush failed")


class EmptyInput(BaseModel):
    pass


class FakeRepo:
    def __init__(self) -> None:
        self.items: list[SimpleNamespace] = []

    async def create_attachment(self, _db, **values):
        attachment_id = values.pop("attachment_id")
        item = SimpleNamespace(
            id=attachment_id,
            **values,
            version=1,
            is_deleted=False,
            updated_by=values["user_id"],
        )
        self.items.append(item)
        return item

    async def list_session_attachments(self, _db, *, session_id, user_id, limit=50):
        return [
            item
            for item in self.items
            if item.session_id == session_id
            and item.user_id == user_id
            and not item.is_deleted
        ][:limit]

    async def get_session_attachment(self, _db, *, session_id, user_id, attachment_ref):
        for item in reversed(self.items):
            if (
                item.session_id == session_id
                and item.user_id == user_id
                and not item.is_deleted
                and attachment_ref in {str(item.id), item.filename}
            ):
                return item
        return None


def _xlsx_bytes() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "销售"
    worksheet.append(["产品", "销量"])
    worksheet.append(["A", 10])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.mark.anyio
async def test_persist_cleans_up_object_when_database_flush_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    objects: dict[str, tuple[bytes, str]] = {}

    def upload_object(_module, key, data, _length, content_type):
        objects[key] = (data, content_type)

    monkeypatch.setattr(module, "is_enabled", lambda: True)
    monkeypatch.setattr(module, "upload_object", upload_object)
    monkeypatch.setattr(module, "get_object", lambda _module, key: objects.get(key))
    monkeypatch.setattr(
        module, "delete_object", lambda _module, key: objects.pop(key, None)
    )

    with pytest.raises(RuntimeError, match="database flush failed"):
        await AgentAttachmentService(FakeRepo()).persist(
            FailingDb(),
            session_id=uuid.uuid4(),
            message_id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            uploads=[
                {
                    "attachment_id": str(uuid.uuid4()),
                    "filename": "销售数据.xlsx",
                    "content_type": (
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    "kind": "document",
                    "data": _xlsx_bytes(),
                    "text": "产品\t销量\nA\t10",
                }
            ],
        )

    assert objects == {}


@pytest.mark.anyio
async def test_persisted_attachment_can_be_read_mutated_and_deleted(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    objects: dict[str, tuple[bytes, str]] = {}

    def upload_object(_module, key, data, _length, content_type):
        objects[key] = (data, content_type)
        return key

    def get_object(_module, key):
        return objects.get(key)

    def delete_object(_module, key):
        objects.pop(key, None)

    monkeypatch.setattr(module, "is_enabled", lambda: True)
    monkeypatch.setattr(module, "upload_object", upload_object)
    monkeypatch.setattr(module, "get_object", get_object)
    monkeypatch.setattr(module, "delete_object", delete_object)

    repo = FakeRepo()
    service = AgentAttachmentService(repo)
    db = FakeDb()
    session_id = uuid.uuid4()
    user_id = uuid.uuid4()
    attachment_id = uuid.uuid4()
    raw = _xlsx_bytes()
    [attachment] = await service.persist(
        db,
        session_id=session_id,
        message_id=uuid.uuid4(),
        user_id=user_id,
        uploads=[
            {
                "attachment_id": str(attachment_id),
                "filename": "销售数据.xlsx",
                "content_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                "kind": "document",
                "data": raw,
                "text": "[工作表: 销售]\n产品\t销量\nA\t10",
            }
        ],
    )

    read = await service.read(
        db,
        session_id=session_id,
        user_id=user_id,
        attachment_ref="销售数据.xlsx",
        offset=0,
        limit=20_000,
    )
    assert "A\t10" in read["content"]
    assert attachment.sha256 == hashlib.sha256(raw).hexdigest()

    result = await service.mutate_tabular(
        db,
        session_id=session_id,
        user_id=user_id,
        attachment_ref=str(attachment_id),
        action="append_row",
        sheet_name="销售",
        row_number=None,
        values=["B", 20],
    )
    assert result["version"] == 2
    stored = load_workbook(BytesIO(objects[attachment.object_key][0]), read_only=True)
    assert list(stored["销售"].values)[-1] == ("B", 20)
    stored.close()

    await service.mutate_tabular(
        db,
        session_id=session_id,
        user_id=user_id,
        attachment_ref="销售数据.xlsx",
        action="update_row",
        sheet_name="销售",
        row_number=3,
        values=["B", 25],
    )
    await service.mutate_tabular(
        db,
        session_id=session_id,
        user_id=user_id,
        attachment_ref="销售数据.xlsx",
        action="delete_row",
        sheet_name="销售",
        row_number=2,
        values=[],
    )
    assert attachment.version == 4

    deleted = await service.delete(
        db,
        session_id=session_id,
        user_id=user_id,
        attachment_ref="销售数据.xlsx",
    )
    assert deleted["deleted"] is True
    assert attachment.is_deleted is True
    assert attachment.object_key not in objects


@pytest.mark.anyio
async def test_mutation_restores_object_when_database_flush_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    objects: dict[str, tuple[bytes, str]] = {}

    def upload_object(_module, key, data, _length, content_type):
        objects[key] = (data, content_type)

    monkeypatch.setattr(module, "is_enabled", lambda: True)
    monkeypatch.setattr(module, "upload_object", upload_object)
    monkeypatch.setattr(module, "get_object", lambda _module, key: objects.get(key))
    monkeypatch.setattr(
        module, "delete_object", lambda _module, key: objects.pop(key, None)
    )

    repo = FakeRepo()
    service = AgentAttachmentService(repo)
    session_id = uuid.uuid4()
    user_id = uuid.uuid4()
    raw = _xlsx_bytes()
    [attachment] = await service.persist(
        FakeDb(),
        session_id=session_id,
        message_id=uuid.uuid4(),
        user_id=user_id,
        uploads=[
            {
                "attachment_id": str(uuid.uuid4()),
                "filename": "销售数据.xlsx",
                "content_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                "kind": "document",
                "data": raw,
                "text": "产品\t销量\nA\t10",
            }
        ],
    )

    with pytest.raises(RuntimeError, match="database flush failed"):
        await service.mutate_tabular(
            FailingDb(),
            session_id=session_id,
            user_id=user_id,
            attachment_ref=str(attachment.id),
            action="append_row",
            sheet_name="销售",
            row_number=None,
            values=["B", 20],
        )

    assert objects[attachment.object_key][0] == raw
    assert attachment.version == 1
    assert attachment.sha256 == hashlib.sha256(raw).hexdigest()


@pytest.mark.anyio
async def test_delete_restores_object_when_database_flush_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    objects: dict[str, tuple[bytes, str]] = {}

    def upload_object(_module, key, data, _length, content_type):
        objects[key] = (data, content_type)

    monkeypatch.setattr(module, "is_enabled", lambda: True)
    monkeypatch.setattr(module, "upload_object", upload_object)
    monkeypatch.setattr(module, "get_object", lambda _module, key: objects.get(key))
    monkeypatch.setattr(
        module, "delete_object", lambda _module, key: objects.pop(key, None)
    )

    repo = FakeRepo()
    service = AgentAttachmentService(repo)
    session_id = uuid.uuid4()
    user_id = uuid.uuid4()
    raw = _xlsx_bytes()
    [attachment] = await service.persist(
        FakeDb(),
        session_id=session_id,
        message_id=uuid.uuid4(),
        user_id=user_id,
        uploads=[
            {
                "attachment_id": str(uuid.uuid4()),
                "filename": "销售数据.xlsx",
                "content_type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                "kind": "document",
                "data": raw,
                "text": "产品\t销量\nA\t10",
            }
        ],
    )

    with pytest.raises(RuntimeError, match="database flush failed"):
        await service.delete(
            FailingDb(),
            session_id=session_id,
            user_id=user_id,
            attachment_ref=str(attachment.id),
        )

    assert objects[attachment.object_key][0] == raw
    assert attachment.is_deleted is False


def test_follow_up_restores_named_or_recent_attachment() -> None:
    attachment = SimpleNamespace(
        id=uuid.uuid4(),
        filename="销售数据.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size=100,
        kind="document",
        extracted_text="产品\t销量\nA\t10",
    )

    named = AgentService._restored_attachments(
        [attachment], "把销售数据.xlsx写入目标文档"
    )
    recent = AgentService._restored_attachments([attachment], "修改刚才上传的文件")

    assert named[0]["text"] == "产品\t销量\nA\t10"
    assert recent[0]["attachment_id"] == str(attachment.id)


@pytest.mark.anyio
async def test_follow_up_materializes_persisted_image_bytes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    raw = b"\x89PNG\r\n\x1a\nimage"
    attachment = SimpleNamespace(
        id=uuid.uuid4(),
        filename="现场照片.png",
        content_type="image/png",
        size=len(raw),
        kind="image",
        object_key="sessions/user/session/image/source",
        sha256=hashlib.sha256(raw).hexdigest(),
        extracted_text=None,
    )
    monkeypatch.setattr(
        module,
        "get_object",
        lambda _module, _key: (raw, "image/png"),
    )

    restored = await AgentAttachmentService().materialize_for_context(
        [attachment],
        text_limit=50_000,
    )

    assert base64.b64decode(restored[0]["data_base64"]) == raw
    assert restored[0]["attachment_id"] == str(attachment.id)


@pytest.mark.anyio
async def test_persist_rejects_unavailable_storage_and_invalid_bytes(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    service = AgentAttachmentService(FakeRepo())
    kwargs = {
        "session_id": uuid.uuid4(),
        "message_id": uuid.uuid4(),
        "user_id": uuid.uuid4(),
        "uploads": [
            {
                "attachment_id": str(uuid.uuid4()),
                "filename": "bad.txt",
                "content_type": "text/plain",
                "kind": "document",
                "data": "not-bytes",
            }
        ],
    }

    monkeypatch.setattr(module, "is_enabled", lambda: False)
    with pytest.raises(HTTPException) as unavailable:
        await service.persist(FakeDb(), **kwargs)
    assert unavailable.value.status_code == 503

    monkeypatch.setattr(module, "is_enabled", lambda: True)
    with pytest.raises(TypeError, match="must be bytes"):
        await service.persist(FakeDb(), **kwargs)


@pytest.mark.anyio
async def test_persist_reports_readback_failure_and_attempts_cleanup(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    cleanup_errors: list[str] = []
    monkeypatch.setattr(module, "is_enabled", lambda: True)
    monkeypatch.setattr(module, "upload_object", lambda *_args: None)
    monkeypatch.setattr(module, "get_object", lambda *_args: None)
    monkeypatch.setattr(
        module,
        "delete_object",
        lambda *_args: (_ for _ in ()).throw(RuntimeError("cleanup failed")),
    )
    monkeypatch.setattr(
        module.logger,
        "exception",
        lambda message, *_args: cleanup_errors.append(message),
    )

    with pytest.raises(HTTPException) as exc_info:
        await AgentAttachmentService(FakeRepo()).persist(
            FakeDb(),
            session_id=uuid.uuid4(),
            message_id=None,
            user_id=uuid.uuid4(),
            uploads=[
                {
                    "attachment_id": str(uuid.uuid4()),
                    "filename": "source.txt",
                    "content_type": "text/plain",
                    "kind": "document",
                    "data": b"source",
                }
            ],
        )

    assert exc_info.value.status_code == 503
    assert cleanup_errors == [
        "Failed to clean up attachment object after persist failure: object_key=%s"
    ]


@pytest.mark.anyio
async def test_materialize_rejects_missing_or_tampered_images_and_bounds_text(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    image = SimpleNamespace(
        id=uuid.uuid4(),
        filename="photo.png",
        content_type="image/png",
        size=4,
        kind="image",
        object_key="image-key",
        sha256=hashlib.sha256(b"safe").hexdigest(),
        extracted_text=None,
    )
    monkeypatch.setattr(module, "get_object", lambda *_args: None)
    with pytest.raises(HTTPException) as missing:
        await AgentAttachmentService().materialize_for_context([image], text_limit=10)
    assert missing.value.status_code == 409

    monkeypatch.setattr(
        module,
        "get_object",
        lambda *_args: (b"evil", "image/png"),
    )
    with pytest.raises(HTTPException, match="校验失败"):
        await AgentAttachmentService().materialize_for_context([image], text_limit=10)

    text_attachment = SimpleNamespace(
        id=uuid.uuid4(),
        filename="empty.txt",
        content_type="text/plain",
        size=0,
        kind="document",
        object_key="text-key",
        sha256=hashlib.sha256(b"").hexdigest(),
        extracted_text=None,
    )
    [restored] = await AgentAttachmentService().materialize_for_context(
        [text_attachment],
        text_limit=4,
    )
    assert restored["text"] == "（未提取"
    assert restored["truncated"] is True


@pytest.mark.anyio
async def test_attachment_lookup_and_storage_failure_boundaries(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import app.modules.agent.attachment_service as module

    repo = FakeRepo()
    service = AgentAttachmentService(repo)
    with pytest.raises(HTTPException) as no_session:
        await service.require(
            FakeDb(),
            session_id=None,
            user_id=uuid.uuid4(),
            attachment_ref="missing",
        )
    assert no_session.value.status_code == 400

    with pytest.raises(HTTPException) as not_found:
        await service.require(
            FakeDb(),
            session_id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            attachment_ref="missing",
        )
    assert not_found.value.status_code == 404

    attachment = SimpleNamespace(
        id=uuid.uuid4(),
        filename="notes.txt",
        content_type="text/plain",
        size=5,
        kind="document",
        object_key="missing-key",
        sha256=hashlib.sha256(b"notes").hexdigest(),
        extracted_text="notes",
        version=1,
        is_deleted=False,
        updated_by=None,
        session_id=uuid.uuid4(),
        user_id=uuid.uuid4(),
    )
    repo.items.append(attachment)
    monkeypatch.setattr(module, "get_object", lambda *_args: None)

    with pytest.raises(HTTPException) as mutate_missing:
        await service.mutate_tabular(
            FakeDb(),
            session_id=attachment.session_id,
            user_id=attachment.user_id,
            attachment_ref=str(attachment.id),
            action="append_row",
            sheet_name=None,
            row_number=None,
            values=["value"],
        )
    assert mutate_missing.value.status_code == 409

    monkeypatch.setattr(
        module,
        "get_object",
        lambda *_args: (b"notes", "text/plain"),
    )
    with pytest.raises(HTTPException) as unsupported:
        await service.mutate_tabular(
            FakeDb(),
            session_id=attachment.session_id,
            user_id=attachment.user_id,
            attachment_ref=str(attachment.id),
            action="append_row",
            sheet_name=None,
            row_number=None,
            values=["value"],
        )
    assert unsupported.value.status_code == 400

    monkeypatch.setattr(module, "get_object", lambda *_args: None)
    with pytest.raises(HTTPException) as delete_missing:
        await service.delete(
            FakeDb(),
            session_id=attachment.session_id,
            user_id=attachment.user_id,
            attachment_ref=str(attachment.id),
        )
    assert delete_missing.value.status_code == 409


def test_tabular_mutation_validation_and_csv_operations() -> None:
    service = AgentAttachmentService()
    raw_xlsx = _xlsx_bytes()

    with pytest.raises(HTTPException, match="工作表不存在"):
        service._mutate_xlsx(
            raw_xlsx,
            action="append_row",
            sheet_name="不存在",
            row_number=None,
            values=["B", 20],
        )
    with pytest.raises(HTTPException, match="新增行"):
        service._mutate_xlsx(
            raw_xlsx,
            action="append_row",
            sheet_name=None,
            row_number=None,
            values=[],
        )
    with pytest.raises(HTTPException, match="row_number"):
        service._mutate_xlsx(
            raw_xlsx,
            action="update_row",
            sheet_name="销售",
            row_number=99,
            values=["B", 20],
        )
    with pytest.raises(HTTPException, match="修改行"):
        service._mutate_xlsx(
            raw_xlsx,
            action="update_row",
            sheet_name="销售",
            row_number=2,
            values=[],
        )

    raw_csv = "产品,销量\r\nA,10\r\n".encode()
    appended = service._mutate_csv(
        raw_csv,
        action="append_row",
        row_number=None,
        values=["B", 20, None],
    )
    updated = service._mutate_csv(
        appended,
        action="update_row",
        row_number=2,
        values=["A", 15],
    )
    deleted = service._mutate_csv(
        updated,
        action="delete_row",
        row_number=2,
        values=[],
    )
    assert "B,20," in appended.decode("utf-8-sig")
    assert "A,15" in updated.decode("utf-8-sig")
    assert "A,15" not in deleted.decode("utf-8-sig")

    with pytest.raises(HTTPException, match="新增行"):
        service._mutate_csv(
            raw_csv,
            action="append_row",
            row_number=None,
            values=[],
        )
    with pytest.raises(HTTPException, match="row_number"):
        service._mutate_csv(
            raw_csv,
            action="update_row",
            row_number=0,
            values=["A", 15],
        )
    with pytest.raises(HTTPException, match="修改行"):
        service._mutate_csv(
            raw_csv,
            action="update_row",
            row_number=2,
            values=[],
        )
    with pytest.raises(HTTPException, match="编码不受支持"):
        service._mutate_csv(
            b"\x81",
            action="append_row",
            row_number=None,
            values=["A"],
        )


@pytest.mark.anyio
async def test_agent_attachment_tool_handlers_enforce_context_and_delegate(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    user = SimpleNamespace(id=uuid.uuid4())
    session_id = uuid.uuid4()
    calls: list[tuple[str, dict[str, object]]] = []
    item = SimpleNamespace(
        id=uuid.uuid4(),
        filename="data.csv",
        content_type="text/csv",
        size=12,
        kind="document",
        version=3,
    )

    class FakeAttachmentService:
        async def list_for_session(self, _db, **kwargs):
            calls.append(("list", kwargs))
            return [item]

        async def read(self, _db, **kwargs):
            calls.append(("read", kwargs))
            return {"content": "row"}

        async def mutate_tabular(self, _db, **kwargs):
            calls.append(("mutate", kwargs))
            return {"version": 4}

        async def delete(self, _db, **kwargs):
            calls.append(("delete", kwargs))
            return {"deleted": True}

    monkeypatch.setattr(agent_tools, "AgentAttachmentService", FakeAttachmentService)

    def context(*, active_user=user, active_session=session_id):
        return ToolContext(
            db=object(),  # type: ignore[arg-type]
            session_id=active_session,
            user_id=getattr(active_user, "id", None),
            user=active_user,  # type: ignore[arg-type]
            reason=None,
            raw_request=SimpleNamespace(operation="agent.list_attachments"),  # type: ignore[arg-type]
        )

    with pytest.raises(HTTPException) as unauthenticated:
        await agent_tools.list_attachments(context(active_user=None), EmptyInput())
    assert unauthenticated.value.status_code == 401

    with pytest.raises(HTTPException) as no_session:
        await agent_tools.list_attachments(context(active_session=None), EmptyInput())
    assert no_session.value.status_code == 400

    listed = await agent_tools.list_attachments(context(), EmptyInput())
    read = await agent_tools.read_attachment(
        context(),
        agent_tools.AttachmentReadInput(attachment_ref="data.csv", offset=1, limit=10),
    )
    mutated = await agent_tools.mutate_tabular_attachment(
        context(),
        agent_tools.TabularAttachmentMutationInput(
            attachment_ref="data.csv",
            action="append_row",
            values=["B", 20],
        ),
    )
    deleted = await agent_tools.delete_attachment(
        context(),
        agent_tools.AttachmentRefInput(attachment_ref="data.csv"),
    )

    assert listed[0]["attachment_id"] == str(item.id)
    assert read == {"content": "row"}
    assert mutated == {"version": 4}
    assert deleted == {"deleted": True}
    assert [name for name, _ in calls] == ["list", "read", "mutate", "delete"]
