import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.modules.procurement import service as procurement_service
from app.modules.procurement.schemas import (
    PurchaseApprovalRequest,
    PurchaseApprovalResult,
    PurchaseApprovalRole,
    PurchaseApprovalView,
    PurchaseRequestCategory,
    PurchaseRequestCreate,
    PurchaseRequestItemInput,
    PurchaseRequestStatus,
    PurchaseRequestUpdate,
)


class FakeDb:
    async def flush(self) -> None:
        return None


class FakePurchaseRequestRepository:
    requests = {}
    items = {}
    approvals = {}

    def __init__(self, session) -> None:
        self.session = session

    @classmethod
    def reset(cls) -> None:
        cls.requests = {}
        cls.items = {}
        cls.approvals = {}

    async def create(self, request, items):
        request.id = uuid.uuid4()
        request.created_at = datetime.now(UTC)
        request.updated_at = request.created_at
        self.requests[request.id] = request
        for item in items:
            item.id = uuid.uuid4()
            item.purchase_request_id = str(request.id)
            item.created_at = request.created_at
            item.updated_at = request.created_at
        self.items[request.id] = items
        self.approvals[request.id] = []
        return request

    async def get(self, request_id):
        return self.requests.get(request_id)

    async def find_by_import_duplicate_key(self, duplicate_key):
        for request in self.requests.values():
            if (
                request.import_duplicate_key == duplicate_key
                and not request.is_deleted
            ):
                return request
        return None

    async def get_for_update(self, request_id):
        return self.requests.get(request_id)

    async def list_items(self, request_id):
        return self.items.get(request_id, [])

    async def list_approvals(self, request_id):
        return list(self.approvals.get(request_id, []))

    async def list_requests(
        self,
        *,
        category=None,
        status=None,
        keyword=None,
        page=1,
        page_size=20,
    ):
        records = list(self.requests.values())
        if category:
            records = [record for record in records if record.category == category]
        if status:
            records = [record for record in records if record.status == status]
        if keyword:
            records = [
                record
                for record in records
                if keyword in record.request_department
            ]
        return records[(page - 1) * page_size : page * page_size], len(records)

    async def list_requests_by_approval(
        self,
        *,
        approval_role,
        result,
        category=None,
        keyword=None,
        page=1,
        page_size=20,
    ):
        matching_ids = []
        for request_id, approvals in self.approvals.items():
            if any(
                approval.approval_role == approval_role
                and approval.result == result
                for approval in approvals
            ):
                matching_ids.append(request_id)

        records = [
            self.requests[request_id]
            for request_id in matching_ids
            if request_id in self.requests
        ]
        if category:
            records = [record for record in records if record.category == category]
        if keyword:
            records = [
                record
                for record in records
                if keyword in record.request_department
            ]
        return records[(page - 1) * page_size : page * page_size], len(records)

    async def list_purchase_order_lines(
        self,
        *,
        start_date,
        end_date,
        status,
        category=None,
        page=None,
        page_size=None,
    ):
        rows = []
        for request_id, request in self.requests.items():
            if request.status != status:
                continue
            if not start_date <= request.request_date < end_date:
                continue
            if category and request.category != category:
                continue
            for item in self.items.get(request_id, []):
                rows.append((request, item))

        rows.sort(
            key=lambda row: (
                row[0].request_date,
                row[0].category,
                row[0].request_department,
                row[1].sequence,
            )
        )
        total = len(rows)
        if page is not None and page_size is not None:
            rows = rows[(page - 1) * page_size : page * page_size]
        return rows, total

    async def replace_items(self, request_id, items):
        for item in items:
            item.id = uuid.uuid4()
            item.purchase_request_id = str(request_id)
            item.created_at = datetime.now(UTC)
            item.updated_at = item.created_at
        self.items[request_id] = items

    async def add_approval(self, approval):
        approval.id = uuid.uuid4()
        approval.created_at = datetime.now(UTC)
        approval.updated_at = approval.created_at
        self.approvals.setdefault(uuid.UUID(approval.purchase_request_id), []).append(
            approval
        )
        return approval

    async def delete(self, request_id):
        request = self.requests.get(request_id)
        if request is None:
            return False
        request.is_deleted = True
        self.items.pop(request_id, None)
        self.approvals.pop(request_id, None)
        return True


@pytest.fixture(autouse=True)
def fake_purchase_request_repository(monkeypatch):
    FakePurchaseRequestRepository.reset()
    monkeypatch.setattr(
        procurement_service,
        "PurchaseRequestRepository",
        FakePurchaseRequestRepository,
    )


def _create_payload() -> PurchaseRequestCreate:
    return PurchaseRequestCreate(
        category=PurchaseRequestCategory.hardware,
        request_department="102一车间",
        request_date=date(2026, 6, 28),
        attachment_note="附件：技术参数表一份",
        items=[
            PurchaseRequestItemInput(
                product_name="碳鼓",
                specification="M1005",
                material_code="MAT-1005",
                material_description="打印机碳鼓",
                rule_model="M1005",
                purpose="更换打印机碳鼓",
                material="",
                brand="惠普",
                quantity=Decimal("2"),
                unit="个",
                unit_price=Decimal("60.005"),
                remarks="申购人：郭娇",
            )
        ],
    )


def _create_payload_for(
    *,
    category: PurchaseRequestCategory = PurchaseRequestCategory.hardware,
    request_department: str = "102一车间",
    request_date: date = date(2026, 6, 28),
    product_name: str = "碳鼓",
) -> PurchaseRequestCreate:
    payload = _create_payload()
    payload.category = category
    payload.request_department = request_department
    payload.request_date = request_date
    payload.items[0].product_name = product_name
    return payload


def _create_approval_payload(
    category: PurchaseRequestCategory,
) -> PurchaseRequestCreate:
    payload = _create_payload_for(category=category)
    if category is PurchaseRequestCategory.urgent:
        payload.items[0].item_category = PurchaseRequestCategory.office
    return payload


async def _approve_role(
    request_id: uuid.UUID,
    approval_role: PurchaseApprovalRole,
    *,
    approver_name: str | None = None,
) -> None:
    await procurement_service.approve_purchase_request(
        FakeDb(),
        request_id,
        PurchaseApprovalRequest(
            approval_role=approval_role,
            approver_name=approver_name or approval_role.value,
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )


async def _approve_request(request_id: uuid.UUID) -> None:
    await procurement_service.submit_purchase_request(FakeDb(), request_id)
    request = FakePurchaseRequestRepository.requests[request_id]
    workflow = procurement_service.get_purchase_approval_workflow(request.category)
    for approval_role in workflow:
        required_count = procurement_service.PURCHASE_APPROVAL_REQUIRED_COUNTS.get(
            approval_role,
            1,
        )
        for index in range(required_count):
            await _approve_role(
                request_id,
                approval_role,
                approver_name=f"{approval_role.value}-{index + 1}",
            )


@pytest.mark.anyio
async def test_purchase_request_amount_and_hardware_approval_flow() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )

    assert created.request_date == date(2026, 6, 28)
    assert created.attachment_note == "附件：技术参数表一份"
    assert created.status == PurchaseRequestStatus.draft
    assert created.total_amount == Decimal("120.01")
    assert created.items[0].total_amount == Decimal("120.01")

    submitted = await procurement_service.submit_purchase_request(
        FakeDb(),
        created.id,
    )
    assert submitted.status == PurchaseRequestStatus.pending_hardware_warehouse

    await _approve_role(
        created.id,
        PurchaseApprovalRole.hardware_warehouse,
        approver_name="五金库",
    )
    request = FakePurchaseRequestRepository.requests[created.id]
    assert request.status == PurchaseRequestStatus.pending_department_head.value

    await _approve_role(created.id, PurchaseApprovalRole.department_head)
    assert request.status == PurchaseRequestStatus.pending_responsible_leader.value
    await _approve_role(created.id, PurchaseApprovalRole.responsible_leader)
    assert request.status == PurchaseRequestStatus.pending_supervising_leader.value
    await _approve_role(created.id, PurchaseApprovalRole.supervising_leader)
    assert request.status == PurchaseRequestStatus.pending_general_manager.value
    approved = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.general_manager,
            approver_name="总经理",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert approved.status == PurchaseRequestStatus.approved
    assert len(approved.approvals) == 5


@pytest.mark.parametrize(
    ("category", "expected_workflow"),
    [
        (
            PurchaseRequestCategory.hardware,
            (
                PurchaseApprovalRole.hardware_warehouse,
                PurchaseApprovalRole.department_head,
                PurchaseApprovalRole.responsible_leader,
                PurchaseApprovalRole.supervising_leader,
                PurchaseApprovalRole.general_manager,
            ),
        ),
        (
            PurchaseRequestCategory.electrical,
            (
                PurchaseApprovalRole.hardware_warehouse,
                PurchaseApprovalRole.equipment_power,
                PurchaseApprovalRole.department_head,
                PurchaseApprovalRole.responsible_leader,
                PurchaseApprovalRole.supervising_leader,
            ),
        ),
        (
            PurchaseRequestCategory.labor_special,
            (
                PurchaseApprovalRole.safety_officer,
                PurchaseApprovalRole.department_head,
                PurchaseApprovalRole.responsible_leader,
            ),
        ),
        (
            PurchaseRequestCategory.urgent,
            (
                PurchaseApprovalRole.hardware_warehouse,
                PurchaseApprovalRole.department_head,
                PurchaseApprovalRole.responsible_leader,
                PurchaseApprovalRole.supervising_leader,
                PurchaseApprovalRole.finance_director,
                PurchaseApprovalRole.general_manager,
            ),
        ),
        (
            PurchaseRequestCategory.office,
            (
                PurchaseApprovalRole.department_head,
                PurchaseApprovalRole.responsible_leader,
                PurchaseApprovalRole.supervising_leader,
            ),
        ),
    ],
)
@pytest.mark.anyio
async def test_purchase_request_workflow_by_category(
    category: PurchaseRequestCategory,
    expected_workflow: tuple[PurchaseApprovalRole, ...],
) -> None:
    assert procurement_service.get_purchase_approval_workflow(category) == (
        expected_workflow
    )

    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_approval_payload(category),
    )
    submitted = await procurement_service.submit_purchase_request(
        FakeDb(),
        created.id,
    )
    assert submitted.status == procurement_service.APPROVAL_ROLE_TO_PENDING_STATUS[
        expected_workflow[0]
    ]

    for index, approval_role in enumerate(expected_workflow):
        required_count = procurement_service.PURCHASE_APPROVAL_REQUIRED_COUNTS.get(
            approval_role,
            1,
        )
        for approval_index in range(required_count):
            await _approve_role(
                created.id,
                approval_role,
                approver_name=f"{approval_role.value}-{approval_index + 1}",
            )

        request = FakePurchaseRequestRepository.requests[created.id]
        if index + 1 < len(expected_workflow):
            assert request.status == (
                procurement_service.APPROVAL_ROLE_TO_PENDING_STATUS[
                    expected_workflow[index + 1]
                ].value
            )
        else:
            assert request.status == PurchaseRequestStatus.approved.value

    assert len(FakePurchaseRequestRepository.approvals[created.id]) == sum(
        procurement_service.PURCHASE_APPROVAL_REQUIRED_COUNTS.get(role, 1)
        for role in expected_workflow
    )


@pytest.mark.anyio
async def test_purchase_request_rejects_approval_role_outside_category_workflow(
) -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)

    with pytest.raises(ValueError, match="不包含此审批步骤"):
        await procurement_service.approve_purchase_request(
            FakeDb(),
            created.id,
            PurchaseApprovalRequest(
                approval_role=PurchaseApprovalRole.safety_officer,
                approver_name="安全员",
                opinion="同意",
                result=PurchaseApprovalResult.approved,
            ),
        )

    assert (
        FakePurchaseRequestRepository.requests[created.id].status
        == PurchaseRequestStatus.pending_hardware_warehouse.value
    )
    assert FakePurchaseRequestRepository.approvals[created.id] == []


@pytest.mark.anyio
async def test_electrical_co_signing_requires_two_current_round_approvals() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_approval_payload(PurchaseRequestCategory.electrical),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)

    first_approval = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.equipment_power,
            approver_name="何学斌",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert first_approval.status == PurchaseRequestStatus.pending_equipment_power
    pending, pending_total = await procurement_service.list_purchase_requests(
        FakeDb(),
        category=PurchaseRequestCategory.electrical.value,
        approval_role=PurchaseApprovalRole.equipment_power,
        approval_view=PurchaseApprovalView.pending,
    )
    assert pending_total == 1
    assert pending[0].id == created.id

    second_approval = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.equipment_power,
            approver_name="安伟",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert second_approval.status == PurchaseRequestStatus.pending_department_head
    assert [approval.approver_name for approval in second_approval.approvals] == [
        "hardware_warehouse",
        "何学斌",
        "安伟",
    ]


@pytest.mark.anyio
async def test_electrical_co_signing_rejection_and_resubmission_isolates_history(
) -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_approval_payload(PurchaseRequestCategory.electrical),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)
    await _approve_role(
        created.id,
        PurchaseApprovalRole.equipment_power,
        approver_name="何学斌",
    )

    rejected = await procurement_service.reject_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.equipment_power,
            approver_name="安伟",
            opinion="参数需要补充",
            result=PurchaseApprovalResult.rejected,
        ),
    )
    assert rejected.status == PurchaseRequestStatus.rejected
    assert rejected.rejected_step == PurchaseApprovalRole.equipment_power

    resubmitted = await procurement_service.submit_purchase_request(
        FakeDb(),
        created.id,
    )
    assert resubmitted.status == PurchaseRequestStatus.pending_hardware_warehouse
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)

    first_new_round_approval = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.equipment_power,
            approver_name="安伟",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert (
        first_new_round_approval.status
        == PurchaseRequestStatus.pending_equipment_power
    )

    second_new_round_approval = await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.equipment_power,
            approver_name="何学斌",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )
    assert (
        second_new_round_approval.status
        == PurchaseRequestStatus.pending_department_head
    )
    assert len(second_new_round_approval.approvals) == 6


@pytest.mark.anyio
async def test_purchase_request_reject_persists_approval_record() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)

    rejected = await procurement_service.reject_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="用途不清",
            result=PurchaseApprovalResult.rejected,
        ),
    )

    assert rejected.status == PurchaseRequestStatus.rejected
    assert rejected.rejected_step == PurchaseApprovalRole.department_head
    assert rejected.approvals[-1].result == PurchaseApprovalResult.rejected
    assert rejected.approvals[-1].opinion == "用途不清"


@pytest.mark.anyio
async def test_purchase_request_role_approval_views() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)
    await procurement_service.approve_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="同意",
            result=PurchaseApprovalResult.approved,
        ),
    )

    department_completed, department_completed_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.department_head,
            approval_view=PurchaseApprovalView.completed,
        )
    )
    leader_pending, leader_pending_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.responsible_leader,
            approval_view=PurchaseApprovalView.pending,
        )
    )

    assert department_completed_total == 1
    assert department_completed[0].id == created.id
    assert (
        department_completed[0].status
        == PurchaseRequestStatus.pending_responsible_leader
    )
    assert leader_pending_total == 1
    assert leader_pending[0].id == created.id


@pytest.mark.anyio
async def test_purchase_request_role_rejected_view() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    await procurement_service.submit_purchase_request(FakeDb(), created.id)
    await _approve_role(created.id, PurchaseApprovalRole.hardware_warehouse)
    await procurement_service.reject_purchase_request(
        FakeDb(),
        created.id,
        PurchaseApprovalRequest(
            approval_role=PurchaseApprovalRole.department_head,
            approver_name="部门负责人",
            opinion="用途不清",
            result=PurchaseApprovalResult.rejected,
        ),
    )

    department_rejected, department_rejected_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.department_head,
            approval_view=PurchaseApprovalView.rejected,
        )
    )
    leader_rejected, leader_rejected_total = (
        await procurement_service.list_purchase_requests(
            FakeDb(),
            category=PurchaseRequestCategory.hardware.value,
            approval_role=PurchaseApprovalRole.responsible_leader,
            approval_view=PurchaseApprovalView.rejected,
        )
    )

    assert department_rejected_total == 1
    assert department_rejected[0].id == created.id
    assert department_rejected[0].status == PurchaseRequestStatus.rejected
    assert leader_rejected_total == 0
    assert leader_rejected == []


@pytest.mark.anyio
async def test_purchase_order_lines_include_only_approved_requests_in_month() -> None:
    approved = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(product_name="当月已通过"),
    )
    await _approve_request(approved.id)

    draft = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(product_name="当月草稿"),
    )
    before_month = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(request_date=date(2026, 5, 31), product_name="上月"),
    )
    after_month = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(request_date=date(2026, 7, 1), product_name="下月"),
    )
    await _approve_request(before_month.id)
    await _approve_request(after_month.id)

    lines, total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        year=2026,
        month=6,
    )

    assert total == 1
    assert lines[0].request_id == approved.id
    assert lines[0].product_name == "当月已通过"
    assert draft.id not in {line.request_id for line in lines}


@pytest.mark.anyio
async def test_purchase_order_lines_filter_category_and_paginate() -> None:
    hardware = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            product_name="五金材料",
        ),
    )
    office = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.office,
            product_name="办公用品",
        ),
    )
    await _approve_request(hardware.id)
    await _approve_request(office.id)

    all_lines, all_total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        year=2026,
        month=6,
        page=1,
        page_size=1,
    )
    office_lines, office_total = await procurement_service.list_purchase_order_lines(
        FakeDb(),
        category=PurchaseRequestCategory.office.value,
        year=2026,
        month=6,
    )

    assert all_total == 2
    assert len(all_lines) == 1
    assert office_total == 1
    assert office_lines[0].category == PurchaseRequestCategory.office
    assert office_lines[0].category_label == "办公用品"


@pytest.mark.anyio
async def test_purchase_order_xlsx_export_uses_reference_layout() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            request_department="102一车间",
            product_name="碳鼓",
        ),
    )
    other_department = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            request_department="103车间",
            product_name="粉盒",
        ),
    )
    await _approve_request(created.id)
    await _approve_request(other_department.id)

    xlsx_bytes = await procurement_service.export_purchase_order_lines_xlsx(
        FakeDb(),
        category=PurchaseRequestCategory.hardware.value,
        year=2026,
        month=6,
    )

    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=False)
    worksheet = workbook.active
    merged_ranges = {str(cell_range) for cell_range in worksheet.merged_cells.ranges}

    assert worksheet.title == "Sheet1"
    assert "A2:L2" in merged_ranges
    assert "A3:L3" in merged_ranges
    assert worksheet["A2"].value == "丽珠集团（宁夏）制药有限公司"
    assert worksheet["A3"].value == "2026年06月份五金材料申购单汇总"
    assert worksheet["A4"].value == "申购部门：102一车间"
    assert [worksheet.cell(5, column).value for column in range(1, 13)] == (
        procurement_service.PURCHASE_ORDER_EXPORT_MATERIAL_HEADERS
    )

    assert [worksheet.cell(6, column).value for column in range(1, 13)] == [
        1,
        "MAT-1005",
        "打印机碳鼓",
        "M1005",
        "更换打印机碳鼓",
        None,
        "惠普",
        2,
        "个",
        60.005,
        120.01,
        "申购人：郭娇",
    ]
    assert worksheet["A7"].value == "合计"
    assert worksheet["K7"].value == "=SUM(K6:K6)"
    assert worksheet["A8"].value == "申购部门：103车间"
    assert worksheet["A12"].value == "总计"
    assert worksheet["K12"].value == "=SUM(K7,K11)"
    assert worksheet["A13"].value.startswith(" 总经理：")
    assert "A13:L13" in merged_ranges
    assert worksheet.column_dimensions["A"].width == 12
    assert worksheet.column_dimensions["B"].width == 18
    assert worksheet["A4"].fill.fgColor.rgb == "00D9E1F4"
    assert worksheet["A4"].alignment.wrap_text is not True
    assert worksheet["A5"].border.left.style == "thin"
    assert worksheet["L6"].alignment.wrap_text is True
    assert worksheet.page_setup.orientation == "landscape"


@pytest.mark.anyio
async def test_mixed_category_order_xlsx_uses_compatibility_columns() -> None:
    hardware = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(
            category=PurchaseRequestCategory.hardware,
            product_name="碳鼓",
        ),
    )
    office_payload = _create_payload_for(
        category=PurchaseRequestCategory.office,
        product_name="办公椅",
    )
    office_payload.items[0].material_code = ""
    office_payload.items[0].material_description = ""
    office_payload.items[0].rule_model = ""
    office = await procurement_service.create_purchase_request(FakeDb(), office_payload)
    await _approve_request(hardware.id)
    await _approve_request(office.id)

    xlsx_bytes = await procurement_service.export_purchase_order_lines_xlsx(
        FakeDb(),
        year=2026,
        month=6,
    )

    worksheet = load_workbook(BytesIO(xlsx_bytes), data_only=False).active
    assert [worksheet.cell(5, column).value for column in range(1, 13)] == (
        procurement_service.PURCHASE_ORDER_EXPORT_COMPATIBILITY_HEADERS
    )
    assert [worksheet.cell(6, column).value for column in range(2, 5)] == [
        "MAT-1005",
        "打印机碳鼓",
        "M1005",
    ]
    assert [worksheet.cell(7, column).value for column in range(2, 5)] == [
        "办公椅",
        "办公椅",
        "M1005",
    ]


@pytest.mark.anyio
async def test_material_field_category_requires_material_code_and_description() -> None:
    missing_code = _create_payload()
    missing_code.items[0].material_code = ""
    with pytest.raises(ValueError, match="缺少物料编码"):
        await procurement_service.create_purchase_request(FakeDb(), missing_code)

    missing_description = _create_payload()
    missing_description.items[0].material_description = ""
    with pytest.raises(ValueError, match="缺少物料说明"):
        await procurement_service.create_purchase_request(
            FakeDb(),
            missing_description,
        )


@pytest.mark.anyio
async def test_urgent_request_supports_mixed_categories_and_item_validation() -> None:
    payload = PurchaseRequestCreate(
        category=PurchaseRequestCategory.urgent,
        request_department="采购部",
        request_date=date(2026, 8, 12),
        attachment_note="加急技术附件",
        items=[
            PurchaseRequestItemInput(
                item_category=PurchaseRequestCategory.hardware,
                material_code="HW-001",
                material_description="不锈钢螺栓",
                quantity=Decimal("2"),
                unit="包",
                unit_price=Decimal("10"),
            ),
            PurchaseRequestItemInput(
                item_category=PurchaseRequestCategory.office,
                product_name="标签纸",
                specification="A4",
                quantity=Decimal("3"),
                unit="包",
                unit_price=Decimal("5.50"),
            ),
        ],
    )

    created = await procurement_service.create_purchase_request(FakeDb(), payload)

    assert created.category == PurchaseRequestCategory.urgent
    assert created.attachment_note == "加急技术附件"
    assert [item.item_category for item in created.items] == [
        PurchaseRequestCategory.hardware,
        PurchaseRequestCategory.office,
    ]
    assert created.total_amount == Decimal("36.50")

    missing_category = payload.model_copy(deep=True)
    missing_category.items[0].item_category = None
    with pytest.raises(ValueError, match="缺少申请类型"):
        await procurement_service.create_purchase_request(FakeDb(), missing_category)

    urgent_item_category = payload.model_copy(deep=True)
    urgent_item_category.items[0].item_category = PurchaseRequestCategory.urgent
    with pytest.raises(ValueError, match="申请类型无效"):
        await procurement_service.create_purchase_request(
            FakeDb(),
            urgent_item_category,
        )


@pytest.mark.anyio
async def test_urgent_order_export_includes_item_category_and_compatibility_columns(
) -> None:
    payload = PurchaseRequestCreate(
        category=PurchaseRequestCategory.urgent,
        request_department="采购部",
        request_date=date(2026, 8, 12),
        items=[
            PurchaseRequestItemInput(
                item_category=PurchaseRequestCategory.fire,
                material_code="FIRE-001",
                material_description="灭火器",
                quantity=Decimal("1"),
                unit="具",
                unit_price=Decimal("50"),
            )
        ],
    )
    created = await procurement_service.create_purchase_request(FakeDb(), payload)
    await _approve_request(created.id)

    xlsx_bytes = await procurement_service.export_purchase_order_lines_xlsx(
        FakeDb(),
        category=PurchaseRequestCategory.urgent.value,
        year=2026,
        month=8,
    )
    worksheet = load_workbook(BytesIO(xlsx_bytes), data_only=False).active

    assert [worksheet.cell(5, column).value for column in range(1, 14)] == (
        procurement_service.PURCHASE_ORDER_EXPORT_URGENT_HEADERS
    )
    assert worksheet.cell(6, 2).value == "消防"
    assert worksheet.cell(6, 3).value == "FIRE-001"


@pytest.mark.anyio
async def test_legacy_category_attachment_update() -> None:
    payload = _create_payload_for(category=PurchaseRequestCategory.office)
    payload.items[0].product_name = ""
    with pytest.raises(ValueError, match="缺少商品名称"):
        await procurement_service.create_purchase_request(FakeDb(), payload)

    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload_for(category=PurchaseRequestCategory.office),
    )
    updated = await procurement_service.update_purchase_request(
        FakeDb(),
        created.id,
        PurchaseRequestUpdate(attachment_note="更新后的附件说明"),
    )
    assert updated.attachment_note == "更新后的附件说明"


@pytest.mark.anyio
async def test_delete_purchase_request_removes_draft() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    request_id = created.id

    deleted = await procurement_service.delete_purchase_request(FakeDb(), request_id)

    assert deleted is True
    assert FakePurchaseRequestRepository.requests[request_id].is_deleted is True
    assert FakePurchaseRequestRepository.items.get(request_id) is None
    assert FakePurchaseRequestRepository.approvals.get(request_id) is None


@pytest.mark.anyio
async def test_delete_purchase_request_rejects_non_draft() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    FakePurchaseRequestRepository.requests[
        created.id
    ].status = PurchaseRequestStatus.pending_department_head.value

    with pytest.raises(ValueError, match="仅草稿状态的采购申请可以删除"):
        await procurement_service.delete_purchase_request(FakeDb(), created.id)


@pytest.mark.anyio
async def test_delete_purchase_request_missing_raises_error() -> None:
    with pytest.raises(ValueError, match="采购申请不存在"):
        await procurement_service.delete_purchase_request(
            FakeDb(),
            uuid.uuid4(),
        )


@pytest.mark.anyio
async def test_submit_purchase_request_rejects_total_amount_mismatch() -> None:
    created = await procurement_service.create_purchase_request(
        FakeDb(),
        _create_payload(),
    )
    FakePurchaseRequestRepository.items[created.id][0].total_amount = Decimal("999.99")

    with pytest.raises(ValueError, match="与数量×单价.*不一致"):
        await procurement_service.submit_purchase_request(FakeDb(), created.id)

    request = FakePurchaseRequestRepository.requests[created.id]
    assert request.status == PurchaseRequestStatus.draft
