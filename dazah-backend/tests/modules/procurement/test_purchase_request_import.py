"""采购申请表格导入（xlsx/xls/csv）解析与导入测试。"""

from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]

from app.modules.procurement import service as procurement_service
from app.modules.procurement.schemas import PurchaseRequestCategory
from tests.modules.procurement.test_purchase_requests import (
    FakeDb,
    FakePurchaseRequestRepository,
)

FIXTURES_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture(autouse=True)
def fake_purchase_request_repository(monkeypatch: Any) -> Any:
    FakePurchaseRequestRepository.reset()
    monkeypatch.setattr(
        procurement_service,
        "PurchaseRequestRepository",
        FakePurchaseRequestRepository,
    )


def _build_xlsx(sheets: list[tuple[str, list[list[object]]]]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, rows in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


HARDWARE_HEADERS = [
    "申购部门",
    "申请日期",
    "采购类型",
    "物料编码",
    "物料说明",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价",
    "备注",
]


def _hardware_row(
    *,
    department: str = "102一车间",
    material_code: str = "HW-001",
    material_description: str = "不锈钢螺栓",
    quantity: object = 200,
    unit_price: object = 1.5,
) -> list[object]:
    return [
        department,
        "2026-08-14",
        "五金材料",
        material_code,
        material_description,
        "M12",
        "设备检修",
        "304不锈钢",
        "固力",
        quantity,
        "个",
        unit_price,
        "急用",
    ]


@pytest.mark.anyio
async def test_import_xlsx_multi_sheet_creates_one_draft_per_sheet() -> None:
    file_bytes = _build_xlsx(
        [
            (
                "五金材料",
                [
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    _hardware_row(),
                    _hardware_row(material_code="HW-002"),
                ],
            ),
            (
                "电气",
                [
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        "设备动力部",
                        "2026/08/14",
                        "电气",
                        "EL-001",
                        "接触器",
                        "CJX2-3210",
                        "配电柜改造",
                        "",
                        "正泰",
                        10,
                        "只",
                        45,
                        "",
                    ],
                ],
            ),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="采购申请.xlsx",
    )

    assert result.total_sheets == 2
    assert result.failed_rows == []
    assert len(result.imported_requests) == 2

    hardware_summary = result.imported_requests[0]
    assert hardware_summary.sheet_name == "五金材料"
    assert hardware_summary.category == PurchaseRequestCategory.hardware
    assert hardware_summary.category_label == "五金材料"
    assert hardware_summary.category_source == "column"
    assert hardware_summary.request_department == "102一车间"
    assert hardware_summary.request_date == date(2026, 8, 14)
    assert hardware_summary.items_count == 2

    electrical_summary = result.imported_requests[1]
    assert electrical_summary.category == PurchaseRequestCategory.electrical
    assert electrical_summary.category_source == "column"
    assert electrical_summary.request_department == "设备动力部"
    assert electrical_summary.request_date == date(2026, 8, 14)
    assert electrical_summary.items_count == 1

    hardware_request = FakePurchaseRequestRepository.requests[
        hardware_summary.request_id
    ]
    assert hardware_request.category == "hardware"
    assert hardware_request.status == "draft"
    assert hardware_request.attachment_note == "通过表格导入（五金材料）"
    assert hardware_request.total_amount == Decimal("600.00")
    hardware_items = FakePurchaseRequestRepository.items[hardware_summary.request_id]
    assert hardware_items[0].material_code == "HW-001"
    assert hardware_items[0].quantity == Decimal("200")
    assert hardware_items[0].unit_price == Decimal("1.5")
    assert hardware_items[1].sequence == 2


@pytest.mark.anyio
async def test_import_category_column_wins_over_sheet_name() -> None:
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [HARDWARE_HEADERS, _hardware_row()],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].category == PurchaseRequestCategory.hardware


@pytest.mark.anyio
async def test_import_mixed_category_column_reports_sheet_error() -> None:
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    _hardware_row(),
                    [
                        "102一车间",
                        "2026-08-14",
                        "电气",
                        "EL-001",
                        "接触器",
                        "",
                        "",
                        "",
                        "",
                        2,
                        "只",
                        10,
                        "",
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    assert result.failed_rows[0].sheet_name == "Sheet1"
    assert result.failed_rows[0].row is None
    assert "采购类型列包含多种类型" in result.failed_rows[0].message


@pytest.mark.anyio
async def test_import_infers_hardware_from_material_fields() -> None:
    headers = list(HARDWARE_HEADERS)
    headers.remove("采购类型")
    row = _hardware_row()
    row.pop(2)
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [headers, row],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.hardware
    assert summary.category_source == "inferred"
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_infers_office_from_product_name() -> None:
    headers = [
        "申购部门",
        "申请日期",
        "商品名称",
        "规格",
        "数量",
        "单位",
        "单价",
        "备注",
    ]
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    headers,  # type: ignore[list-item]
                    ["行政部", "2026-08-14", "打印纸", "A4", 5, "箱", 20, ""],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.office
    assert summary.category_source == "inferred"
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_infers_category_from_uniform_item_category_column() -> None:
    headers = [
        "申购部门",
        "申请日期",
        "申请类型",
        "商品名称",
        "规格",
        "数量",
        "单位",
        "单价",
        "备注",
    ]
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    headers,  # type: ignore[list-item]
                    [
                        "行政部",
                        "2026-08-14",
                        "办公用品",
                        "打印纸",
                        "A4",
                        5,
                        "箱",
                        20,
                        "",
                    ],
                    [
                        "行政部",
                        "2026-08-14",
                        "办公用品",
                        "签字笔",
                        "0.5mm",
                        50,
                        "支",
                        2,
                        "",
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.office
    assert summary.category_source == "inferred"
    assert summary.items_count == 2
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_sheet_without_recognizable_fields_reports_error() -> None:
    headers = ["申购部门", "数量", "单位", "单价"]
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    headers,  # type: ignore[list-item]
                    ["102一车间", 10, "个", 5],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    assert result.failed_rows[0].row is None
    assert "无法识别采购类型" in result.failed_rows[0].message


@pytest.mark.anyio
async def test_import_any_row_error_blocks_whole_sheet() -> None:
    """任一明细错误（缺编码/数量无效/数量为负）→ 整张工作表不导入。"""
    file_bytes = _build_xlsx(
        [
            (
                "五金材料",
                [
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    _hardware_row(),
                    _hardware_row(material_code="", material_description="缺编码行"),
                    _hardware_row(quantity="abc", material_code="HW-003"),
                    _hardware_row(quantity=-1, material_code="HW-005"),
                    _hardware_row(material_code="HW-004"),
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    error = result.failed_rows[0]
    assert error.row is None
    assert "整张工作表未导入" in error.message
    assert "请修改后重新提交" in error.message
    assert "缺少物料编码" in error.message
    assert "数量无效" in error.message
    assert "数量不能为负数" in error.message


@pytest.mark.anyio
async def test_import_urgent_sheet_with_item_category_per_row() -> None:
    urgent_headers = [
        "申购部门",
        "申请日期",
        "申请类型",
        "商品名称",
        "规格",
        "用途",
        "数量",
        "单位",
        "单价",
        "备注",
    ]
    file_bytes = _build_xlsx(
        [
            (
                "加急单",
                [
                    urgent_headers,  # type: ignore[list-item]
                    [
                        "采购部",
                        "2026-08-14",
                        "电脑材料",
                        "键盘",
                        "USB",
                        "办公",
                        3,
                        "个",
                        30,
                        "",
                    ],
                    [
                        "采购部",
                        "2026-08-14",
                        "办公用品",
                        "打印纸",
                        "A4",
                        "办公",
                        5,
                        "箱",
                        20,
                        "",
                    ],
                    [
                        "采购部",
                        "2026-08-14",
                        "",
                        "标签纸",
                        "A4",
                        "办公",
                        2,
                        "卷",
                        8,
                        "",
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="加急申请.xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    error = result.failed_rows[0]
    assert error.row is None
    assert "缺少申请类型" in error.message
    assert "整张工作表未导入" in error.message


@pytest.mark.anyio
async def test_import_normal_sheet_item_category_mismatch_blocks_whole_sheet() -> None:
    headers = HARDWARE_HEADERS + ["申请类型"]
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    headers,  # type: ignore[list-item]
                    _hardware_row() + ["办公用品"],
                    _hardware_row(material_code="HW-002") + [""],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    error = result.failed_rows[0]
    assert error.row is None
    assert "与采购分类不一致" in error.message
    assert "整张工作表未导入" in error.message


@pytest.mark.anyio
async def test_import_csv_supports_utf8_and_gb18030() -> None:
    csv_text = (
        "申购部门,申请日期,采购类型,商品名称,规格,数量,单位,单价,备注\n"
        "行政部,2026-08-14,办公用品,打印纸,A4,5,箱,20,加急\n"
        "行政部,2026-08-14,办公用品,签字笔,0.5mm,50,支,2,\n"
    )
    for encoding in ("utf-8-sig", "gb18030"):
        file_bytes = csv_text.encode(encoding)

        result = await procurement_service.import_purchase_request_table_file(
            cast(Any, FakeDb)(),
            file_bytes,
            file_name="办公用品.csv",
        )

        assert len(result.imported_requests) == 1, encoding
        summary = result.imported_requests[0]
        assert summary.category == PurchaseRequestCategory.office
        assert summary.category_source == "column"
        assert summary.items_count == 2
        assert result.failed_rows == [], encoding


@pytest.mark.anyio
async def test_import_xls_sheet_with_date_cells() -> None:
    fixture_path = FIXTURES_DIR / "purchase_requests_sample.xls"
    file_bytes = fixture_path.read_bytes()

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="采购申请.xls",
    )

    assert len(result.imported_requests) == 2
    hardware_summary = result.imported_requests[0]
    assert hardware_summary.category == PurchaseRequestCategory.hardware
    assert hardware_summary.category_source == "sheet_name"
    assert hardware_summary.request_date == date(2026, 8, 14)
    assert hardware_summary.request_department == "102一车间"
    assert hardware_summary.items_count == 2

    electrical_summary = result.imported_requests[1]
    assert electrical_summary.category == PurchaseRequestCategory.electrical
    assert electrical_summary.category_source == "sheet_name"
    assert electrical_summary.request_date == date(2026, 8, 14)

    hardware_request = FakePurchaseRequestRepository.requests[
        hardware_summary.request_id
    ]
    assert hardware_request.total_amount == Decimal("444.00")


@pytest.mark.anyio
async def test_import_number_parsing_currency_and_thousands() -> None:
    file_bytes = _build_xlsx(
        [
            (
                "Sheet1",
                [
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    _hardware_row(quantity="1,200", unit_price="¥50.5"),
                    _hardware_row(quantity=3, unit_price=""),
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.items_count == 2
    assert result.failed_rows == []

    items = FakePurchaseRequestRepository.items[summary.request_id]
    assert items[0].quantity == Decimal("1200")
    assert items[0].unit_price == Decimal("50.5")
    assert items[1].unit_price == Decimal("0")


@pytest.mark.anyio
async def test_import_sheet_name_contains_category_keyword() -> None:
    headers = list(HARDWARE_HEADERS)
    headers.remove("采购类型")
    row = _hardware_row()
    row.pop(2)
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸电气",
                [headers, row],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.electrical
    assert summary.category_source == "sheet_name"
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_real_world_sheet_with_category_column_no_department() -> None:
    headers = [
        "采购类型",
        "申请日期",
        "物料编码",
        "物料说明",
        "规格型号",
        "用途",
        "数量",
        "单位",
        "单价",
    ]
    file_bytes = _build_xlsx(
        [
            (
                "MC原辅料",
                [
                    headers,  # type: ignore[list-item]
                    [
                        "原辅料",
                        "2026-08-14",
                        "RA-001",
                        "葡萄糖",
                        "食品级",
                        "生产投料",
                        500,
                        "kg",
                        4.2,
                    ],
                    [
                        "原辅料",
                        "2026-08-14",
                        "RA-002",
                        "柠檬酸",
                        "食品级",
                        "生产投料",
                        200,
                        "kg",
                        6.5,
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="采购申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.raw_auxiliary
    assert summary.category_source == "column"
    assert summary.request_department == "MC"
    assert summary.items_count == 2
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_missing_department_falls_back_to_sheet_name_prefix() -> None:
    headers = list(HARDWARE_HEADERS)
    headers.remove("申购部门")
    headers.remove("采购类型")
    row = _hardware_row()
    row.pop(2)
    row.pop(0)
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [headers, row],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.hardware
    assert summary.category_source == "sheet_name"
    assert summary.request_department == "霉酚酸"
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_missing_department_without_keyword_uses_placeholder() -> None:
    headers = list(HARDWARE_HEADERS)
    headers.remove("申购部门")
    headers.remove("采购类型")
    row = _hardware_row()
    row.pop(2)
    row.pop(0)
    file_bytes = _build_xlsx(
        [
            (
                "五金材料",
                [headers, row],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].request_department == "未填写"


@pytest.mark.anyio
async def test_import_csv_missing_department_uses_file_stem() -> None:
    csv_text = "采购类型,商品名称,规格,数量,单位,单价\n办公用品,打印纸,A4,5,箱,20\n"
    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        csv_text.encode("utf-8-sig"),
        file_name="霉酚酸办公采购.csv",
    )

    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.office
    assert summary.request_department == "霉酚酸办公采购"
    assert result.failed_rows == []


@pytest.mark.anyio
async def test_import_request_date_defaults_to_today() -> None:
    headers = list(HARDWARE_HEADERS)
    headers.remove("申请日期")
    row = _hardware_row()
    row.pop(1)
    file_bytes = _build_xlsx(
        [
            (
                "五金材料",
                [headers, row],  # type: ignore[list-item]
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].request_date == datetime.now(UTC).date()


@pytest.mark.anyio
async def test_import_skips_blank_sheets_and_reports_header_only_sheet() -> None:
    workbook = Workbook()
    hardware_sheet = workbook.active
    hardware_sheet.title = "五金材料"
    for row in [HARDWARE_HEADERS, _hardware_row()]:
        hardware_sheet.append(row)
    workbook.create_sheet("空表")
    header_only = workbook.create_sheet("只有表头")
    header_only.append(HARDWARE_HEADERS)
    output = BytesIO()
    workbook.save(output)

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        output.getvalue(),
        file_name="申请.xlsx",
    )

    assert result.total_sheets == 2
    assert len(result.imported_requests) == 1
    assert len(result.failed_rows) == 1
    assert result.failed_rows[0].sheet_name == "只有表头"
    assert "没有可导入的数据行" in result.failed_rows[0].message


@pytest.mark.anyio
async def test_import_rejects_empty_file_and_unsupported_extension() -> None:
    with pytest.raises(ValueError, match="上传文件为空"):
        await procurement_service.import_purchase_request_table_file(
            cast(Any, FakeDb)(),
            b"",
            file_name="申请.xlsx",
        )

    with pytest.raises(ValueError, match="暂不支持该文件类型"):
        await procurement_service.import_purchase_request_table_file(
            cast(Any, FakeDb)(),
            b"abc",
            file_name="申请.docx",
        )


@pytest.mark.anyio
async def test_import_xlsx_date_cell_for_request_date() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "五金材料"
    sheet.append(HARDWARE_HEADERS)
    sheet.append(
        [
            "102一车间",
            datetime(2026, 8, 14, 9, 30),
            "五金材料",
            "HW-001",
            "不锈钢螺栓",
            "M12",
            "设备检修",
            "304不锈钢",
            "固力",
            200,
            "个",
            1.5,
            "",
        ]
    )
    output = BytesIO()
    workbook.save(output)

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        output.getvalue(),
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].request_date == date(2026, 8, 14)


@pytest.mark.anyio
async def test_import_sheet_failure_does_not_block_other_sheets() -> None:
    file_bytes = _build_xlsx(
        [
            ("五金材料", [HARDWARE_HEADERS, _hardware_row()]),  # type: ignore[list-item]
            ("Sheet1", [["只有一列", "数据"]]),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert len(result.imported_requests) == 1
    assert len(result.failed_rows) == 1
    assert result.failed_rows[0].sheet_name == "Sheet1"


# ---- 车间计划表（2026.9月201二车间计划）真实表格结构回归测试 ----

_TITLE_ROW = ["   丽珠集团（宁夏）制药有限公司"]
_TITLE_ROW_WITH_AVERAGE = _TITLE_ROW + [None] * 11 + ["月均：53337"]

_HARDWARE_HEADERS = [
    "序号",
    "物料编码",
    "物料描述",
    "规格型号",
    "单位",
    "请购数量",
    "预估单价",
    "预估总价",
    "用途",
    "材质",
    "品牌",
    "备注",
]
_ELECTRICAL_HEADERS = [
    "序号",
    "物料编码",
    "物料描述",
    "规格型号",
    "单位",
    "请购数量",
    "预估单价（元）",
    "预估总价（元）",
    "用途",
    "材质",
    "品牌",
    "备注",
]
_RAW_MATERIAL_HEADERS = [
    "序号",
    "物料编码",
    "商品名称",
    "规格",
    "单位",
    "需求数量(kg)",
    "用途",
    "供应商要求",
    "质量要求/到货要求",
    "备注",
]
_PACKAGING_HEADERS = [
    "序号",
    "物料代码",
    "商品名称",
    "规格",
    "单位",
    "单批用量",
    "月生产产量",
    "月需求数量",
    "用途",
    "供应商要求",
    "质量要求/到货要求",
    "备注",
]
_DR_PACKAGING_HEADERS = [
    "序号",
    "物料代码",
    "商品名称",
    "规格",
    "单位",
    "单批用量",
    "月生产产量",
    "1个月需求数量",
    "用途",
    "供应商要求",
    "质量要求/到货要求",
    "备注",
]
_OFFICE_HEADERS = _HARDWARE_HEADERS
_SIGNATURE_ROW = [
    "副总经理： 主管领导： 设备动力部： 部门负责人： 五金库： 统计人："
] + [None] * 11


@pytest.mark.anyio
async def test_import_workshop_title_rows_not_mistaken_for_headers() -> None:
    """标题行（公司名+月均分散两格）不应被识别为表头；五金/电气表头别名可用。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "酸化打料泵机封更换",
                        "",
                        "",
                        "杨冬15009523102",
                    ],
                    [
                        2,
                        51065615,
                        "机封,VIC50-160",
                        "VIC50-160",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                    [3, None, "", "", "", "", "", "", "", "", "", ""],  # 仅序号
                    [
                        "合计",
                        None,
                        None,
                        4104,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    _SIGNATURE_ROW,  # type: ignore[list-item]
                ],
            ),
            (
                "霉酚酸电气",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（电气）申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _ELECTRICAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51132617,
                        "不锈钢耐震压力表",
                        "量程0-1Mpa",
                        "块",
                        3,
                        167,
                        501,
                        "压滤打料泵压力表坏",
                        "",
                        "",
                        "张彩红18295228807",
                    ],
                    [
                        "合计",
                        None,
                        None,
                        501,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    _SIGNATURE_ROW,  # type: ignore[list-item]
                ],
            ),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.total_sheets == 2
    assert result.failed_rows == []
    assert len(result.imported_requests) == 2

    hardware_summary = result.imported_requests[0]
    assert hardware_summary.sheet_name == "霉酚酸五金"
    assert hardware_summary.category == PurchaseRequestCategory.hardware
    assert hardware_summary.category_source == "sheet_name"
    assert hardware_summary.request_department == "201二车间(霉酚酸)"
    assert hardware_summary.items_count == 2
    hardware_items = FakePurchaseRequestRepository.items[hardware_summary.request_id]
    assert hardware_items[0].material_code == "51040070"
    assert hardware_items[0].quantity == Decimal("4")
    assert hardware_items[0].unit_price == Decimal("513")

    electrical_summary = result.imported_requests[1]
    assert electrical_summary.category == PurchaseRequestCategory.electrical
    assert electrical_summary.items_count == 1
    electrical_items = FakePurchaseRequestRepository.items[
        electrical_summary.request_id
    ]
    assert electrical_items[0].unit_price == Decimal("167")


@pytest.mark.anyio
async def test_import_raw_material_and_packaging_workshop_headers() -> None:
    """原辅料/包材表头：物料代码、商品名称回退物料说明、需求数量类别名。"""
    file_bytes = _build_xlsx(
        [
            (
                "MC原辅料",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月份 原辅料申购单"],
                    ["申购部门：201二车间（霉酚酸）"],
                    _RAW_MATERIAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        12002413,
                        "硫酸,kg,槽车",
                        "98%、槽车",
                        "kg",
                        64000,
                        "岗位生产使用",
                        "",
                        "通知到货",
                        "",
                    ],
                    [
                        2,
                        11003003,
                        "液碱,kg,槽车",
                        "32%、槽车",
                        "kg",
                        198000,
                        "",
                        "",
                        "",
                        "",
                    ],
                    [
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        ".",
                    ],  # 占位行
                ],
            ),
            (
                "MC包材",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月份 霉酚酸包材 申购单"],
                    ["申请部门：201二车间"],
                    _PACKAGING_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        60071312,
                        "聚乙烯袋",
                        "0.1*600*850",
                        "条",
                        "",
                        35000,
                        3800,
                        "包装使用",
                        "",
                        "",
                        "",
                    ],
                    [
                        "合计",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                ],
            ),
            (
                "DR包材",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月份 多拉包材 申购单"],
                    ["申请部门：201二车间"],
                    _DR_PACKAGING_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        60071379,
                        "药用低密度聚乙烯袋（内外袋）",
                        "内袋：590*950*0.05",
                        "条",
                        "",
                        "1000kg",
                        200,
                        "高规GB 5kg成品包装使用",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.failed_rows == []
    assert len(result.imported_requests) == 3
    by_sheet = {summary.sheet_name: summary for summary in result.imported_requests}

    raw_summary = by_sheet["MC原辅料"]
    assert raw_summary.category == PurchaseRequestCategory.raw_auxiliary
    assert raw_summary.request_department == "201二车间（霉酚酸）"
    assert raw_summary.items_count == 2


@pytest.mark.anyio
async def test_import_office_sheet_uses_material_description_column() -> None:
    """其他商品表头用“物料描述”代替“商品名称”，数量/单价别名同样适用。"""
    file_bytes = _build_xlsx(
        [
            (
                "其他商品",
                [
                    ["上报时请备注上报人，以便核对信息"],
                    ["2026年其它申购单"],
                    ["   申请部门:201-2车间"],
                    _OFFICE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        "",
                        "安全警示标识贴纸",
                        "30*40cm",
                        "张",
                        80,
                        5,
                        400,
                        "更褪色老化安全标识",
                        "（可易撕、不留痕）背胶+内容+覆膜",
                        "",
                        "姚淇淋",
                    ],
                    [
                        2,
                        85001645,
                        "防化服,耐酸碱",
                        "XL：4套，XXL：2套",
                        "套",
                        6,
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                    ],
                    [
                        "总计（元）",
                        None,
                        None,
                        400,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        " 分管领导： 主管领导： 部门负责人： 五金库： 统计人： ",
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.failed_rows == []
    assert len(result.imported_requests) == 1
    summary = result.imported_requests[0]
    assert summary.category == PurchaseRequestCategory.office
    assert summary.category_source == "sheet_name"
    assert summary.items_count == 2

    items = FakePurchaseRequestRepository.items[summary.request_id]
    assert items[0].product_name == "安全警示标识贴纸"
    assert items[0].quantity == Decimal("80")
    assert items[0].unit_price == Decimal("5")
    assert items[1].product_name == "防化服,耐酸碱"
    assert items[1].material_code == "85001645"


@pytest.mark.anyio
async def test_import_workshop_sheets_without_data_report_sheet_error() -> None:
    """无采购明细的工作表（只有表头/仅序号行）报工作表级错误，不阻断其他表。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
            (
                "MC化玻",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    ["2026年9月（化玻试剂）申购单"],
                    ["  申请部门:201二车间（霉酚酸）"],
                    _ELECTRICAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        "总计（元）",
                        None,
                        None,
                        0,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    _SIGNATURE_ROW,  # type: ignore[list-item]
                ],
            ),
            (
                "DR消防",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月（消防）正常申购单"],
                    ["申请部门:201二车间(多拉)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        2,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                    [
                        "合计",
                        None,
                        None,
                        0,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                ],
            ),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].sheet_name == "霉酚酸五金"
    assert {error.sheet_name for error in result.failed_rows} == {"MC化玻", "DR消防"}
    assert all("没有可导入的数据行" in error.message for error in result.failed_rows)


@pytest.mark.anyio
async def test_import_department_from_title_row() -> None:
    """申购部门取自标题行“申请部门/申购部门：xxx”，半角/全角冒号均可。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
            (
                "MC原辅料",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月份 原辅料申购单"],
                    ["申购部门：201二车间（霉酚酸）"],
                    _RAW_MATERIAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        12002413,
                        "硫酸,kg,槽车",
                        "98%、槽车",
                        "kg",
                        64000,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
            (
                "多拉化玻",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    ["2026年9月（化玻试剂）申购单"],
                    ["申请部门:201二车间（多拉菌素）"],
                    _ELECTRICAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        80005889,
                        "红水温度计,0-100℃",
                        "0-100℃",
                        "支",
                        20,
                        10,
                        200,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.failed_rows == []
    by_sheet = {summary.sheet_name: summary for summary in result.imported_requests}
    assert by_sheet["霉酚酸五金"].request_department == "201二车间(霉酚酸)"
    assert by_sheet["MC原辅料"].request_department == "201二车间（霉酚酸）"
    assert by_sheet["多拉化玻"].request_department == "201二车间（多拉菌素）"


@pytest.mark.anyio
async def test_import_column_department_wins_over_title_row() -> None:
    """表内“申请部门”列有值时优先取列值，标题行仅作回退。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    HARDWARE_HEADERS,  # type: ignore[list-item]
                    _hardware_row(department="102一车间"),
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.failed_rows == []
    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].request_department == "102一车间"


@pytest.mark.anyio
async def test_import_validates_total_amount_against_quantity_price() -> None:
    """任一明细预估总价与数量×单价不一致 → 整张工作表不导入。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                    [
                        2,
                        51031459,
                        "四氟盘根,φ12",
                        "φ12",
                        "卷",
                        2,
                        300,
                        300,
                        "",
                        "",
                        "",
                        "",
                    ],
                    [
                        3,
                        51146151,
                        "加重型六棱撬棍",
                        "28*1200mm",
                        "根",
                        1,
                        80,
                        160,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    error = result.failed_rows[0]
    assert error.row is None
    assert "整张工作表未导入" in error.message
    assert "请修改后重新提交" in error.message
    assert "预估总价（300）与数量×单价（600）不一致" in error.message
    assert "预估总价（160）与数量×单价（80）不一致" in error.message


@pytest.mark.anyio
async def test_import_skips_total_amount_check_without_unit_price() -> None:
    """表格有预估总价但无单价列时跳过校验，避免无法核对而误报。"""
    headers = ["序号", "物料编码", "物料说明", "数量", "单位", "预估总价"]
    file_bytes = _build_xlsx(
        [
            (
                "五金材料",
                [
                    headers,  # type: ignore[list-item]
                    [1, "HW-001", "不锈钢螺栓", 200, "个", 1000],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )

    assert result.failed_rows == []
    assert len(result.imported_requests) == 1
    assert result.imported_requests[0].items_count == 1


@pytest.mark.anyio
async def test_import_electrical_total_amount_header_with_parentheses() -> None:
    """电气表头“预估总价（元）”归一化后可识别并校验。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸电气",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（电气）申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _ELECTRICAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51132617,
                        "不锈钢耐震压力表",
                        "量程0-1Mpa",
                        "块",
                        3,
                        167,
                        501,
                        "",
                        "",
                        "",
                        "",
                    ],
                    [
                        2,
                        "51142112",
                        "LED灯管,T8-1.2米/ 36W",
                        "T8-1.2米/ 36W",
                        "根",
                        10,
                        30,
                        400,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            )
        ]
    )

    result = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )

    assert result.imported_requests == []
    assert len(result.failed_rows) == 1
    error = result.failed_rows[0]
    assert error.row is None
    assert "整张工作表未导入" in error.message
    assert "预估总价（400）与数量×单价（300）不一致" in error.message


@pytest.mark.anyio
async def test_import_same_file_twice_skips_duplicate_sheets() -> None:
    """同一文件（内容哈希相同）重复导入时整张工作表跳过，不生成重复草稿。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
            (
                "MC原辅料",
                [
                    _TITLE_ROW,  # type: ignore[list-item]
                    ["2026年9月份 原辅料申购单"],
                    ["申购部门：201二车间（霉酚酸）"],
                    _RAW_MATERIAL_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        12002413,
                        "硫酸,kg,槽车",
                        "98%、槽车",
                        "kg",
                        64000,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            ),
        ]
    )

    first = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )
    assert len(first.imported_requests) == 2
    assert first.failed_rows == []

    second = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="2026.9月201二车间计划(1).xlsx",
    )
    assert second.imported_requests == []
    assert len(second.failed_rows) == 2
    assert all(
        error.row is None and "已导入过" in error.message
        for error in second.failed_rows
    )


@pytest.mark.anyio
async def test_import_modified_file_allows_reimport() -> None:
    """修正表格内容（哈希变化）后可重新导入，对应“修改后重新提交”场景。"""

    def build(quantity: int) -> bytes:
        return _build_xlsx(
            [
                (
                    "霉酚酸五金",
                    [
                        _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                        [" 2026年9月（五金）正常申购单"],
                        [" 申请部门:201二车间(霉酚酸)"],
                        _HARDWARE_HEADERS,  # type: ignore[list-item]
                        [
                            1,
                            51040070,
                            "机封,VIC50-250",
                            "VIC50-250",
                            "套",
                            quantity,
                            513,
                            513 * quantity,
                            "",
                            "",
                            "",
                            "",
                        ],
                    ],
                )
            ]
        )

    first = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        build(4),
        file_name="2026.9月201二车间计划(1).xlsx",
    )
    assert len(first.imported_requests) == 1

    second = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        build(8),
        file_name="2026.9月201二车间计划(1).xlsx",
    )
    assert len(second.imported_requests) == 1
    assert second.imported_requests[0].items_count == 1
    assert second.failed_rows == []


@pytest.mark.anyio
async def test_import_same_file_after_soft_delete_allows_reimport() -> None:
    """同一文件导入后软删除草稿，再次导入同一文件应允许（查重排除已删除）。"""
    file_bytes = _build_xlsx(
        [
            (
                "霉酚酸五金",
                [
                    _TITLE_ROW_WITH_AVERAGE,  # type: ignore[list-item]
                    [" 2026年9月（五金）正常申购单"],
                    [" 申请部门:201二车间(霉酚酸)"],
                    _HARDWARE_HEADERS,  # type: ignore[list-item]
                    [
                        1,
                        51040070,
                        "机封,VIC50-250",
                        "VIC50-250",
                        "套",
                        4,
                        513,
                        2052,
                        "",
                        "",
                        "",
                        "",
                    ],
                ],
            )
        ]
    )

    first = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )
    assert len(first.imported_requests) == 1
    deleted_request = FakePurchaseRequestRepository.requests[
        first.imported_requests[0].request_id
    ]
    deleted_request.is_deleted = True

    second = await procurement_service.import_purchase_request_table_file(
        cast(Any, FakeDb)(),
        file_bytes,
        file_name="申请.xlsx",
    )
    assert len(second.imported_requests) == 1
    assert second.failed_rows == []
