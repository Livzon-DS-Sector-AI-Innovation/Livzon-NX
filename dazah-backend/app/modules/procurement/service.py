"""Procurement business workflows live here."""

import csv
import re
from collections.abc import Sequence
from datetime import UTC, date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

import xlrd  # type: ignore[import-untyped]
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.page import PageMargins  # type: ignore[import-untyped]
from openpyxl.worksheet.properties import (  # type: ignore[import-untyped]
    PageSetupProperties,
)
from pydantic import ValidationError
from pypdf import PdfReader
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.storage import get_object, upload_object
from app.core.storage import is_enabled as minio_enabled
from app.modules.procurement.contract_generator import generate_contract
from app.modules.procurement.models import (
    ContractRecord,
    InvoiceRecognitionRecord,
    PurchaseRequest,
    PurchaseRequestApproval,
    PurchaseRequestItem,
    Supplier,
)
from app.modules.procurement.repository import (
    ContractRecordRepository,
    InvoiceRecognitionRepository,
    PurchaseRequestRepository,
    SupplierRepository,
)
from app.modules.procurement.schemas import (
    MATERIAL_FIELD_PURCHASE_CATEGORIES,
    NORMAL_PURCHASE_CATEGORIES,
    ContractGenerateRequest,
    InvoiceLineItem,
    InvoiceRecognitionResult,
    PurchaseApprovalRequest,
    PurchaseApprovalResult,
    PurchaseApprovalRole,
    PurchaseApprovalView,
    PurchaseOrderLineResponse,
    PurchaseRequestCategory,
    PurchaseRequestCreate,
    PurchaseRequestImportError,
    PurchaseRequestImportResult,
    PurchaseRequestImportSummary,
    PurchaseRequestItemInput,
    PurchaseRequestResponse,
    PurchaseRequestStatus,
    PurchaseRequestUpdate,
    SupplierImportResult,
    SupplierResponse,
)

MONEY_PATTERN = r"[¥￥]?\s*([0-9]+(?:\.[0-9]+)?)"
NUMBER_PATTERN = r"[0-9]+(?:\.[0-9]+)?"
MONEY_QUANT = Decimal("0.01")

APPROVAL_ROLE_TO_PENDING_STATUS = {
    PurchaseApprovalRole.hardware_warehouse: (
        PurchaseRequestStatus.pending_hardware_warehouse
    ),
    PurchaseApprovalRole.equipment_power: PurchaseRequestStatus.pending_equipment_power,
    PurchaseApprovalRole.safety_officer: PurchaseRequestStatus.pending_safety_officer,
    PurchaseApprovalRole.department_head: PurchaseRequestStatus.pending_department_head,
    PurchaseApprovalRole.responsible_leader: (
        PurchaseRequestStatus.pending_responsible_leader
    ),
    PurchaseApprovalRole.supervising_leader: (
        PurchaseRequestStatus.pending_supervising_leader
    ),
    PurchaseApprovalRole.finance_director: (
        PurchaseRequestStatus.pending_finance_director
    ),
    PurchaseApprovalRole.general_manager: (
        PurchaseRequestStatus.pending_general_manager
    ),
}

DEFAULT_PURCHASE_APPROVAL_WORKFLOW = (
    PurchaseApprovalRole.department_head,
    PurchaseApprovalRole.responsible_leader,
    PurchaseApprovalRole.supervising_leader,
)

PURCHASE_APPROVAL_WORKFLOWS = {
    "hardware": (
        PurchaseApprovalRole.hardware_warehouse,
        PurchaseApprovalRole.department_head,
        PurchaseApprovalRole.responsible_leader,
        PurchaseApprovalRole.supervising_leader,
        PurchaseApprovalRole.general_manager,
    ),
    "electrical": (
        PurchaseApprovalRole.hardware_warehouse,
        PurchaseApprovalRole.equipment_power,
        PurchaseApprovalRole.department_head,
        PurchaseApprovalRole.responsible_leader,
        PurchaseApprovalRole.supervising_leader,
    ),
    "labor-special": (
        PurchaseApprovalRole.safety_officer,
        PurchaseApprovalRole.department_head,
        PurchaseApprovalRole.responsible_leader,
    ),
    "urgent": (
        PurchaseApprovalRole.hardware_warehouse,
        PurchaseApprovalRole.department_head,
        PurchaseApprovalRole.responsible_leader,
        PurchaseApprovalRole.supervising_leader,
        PurchaseApprovalRole.finance_director,
        PurchaseApprovalRole.general_manager,
    ),
}

PURCHASE_APPROVAL_REQUIRED_COUNTS = {
    PurchaseApprovalRole.equipment_power: 2,
}

PURCHASE_CATEGORY_LABELS = {
    "hardware": "五金材料",
    "computer": "电脑材料",
    "office": "办公用品",
    "raw-auxiliary": "原辅料",
    "chemical-glass": "化玻",
    "electrical": "电气",
    "advertising-printing": "广告/印刷",
    "fire": "消防",
    "packaging": "包材",
    "labor-special": "特防",
    "labor-miscellaneous": "杂品",
    "urgent": "加急单",
}

SUPPORTED_PURCHASE_REQUEST_IMPORT_EXTENSIONS = {".xlsx", ".xls", ".csv"}

# 导入时用于识别采购类型的中文名称/别名（键在查找时经 _normalize_header 归一化，
# 支持包含式匹配：如工作表名“霉酚酸五金”可匹配“五金”）
PURCHASE_REQUEST_IMPORT_CATEGORY_ALIASES: dict[str, str] = {
    "五金材料": "hardware",
    "五金件": "hardware",
    "五金工具": "hardware",
    "五金": "hardware",
    "电脑材料": "computer",
    "电脑": "computer",
    "办公用品": "office",
    "办公": "office",
    "其他商品": "office",
    "原辅料": "raw-auxiliary",
    "原辅材料": "raw-auxiliary",
    "原料": "raw-auxiliary",
    "辅料": "raw-auxiliary",
    "化玻": "chemical-glass",
    "化学玻璃": "chemical-glass",
    "化学试剂": "chemical-glass",
    "玻璃仪器": "chemical-glass",
    "电气": "electrical",
    "电器": "electrical",
    "电气材料": "electrical",
    "电气件": "electrical",
    "广告印刷": "advertising-printing",
    "广告": "advertising-printing",
    "印刷": "advertising-printing",
    "消防": "fire",
    "消防器材": "fire",
    "消防用品": "fire",
    "包材": "packaging",
    "包装材料": "packaging",
    "特防": "labor-special",
    "劳保特防": "labor-special",
    "杂品": "labor-miscellaneous",
    "劳保杂品": "labor-miscellaneous",
    "加急单": "urgent",
    "加急": "urgent",
}

# 导入表头别名 → 采购申请明细字段
PURCHASE_REQUEST_IMPORT_ITEM_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "item_category": ("申请类型", "明细申请类型", "明细采购类型"),
    "product_name": ("商品名称", "名称", "物品名称"),
    "specification": ("规格",),
    "material_code": ("物料编码", "物料编号", "编码", "物料代码"),
    "material_description": ("物料说明", "物料描述", "物料名称"),
    "rule_model": ("规格型号", "型号"),
    "purpose": ("用途",),
    "material": ("材质",),
    "brand": ("品牌",),
    "quantity": (
        "数量",
        "请购数量",
        "申购数量",
        "需求数量",
        "需求数量(kg)",
        "月需求数量",
        "月需求数量(kg)",
        "1个月需求数量",
    ),
    "unit": ("单位",),
    "unit_price": (
        "单价",
        "单价（元）",
        "单价(元)",
        "预估单价",
        "预估单价（元）",
        "预估单价(元)",
        "预算单价",
    ),
    "total_amount": (
        "预估总价",
        "预估总价（元）",
        "预估总价(元)",
        "总价",
        "总价（元）",
        "总额",
        "总额（元）",
        "合计金额",
    ),
    "remarks": ("备注",),
}

# 导入表头别名 → 采购申请单级字段（取自工作表内首个非空值）
PURCHASE_REQUEST_IMPORT_SHEET_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "category": ("采购类型", "采购类别", "类别"),
    "request_department": (
        "申购部门",
        "申请部门",
        "部门",
        "部门名称",
        "使用部门",
        "需求部门",
        "申购部门名称",
        "申请部门名称",
    ),
    "request_date": ("申请日期", "申购日期", "日期"),
}

MATERIAL_FIELD_PURCHASE_CATEGORY_VALUES = {
    category.value for category in MATERIAL_FIELD_PURCHASE_CATEGORIES
}

PURCHASE_ORDER_EXPORT_HEADERS = [
    "序号",
    "商品名称",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价（元）",
    "总额（元）",
    "备注",
]

PURCHASE_ORDER_EXPORT_MATERIAL_HEADERS = [
    "序号",
    "物料编码",
    "物料说明",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价（元）",
    "总额（元）",
    "备注",
]

PURCHASE_ORDER_EXPORT_COMPATIBILITY_HEADERS = [
    "序号",
    "物料编码/商品名称",
    "物料说明/商品名称",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价（元）",
    "总额（元）",
    "备注",
]

PURCHASE_ORDER_EXPORT_URGENT_HEADERS = [
    "序号",
    "申请类型",
    "物料编码/商品名称",
    "物料说明/商品名称",
    "规格型号",
    "用途",
    "材质",
    "品牌",
    "数量",
    "单位",
    "单价（元）",
    "总额（元）",
    "备注",
]

PURCHASE_ORDER_EXPORT_COLUMN_WIDTHS = {
    "A": 18,
    "B": 19.33,
    "C": 13.08,
    "D": 17.83,
    "E": 5.5,
    "F": 9.5,
    "G": 5.5,
    "H": 10.25,
    "I": 13,
    "J": 11.58,
    "K": 29.75,
}

PURCHASE_ORDER_EXPORT_MATERIAL_COLUMN_WIDTHS = {
    "A": 12,
    "B": 18,
    "C": 22,
    "D": 18,
    "E": 17.83,
    "F": 10,
    "G": 10,
    "H": 8,
    "I": 10.25,
    "J": 13,
    "K": 13,
    "L": 29.75,
}

PURCHASE_ORDER_EXPORT_URGENT_COLUMN_WIDTHS = {
    "A": 12,
    "B": 16,
    "C": 18,
    "D": 22,
    "E": 18,
    "F": 17.83,
    "G": 10,
    "H": 10,
    "I": 8,
    "J": 10,
    "K": 10.25,
    "L": 13,
    "M": 29.75,
}

SUPPORTED_SUPPLIER_TABLE_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".tsv"}
SUPPLIER_FIELD_ALIASES = {
    "supplier_code": {"供应商代码", "供应商编码", "供应商编号"},
    "supplier_name": {"供应商名称", "供应商", "供应商全称"},
    "material_code": {"物料编码", "物料代码", "物料编号"},
    "material_name": {"物料名称", "物料", "品名"},
    "manufacturer_code": {"生产厂家编码", "厂家编码", "生产商编码"},
    "manufacturer_name": {"生产厂家名称", "生产厂家", "厂家名称", "生产商"},
    "purchase_category": {"采购品类名称", "采购品类", "品类", "采购类别"},
    "last_updated_by": {"最后更新人", "更新人"},
    "last_updated_date": {"最后更新日期", "更新日期", "最后更新时间"},
}

CONTRACT_STORAGE_DIR = "contracts"


class DuplicateInvoiceError(ValueError):
    def __init__(self, existing_record: InvoiceRecognitionRecord) -> None:
        self.existing_record = existing_record
        super().__init__(
            "发票已识别过，"
            f"记录 ID：{existing_record.id}，文件：{existing_record.file_name}"
        )


def recognize_invoice_pdf(
    pdf_bytes: bytes,
    *,
    include_details: bool = False,
) -> InvoiceRecognitionResult:
    """Extract invoice fields from an electronic VAT invoice PDF."""
    reader = PdfReader(BytesIO(pdf_bytes))
    text_parts: list[str] = []
    for page in reader.pages:
        text_parts.append(page.extract_text(extraction_mode="layout") or "")

    raw_text = "\n".join(text_parts).strip()
    if not raw_text:
        raise ValueError("未能从 PDF 中提取到文本，请确认文件为电子发票 PDF")

    return _parse_invoice_text(raw_text, include_details=include_details)


async def recognize_and_store_invoice_pdf(
    db: AsyncSession,
    pdf_bytes: bytes,
    *,
    file_name: str,
    include_details: bool = False,
) -> InvoiceRecognitionRecord:
    result = recognize_invoice_pdf(pdf_bytes, include_details=include_details)
    file_sha256 = sha256(pdf_bytes).hexdigest()
    duplicate_key = _build_invoice_duplicate_key(result)
    repository = InvoiceRecognitionRepository(db)
    duplicate = await repository.find_duplicate(
        duplicate_key=duplicate_key,
        source_file_sha256=file_sha256,
    )
    if duplicate:
        raise DuplicateInvoiceError(duplicate)

    record = InvoiceRecognitionRecord(
        file_name=file_name,
        include_details=include_details,
        invoice_number=result.invoice_number,
        duplicate_key=duplicate_key,
        source_file_sha256=file_sha256,
        invoice_date=result.invoice_date,
        seller_name=result.seller_name,
        total_tax_amount=result.total_tax_amount,
        total_amount_with_tax_small=result.total_amount_with_tax_small,
        line_items=[
            item.model_dump(mode="json", exclude_none=False)
            for item in result.line_items
        ],
        raw_text=result.raw_text,
    )
    return await repository.create(record)


async def list_invoice_recognition_records(
    db: AsyncSession,
    *,
    keyword: str | None = None,
    seller_name: str | None = None,
    invoice_number: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[InvoiceRecognitionRecord], int]:
    return await InvoiceRecognitionRepository(db).list_records(
        keyword=keyword,
        seller_name=seller_name,
        invoice_number=invoice_number,
        page=page,
        page_size=page_size,
    )


async def delete_invoice_recognition_record(
    db: AsyncSession,
    record_id: UUID,
) -> bool:
    return await InvoiceRecognitionRepository(db).delete_record(record_id)


async def batch_delete_invoice_recognition_records(
    db: AsyncSession,
    record_ids: list[UUID],
) -> int:
    return await InvoiceRecognitionRepository(db).batch_delete_records(record_ids)


async def import_supplier_table_file(
    db: AsyncSession,
    file_bytes: bytes,
    *,
    file_name: str,
) -> SupplierImportResult:
    if not file_bytes:
        raise ValueError("上传文件为空")

    columns, rows, sheet_name = _parse_supplier_table_file(file_bytes, file_name)
    suppliers = [
        _build_supplier_from_row(
            raw_data=row_data,
            columns=columns,
            file_name=file_name,
            sheet_name=sheet_name,
            row_number=row_number,
        )
        for row_number, row_data in rows
    ]
    imported_count = await SupplierRepository(db).replace_all(suppliers)
    return SupplierImportResult(
        imported_count=imported_count,
        columns=columns,
        file_name=file_name,
        sheet_name=sheet_name,
    )


async def import_purchase_request_table_file(
    db: AsyncSession,
    file_bytes: bytes,
    *,
    file_name: str,
) -> PurchaseRequestImportResult:
    """从 xlsx/xls/csv 表格导入采购申请。

    每个工作表生成一份采购申请草稿；行级与工作表级错误分别收集，
    一个工作表失败不影响其他工作表。同一文件（内容哈希相同）的同一
    工作表重复导入时跳过，避免生成重复申请草稿。
    """
    if not file_bytes:
        raise ValueError("上传文件为空")

    suffix = _get_file_suffix(file_name)
    if suffix not in SUPPORTED_PURCHASE_REQUEST_IMPORT_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_PURCHASE_REQUEST_IMPORT_EXTENSIONS))
        raise ValueError(f"暂不支持该文件类型，请上传 {supported} 文件")

    sheets = _parse_purchase_request_import_sheets(file_bytes, suffix)
    result = PurchaseRequestImportResult(file_name=file_name, total_sheets=len(sheets))
    for sheet_name, rows, parse_error, title_department in sheets:
        if parse_error is not None:
            result.failed_rows.append(
                PurchaseRequestImportError(
                    sheet_name=sheet_name,
                    row=None,
                    message=parse_error,
                )
            )
            continue
        if not rows:
            continue
        duplicate_key = sha256(file_bytes + f"|{sheet_name}".encode()).hexdigest()
        existing = await PurchaseRequestRepository(db).find_by_import_duplicate_key(
            duplicate_key
        )
        if existing is not None:
            result.failed_rows.append(
                PurchaseRequestImportError(
                    sheet_name=sheet_name,
                    row=None,
                    message=f"工作表“{sheet_name}”已导入过（相同文件），已跳过重复导入",
                )
            )
            continue
        summary, errors = await _import_purchase_request_sheet(
            db,
            sheet_name=sheet_name,
            rows=rows,
            file_name=file_name,
            title_department=title_department,
            duplicate_key=duplicate_key,
        )
        result.failed_rows.extend(errors)
        if summary is not None:
            result.imported_requests.append(summary)
    return result


async def _import_purchase_request_sheet(
    db: AsyncSession,
    *,
    sheet_name: str,
    rows: list[tuple[int, dict[str, object]]],
    file_name: str,
    title_department: str = "",
    duplicate_key: str = "",
) -> tuple[PurchaseRequestImportSummary | None, list[PurchaseRequestImportError]]:
    errors: list[PurchaseRequestImportError] = []

    try:
        category_value, category_source = _resolve_import_category(sheet_name, rows)
    except ValueError as exc:
        return None, [
            PurchaseRequestImportError(
                sheet_name=sheet_name,
                row=None,
                message=str(exc),
            )
        ]

    # 申购部门优先级：表内“申请部门”列 → 标题行“申请部门：xxx” → 回退
    request_department = _first_sheet_field_value(rows, "request_department")
    if not request_department:
        request_department = title_department
    if not request_department:
        request_department = _fallback_import_department(sheet_name, file_name)
    request_date = _parse_import_request_date(rows)

    items: list[PurchaseRequestItemInput] = []
    for row_number, raw_data in rows:
        try:
            items.append(
                _build_import_item(
                    raw_data,
                    row_number=row_number,
                    category_value=category_value,
                )
            )
        except (ValueError, ValidationError) as exc:
            errors.append(
                PurchaseRequestImportError(
                    sheet_name=sheet_name,
                    row=row_number,
                    message=str(exc),
                )
            )

    if errors:
        # 任一明细错误（如总额计算不一致）→ 整张工作表不导入，
        # 避免错误明细被静默跳过导致提交后才发现问题
        detail = "；".join(error.message for error in errors[:5])
        if len(errors) > 5:
            detail = f"{detail}；等共{len(errors)}处"
        return None, [
            PurchaseRequestImportError(
                sheet_name=sheet_name,
                row=None,
                message=(
                    f"工作表“{sheet_name}”存在{len(errors)}处明细错误，"
                    f"整张工作表未导入，请修改后重新提交：{detail}"
                ),
            )
        ]

    if not items:
        return None, errors + [
            PurchaseRequestImportError(
                sheet_name=sheet_name,
                row=None,
                message="该工作表没有可导入的明细行",
            )
        ]

    try:
        created = await create_purchase_request(
            db,
            PurchaseRequestCreate(
                category=PurchaseRequestCategory(category_value),
                request_department=request_department,
                request_date=request_date,
                attachment_note=f"通过表格导入（{sheet_name}）",
                import_duplicate_key=duplicate_key or None,
                items=items,
            ),
        )
    except (ValueError, ValidationError) as exc:
        return None, errors + [
            PurchaseRequestImportError(
                sheet_name=sheet_name,
                row=None,
                message=f"申请生成失败：{exc}",
            )
        ]

    summary = PurchaseRequestImportSummary(
        request_id=created.id,
        sheet_name=sheet_name,
        category=PurchaseRequestCategory(category_value),
        category_label=PURCHASE_CATEGORY_LABELS.get(category_value, category_value),
        category_source=category_source,
        request_department=request_department,
        request_date=request_date,
        items_count=len(items),
    )
    return summary, errors


def _resolve_import_category(
    sheet_name: str,
    rows: list[tuple[int, dict[str, object]]],
) -> tuple[str, str]:
    """识别采购类型，返回 (采购类型, 来源)。

    来源依次为：表内“采购类型”列（识别值须一致）→ 工作表名称（包含式
    匹配，如“霉酚酸五金”）→ 明细“申请类型”列（识别值一致）→ 按明细
    字段自动推断（物料编码+物料说明 → 五金材料；商品名称 → 办公用品）。
    无法识别且无法映射的值不阻断，全部失败时抛出 ValueError。
    """
    normalized_values: list[tuple[int, str]] = []
    for row_number, raw_data in rows:
        value = _get_import_field(raw_data, "category")
        if not value:
            continue
        normalized = _match_import_category(value)
        if normalized is not None:
            normalized_values.append((row_number, normalized))

    if normalized_values:
        first_category = normalized_values[0][1]
        if any(value != first_category for _, value in normalized_values):
            mixed = "、".join(dict.fromkeys(value for _, value in normalized_values))
            raise ValueError(
                f"采购类型列包含多种类型（{mixed}），请按类型拆分到不同工作表"
            )
        return first_category, "column"

    normalized_name = _match_import_category(sheet_name)
    if normalized_name is not None:
        return normalized_name, "sheet_name"

    item_category_values: list[tuple[int, str]] = []
    for row_number, raw_data in rows:
        value = _get_import_field(raw_data, "item_category")
        if not value:
            continue
        normalized = _match_import_category(value)
        if normalized is not None:
            item_category_values.append((row_number, normalized))
    if item_category_values:
        first_item_category = item_category_values[0][1]
        if all(value == first_item_category for _, value in item_category_values):
            return first_item_category, "inferred"

    inferred = _infer_import_category(rows)
    if inferred is not None:
        return inferred, "inferred"

    raise ValueError(
        "无法识别采购类型：请在表头添加“采购类型”列，"
        "或将工作表命名为含类别名称的表名（如“霉酚酸五金”“加急单”）；"
        "也可以提供物料编码/物料说明或商品名称列，由系统自动判断"
    )


def _match_import_category(value: str) -> str | None:
    """类别别名匹配：先精确匹配，再按别名长度降序做包含匹配。"""
    key = _normalize_header(value)
    if not key:
        return None
    exact = PURCHASE_REQUEST_IMPORT_CATEGORY_ALIASES.get(key)
    if exact is not None:
        return exact
    for alias in sorted(
        PURCHASE_REQUEST_IMPORT_CATEGORY_ALIASES,
        key=len,
        reverse=True,
    ):
        normalized_alias = _normalize_header(alias)
        if normalized_alias and normalized_alias in key:
            return PURCHASE_REQUEST_IMPORT_CATEGORY_ALIASES[alias]
    return None


def _infer_import_category(
    rows: list[tuple[int, dict[str, object]]],
) -> str | None:
    """按明细字段存在性推断采购类型；字段需要有可导入的数据。"""
    has_material_code = any(
        _get_import_field(raw_data, "material_code") for _, raw_data in rows
    )
    has_material_description = any(
        _get_import_field(raw_data, "material_description") for _, raw_data in rows
    )
    if has_material_code and has_material_description:
        return PurchaseRequestCategory.hardware.value
    has_product_name = any(
        _get_import_field(raw_data, "product_name") for _, raw_data in rows
    )
    if has_product_name:
        return PurchaseRequestCategory.office.value
    return None


def _fallback_import_department(sheet_name: str, file_name: str) -> str:
    """申购部门缺失时的回退值：工作表名去掉类别关键字（如“霉酚酸五金”→“霉酚酸”）；
    无法提取时 CSV 使用文件名，其他使用“未填写”。"""
    normalized = _normalize_header(sheet_name)
    if normalized:
        for alias in sorted(
            PURCHASE_REQUEST_IMPORT_CATEGORY_ALIASES,
            key=len,
            reverse=True,
        ):
            normalized_alias = _normalize_header(alias)
            if normalized_alias and normalized_alias in normalized:
                candidate = re.sub(
                    r"[\s_\-—:：]+$",
                    "",
                    sheet_name.replace(alias, ""),
                ).strip()
                if candidate:
                    return candidate
                return "未填写"
    if sheet_name in {"CSV", "csv"}:
        stem = Path(file_name).stem.strip()
        if stem:
            return stem
    return "未填写"


def _build_import_item(
    raw_data: dict[str, object],
    *,
    row_number: int,
    category_value: str,
) -> PurchaseRequestItemInput:
    """按行构建明细；任何字段问题抛出带行号的 ValueError。"""
    item_category: PurchaseRequestCategory | None = None
    item_category_value = _get_import_field(raw_data, "item_category")
    if item_category_value:
        normalized_item_category = _match_import_category(item_category_value)
        if normalized_item_category is None:
            raise ValueError(f"第{row_number}行申请类型无法识别：{item_category_value}")
        item_category = PurchaseRequestCategory(normalized_item_category)

    if category_value == PurchaseRequestCategory.urgent.value:
        if item_category is None:
            raise ValueError(f"第{row_number}行加急单明细缺少申请类型")
    elif item_category is not None and item_category.value != category_value:
        raise ValueError(
            f"第{row_number}行申请类型“{item_category.value}”与采购分类不一致"
        )

    item_uses_material_fields = (
        item_category.value if item_category is not None else category_value
    ) in MATERIAL_FIELD_PURCHASE_CATEGORY_VALUES
    product_name = _get_import_field(raw_data, "product_name")
    material_code = _get_import_field(raw_data, "material_code")
    material_description = _get_import_field(raw_data, "material_description")
    if item_uses_material_fields:
        if not material_code:
            raise ValueError(f"第{row_number}行缺少物料编码")
        # 原辅料/包材等表格常见“商品名称”列代替“物料说明”列
        material_description = material_description or product_name
        if not material_description:
            raise ValueError(f"第{row_number}行缺少物料说明")
    elif not product_name:
        # 办公用品等表格常见“物料描述”列代替“商品名称”列
        product_name = material_description
        if not product_name:
            raise ValueError(f"第{row_number}行缺少商品名称")

    quantity = _parse_import_number(_get_import_field(raw_data, "quantity"))
    if quantity is None:
        raise ValueError(f"第{row_number}行数量无效")
    if quantity < 0:
        raise ValueError(f"第{row_number}行数量不能为负数")
    unit_price_value = _get_import_field(raw_data, "unit_price")
    unit_price = _parse_import_number(unit_price_value)
    if unit_price is None:
        if unit_price_value:
            raise ValueError(f"第{row_number}行单价无效")
        unit_price = Decimal("0")
    elif unit_price < 0:
        raise ValueError(f"第{row_number}行单价不能为负数")

    total_amount_value = _get_import_field(raw_data, "total_amount")
    if total_amount_value:
        table_total = _parse_import_number(total_amount_value)
        if table_total is None:
            raise ValueError(f"第{row_number}行预估总价无效")
        if unit_price_value:
            calculated_total = _calculate_line_amount(quantity, unit_price)
            if calculated_total != table_total:
                raise ValueError(
                    f"第{row_number}行预估总价（{_format_decimal(table_total)}）"
                    f"与数量×单价（{_format_decimal(calculated_total)}）不一致"
                )

    return PurchaseRequestItemInput(
        item_category=item_category,
        product_name=product_name,
        specification=_get_import_field(raw_data, "specification"),
        material_code=material_code,
        material_description=material_description,
        rule_model=_get_import_field(raw_data, "rule_model"),
        purpose=_get_import_field(raw_data, "purpose"),
        material=_get_import_field(raw_data, "material"),
        brand=_get_import_field(raw_data, "brand"),
        quantity=quantity,
        unit=_get_import_field(raw_data, "unit"),
        unit_price=unit_price,
        remarks=_get_import_field(raw_data, "remarks"),
    )


def _get_import_field(raw_data: dict[str, object], field_name: str) -> str:
    aliases = PURCHASE_REQUEST_IMPORT_ITEM_FIELD_ALIASES.get(
        field_name,
        PURCHASE_REQUEST_IMPORT_SHEET_FIELD_ALIASES.get(field_name, ()),
    )
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for column, value in raw_data.items():
        if _normalize_header(column) in normalized_aliases:
            return _format_cell_string(value)
    return ""


def _first_sheet_field_value(
    rows: list[tuple[int, dict[str, object]]],
    field_name: str,
) -> str:
    for _, raw_data in rows:
        value = _get_import_field(raw_data, field_name)
        if value:
            return value
    return ""


def _parse_import_number(value: str) -> Decimal | None:
    text = value.strip().replace(",", "").replace("¥", "").replace("￥", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _parse_import_request_date(
    rows: list[tuple[int, dict[str, object]]],
) -> date:
    raw_value = _first_sheet_field_value(rows, "request_date")
    if raw_value:
        parsed = _parse_supplier_date(raw_value)
        if parsed is not None:
            return parsed
    return datetime.now(UTC).date()


async def list_suppliers(
    db: AsyncSession,
    *,
    keyword: str | None = None,
    supplier_name: str | None = None,
    material_name: str | None = None,
    purchase_category: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[SupplierResponse], int, list[str]]:
    suppliers, total, columns = await SupplierRepository(db).list_suppliers(
        keyword=keyword,
        supplier_name=supplier_name,
        material_name=material_name,
        purchase_category=purchase_category,
        page=page,
        page_size=page_size,
    )
    return (
        [SupplierResponse.model_validate(supplier) for supplier in suppliers],
        total,
        columns,
    )


async def create_purchase_request(
    db: AsyncSession,
    data: PurchaseRequestCreate,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    items, total_amount = _build_purchase_request_items(
        data.items,
        category=data.category.value,
    )
    now = datetime.now(UTC)
    request = PurchaseRequest(
        category=data.category.value,
        request_department=data.request_department,
        request_date=data.request_date,
        attachment_note=data.attachment_note,
        import_duplicate_key=data.import_duplicate_key or None,
        status=PurchaseRequestStatus.draft.value,
        total_amount=total_amount,
        status_updated_at=now,
    )
    created = await repository.create(request, items)
    return await _get_purchase_request_response(repository, created.id)


async def update_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseRequestUpdate,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    if request.status not in {
        PurchaseRequestStatus.draft.value,
        PurchaseRequestStatus.rejected.value,
    }:
        raise ValueError("只有草稿或已驳回的采购申请可以编辑")

    if data.request_department is not None:
        request.request_department = data.request_department
    if data.request_date is not None:
        request.request_date = data.request_date
    if data.attachment_note is not None:
        request.attachment_note = data.attachment_note
    if data.items is not None:
        items, total_amount = _build_purchase_request_items(
            data.items,
            category=request.category,
        )
        await repository.replace_items(request_id, items)
        request.total_amount = total_amount
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


async def get_purchase_request(
    db: AsyncSession,
    request_id: UUID,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    return await _get_purchase_request_response(repository, request_id)


async def delete_purchase_request(
    db: AsyncSession,
    request_id: UUID,
) -> bool:
    """删除采购申请草稿（软删除，含明细与审批记录）。

    仅草稿状态允许删除；已提交或已审批的申请不可删除。
    """
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    if request.status != PurchaseRequestStatus.draft.value:
        raise ValueError("仅草稿状态的采购申请可以删除")
    return await repository.delete(request_id)


async def list_purchase_requests(
    db: AsyncSession,
    *,
    category: str | None = None,
    status: str | None = None,
    approval_role: PurchaseApprovalRole | None = None,
    approval_view: PurchaseApprovalView = PurchaseApprovalView.pending,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[PurchaseRequestResponse], int]:
    repository = PurchaseRequestRepository(db)
    if approval_role:
        if approval_view == PurchaseApprovalView.pending:
            status = APPROVAL_ROLE_TO_PENDING_STATUS[approval_role].value
        else:
            approval_result = (
                PurchaseApprovalResult.approved
                if approval_view == PurchaseApprovalView.completed
                else PurchaseApprovalResult.rejected
            )
            requests, total = await repository.list_requests_by_approval(
                approval_role=approval_role.value,
                result=approval_result.value,
                category=category,
                keyword=keyword,
                page=page,
                page_size=page_size,
            )
            responses = [
                await _get_purchase_request_response(repository, request.id)
                for request in requests
            ]
            return responses, total

    requests, total = await repository.list_requests(
        category=category,
        status=status,
        keyword=keyword,
        page=page,
        page_size=page_size,
    )
    responses = [
        await _get_purchase_request_response(repository, request.id)
        for request in requests
    ]
    return responses, total


async def list_purchase_order_lines(
    db: AsyncSession,
    *,
    category: str | None = None,
    year: int,
    month: int,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[PurchaseOrderLineResponse], int]:
    repository = PurchaseRequestRepository(db)
    start_date, end_date = _month_date_range(year, month)
    rows, total = await repository.list_purchase_order_lines(
        start_date=start_date,
        end_date=end_date,
        status=PurchaseRequestStatus.approved.value,
        category=category,
        page=page,
        page_size=page_size,
    )
    return [_build_purchase_order_line(request, item) for request, item in rows], total


async def export_purchase_order_lines_xlsx(
    db: AsyncSession,
    *,
    category: str | None = None,
    year: int,
    month: int,
) -> bytes:
    repository = PurchaseRequestRepository(db)
    start_date, end_date = _month_date_range(year, month)
    rows, _ = await repository.list_purchase_order_lines(
        start_date=start_date,
        end_date=end_date,
        status=PurchaseRequestStatus.approved.value,
        category=category,
    )
    lines = [_build_purchase_order_line(request, item) for request, item in rows]
    category_label = PURCHASE_CATEGORY_LABELS.get(category or "", "全部类别")
    material_field_mode: bool | None | str = (
        "urgent"
        if category == PurchaseRequestCategory.urgent.value
        else (
            category in MATERIAL_FIELD_PURCHASE_CATEGORY_VALUES
            if category is not None
            else None
        )
    )
    workbook = _build_purchase_order_workbook(
        lines,
        year=year,
        month=month,
        category_label=category_label,
        material_field_mode=material_field_mode,
    )
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def generate_and_store_contract(
    db: AsyncSession,
    payload: ContractGenerateRequest,
) -> tuple[BytesIO, str, str, ContractRecord]:
    buffer, filename, content_type = generate_contract(payload)
    file_bytes = buffer.getvalue()
    record_id = uuid4()
    stored_path = _store_contract_file(
        record_id=record_id,
        filename=filename,
        content=file_bytes,
        content_type=content_type,
    )
    record = ContractRecord(
        id=record_id,
        title=payload.title.strip(),
        category=payload.category.value,
        contract_number=payload.contract_number,
        contract_date=payload.contract_date,
        seller_name=payload.seller.name,
        filename=filename,
        file_path=stored_path,
        content_type=content_type,
        file_size=len(file_bytes),
        payload=payload.model_dump(mode="json"),
    )
    created = await ContractRecordRepository(db).create(record)
    return BytesIO(file_bytes), filename, content_type, created


async def list_contract_records(
    db: AsyncSession,
    *,
    keyword: str | None = None,
    page: int = 1,
    page_size: int = 20,
) -> tuple[list[ContractRecord], int]:
    return await ContractRecordRepository(db).list_records(
        keyword=keyword,
        page=page,
        page_size=page_size,
    )


async def get_contract_record(
    db: AsyncSession,
    record_id: UUID,
) -> ContractRecord:
    record = await ContractRecordRepository(db).get(record_id)
    if not record:
        raise ValueError("合同记录不存在")
    return record


async def get_contract_record_file(
    db: AsyncSession,
    record_id: UUID,
) -> tuple[bytes, str, str]:
    record = await get_contract_record(db, record_id)
    if minio_enabled():
        result = get_object("procurement", record.file_path)
        if result is None:
            raise ValueError("合同文件不存在")
        data, content_type = result
        return data, content_type or record.content_type, record.filename

    file_path = Path(record.file_path)
    if not file_path.exists():
        raise ValueError("合同文件不存在")
    return file_path.read_bytes(), record.content_type, record.filename


def _store_contract_file(
    *,
    record_id: UUID,
    filename: str,
    content: bytes,
    content_type: str,
) -> str:
    safe_filename = Path(filename).name
    if minio_enabled():
        object_key = f"{CONTRACT_STORAGE_DIR}/{record_id}/{safe_filename}"
        return upload_object(
            module="procurement",
            object_key=object_key,
            data=content,
            length=len(content),
            content_type=content_type,
        )

    storage_root = Path(get_settings().STORAGE_ROOT)
    directory = storage_root / "procurement" / CONTRACT_STORAGE_DIR / str(record_id)
    directory.mkdir(parents=True, exist_ok=True)
    file_path = directory / safe_filename
    file_path.write_bytes(content)
    return str(file_path)


async def submit_purchase_request(
    db: AsyncSession,
    request_id: UUID,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    if request.status not in {
        PurchaseRequestStatus.draft.value,
        PurchaseRequestStatus.rejected.value,
    }:
        raise ValueError("只有草稿或已驳回的采购申请可以提交")
    items = await repository.list_items(request_id)
    if not items:
        raise ValueError("采购申请至少需要一条明细")
    for item in items:
        calculated_total = _calculate_line_amount(item.quantity, item.unit_price)
        if calculated_total != item.total_amount:
            raise ValueError(
                f"第{item.sequence}条明细总额（{_format_decimal(item.total_amount)}）"
                f"与数量×单价（{_format_decimal(calculated_total)}）不一致，"
                "请修改后重新提交"
            )

    now = datetime.now(UTC)
    first_role = get_purchase_approval_workflow(request.category)[0]
    request.status = APPROVAL_ROLE_TO_PENDING_STATUS[first_role].value
    request.rejected_step = None
    request.status_updated_at = now
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


async def approve_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    if data.result != PurchaseApprovalResult.approved:
        raise ValueError("审批通过接口的结果必须为 approved")
    return await _review_purchase_request(db, request_id, data)


async def reject_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    if data.result != PurchaseApprovalResult.rejected:
        raise ValueError("审批驳回接口的结果必须为 rejected")
    return await _review_purchase_request(db, request_id, data)


async def _review_purchase_request(
    db: AsyncSession,
    request_id: UUID,
    data: PurchaseApprovalRequest,
) -> PurchaseRequestResponse:
    repository = PurchaseRequestRepository(db)
    request = await repository.get_for_update(request_id)
    if not request:
        raise ValueError("采购申请不存在")

    workflow = get_purchase_approval_workflow(request.category)
    try:
        workflow_index = workflow.index(data.approval_role)
    except ValueError as exc:
        raise ValueError("该采购类型不包含此审批步骤") from exc

    expected_status = APPROVAL_ROLE_TO_PENDING_STATUS[data.approval_role].value
    if request.status != expected_status:
        raise ValueError("当前采购申请不在该审批步骤")

    approvals = await repository.list_approvals(request_id)
    now = datetime.now(UTC)
    approval = PurchaseRequestApproval(
        purchase_request_id=str(request_id),
        approval_role=data.approval_role.value,
        result=data.result.value,
        opinion=data.opinion,
        approver_name=data.approver_name,
        approval_time=now,
    )
    await repository.add_approval(approval)

    if data.result == PurchaseApprovalResult.rejected:
        request.status = PurchaseRequestStatus.rejected.value
        request.rejected_step = data.approval_role.value
    else:
        approvals.append(approval)
        required_count = PURCHASE_APPROVAL_REQUIRED_COUNTS.get(
            data.approval_role,
            1,
        )
        approved_count = _count_current_round_approvals(
            approvals,
            approval_role=data.approval_role,
        )
        if approved_count < required_count:
            request.status = expected_status
            request.rejected_step = None
        elif workflow_index + 1 < len(workflow):
            next_role = workflow[workflow_index + 1]
            request.status = APPROVAL_ROLE_TO_PENDING_STATUS[next_role].value
            request.rejected_step = None
        else:
            request.status = PurchaseRequestStatus.approved.value
            request.rejected_step = None
    request.status_updated_at = now
    await db.flush()
    return await _get_purchase_request_response(repository, request_id)


def get_purchase_approval_workflow(
    category: str | PurchaseRequestCategory,
) -> tuple[PurchaseApprovalRole, ...]:
    category_value = (
        category.value if isinstance(category, PurchaseRequestCategory) else category
    )
    return PURCHASE_APPROVAL_WORKFLOWS.get(
        category_value,
        DEFAULT_PURCHASE_APPROVAL_WORKFLOW,
    )


def _count_current_round_approvals(
    approvals: list[PurchaseRequestApproval],
    *,
    approval_role: PurchaseApprovalRole,
) -> int:
    rejected_times = [
        approval.approval_time
        for approval in approvals
        if approval.result == PurchaseApprovalResult.rejected.value
    ]
    latest_rejection = max(rejected_times) if rejected_times else None
    return sum(
        approval.approval_role == approval_role.value
        and approval.result == PurchaseApprovalResult.approved.value
        and (latest_rejection is None or approval.approval_time > latest_rejection)
        for approval in approvals
    )


def _build_purchase_request_items(
    item_inputs: list[PurchaseRequestItemInput],
    *,
    category: str | PurchaseRequestCategory,
) -> tuple[list[PurchaseRequestItem], Decimal]:
    category_value = (
        category.value if isinstance(category, PurchaseRequestCategory) else category
    )
    items: list[PurchaseRequestItem] = []
    total_amount = Decimal("0")
    for index, item in enumerate(item_inputs, start=1):
        item_category = _resolve_item_category(
            item.item_category,
            request_category=category_value,
            index=index,
        )
        item_uses_material_fields = (
            item_category in MATERIAL_FIELD_PURCHASE_CATEGORY_VALUES
        )
        if item_uses_material_fields:
            if not item.material_code.strip():
                raise ValueError(f"第{index}条明细缺少物料编码")
            if not item.material_description.strip():
                raise ValueError(f"第{index}条明细缺少物料说明")
        elif not item.product_name.strip():
            raise ValueError(f"第{index}条明细缺少商品名称")

        line_amount = _calculate_line_amount(item.quantity, item.unit_price)
        total_amount += line_amount
        items.append(
            PurchaseRequestItem(
                purchase_request_id="",
                sequence=index,
                item_category=item_category,
                product_name=item.product_name,
                specification=item.specification,
                material_code=item.material_code,
                material_description=item.material_description,
                rule_model=item.rule_model,
                purpose=item.purpose,
                material=item.material,
                brand=item.brand,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                total_amount=line_amount,
                remarks=item.remarks,
            )
        )
    return items, total_amount.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _resolve_item_category(
    item_category: PurchaseRequestCategory | None,
    *,
    request_category: str,
    index: int,
) -> str:
    resolved = item_category.value if item_category else None
    if request_category == PurchaseRequestCategory.urgent.value:
        if not resolved:
            raise ValueError(f"第{index}条明细缺少申请类型")
        if item_category not in NORMAL_PURCHASE_CATEGORIES:
            raise ValueError(f"第{index}条明细申请类型无效")
        return resolved

    if resolved and resolved != request_category:
        raise ValueError(f"第{index}条明细申请类型与采购分类不一致")
    return request_category


def _calculate_line_amount(quantity: Decimal, unit_price: Decimal) -> Decimal:
    return (quantity * unit_price).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def _month_date_range(year: int, month: int) -> tuple[date, date]:
    start_date = date(year, month, 1)
    if month == 12:
        return start_date, date(year + 1, 1, 1)
    return start_date, date(year, month + 1, 1)


def _build_purchase_order_line(
    request: PurchaseRequest,
    item: PurchaseRequestItem,
) -> PurchaseOrderLineResponse:
    category_label = PURCHASE_CATEGORY_LABELS.get(request.category, request.category)
    return PurchaseOrderLineResponse.model_validate(
        {
            "request_id": request.id,
            "category": request.category,
            "category_label": category_label,
            "request_department": request.request_department,
            "request_date": request.request_date,
            "item_id": item.id,
            "item_sequence": item.sequence,
            "item_category": item.item_category or request.category,
            "product_name": item.product_name,
            "specification": item.specification,
            "material_code": item.material_code,
            "material_description": item.material_description,
            "rule_model": item.rule_model,
            "purpose": item.purpose,
            "material": item.material,
            "brand": item.brand,
            "quantity": item.quantity,
            "unit": item.unit,
            "unit_price": item.unit_price,
            "total_amount": item.total_amount,
            "remarks": item.remarks,
        }
    )


def _build_purchase_order_workbook(
    lines: list[PurchaseOrderLineResponse],
    *,
    year: int,
    month: int,
    category_label: str,
    material_field_mode: bool | None | str,
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    headers = _get_purchase_order_export_headers(material_field_mode)
    column_count = len(headers)
    amount_column = headers.index("总额（元）") + 1
    _apply_purchase_order_sheet_setup(
        worksheet,
        _get_purchase_order_export_column_widths(material_field_mode),
    )

    worksheet.row_dimensions[1].height = 9
    _write_merged_title(
        worksheet,
        2,
        "丽珠集团（宁夏）制药有限公司",
        column_count,
    )
    _write_merged_title(
        worksheet,
        3,
        f"{year}年{month:02d}月份{category_label}申购单汇总",
        column_count,
    )

    row_index = 4
    total_rows: list[int] = []
    for department, department_lines in _group_purchase_order_lines_by_department(
        lines
    ):
        _write_department_row(worksheet, row_index, department, column_count)
        row_index += 1
        _write_purchase_order_header_row(worksheet, row_index, headers)
        row_index += 1

        first_detail_row = row_index
        for index, line in enumerate(department_lines, start=1):
            _write_purchase_order_detail_row(
                worksheet,
                row_index,
                index,
                line,
                material_field_mode,
                amount_column,
            )
            row_index += 1

        total_row = row_index
        total_rows.append(total_row)
        _write_purchase_order_total_row(
            worksheet,
            total_row,
            "合计",
            f"=SUM({get_column_letter(amount_column)}{first_detail_row}:"
            f"{get_column_letter(amount_column)}{row_index - 1})",
            column_count,
            amount_column,
        )
        row_index += 1

    total_formula = (
        "=SUM("
        f"{','.join(f'{get_column_letter(amount_column)}{row}' for row in total_rows)}"
        ")"
        if total_rows
        else "0"
    )
    _write_purchase_order_total_row(
        worksheet,
        row_index,
        "总计",
        total_formula,
        column_count,
        amount_column,
    )
    row_index += 1
    _write_signature_row(worksheet, row_index, column_count)
    worksheet.print_area = f"A1:{get_column_letter(column_count)}{row_index}"
    return workbook


def _build_purchase_order_row_values(
    index: int,
    line: PurchaseOrderLineResponse,
    material_field_mode: bool | None | str,
) -> list[str | int | Decimal]:
    if material_field_mode is True:
        return [
            index,
            line.material_code,
            line.material_description,
            line.rule_model,
            line.purpose,
            line.material,
            line.brand,
            line.quantity,
            line.unit,
            line.unit_price,
            line.total_amount,
            line.remarks,
        ]
    if material_field_mode is None:
        return [
            index,
            line.material_code or line.product_name,
            line.material_description or line.product_name,
            line.rule_model or line.specification,
            line.purpose,
            line.material,
            line.brand,
            line.quantity,
            line.unit,
            line.unit_price,
            line.total_amount,
            line.remarks,
        ]
    if material_field_mode == "urgent":
        item_category_label = PURCHASE_CATEGORY_LABELS.get(
            line.item_category or "",
            line.item_category or "",
        )
        return [
            index,
            item_category_label,
            line.material_code or line.product_name,
            line.material_description or line.product_name,
            line.rule_model or line.specification,
            line.purpose,
            line.material,
            line.brand,
            line.quantity,
            line.unit,
            line.unit_price,
            line.total_amount,
            line.remarks,
        ]
    return [
        index,
        line.product_name,
        line.specification,
        line.purpose,
        line.material,
        line.brand,
        line.quantity,
        line.unit,
        line.unit_price,
        line.total_amount,
        line.remarks,
    ]


def _group_purchase_order_lines_by_department(
    lines: list[PurchaseOrderLineResponse],
) -> list[tuple[str, list[PurchaseOrderLineResponse]]]:
    groups: dict[str, list[PurchaseOrderLineResponse]] = {}
    for line in sorted(
        lines,
        key=lambda item: (
            item.request_department,
            item.request_date,
            item.category,
            str(item.request_id),
            item.item_sequence,
        ),
    ):
        groups.setdefault(line.request_department, []).append(line)
    return list(groups.items())


def _get_purchase_order_export_headers(
    material_field_mode: bool | None | str,
) -> list[str]:
    if material_field_mode is True:
        return PURCHASE_ORDER_EXPORT_MATERIAL_HEADERS
    if material_field_mode is None:
        return PURCHASE_ORDER_EXPORT_COMPATIBILITY_HEADERS
    if material_field_mode == "urgent":
        return PURCHASE_ORDER_EXPORT_URGENT_HEADERS
    return PURCHASE_ORDER_EXPORT_HEADERS


def _get_purchase_order_export_column_widths(
    material_field_mode: bool | None | str,
) -> dict[str, float]:
    if material_field_mode == "urgent":
        return PURCHASE_ORDER_EXPORT_URGENT_COLUMN_WIDTHS
    if material_field_mode is True or material_field_mode is None:
        return PURCHASE_ORDER_EXPORT_MATERIAL_COLUMN_WIDTHS
    return PURCHASE_ORDER_EXPORT_COLUMN_WIDTHS


def _apply_purchase_order_sheet_setup(
    worksheet: Any,
    column_widths: dict[str, float],
) -> None:
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    worksheet.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )


def _write_merged_title(
    worksheet: Any,
    row_index: int,
    value: str,
    column_count: int,
) -> None:
    worksheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=column_count,
    )
    worksheet.row_dimensions[row_index].height = 27
    cell = worksheet.cell(row_index, 1, value)
    cell.font = Font(name="宋体", size=12, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_department_row(
    worksheet: Any,
    row_index: int,
    department: str,
    column_count: int,
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    worksheet.cell(row_index, 1, f"申购部门：{department}")
    fill = PatternFill("solid", fgColor="D9E1F4")
    border = _purchase_order_border(top=True, bottom=True)
    for cell in worksheet[row_index][0:column_count]:
        cell.font = Font(name="宋体", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=False,
        )
        cell.fill = fill
        cell.border = border


def _write_purchase_order_header_row(
    worksheet: Any,
    row_index: int,
    headers: list[str],
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row_index, column_index, header)
        cell.font = Font(name="宋体", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = _purchase_order_border()


def _write_purchase_order_detail_row(
    worksheet: Any,
    row_index: int,
    index: int,
    line: PurchaseOrderLineResponse,
    material_field_mode: bool | None | str,
    amount_column: int,
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    for column_index, value in enumerate(
        _build_purchase_order_row_values(index, line, material_field_mode),
        start=1,
    ):
        cell = worksheet.cell(row_index, column_index, value)
        cell.font = Font(name="宋体", size=12)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = _purchase_order_border()
    text_columns = (
        (2, 3, 4, 5, 6, 11) if material_field_mode is False else (2, 3, 4, 5, 6, 7, 12)
    )
    for column_index in text_columns:
        worksheet.cell(row_index, column_index).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
    worksheet.cell(row_index, amount_column - 1).number_format = "0.00"
    worksheet.cell(row_index, amount_column).number_format = "0.00"


def _write_purchase_order_total_row(
    worksheet: Any,
    row_index: int,
    label: str,
    total_formula: str,
    column_count: int,
    amount_column: int,
) -> None:
    worksheet.row_dimensions[row_index].height = 32.15
    worksheet.cell(row_index, 1, label)
    worksheet.cell(row_index, amount_column, total_formula)
    for cell in worksheet[row_index][0:column_count]:
        cell.font = Font(name="宋体", size=12, bold=(label == "合计"))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _purchase_order_border(top=True, bottom=True)
    worksheet.cell(row_index, amount_column).font = Font(
        name="宋体",
        size=12,
        bold=True,
    )
    worksheet.cell(row_index, amount_column).number_format = "0.00"


def _write_signature_row(worksheet: Any, row_index: int, column_count: int) -> None:
    worksheet.merge_cells(
        start_row=row_index,
        start_column=1,
        end_row=row_index,
        end_column=column_count,
    )
    worksheet.row_dimensions[row_index].height = 32.15
    cell = worksheet.cell(
        row_index,
        1,
        " 总经理：                       财务总监："
        "                          部门负责人："
        "                       统计人：",
    )
    cell.font = Font(name="宋体", size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _purchase_order_border(*, top: bool = True, bottom: bool = True) -> Border:
    side = Side(style="thin", color="000000")
    return Border(
        left=side,
        right=side,
        top=side if top else Side(style=None),
        bottom=side if bottom else Side(style=None),
    )


async def _get_purchase_request_response(
    repository: PurchaseRequestRepository,
    request_id: UUID,
) -> PurchaseRequestResponse:
    request = await repository.get(request_id)
    if not request:
        raise ValueError("采购申请不存在")
    items = await repository.list_items(request_id)
    approvals = await repository.list_approvals(request_id)
    serialized_items = [
        {
            **item.__dict__,
            "item_category": item.item_category or request.category,
        }
        for item in items
    ]
    return PurchaseRequestResponse.model_validate(
        {
            **request.__dict__,
            "items": serialized_items,
            "approvals": approvals,
        }
    )


def _parse_supplier_table_file(
    file_bytes: bytes,
    file_name: str,
) -> tuple[list[str], list[tuple[int, dict[str, object]]], str]:
    suffix = _get_file_suffix(file_name)
    if suffix not in SUPPORTED_SUPPLIER_TABLE_EXTENSIONS:
        supported = "、".join(sorted(SUPPORTED_SUPPLIER_TABLE_EXTENSIONS))
        raise ValueError(f"暂不支持该文件类型，请上传 {supported} 文件")

    if suffix in {".xlsx", ".xlsm"}:
        return _parse_supplier_workbook(file_bytes)
    delimiter = "\t" if suffix == ".tsv" else ","
    return _parse_supplier_text_table(file_bytes, delimiter=delimiter)


def _parse_supplier_workbook(
    file_bytes: bytes,
) -> tuple[list[str], list[tuple[int, dict[str, object]]], str]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    table_rows = list(worksheet.iter_rows(values_only=True))
    return _build_supplier_rows(table_rows, worksheet.title)


def _parse_supplier_text_table(
    file_bytes: bytes,
    *,
    delimiter: str,
) -> tuple[list[str], list[tuple[int, dict[str, object]]], str]:
    text = _decode_table_text(file_bytes)
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    return _build_supplier_rows(list(reader), "CSV" if delimiter == "," else "TSV")


def _parse_purchase_request_import_sheets(
    file_bytes: bytes,
    suffix: str,
) -> list[
    tuple[
        str,
        list[tuple[int, dict[str, object]]] | None,
        str | None,
        str,
    ]
]:
    """按格式解析全部工作表，返回 [(sheet_name, rows, error, title_department), ...]。

    完全空白的工作表直接跳过；有内容但缺少表头或没有数据行的工作表
    记录为工作表级错误（rows 为 None），不中断其他工作表。
    title_department 为表头前标题行中“申请部门/申购部门：xxx”提取的申购部门。
    """
    if suffix == ".xlsx":
        return _parse_purchase_request_xlsx_sheets(file_bytes)
    if suffix == ".xls":
        return _parse_purchase_request_xls_sheets(file_bytes)
    return _parse_purchase_request_csv_sheet(file_bytes)


def _parse_purchase_request_xlsx_sheets(
    file_bytes: bytes,
) -> list[
    tuple[
        str,
        list[tuple[int, dict[str, object]]] | None,
        str | None,
        str,
    ]
]:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheets: list[
        tuple[
            str,
            list[tuple[int, dict[str, object]]] | None,
            str | None,
            str,
        ]
    ] = []
    for worksheet in workbook.worksheets:
        table_rows = list(worksheet.iter_rows(values_only=True))
        if not table_rows:
            continue
        try:
            _, rows, _ = _build_supplier_rows(table_rows, worksheet.title)
        except ValueError as exc:
            sheets.append((worksheet.title, None, str(exc), ""))
            continue
        title_department = _extract_import_request_department(table_rows)
        sheets.append((worksheet.title, rows, None, title_department))
    return sheets


def _parse_purchase_request_xls_sheets(
    file_bytes: bytes,
) -> list[
    tuple[
        str,
        list[tuple[int, dict[str, object]]] | None,
        str | None,
        str,
    ]
]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheets: list[
        tuple[
            str,
            list[tuple[int, dict[str, object]]] | None,
            str | None,
            str,
        ]
    ] = []
    for worksheet in workbook.sheets():
        table_rows = [
            _xls_row_values(worksheet, row_index)
            for row_index in range(worksheet.nrows)
        ]
        if not table_rows:
            continue
        try:
            _, rows, _ = _build_supplier_rows(table_rows, worksheet.name)
        except ValueError as exc:
            sheets.append((worksheet.name, None, str(exc), ""))
            continue
        title_department = _extract_import_request_department(table_rows)
        sheets.append((worksheet.name, rows, None, title_department))
    return sheets


def _xls_row_values(worksheet: xlrd.sheet.Sheet, row_index: int) -> tuple[object, ...]:
    values: list[object] = []
    for col_index in range(worksheet.ncols):
        cell = worksheet.cell(row_index, col_index)
        if cell.ctype == xlrd.XL_CELL_DATE:
            values.append(
                xlrd.xldate.xldate_as_datetime(cell.value, worksheet.book.datemode)
            )
        else:
            values.append(cell.value)
    return tuple(values)


def _parse_purchase_request_csv_sheet(
    file_bytes: bytes,
) -> list[
    tuple[
        str,
        list[tuple[int, dict[str, object]]] | None,
        str | None,
        str,
    ]
]:
    text = _decode_table_text(file_bytes)
    reader = csv.reader(StringIO(text), delimiter=",")
    table_rows = list(reader)
    if not table_rows:
        return [("CSV", None, "表格中没有可导入的数据行", "")]
    _, rows, _ = _build_supplier_rows(table_rows, "CSV")
    return [("CSV", rows, None, "")]


def _extract_import_request_department(
    table_rows: Sequence[Sequence[object]],
) -> str:
    """从表头前的标题行中提取“申请部门/申购部门：xxx”的申购部门。"""
    header_index = _find_header_row_index(table_rows)
    if header_index is None:
        return ""
    for row in table_rows[:header_index]:
        for value in row:
            text = _format_cell_string(value)
            if not text:
                continue
            match = re.search(
                r"(?:申请部门|申购部门)\s*[:：]\s*([^\n]+)",
                text,
            )
            if match:
                return match.group(1).strip()
    return ""


def _decode_table_text(file_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法识别文本编码，请使用 UTF-8 或 GB18030 编码")


def _build_supplier_rows(
    table_rows: Sequence[Sequence[object]],
    sheet_name: str,
) -> tuple[list[str], list[tuple[int, dict[str, object]]], str]:
    header_index = _find_header_row_index(table_rows)
    if header_index is None:
        raise ValueError("未找到表头行")

    raw_headers = [_format_cell_string(value) for value in table_rows[header_index]]
    columns = _deduplicate_headers(raw_headers)
    if not columns:
        raise ValueError("表头为空")

    rows: list[tuple[int, dict[str, object]]] = []
    for row_index, row in enumerate(
        table_rows[header_index + 1 :],
        start=header_index + 2,
    ):
        row_values = list(row)
        raw_data = {
            column: _cell_to_json_value(
                row_values[index] if index < len(row_values) else None
            )
            for index, column in enumerate(columns)
        }
        if all(value in ("", None) for value in raw_data.values()):
            continue
        if _is_import_non_item_row(raw_data):
            continue
        rows.append((row_index, raw_data))

    if not rows:
        raise ValueError("表格中没有可导入的数据行")
    return columns, rows, sheet_name


def _is_import_non_item_row(raw_data: dict[str, object]) -> bool:
    """跳过表格中的合计、签名等非明细行（采购申请导入与供应商导入共用）。

    判断规则：行内出现“合计/总计/小计”；或整行只有一个非空单元格，
    且内容为签字栏关键词；或仅序号/行号列有值（如只有“序号 1”的占位行）。
    """
    non_empty = [
        (column, value) for column, value in raw_data.items() if value not in ("", None)
    ]
    if not non_empty:
        return True
    text = " ".join(str(value) for _, value in non_empty)
    if re.search(r"合计|总计|小计", text):
        return True
    if len(non_empty) == 1:
        column, value = non_empty[0]
        if re.search(r"领导|负责人|经理|主管|统计人|申请人|五金库|签字|签名", text):
            return True
        if _normalize_header(column) in {"序号", "行号"}:
            return True
        # 仅含标点/空白的占位行（如“.”）
        if not re.search(r"[\u4e00-\u9fffA-Za-z0-9]", text):
            return True
    return False


def _find_header_row_index(
    table_rows: Sequence[Sequence[object]],
) -> int | None:
    for index, row in enumerate(table_rows):
        non_empty_count = sum(1 for value in row if _format_cell_string(value))
        if non_empty_count < 2:
            continue
        # 表头字段应连续分布；排除“公司名+月均”这类分散单元格的标题行
        max_run = 0
        run = 0
        for value in row:
            if _format_cell_string(value):
                run += 1
                max_run = max(max_run, run)
            else:
                run = 0
        if max_run >= 2:
            return index
    return None


def _deduplicate_headers(raw_headers: list[str]) -> list[str]:
    columns: list[str] = []
    seen: dict[str, int] = {}
    for index, raw_header in enumerate(raw_headers, start=1):
        header = raw_header or f"未命名字段{index}"
        next_count = seen.get(header, 0) + 1
        seen[header] = next_count
        columns.append(header if next_count == 1 else f"{header}_{next_count}")
    return columns


def _build_supplier_from_row(
    *,
    raw_data: dict[str, object],
    columns: list[str],
    file_name: str,
    sheet_name: str,
    row_number: int,
) -> Supplier:
    return Supplier(
        supplier_code=_get_supplier_field(raw_data, "supplier_code"),
        supplier_name=_get_supplier_field(raw_data, "supplier_name"),
        material_code=_get_supplier_field(raw_data, "material_code"),
        material_name=_get_supplier_field(raw_data, "material_name"),
        manufacturer_code=_get_supplier_field(raw_data, "manufacturer_code"),
        manufacturer_name=_get_supplier_field(raw_data, "manufacturer_name"),
        purchase_category=_get_supplier_field(raw_data, "purchase_category"),
        last_updated_by=_get_supplier_field(raw_data, "last_updated_by"),
        last_updated_date=_parse_supplier_date(
            _get_supplier_field(raw_data, "last_updated_date")
        ),
        import_file_name=file_name,
        import_sheet_name=sheet_name,
        import_row_number=row_number,
        import_columns=columns,
        raw_data=raw_data,
    )


def _get_supplier_field(raw_data: dict[str, object], field_name: str) -> str:
    aliases = SUPPLIER_FIELD_ALIASES[field_name]
    normalized_aliases = {_normalize_header(alias) for alias in aliases}
    for column, value in raw_data.items():
        if _normalize_header(column) in normalized_aliases:
            return _format_cell_string(value)
    return ""


def _parse_supplier_date(value: str) -> date | None:
    if not value:
        return None
    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value).date()
    except ValueError:
        return None


def _get_file_suffix(file_name: str) -> str:
    normalized = file_name.strip().lower()
    if "." not in normalized:
        return ""
    return normalized[normalized.rfind(".") :]


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_（）()：:]+", "", value).lower()


def _cell_to_json_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return _format_decimal(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _format_cell_string(value: object) -> str:
    json_value = _cell_to_json_value(value)
    if json_value is None:
        return ""
    if isinstance(json_value, str):
        return json_value.strip()
    return str(json_value).strip()


def _parse_invoice_text(
    raw_text: str,
    *,
    include_details: bool = False,
) -> InvoiceRecognitionResult:
    lines = [line.rstrip() for line in raw_text.splitlines() if line.strip()]

    return InvoiceRecognitionResult(
        invoice_number=_search_first(r"发票号码[:：]\s*([0-9]{8,})", raw_text),
        invoice_date=_search_first(
            r"开票日期[:：]\s*([0-9]{4}年[0-9]{1,2}月[0-9]{1,2}日)",
            raw_text,
        ),
        seller_name=_extract_seller_name(lines),
        total_tax_amount=_extract_total_tax_amount(lines),
        total_amount_with_tax_small=_extract_total_amount(raw_text),
        line_items=_extract_line_items(lines) if include_details else [],
        raw_text=raw_text,
    )


def _build_invoice_duplicate_key(result: InvoiceRecognitionResult) -> str | None:
    if not result.invoice_number:
        return None

    parts = [
        _normalize_duplicate_part(result.invoice_number),
        _normalize_duplicate_part(result.invoice_date),
        _normalize_duplicate_part(result.seller_name),
        _normalize_duplicate_part(
            _format_money_for_key(result.total_amount_with_tax_small)
        ),
    ]
    return "|".join(parts)


def _normalize_duplicate_part(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", value).strip().lower()


def _format_money_for_key(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return f"{value.quantize(Decimal('0.01')):.2f}"


def _search_first(pattern: str, text: str) -> str | None:
    match = re.search(pattern, text)
    if not match:
        return None
    return match.group(1).strip()


def _extract_seller_name(lines: list[str]) -> str | None:
    for line in lines:
        if "名称" not in line:
            continue
        names = re.findall(
            r"名称[:：]\s*([^\n]+?)(?=\s{2,}\S+\s+名称[:：]|$)",
            line,
        )
        if len(names) >= 2:
            return str(names[1].strip())

        seller_name = _extract_seller_name_before_marker(line)
        if seller_name:
            return seller_name
    return None


def _extract_seller_name_before_marker(line: str) -> str | None:
    match = re.search(r"(.+?)\s*销\s*名称[:：]", line)
    if not match:
        return None

    candidates = [
        part.strip()
        for part in re.split(r"\s{2,}", match.group(1).strip())
        if part.strip()
    ]
    if not candidates:
        return None

    return candidates[-1]


def _extract_total_amount(text: str) -> Decimal | None:
    for line in text.splitlines():
        if "小写" not in line:
            continue

        amounts = re.findall(MONEY_PATTERN, line)
        if amounts:
            return _to_decimal(amounts[-1])

    return None


def _extract_total_tax_amount(lines: list[str]) -> Decimal | None:
    total_tax_amount: Decimal | None = None
    for line in lines:
        compact_line = re.sub(r"\s+", "", line)
        if "合计" not in compact_line or "小计" in compact_line:
            continue

        amounts = re.findall(MONEY_PATTERN, line)
        if len(amounts) < 2:
            continue

        total_tax_amount = _to_decimal(amounts[-1])

    return total_tax_amount


def _extract_line_items(lines: list[str]) -> list[InvoiceLineItem]:
    items: list[InvoiceLineItem] = []
    for line in lines:
        stripped_line = line.strip()
        if not stripped_line.startswith("*"):
            continue

        item = _parse_line_item(stripped_line)
        if item:
            items.append(item)

    return items


def _parse_line_item(line: str) -> InvoiceLineItem | None:
    parts = [part.strip() for part in re.split(r"\s{2,}", line) if part.strip()]
    tax_rate_index = _find_tax_rate_index(parts)
    if len(parts) < 4 or tax_rate_index is None:
        return None

    pre_tax_parts = parts[: tax_rate_index + 1]
    numeric_start = len(pre_tax_parts)
    for index in range(len(pre_tax_parts) - 1, -1, -1):
        if not _is_numeric_cluster(pre_tax_parts[index], allow_percent=True):
            break
        numeric_start = index

    if numeric_start == 0 or numeric_start >= len(pre_tax_parts):
        return None

    tax_amount = _to_decimal(
        _last_match(NUMBER_PATTERN, " ".join(parts[tax_rate_index:])),
    )
    tax_rate = _extract_tax_rate(parts[tax_rate_index])
    numeric_text = " ".join(pre_tax_parts[numeric_start:])
    dense_quantity = _extract_dense_quantity(numeric_text, tax_amount, tax_rate)
    quantity: Decimal | None
    if dense_quantity is not None:
        quantity = dense_quantity
    else:
        numeric_values = re.findall(
            NUMBER_PATTERN,
            _remove_tax_rate(numeric_text, tax_rate),
        )
        if len(numeric_values) < 2:
            return None

        amount = _to_decimal(numeric_values[-1])
        quantity = (
            _to_decimal(numeric_values[0])
            if len(numeric_values) >= 3
            else _extract_quantity(numeric_values[0], amount)
        )

    return InvoiceLineItem(
        project_name=parts[0],
        unit=pre_tax_parts[numeric_start - 1],
        quantity=quantity,
    )


def _find_tax_rate_index(parts: list[str]) -> int | None:
    for index, part in enumerate(parts):
        if "%" in part:
            return index
    return None


def _extract_tax_rate(value: str) -> Decimal | None:
    compact_value = re.sub(r"\s+", "", value)
    if "%" not in compact_value:
        return None

    before_percent = compact_value.split("%", 1)[0]
    for rate in ("13", "9", "6", "5", "3", "1", "0"):
        if before_percent.endswith(rate):
            return Decimal(rate)

    return None


def _extract_dense_quantity(
    numeric_text: str,
    tax_amount: Decimal | None,
    tax_rate: Decimal | None,
) -> Decimal | None:
    if re.search(r"\s", numeric_text.strip()):
        return None

    compact_text = re.sub(r"\s+", "", numeric_text)
    if compact_text.count(".") < 2 or tax_amount is None or not tax_rate:
        return None

    before_percent = compact_text.split("%", 1)[0]
    rate_text = _format_decimal(tax_rate)
    if not before_percent.endswith(rate_text):
        return None

    dense_text = before_percent[: -len(rate_text)]
    estimated_amount = tax_amount / (tax_rate / Decimal(100))
    best_quantity: Decimal | None = None
    best_delta: Decimal | None = None

    for quantity_end in range(1, len(dense_text)):
        quantity_text = dense_text[:quantity_end]
        if quantity_text.count(".") != 1:
            continue
        decimal_part = quantity_text.rsplit(".", 1)[1]
        if len(decimal_part) > 4:
            continue

        quantity = _to_decimal(quantity_text)
        if quantity is None or quantity <= 0:
            continue

        remaining_text = dense_text[quantity_end:]
        for unit_price_end in range(1, len(remaining_text) + 1):
            unit_price_text = remaining_text[:unit_price_end]
            if unit_price_text.count(".") != 1:
                continue

            unit_price = _to_decimal(unit_price_text)
            if unit_price is None:
                continue

            delta = abs(quantity * unit_price - estimated_amount)
            if delta > Decimal("0.2"):
                continue

            if best_delta is None or delta < best_delta:
                best_delta = delta
                best_quantity = quantity

    return best_quantity


def _remove_tax_rate(numeric_text: str, tax_rate: Decimal | None) -> str:
    if tax_rate is None or "%" not in numeric_text:
        return numeric_text

    rate_text = _format_decimal(tax_rate)
    before_percent, after_percent = numeric_text.split("%", 1)
    return before_percent.removesuffix(rate_text) + after_percent


def _last_match(pattern: str, text: str) -> str | None:
    matches = re.findall(pattern, text)
    return matches[-1] if matches else None


def _format_decimal(value: Decimal) -> str:
    return format(value.normalize(), "f")


def _is_numeric_cluster(value: str, *, allow_percent: bool = False) -> bool:
    pattern = r"[0-9¥￥\s.,%]+"
    if not allow_percent:
        pattern = r"[0-9¥￥\s.,]+"
    return bool(re.fullmatch(pattern, value))


def _extract_quantity(
    quantity_and_price: str,
    amount: Decimal | None,
) -> Decimal | None:
    numbers = re.findall(NUMBER_PATTERN, quantity_and_price.replace(",", ""))
    if not numbers:
        return None

    if len(numbers) >= 2:
        return _to_decimal(numbers[0])

    token = numbers[0]
    if "." not in token:
        return _to_decimal(token)

    return _infer_conjoined_quantity(token, amount)


def _infer_conjoined_quantity(token: str, amount: Decimal | None) -> Decimal | None:
    if amount is None:
        return None

    decimal_index = token.find(".")
    best_quantity: Decimal | None = None
    best_delta: Decimal | None = None
    for split_index in range(1, decimal_index):
        quantity = _to_decimal(token[:split_index])
        unit_price = _to_decimal(token[split_index:])
        if quantity is None or unit_price is None:
            continue

        delta = abs(quantity * unit_price - amount)
        if best_delta is None or delta < best_delta:
            best_delta = delta
            best_quantity = quantity

    if best_delta is not None and best_delta <= Decimal("0.02"):
        return best_quantity

    return None


def _to_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    try:
        return Decimal(value.replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None
