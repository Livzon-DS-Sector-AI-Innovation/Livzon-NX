"""Certificate management service."""

from __future__ import annotations

import asyncio
import hashlib
import logging
import re
import shutil
import tempfile
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from uuid import UUID

from fastapi import UploadFile
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import MergedCell  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.exceptions import AppException, NotFoundException
from app.core.upload_security import read_upload_secure
from app.modules.registration.models import (
    RegistrationCertificateEntry,
)
from app.modules.registration.repository import RegistrationCertificateRepository
from app.modules.registration.schemas.certificate import (
    CertificateColumn,
    CertificateEntryCreate,
    CertificateEntryResponse,
    CertificateEntryUpdate,
    CertificateRecordSummary,
    CertificateReminderRecipientOption,
    CertificateReminderSettingResponse,
    CertificateReminderSettingUpdate,
    CertificateSheetDetail,
    CertificateSheetRow,
    CertificateSheetSummary,
    CertificateWorkbookDetail,
    CertificateWorkbookImportResult,
    CertificateWorkbookOverview,
    CertificateWorkbookSheet,
)

logger = logging.getLogger(__name__)

CERTIFICATE_WORKBOOK_NAME = "2. 药政证书台账.xlsx"
CERTIFICATE_SHEET_CONFIG: tuple[dict[str, object], ...] = (
    {
        "key": "international-registration",
        "name": "国外注册",
        "columns": [
            "证照名称",
            "证书编号",
            "国家/发证机关",
            "发证日期",
            "有效期/复验期",
            "产品范围",
            "质量标准",
            "页数",
            "备注",
        ],
    },
    {
        "key": "domestic-registration",
        "name": "国内注册",
        "columns": [
            "证照名称",
            "受理号",
            "批件号",
            "编号",
            "发证机关",
            "发证日期",
            "有效期/复验期",
            "产品范围",
            "质量标准",
            "页数",
            "备注",
        ],
    },
    {
        "key": "domestic-gmp",
        "name": "国内GMP",
        "columns": [
            "证照名称",
            "编号",
            "发证机关",
            "发证日期",
            "有效期/复验期",
            "产品范围",
            "质量标准",
            "页数",
            "备注",
        ],
    },
    {
        "key": "international-gmp",
        "name": "国际GMP",
        "columns": [
            "证照名称",
            "编号",
            "国家/发证机关",
            "发证日期",
            "有效期/复验期",
            "产品范围",
            "质量标准",
            "页数",
            "备注",
        ],
    },
)
PLACEHOLDER_VALUES = {"", "-", "--", "——", "—", "/"}


def _get_certificate_workbook_path() -> Path:
    """获取证书工作簿路径，优先使用配置路径，回退到默认位置。"""
    settings = get_settings()
    workbook_dir = Path(settings.REGISTRATION_WORKBOOK_DIR)
    workbook_path = workbook_dir / settings.REGISTRATION_CERTIFICATE_WORKBOOK_NAME

    if not workbook_path.exists():
        desktop = Path.home() / "Desktop"
        legacy_path = desktop / "注册相关文件" / CERTIFICATE_WORKBOOK_NAME
        if legacy_path.exists():
            logger.warning(
                "使用桌面路径: %s，建议配置 REGISTRATION_WORKBOOK_DIR", legacy_path
            )
            return legacy_path

    return workbook_path


def _get_sheet_headers(sheet_meta: dict[str, object]) -> list[str]:
    return ["序号", *_get_sheet_columns(sheet_meta)]


def _get_sheet_columns(sheet_meta: dict[str, object]) -> list[str]:
    columns = sheet_meta.get("columns")
    if not isinstance(columns, list):
        return []
    return [str(item) for item in columns]


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, date):
        return value.strftime("%Y.%m.%d")
    text = str(value).replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def _preserve_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, date):
        return value.strftime("%Y.%m.%d")
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _slugify(text: str) -> str:
    ascii_text = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    if ascii_text:
        return ascii_text
    return hashlib.md5(text.encode("utf-8")).hexdigest()[:12]


def _normalize_optional(value: str | None) -> str | None:
    if value is None:
        return None
    text = value.strip()
    return None if text in PLACEHOLDER_VALUES else text


def _preserve_optional(value: str | None) -> str | None:
    if value is None:
        return None
    return value


def _is_template_placeholder_value(value: object | None) -> bool:
    if value is None:
        return False
    text = str(value)
    if not text:
        return False
    stripped = text.strip()
    if not stripped:
        return True
    if stripped in PLACEHOLDER_VALUES:
        return True
    return bool(re.fullmatch(r"[_\-\u2014/\s]+", text))


def _extract_date(value: str | None) -> date | None:
    if not value:
        return None
    normalized = value.strip()
    if normalized in PLACEHOLDER_VALUES or "长期" in normalized:
        return None
    matches = re.findall(r"(\d{4}[./-]\d{1,2}[./-]\d{1,2})", normalized)
    if not matches:
        return None
    candidate = matches[-1].replace(".", "-").replace("/", "-")
    try:
        return datetime.strptime(candidate, "%Y-%m-%d").date()
    except ValueError:
        return None


def _format_date(value: date | None) -> str | None:
    return value.strftime("%Y.%m.%d") if value else None


def _calculate_expiry_status(expiry_date: str | None) -> str:
    parsed = _extract_date(expiry_date)
    if parsed is None:
        return "无明确期限"
    delta_days = (parsed - date.today()).days
    if delta_days < 0:
        return "已过期"
    if delta_days <= 90:
        return "90天内到期"
    return "有效"


def _parse_page_count(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"\d+", value)
    return int(match.group()) if match else None


def _split_lines(value: str | None) -> list[str]:
    if not value:
        return []
    return [
        line.strip()
        for line in value.split("\n")
        if line.strip() and line.strip() not in PLACEHOLDER_VALUES
    ]


def _normalize_group_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip().lower()


def _normalize_department_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", "", value).upper()


def _is_qa_department(value: str | None) -> bool:
    normalized = _normalize_department_text(value)
    return "QA" in normalized or "质量保证" in normalized


def _get_sheet_meta(sheet_key: str) -> dict[str, object]:
    for item in CERTIFICATE_SHEET_CONFIG:
        if item["key"] == sheet_key:
            return item
    raise NotFoundException("证书子表", sheet_key)


def _get_sheet_meta_by_name(sheet_name: str) -> dict[str, object]:
    for item in CERTIFICATE_SHEET_CONFIG:
        if item["name"] == sheet_name:
            return item
    raise NotFoundException("证书子表", sheet_name)


def _extract_sequence(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = _normalize_text(value).strip()
    if text.isdigit():
        return int(text)
    return None


def _map_values_to_entry_payload(
    sheet_key: str,
    sheet_name: str,
    sheet_title: str,
    values: dict[str, str | None],
    *,
    source_sequence: int | None,
) -> dict[str, object | None]:
    issuing_authority = values.get("发证机关") or values.get("国家/发证机关")
    certificate_number = values.get("证书编号") or values.get("编号")
    validity_period = values.get("有效期/复验期")
    return {
        "sheet_key": sheet_key,
        "sheet_name": sheet_name,
        "sheet_title": sheet_title,
        "source_sequence": source_sequence,
        "certificate_name": _preserve_optional(values.get("证照名称")),
        "acceptance_number": _preserve_optional(values.get("受理号")),
        "approval_number": _preserve_optional(values.get("批件号")),
        "certificate_number": _preserve_optional(certificate_number),
        "issuing_authority": _preserve_optional(issuing_authority),
        "issue_date": _preserve_optional(values.get("发证日期")),
        "validity_period": _preserve_optional(validity_period),
        "expiry_date": _format_date(_extract_date(validity_period)),
        "product_scope": _preserve_optional(values.get("产品范围")),
        "quality_standard": _preserve_optional(values.get("质量标准")),
        "page_count": _parse_page_count(values.get("页数")),
        "remarks": _preserve_optional(values.get("备注")),
    }


def _build_row_values(entry: RegistrationCertificateEntry) -> dict[str, str | None]:
    if entry.sheet_key == "international-registration":
        return {
            "证照名称": entry.certificate_name,
            "证书编号": entry.certificate_number,
            "国家/发证机关": entry.issuing_authority,
            "发证日期": entry.issue_date,
            "有效期/复验期": entry.validity_period,
            "产品范围": entry.product_scope,
            "质量标准": entry.quality_standard,
            "页数": str(entry.page_count) if entry.page_count is not None else None,
            "备注": entry.remarks,
        }
    if entry.sheet_key == "domestic-registration":
        return {
            "证照名称": entry.certificate_name,
            "受理号": entry.acceptance_number,
            "批件号": entry.approval_number,
            "编号": entry.certificate_number,
            "发证机关": entry.issuing_authority,
            "发证日期": entry.issue_date,
            "有效期/复验期": entry.validity_period,
            "产品范围": entry.product_scope,
            "质量标准": entry.quality_standard,
            "页数": str(entry.page_count) if entry.page_count is not None else None,
            "备注": entry.remarks,
        }
    if entry.sheet_key == "domestic-gmp":
        return {
            "证照名称": entry.certificate_name,
            "编号": entry.certificate_number,
            "发证机关": entry.issuing_authority,
            "发证日期": entry.issue_date,
            "有效期/复验期": entry.validity_period,
            "产品范围": entry.product_scope,
            "质量标准": entry.quality_standard,
            "页数": str(entry.page_count) if entry.page_count is not None else None,
            "备注": entry.remarks,
        }
    return {
        "证照名称": entry.certificate_name,
        "编号": entry.certificate_number,
        "国家/发证机关": entry.issuing_authority,
        "发证日期": entry.issue_date,
        "有效期/复验期": entry.validity_period,
        "产品范围": entry.product_scope,
        "质量标准": entry.quality_standard,
        "页数": str(entry.page_count) if entry.page_count is not None else None,
        "备注": entry.remarks,
    }


def _build_sheet_row(entry: RegistrationCertificateEntry) -> CertificateSheetRow:
    return CertificateSheetRow(
        id=str(entry.id),
        sequence=entry.source_sequence or 0,
        values=_build_row_values(entry),
        issue_date=entry.issue_date,
        expiry_date=entry.expiry_date,
        expiry_status=_calculate_expiry_status(entry.expiry_date),
    )


def _build_record_summary(
    entry: RegistrationCertificateEntry,
) -> CertificateRecordSummary:
    return CertificateRecordSummary(
        id=str(entry.id),
        sheet_key=entry.sheet_key,
        sheet_name=entry.sheet_name,
        certificate_name=entry.certificate_name,
        certificate_number=entry.certificate_number,
        authority=entry.issuing_authority,
        issue_date=entry.issue_date,
        expiry_date=entry.expiry_date,
        expiry_status=_calculate_expiry_status(entry.expiry_date),
        product_scope=entry.product_scope,
        remarks=entry.remarks,
    )


def _build_sheet_summary(
    sheet_key: str,
    entries: list[RegistrationCertificateEntry],
) -> CertificateSheetSummary:
    meta = _get_sheet_meta(sheet_key)
    issuers = {entry.issuing_authority for entry in entries if entry.issuing_authority}
    products = {line for entry in entries for line in _split_lines(entry.product_scope)}
    return CertificateSheetSummary(
        sheet_key=sheet_key,
        sheet_name=str(meta["name"]),
        title=entries[0].sheet_title if entries else str(meta["name"]),
        total_records=len(entries),
        issuer_count=len(issuers),
        product_count=len(products),
        expired_count=sum(
            1
            for entry in entries
            if _calculate_expiry_status(entry.expiry_date) == "已过期"
        ),
        due_90_count=sum(
            1
            for entry in entries
            if _calculate_expiry_status(entry.expiry_date) == "90天内到期"
        ),
        total_pages=sum(entry.page_count or 0 for entry in entries),
    )


def _build_entry_response(
    entry: RegistrationCertificateEntry,
) -> CertificateEntryResponse:
    return CertificateEntryResponse(
        id=entry.id,
        sheet_key=entry.sheet_key,
        sheet_name=entry.sheet_name,
        sheet_title=entry.sheet_title,
        source_sequence=entry.source_sequence,
        certificate_name=entry.certificate_name,
        acceptance_number=entry.acceptance_number,
        approval_number=entry.approval_number,
        certificate_number=entry.certificate_number,
        issuing_authority=entry.issuing_authority,
        issue_date=entry.issue_date,
        validity_period=entry.validity_period,
        product_scope=entry.product_scope,
        quality_standard=entry.quality_standard,
        page_count=entry.page_count,
        remarks=entry.remarks,
        expiry_date=entry.expiry_date,
        expiry_status=_calculate_expiry_status(entry.expiry_date),
        created_at=entry.created_at,
        updated_at=entry.updated_at,
    )


def _build_reminder_content(
    entries: list[RegistrationCertificateEntry],
    reminder_days: int,
) -> str:
    lines = [
        f"以下证书已进入**到期前 {reminder_days} 天**提醒窗口，请及时处理：",
        "",
    ]
    preview_entries = entries[:10]
    for index, entry in enumerate(preview_entries, start=1):
        lines.append(
            "\n".join(
                [
                    f"{index}. **[{entry.sheet_name}] {entry.certificate_name}**",
                    f"   - 到期日期：{entry.expiry_date or '—'}",
                    f"   - 发证机关：{entry.issuing_authority or '—'}",
                    f"   - 产品范围：{entry.product_scope or '—'}",
                ]
            )
        )
    if len(entries) > len(preview_entries):
        lines.extend(
            [
                "",
                f"其余还有 **{len(entries) - len(preview_entries)}** 份，"
                "请到系统 `注册管理 -> 证书管理` 查看。",
            ]
        )
    return "\n".join(lines)


def _count_due_entries_for_setting(
    entries: list[RegistrationCertificateEntry],
    reminder_days: int,
) -> int:
    count = 0
    for entry in entries:
        expiry = _extract_date(entry.expiry_date)
        if expiry is None:
            continue
        days_until_expiry = (expiry - date.today()).days
        if 0 <= days_until_expiry <= reminder_days:
            count += 1
    return count


def _parse_sheet_entries(
    worksheet: Worksheet,
    sheet_meta: dict[str, object],
) -> list[RegistrationCertificateEntry]:
    expected_headers = _get_sheet_headers(sheet_meta)
    header_row = [
        _normalize_text(worksheet.cell(row=2, column=index).value)
        for index in range(1, len(expected_headers) + 1)
    ]
    if header_row != expected_headers:
        raise AppException(message=f"子表《{worksheet.title}》表头不匹配")

    sheet_key = str(sheet_meta["key"])
    sheet_name = str(sheet_meta["name"])
    sheet_title = _normalize_text(worksheet.cell(row=1, column=1).value) or sheet_name
    entries: list[RegistrationCertificateEntry] = []

    for row_index in range(3, worksheet.max_row + 1):
        raw_values = [
            _preserve_text(worksheet.cell(row=row_index, column=index).value)
            for index in range(1, len(expected_headers) + 1)
        ]
        normalized_values = [_normalize_text(value) for value in raw_values]
        if not any(normalized_values):
            continue

        non_sequence_values = normalized_values[1:]
        if not any(non_sequence_values):
            continue

        sequence = _extract_sequence(worksheet.cell(row=row_index, column=1).value)
        if sequence is None:
            continue

        values: dict[str, str | None] = {
            header: raw_values[index]
            for index, header in enumerate(expected_headers[1:], start=1)
        }
        payload = _map_values_to_entry_payload(
            sheet_key,
            sheet_name,
            sheet_title,
            values,
            source_sequence=sequence,
        )
        if not payload["certificate_name"]:
            raise AppException(
                message=f"子表《{sheet_name}》第 {row_index} 行证照名称为空"
            )
        entries.append(RegistrationCertificateEntry(**payload))

    return entries


def _clear_sheet_data_area(
    worksheet: Worksheet,
    *,
    row_indexes: set[int],
    column_count: int,
) -> None:
    if not row_indexes:
        return

    target_min_row = min(row_indexes)
    target_max_row = max(row_indexes)
    for merged_range in list(worksheet.merged_cells.ranges):
        if (
            merged_range.max_row < target_min_row
            or merged_range.min_row > target_max_row
        ):
            continue
        if merged_range.min_col > column_count or merged_range.max_col < 1:
            continue
        if not any(
            merged_range.min_row <= row_index <= merged_range.max_row
            for row_index in row_indexes
        ):
            continue
        worksheet.unmerge_cells(str(merged_range))

    for row_index in sorted(row_indexes):
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _copy_row_style(
    worksheet: Worksheet, source_row: int, target_row: int, column_count: int
) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[
        source_row
    ].height
    for column_index in range(1, column_count + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format


def _fill_certificate_sheet(
    worksheet: Worksheet,
    rows: list[dict[str, int | str | None]],
    *,
    headers: list[str],
    start_row: int,
    template_row: int,
) -> None:
    existing_row_values: dict[int, list[object | None]] = {
        row_index: [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, len(headers) + 1)
        ]
        for row_index in range(start_row, worksheet.max_row + 1)
    }
    template_fillable_rows: list[int] = []
    sequence_to_row: dict[int, int] = {}
    template_data_rows: set[int] = set()
    for row_index in range(start_row, worksheet.max_row + 1):
        sequence = _extract_sequence(worksheet.cell(row=row_index, column=1).value)
        normalized_row_values = [
            _normalize_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, len(headers) + 1)
        ]
        is_repeat_header = normalized_row_values == headers

        if is_repeat_header:
            continue

        template_fillable_rows.append(row_index)
        if sequence is not None:
            sequence_to_row.setdefault(sequence, row_index)

        normalized_values = [
            _normalize_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(2, len(headers) + 1)
        ]
        if any(normalized_values):
            template_data_rows.add(row_index)

    style_source_row = max(template_data_rows) if template_data_rows else template_row

    current_max_row = worksheet.max_row
    target_rows: list[int] = []
    used_template_rows: set[int] = set()
    rows_requiring_style_reset: set[int] = set()
    next_fallback_index = 0
    for row_values in rows:
        sequence = _extract_sequence(row_values.get("序号"))
        if sequence is not None:
            matched_row = sequence_to_row.get(sequence)
            if matched_row is not None and matched_row not in used_template_rows:
                target_rows.append(matched_row)
                used_template_rows.add(matched_row)
                if matched_row not in template_data_rows:
                    rows_requiring_style_reset.add(matched_row)
                continue

        while next_fallback_index < len(template_fillable_rows):
            candidate_row = template_fillable_rows[next_fallback_index]
            next_fallback_index += 1
            if candidate_row in used_template_rows:
                continue
            target_rows.append(candidate_row)
            used_template_rows.add(candidate_row)
            if candidate_row not in template_data_rows:
                rows_requiring_style_reset.add(candidate_row)
            break
        else:
            target_row = current_max_row + 1
            _copy_row_style(worksheet, style_source_row, target_row, len(headers))
            current_max_row = target_row
            target_rows.append(target_row)
            rows_requiring_style_reset.add(target_row)
    rows_to_clear = set(target_rows)
    rows_to_clear.update(
        row_index for row_index in template_data_rows if row_index not in target_rows
    )
    _clear_sheet_data_area(
        worksheet, row_indexes=rows_to_clear, column_count=len(headers)
    )

    for row_index in rows_requiring_style_reset:
        _copy_row_style(worksheet, style_source_row, row_index, len(headers))

    for row_values, target_row in zip(rows, target_rows, strict=False):
        for column_index, header in enumerate(headers, start=1):
            value = row_values.get(header)
            if value is None and target_row in existing_row_values:
                template_value = existing_row_values[target_row][column_index - 1]
                if _is_template_placeholder_value(template_value):
                    value = (
                        template_value
                        if isinstance(template_value, (int, str))
                        else str(template_value)
                    )
            worksheet.cell(row=target_row, column=column_index).value = value


def _load_entries_from_workbook_bytes(
    content: bytes,
) -> list[RegistrationCertificateEntry]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    try:
        entries: list[RegistrationCertificateEntry] = []
        for sheet_meta in CERTIFICATE_SHEET_CONFIG:
            sheet_name = str(sheet_meta["name"])
            if sheet_name not in workbook.sheetnames:
                raise AppException(message=f"导入失败：缺少子表《{sheet_name}》")
            entries.extend(_parse_sheet_entries(workbook[sheet_name], sheet_meta))
        return entries
    finally:
        workbook.close()


def _load_seed_entries() -> list[RegistrationCertificateEntry]:
    workbook_path = _get_certificate_workbook_path()
    if not workbook_path.exists():
        logger.warning(
            "药政证书台账种子文件不存在（%s），跳过种子加载；请先通过页面导入台账。",
            workbook_path,
        )
        return []
    return _load_entries_from_workbook_bytes(workbook_path.read_bytes())


class CertificateWorkbookService:
    """证书管理业务服务。"""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.repository = RegistrationCertificateRepository(session)

    async def _list_qa_reminder_recipient_options(
        self,
    ) -> list[CertificateReminderRecipientOption]:
        from app.modules.quality.public_api import (
            get_department_contact_list_from_feishu,
        )

        try:
            result = await get_department_contact_list_from_feishu(
                self.session,
                page=1,
                page_size=1000,
            )
        except Exception:
            # 降级策略：获取飞书联系人列表失败时返回空列表，不影响主流程
            logger.exception(
                "Failed to fetch QA reminder recipients from quality module"
            )
            return []

        options_by_open_id: dict[str, CertificateReminderRecipientOption] = {}
        for item in result.get("items", []):
            open_id = str(item.get("open_id") or "").strip()
            if not open_id:
                continue

            department = str(item.get("department") or "").strip() or None
            if not _is_qa_department(department):
                continue

            name = str(item.get("name") or "").strip() or "未命名联系人"
            enterprise_email = str(item.get("enterprise_email") or "").strip() or None
            options_by_open_id[open_id] = CertificateReminderRecipientOption(
                open_id=open_id,
                name=name,
                department=department,
                enterprise_email=enterprise_email,
            )

        return sorted(
            options_by_open_id.values(),
            key=lambda item: ((item.department or ""), item.name, item.open_id),
        )

    async def _get_qa_reminder_recipient_by_open_id(
        self,
        open_id: str,
    ) -> CertificateReminderRecipientOption | None:
        options = await self._list_qa_reminder_recipient_options()
        for item in options:
            if item.open_id == open_id:
                return item
        return None

    async def ensure_seeded(self) -> None:
        if await self.repository.count_entries() == 0:
            seed_entries = _load_seed_entries()
            if seed_entries:
                await self.repository.create_entries(seed_entries)
                await self.session.commit()

    async def get_reminder_settings(self) -> CertificateReminderSettingResponse:
        await self.ensure_seeded()
        setting = await self.repository.get_reminder_setting()

        if setting is None:
            return CertificateReminderSettingResponse(
                is_enabled=False,
                reminder_days=90,
                recipient_open_id=None,
                recipient_name=None,
                recipient_department=None,
                pending_count=0,
            )

        pending_count = 0
        if setting.is_enabled and setting.recipient_open_id:
            due_batches = await self.find_due_reminder_batch()
            entries = due_batches[0].get("entries") if due_batches else None
            pending_count = len(entries) if isinstance(entries, list) else 0
        return CertificateReminderSettingResponse(
            is_enabled=setting.is_enabled,
            reminder_days=setting.reminder_days,
            recipient_open_id=setting.recipient_open_id,
            recipient_name=setting.recipient_name,
            recipient_department=setting.recipient_department,
            pending_count=pending_count,
        )

    async def list_reminder_recipient_options(
        self,
    ) -> list[CertificateReminderRecipientOption]:
        return await self._list_qa_reminder_recipient_options()

    async def update_reminder_settings(
        self,
        data: CertificateReminderSettingUpdate,
    ) -> CertificateReminderSettingResponse:
        await self.ensure_seeded()
        recipient_open_id = (data.recipient_open_id or "").strip() or None
        recipient_name: str | None = None
        recipient_department: str | None = None

        if data.is_enabled:
            if not recipient_open_id:
                raise AppException(message="启用自动提醒时必须选择通知人")
            recipient = await self._get_qa_reminder_recipient_by_open_id(
                recipient_open_id
            )
            if recipient is None:
                raise AppException(message="所选通知人不在 QA 联系人范围内")
            recipient_name = recipient.name
            recipient_department = recipient.department
        else:
            recipient_open_id = None

        setting = await self.repository.get_reminder_setting()
        await self.repository.save_reminder_setting(
            setting=setting,
            is_enabled=data.is_enabled,
            reminder_days=data.reminder_days,
            recipient_open_id=recipient_open_id,
            recipient_name=recipient_name,
            recipient_department=recipient_department,
        )
        await self.session.commit()
        return await self.get_reminder_settings()

    async def get_overview(self) -> CertificateWorkbookOverview:
        await self.ensure_seeded()
        workbook_path = _get_certificate_workbook_path()
        entries = await self.repository.list_entries()

        sheet_summaries: list[CertificateSheetSummary] = []
        for sheet_meta in CERTIFICATE_SHEET_CONFIG:
            sheet_key = str(sheet_meta["key"])
            sheet_entries = [entry for entry in entries if entry.sheet_key == sheet_key]
            sheet_summaries.append(_build_sheet_summary(sheet_key, sheet_entries))

        record_summaries = [_build_record_summary(entry) for entry in entries]
        upcoming_expirations = [
            record
            for record in record_summaries
            if _extract_date(record.expiry_date) is not None
        ]
        upcoming_expirations.sort(
            key=lambda item: _extract_date(item.expiry_date) or date.max
        )

        recent_issued = [
            record
            for record in record_summaries
            if _extract_date(record.issue_date) is not None
        ]
        recent_issued.sort(
            key=lambda item: _extract_date(item.issue_date) or date.min, reverse=True
        )

        return CertificateWorkbookOverview(
            workbook_name=workbook_path.name,
            updated_at=datetime.fromtimestamp(workbook_path.stat().st_mtime)
            if workbook_path.exists()
            else None,
            total_records=len(entries),
            sheet_count=len(CERTIFICATE_SHEET_CONFIG),
            issuer_count=len(
                {
                    entry.issuing_authority
                    for entry in entries
                    if entry.issuing_authority
                }
            ),
            product_count=len(
                {
                    line
                    for entry in entries
                    for line in _split_lines(entry.product_scope)
                }
            ),
            expired_count=sum(
                1
                for entry in entries
                if _calculate_expiry_status(entry.expiry_date) == "已过期"
            ),
            due_90_count=sum(
                1
                for entry in entries
                if _calculate_expiry_status(entry.expiry_date) == "90天内到期"
            ),
            total_pages=sum(entry.page_count or 0 for entry in entries),
            sheet_summaries=sheet_summaries,
            upcoming_expirations=upcoming_expirations[:8],
            recent_issued=recent_issued[:8],
        )

    async def get_workbook_detail(self) -> CertificateWorkbookDetail:
        await self.ensure_seeded()
        workbook_path = _get_certificate_workbook_path()
        entries = await self.repository.list_entries()

        sheets: list[CertificateWorkbookSheet] = []
        for sheet_meta in CERTIFICATE_SHEET_CONFIG:
            sheet_key = str(sheet_meta["key"])
            sheet_name = str(sheet_meta["name"])
            sheet_entries = [entry for entry in entries if entry.sheet_key == sheet_key]
            sheets.append(
                CertificateWorkbookSheet(
                    sheet_key=sheet_key,
                    sheet_name=sheet_name,
                    title=sheet_entries[0].sheet_title if sheet_entries else sheet_name,
                    columns=[
                        CertificateColumn(key=_slugify(label), label=label)
                        for label in _get_sheet_columns(sheet_meta)
                    ],
                    rows=[_build_sheet_row(entry) for entry in sheet_entries],
                    summary=_build_sheet_summary(sheet_key, sheet_entries),
                )
            )

        return CertificateWorkbookDetail(
            workbook_name=workbook_path.name,
            updated_at=datetime.fromtimestamp(workbook_path.stat().st_mtime)
            if workbook_path.exists()
            else None,
            sheets=sheets,
        )

    async def import_workbook(
        self, upload_file: UploadFile
    ) -> CertificateWorkbookImportResult:
        filename, content = await read_upload_secure(
            upload_file,
            max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
            allowed_extensions={".xlsx"},
            what="药政证书台账",
        )

        await self.ensure_seeded()
        old_entries = await self.repository.list_entries()
        parsed_entries = await asyncio.to_thread(
            _load_entries_from_workbook_bytes, content
        )

        try:
            await self.repository.soft_delete_many(old_entries)
            if parsed_entries:
                await self.repository.create_entries(parsed_entries)
            await self.session.commit()
        except Exception:
            await self.session.rollback()
            raise

        # 配置文件缺失时，将本次上传落盘到配置路径（create-if-missing，绝不覆盖）
        config_path = _get_certificate_workbook_path()
        if not config_path.exists():
            try:
                config_path.parent.mkdir(parents=True, exist_ok=True)
                config_path.write_bytes(content)
                logger.info("药政证书台账配置文件已从上传创建: %s", config_path)
            except OSError as exc:
                logger.warning(
                    "药政证书台账配置文件落盘失败（%s），数据已入库但导出/种子将不可用",
                    exc,
                )

        return CertificateWorkbookImportResult(
            workbook_name=filename,
            imported_sheet_count=len(CERTIFICATE_SHEET_CONFIG),
            imported_record_count=len(parsed_entries),
            replaced_record_count=len(old_entries),
        )

    async def export_workbook(self) -> tuple[Path, str]:
        await self.ensure_seeded()
        workbook_path = _get_certificate_workbook_path()
        if not workbook_path.exists():
            raise NotFoundException(
                "药政证书台账文件",
                f"{workbook_path}（请先在页面导入台账文件后再导出）",
            )

        temp_dir = Path(tempfile.mkdtemp(prefix="certificate-workbook-export-"))
        export_path = temp_dir / "药政证书台账-导出.xlsx"
        await asyncio.to_thread(shutil.copyfile, workbook_path, export_path)

        workbook = await asyncio.to_thread(load_workbook, export_path)
        try:
            entries = await self.repository.list_entries()
            for sheet_meta in CERTIFICATE_SHEET_CONFIG:
                sheet_key = str(sheet_meta["key"])
                sheet_name = str(sheet_meta["name"])
                headers = _get_sheet_headers(sheet_meta)
                worksheet = workbook[sheet_name]
                sheet_entries = [
                    entry for entry in entries if entry.sheet_key == sheet_key
                ]
                row_values = [
                    {"序号": entry.source_sequence or index, **_build_row_values(entry)}
                    for index, entry in enumerate(sheet_entries, start=1)
                ]
                _fill_certificate_sheet(
                    worksheet,
                    row_values,
                    headers=headers,
                    start_row=3,
                    template_row=3,
                )
            await asyncio.to_thread(workbook.save, export_path)
        finally:
            workbook.close()

        return export_path, "药政证书台账-导出.xlsx"

    async def get_sheet_detail(self, sheet_key: str) -> CertificateSheetDetail:
        await self.ensure_seeded()
        meta = _get_sheet_meta(sheet_key)
        entries = await self.repository.list_entries(sheet_key=sheet_key)
        summary = _build_sheet_summary(sheet_key, entries)
        return CertificateSheetDetail(
            sheet_key=sheet_key,
            sheet_name=str(meta["name"]),
            title=entries[0].sheet_title if entries else str(meta["name"]),
            source_file_name=CERTIFICATE_WORKBOOK_NAME,
            updated_at=datetime.fromtimestamp(
                _get_certificate_workbook_path().stat().st_mtime
            )
            if _get_certificate_workbook_path().exists()
            else None,
            columns=[
                CertificateColumn(key=_slugify(label), label=label)
                for label in _get_sheet_columns(meta)
            ],
            rows=[_build_sheet_row(entry) for entry in entries],
            summary=summary,
        )

    async def create_entry(
        self, data: CertificateEntryCreate
    ) -> CertificateEntryResponse:
        await self.ensure_seeded()
        meta = _get_sheet_meta(data.sheet_key)
        payload = data.model_dump()
        validity_period = payload.get("validity_period")
        source_sequence = payload.get("source_sequence")
        if not source_sequence or source_sequence < 1:
            source_sequence = await self.repository.get_next_source_sequence(
                data.sheet_key
            )
        entry = RegistrationCertificateEntry(
            sheet_key=data.sheet_key,
            sheet_name=str(meta["name"]),
            sheet_title=str(meta["name"]),
            source_sequence=source_sequence,
            certificate_name=data.certificate_name,
            acceptance_number=payload.get("acceptance_number"),
            approval_number=payload.get("approval_number"),
            certificate_number=payload.get("certificate_number"),
            issuing_authority=payload.get("issuing_authority"),
            issue_date=payload.get("issue_date"),
            validity_period=validity_period,
            expiry_date=_format_date(_extract_date(validity_period)),
            product_scope=payload.get("product_scope"),
            quality_standard=payload.get("quality_standard"),
            page_count=payload.get("page_count"),
            remarks=payload.get("remarks"),
        )
        created = await self.repository.create_entry(entry)
        await self.session.flush()
        await self.session.commit()
        return _build_entry_response(created)

    async def update_entry(
        self, entry_id: UUID, data: CertificateEntryUpdate
    ) -> CertificateEntryResponse:
        await self.ensure_seeded()
        entry = await self.repository.get_by_id(entry_id)
        if entry is None:
            raise NotFoundException("证书台账记录", str(entry_id))

        payload = data.model_dump(exclude_unset=True)
        if "validity_period" in payload:
            payload["expiry_date"] = _format_date(
                _extract_date(payload.get("validity_period"))
            )

        updated = await self.repository.update_entry(entry, payload)
        await self.session.commit()
        return _build_entry_response(updated)

    async def delete_entry(self, entry_id: UUID) -> None:
        await self.ensure_seeded()
        entry = await self.repository.get_by_id(entry_id)
        if entry is None:
            raise NotFoundException("证书台账记录", str(entry_id))
        await self.repository.soft_delete(entry)
        await self.session.commit()

    async def find_due_reminder_batch(self) -> list[dict[str, object]]:
        # 独立通知应用尚未接入；不生成任务或写入已发送记录。
        return []

    async def send_due_reminder_batch(self, item: dict[str, object]) -> None:
        raise AppException(
            status_code=503,
            message="注册证书飞书提醒暂未启用，待配置独立业务应用后恢复",
        )


async def find_due_certificate_reminder_batches(
    session: AsyncSession,
) -> list[dict[str, object]]:
    return await CertificateWorkbookService(session).find_due_reminder_batch()


async def send_certificate_reminder_batch(
    session: AsyncSession,
    item: dict[str, object],
) -> None:
    await CertificateWorkbookService(session).send_due_reminder_batch(item)
