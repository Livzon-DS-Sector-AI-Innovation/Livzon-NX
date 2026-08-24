"""Project ledger workbook service."""

from __future__ import annotations

import hashlib
import logging
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import MergedCell  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.exceptions import AppException, NotFoundException
from app.core.upload_security import read_upload_secure
from app.modules.registration.models import RegistrationProjectLedgerVersion
from app.modules.registration.repository import RegistrationProjectLedgerRepository
from app.modules.registration.schemas.project_ledger import (
    ProjectLedgerColumn,
    ProjectLedgerEntryInput,
    ProjectLedgerEntryResponse,
    ProjectLedgerHistoryRecord,
    ProjectLedgerRecord,
    ProjectLedgerRecordHistory,
    ProjectLedgerSheetDetail,
    ProjectLedgerSheetSummary,
    ProjectLedgerWorkbookDetail,
    ProjectLedgerWorkbookImportResult,
    ProjectLedgerWorkbookOverview,
)

logger = logging.getLogger(__name__)

PROJECT_LEDGER_WORKBOOK_NAME = "1. 注册台账.xlsx"
PROJECT_LEDGER_SHEET_CONFIG: dict[str, tuple[str, str]] = {
    "1.国际注册（关联审评机制）": (
        "international-associated-review",
        "国际注册（关联审评机制）",
    ),
    "2.国际注册（原料药单独审评机制）": (
        "international-standalone-review",
        "国际注册（原料药单独审评机制）",
    ),
    "3.国内注册（关联审评机制）": (
        "domestic-associated-review",
        "国内注册（关联审评机制）",
    ),
    "4.国内注册（原料药单独审评机制）": (
        "domestic-standalone-review",
        "国内注册（原料药单独审评机制）",
    ),
}


@dataclass(slots=True)
class ProjectLedgerSheetDefinition:
    sheet_key: str
    sheet_name: str
    sheet_title: str
    worksheet_title: str
    columns: list[ProjectLedgerColumn]


def _get_workbook_path() -> Path:
    """获取申报台账工作簿路径，优先使用配置路径，回退到默认位置。"""
    settings = get_settings()
    workbook_dir = Path(settings.REGISTRATION_WORKBOOK_DIR)
    workbook_path = workbook_dir / settings.REGISTRATION_PROJECT_LEDGER_WORKBOOK_NAME

    # 如果配置路径不存在，回退到桌面路径（向后兼容）
    if not workbook_path.exists():
        desktop = Path.home() / "Desktop"
        legacy_path = desktop / "注册相关文件" / PROJECT_LEDGER_WORKBOOK_NAME
        if legacy_path.exists():
            logger.warning(
                "使用桌面路径: %s，建议配置 REGISTRATION_WORKBOOK_DIR", legacy_path
            )
            return legacy_path

    return workbook_path


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def _slugify(text: str) -> str:
    ascii_text = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    if ascii_text:
        return ascii_text
    return hashlib.md5(text.encode("utf-8")).hexdigest()[:12]


def _build_column_key(label: str, used_keys: set[str]) -> str:
    base_key = _slugify(label)
    candidate = base_key
    suffix = 2
    while candidate in used_keys:
        candidate = f"{base_key}_{suffix}"
        suffix += 1
    used_keys.add(candidate)
    return candidate


def _normalize_header_label(
    raw_label: str, previous_label: str | None, index: int
) -> str:
    label = raw_label.strip()
    if label:
        return label
    if previous_label and "药政活动" in previous_label:
        return "药政活动说明"
    return f"补充信息{index + 1}"


def _normalize_header_labels(raw_labels: list[str]) -> list[str]:
    labels: list[str] = []
    previous_label: str | None = None
    for index, raw_label in enumerate(raw_labels):
        label = _normalize_header_label(raw_label, previous_label, index)
        labels.append(label)
        previous_label = label
    return labels


def _extract_sequence_number(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"\d+", value)
    return int(match.group()) if match else None


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = value.strip()
    return text or None


def _resolve_sheet_key_and_name(worksheet_title: str) -> tuple[str, str]:
    return PROJECT_LEDGER_SHEET_CONFIG.get(
        worksheet_title,
        (_slugify(worksheet_title), worksheet_title),
    )


def _parse_workbook_definitions() -> tuple[
    list[ProjectLedgerSheetDefinition], datetime
]:
    workbook_path = _get_workbook_path()
    if not workbook_path.exists():
        raise NotFoundException("注册台账文件", str(workbook_path))

    workbook = load_workbook(workbook_path, data_only=True)
    try:
        definitions: list[ProjectLedgerSheetDefinition] = []
        for worksheet in workbook.worksheets:
            raw_rows = [
                [_normalize_text(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            visible_rows = [row for row in raw_rows if any(cell for cell in row)]
            if len(visible_rows) < 2:
                continue

            last_used_index = 0
            for row in visible_rows[1:]:
                for index, value in enumerate(row, start=1):
                    if value:
                        last_used_index = max(last_used_index, index)
            if last_used_index == 0:
                continue

            title = visible_rows[0][0] or worksheet.title
            header_row = visible_rows[1][:last_used_index]
            used_keys: set[str] = set()
            columns = [
                ProjectLedgerColumn(
                    key=_build_column_key(label, used_keys),
                    label=label,
                )
                for label in _normalize_header_labels(header_row)
            ]

            sheet_key, sheet_name = _resolve_sheet_key_and_name(worksheet.title)
            definitions.append(
                ProjectLedgerSheetDefinition(
                    sheet_key=sheet_key,
                    sheet_name=sheet_name,
                    sheet_title=title,
                    worksheet_title=worksheet.title,
                    columns=columns,
                )
            )

        return definitions, datetime.fromtimestamp(workbook_path.stat().st_mtime)
    finally:
        workbook.close()


def _find_sheet_definition(sheet_key: str) -> ProjectLedgerSheetDefinition:
    definitions, _ = _parse_workbook_definitions()
    for definition in definitions:
        if definition.sheet_key == sheet_key:
            return definition
    raise NotFoundException("申报台账子表", sheet_key)


def _load_seed_versions() -> list[RegistrationProjectLedgerVersion]:
    return _load_versions_from_workbook_path(_get_workbook_path())


def _load_versions_from_worksheet(
    definition: ProjectLedgerSheetDefinition,
    worksheet: Any,
) -> list[RegistrationProjectLedgerVersion]:
    expected_labels = [column.label for column in definition.columns]
    header_row = [
        _normalize_text(worksheet.cell(row=2, column=index).value)
        for index in range(1, len(definition.columns) + 1)
    ]
    if _normalize_header_labels(header_row) != expected_labels:
        raise AppException(message=f"子表《{worksheet.title}》表头不匹配")

    grouped_rows: list[list[tuple[int, dict[str, str | None]]]] = []
    current_group: list[tuple[int, dict[str, str | None]]] = []
    sequence_key = definition.columns[0].key
    project_key = next(
        (column.key for column in definition.columns if "项目" in column.label),
        None,
    )
    product_key = next(
        (column.key for column in definition.columns if column.label == "产品"),
        None,
    )

    for source_row_number in range(3, worksheet.max_row + 1):
        values: dict[str, str | None] = {}
        for index, column in enumerate(definition.columns, start=1):
            cell_value = _normalize_text(
                worksheet.cell(row=source_row_number, column=index).value
            )
            values[column.key] = _normalize_optional_text(cell_value)

        if not any(value is not None for value in values.values()):
            continue

        # Some workbook templates contain trailing scaffold rows where only the
        # sequence cell is populated by a merged block and the business cells
        # are otherwise empty. Those rows are not real records and must be
        # ignored, otherwise later exports get shifted by one full record block.
        non_sequence_values = [
            value
            for key, value in values.items()
            if key != sequence_key and value is not None
        ]
        if values.get(sequence_key) is not None and not non_sequence_values:
            continue

        if values.get(sequence_key) is not None:
            if current_group:
                grouped_rows.append(current_group)
            current_group = [(source_row_number, values)]
            continue

        if current_group:
            current_group.append((source_row_number, values))

    if current_group:
        grouped_rows.append(current_group)

    versions: list[RegistrationProjectLedgerVersion] = []
    for fallback_sequence, group_rows in enumerate(grouped_rows, start=1):
        record_group_id = RegistrationProjectLedgerRepository.generate_group_id()
        snapshot_values: dict[str, str | None] = {}
        first_row_values = group_rows[0][1]
        source_sequence = (
            _extract_sequence_number(first_row_values.get(sequence_key))
            or fallback_sequence
        )

        for version_index, (source_row_number, row_values) in enumerate(
            group_rows, start=1
        ):
            current_snapshot = snapshot_values.copy()
            for column in definition.columns:
                next_value = row_values.get(column.key)
                if next_value is not None:
                    current_snapshot[column.key] = next_value

            current_snapshot[sequence_key] = str(source_sequence)
            versions.append(
                RegistrationProjectLedgerVersion(
                    record_group_id=record_group_id,
                    sheet_key=definition.sheet_key,
                    sheet_name=definition.sheet_name,
                    sheet_title=definition.sheet_title,
                    source_sequence=source_sequence,
                    version_number=version_index,
                    source_row_number=source_row_number,
                    values_data=current_snapshot.copy(),
                    project_name=current_snapshot.get(project_key)
                    if project_key
                    else None,
                    product_name=current_snapshot.get(product_key)
                    if product_key
                    else None,
                )
            )
            snapshot_values = current_snapshot

    return versions


def _load_versions_from_workbook_path(
    workbook_path: Path,
) -> list[RegistrationProjectLedgerVersion]:
    definitions, _ = _parse_workbook_definitions()
    definition_map = {
        definition.worksheet_title: definition for definition in definitions
    }

    workbook = load_workbook(workbook_path, data_only=True)
    try:
        missing_sheet_titles = [
            definition.worksheet_title
            for definition in definitions
            if definition.worksheet_title not in workbook.sheetnames
        ]
        if missing_sheet_titles:
            raise AppException(
                message=f"导入失败：缺少子表《{missing_sheet_titles[0]}》"
            )

        unexpected_sheet_titles = [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.title not in definition_map
            and any(
                _normalize_text(cell)
                for row in worksheet.iter_rows(values_only=True)
                for cell in row
            )
        ]
        if unexpected_sheet_titles:
            raise AppException(
                message=f"导入失败：存在未定义子表《{unexpected_sheet_titles[0]}》"
            )

        versions: list[RegistrationProjectLedgerVersion] = []
        for definition in definitions:
            versions.extend(
                _load_versions_from_worksheet(
                    definition,
                    workbook[definition.worksheet_title],
                )
            )
        return versions
    finally:
        workbook.close()


def _build_history_record(
    version: RegistrationProjectLedgerVersion,
) -> ProjectLedgerHistoryRecord:
    return ProjectLedgerHistoryRecord(
        entry_id=version.id,
        version=version.version_number,
        source_row_number=version.source_row_number,
        values=version.values_data,
    )


def _build_record(
    versions: list[RegistrationProjectLedgerVersion],
    *,
    include_history_records: bool = True,
) -> ProjectLedgerRecord:
    latest = versions[-1]
    return ProjectLedgerRecord(
        record_id=latest.record_group_id,
        record_key=str(latest.record_group_id),
        sequence=latest.source_sequence,
        latest_values=latest.values_data,
        history_count=len(versions),
        history_records=(
            [_build_history_record(version) for version in versions]
            if include_history_records
            else []
        ),
    )


def _build_sheet_summary(
    definition: ProjectLedgerSheetDefinition,
    records: list[ProjectLedgerRecord],
) -> ProjectLedgerSheetSummary:
    return ProjectLedgerSheetSummary(
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        title=definition.sheet_title,
        total_records=len(records),
        records_with_history=sum(1 for item in records if item.history_count > 1),
        total_history_versions=sum(max(item.history_count - 1, 0) for item in records),
    )


def _build_sheet_detail(
    definition: ProjectLedgerSheetDefinition,
    versions: list[RegistrationProjectLedgerVersion],
    *,
    include_columns: bool = True,
    include_records: bool = True,
    include_history_records: bool = True,
) -> ProjectLedgerSheetDetail:
    grouped_versions: dict[object, list[RegistrationProjectLedgerVersion]] = (
        defaultdict(list)
    )
    for version in versions:
        grouped_versions[version.record_group_id].append(version)

    records = (
        [
            _build_record(group, include_history_records=include_history_records)
            for group in grouped_versions.values()
        ]
        if include_records
        else []
    )
    records.sort(key=lambda item: item.sequence)
    summary = _build_sheet_summary(
        definition,
        records
        if include_records
        else [
            ProjectLedgerRecord(
                record_id=group[-1].record_group_id,
                record_key=str(group[-1].record_group_id),
                sequence=group[-1].source_sequence,
                latest_values={},
                history_count=len(group),
                history_records=[],
            )
            for group in grouped_versions.values()
        ],
    )
    return ProjectLedgerSheetDetail(
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        title=definition.sheet_title,
        columns=definition.columns if include_columns else [],
        records=records,
        summary=summary,
    )


def _normalize_entry_values(
    definition: ProjectLedgerSheetDefinition,
    values: dict[str, str | None],
) -> dict[str, str | None]:
    normalized: dict[str, str | None] = {}
    for column in definition.columns:
        normalized[column.key] = _normalize_optional_text(values.get(column.key))
    return normalized


def _build_entry_response(
    version: RegistrationProjectLedgerVersion,
) -> ProjectLedgerEntryResponse:
    return ProjectLedgerEntryResponse(
        record_id=version.record_group_id,
        sheet_key=version.sheet_key,
        version_number=version.version_number,
        sequence=version.source_sequence,
        values=version.values_data,
    )


def _get_project_name_key(definition: ProjectLedgerSheetDefinition) -> str | None:
    return next(
        (column.key for column in definition.columns if "项目" in column.label), None
    )


def _get_product_name_key(definition: ProjectLedgerSheetDefinition) -> str | None:
    return next(
        (column.key for column in definition.columns if column.label == "产品"), None
    )


def _copy_project_ledger_row_style(
    worksheet: Any,
    source_row: int,
    target_row: int,
    column_count: int,
) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[
        source_row
    ].height
    for column_index in range(1, column_count + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format


def _reset_project_ledger_merges(
    worksheet: Any,
    *,
    start_row: int,
    end_row: int,
    column_count: int,
) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        if merged_range.max_col < 1 or merged_range.min_col > column_count:
            continue
        worksheet.unmerge_cells(str(merged_range))


def _apply_project_ledger_vertical_merges(
    worksheet: Any,
    definition: ProjectLedgerSheetDefinition,
    rows: list[dict[str, str | None]],
    *,
    start_row: int,
) -> None:
    if not rows:
        return

    sequence_key = definition.columns[0].key
    group_start = 0
    total_rows = len(rows)

    for row_index in range(1, total_rows + 1):
        next_is_group_start = (
            row_index == total_rows or rows[row_index].get(sequence_key) is not None
        )
        if not next_is_group_start:
            continue

        group_end = row_index - 1
        if group_end > group_start:
            sheet_start_row = start_row + group_start
            sheet_end_row = start_row + group_end
            for column_index, column in enumerate(definition.columns, start=1):
                first_value = rows[group_start].get(column.key)
                if first_value in (None, ""):
                    continue
                if all(
                    rows[item_index].get(column.key) in (None, "")
                    for item_index in range(group_start + 1, group_end + 1)
                ):
                    worksheet.merge_cells(
                        start_row=sheet_start_row,
                        start_column=column_index,
                        end_row=sheet_end_row,
                        end_column=column_index,
                    )

        group_start = row_index


def _build_export_rows(
    definition: ProjectLedgerSheetDefinition,
    versions: list[RegistrationProjectLedgerVersion],
) -> list[dict[str, str | None]]:
    grouped_versions: dict[object, list[RegistrationProjectLedgerVersion]] = (
        defaultdict(list)
    )
    for version in versions:
        grouped_versions[version.record_group_id].append(version)

    sequence_key = definition.columns[0].key
    rows: list[dict[str, str | None]] = []
    for group_versions in grouped_versions.values():
        previous_values: dict[str, str | None] = {}
        for index, version in enumerate(group_versions):
            row_values: dict[str, str | None] = {}
            for column in definition.columns:
                current_value = version.values_data.get(column.key)
                if index == 0:
                    row_values[column.key] = current_value
                    continue
                if column.key == sequence_key:
                    row_values[column.key] = None
                    continue
                row_values[column.key] = (
                    current_value
                    if current_value is not None
                    and current_value != previous_values.get(column.key)
                    else None
                )
            row_values[sequence_key] = (
                str(version.source_sequence) if index == 0 else None
            )
            rows.append(row_values)
            previous_values = version.values_data

    return rows


def _fill_project_ledger_sheet(
    worksheet: Any,
    definition: ProjectLedgerSheetDefinition,
    rows: list[dict[str, str | None]],
    *,
    start_row: int = 3,
    template_row: int = 3,
) -> None:
    column_count = len(definition.columns)
    populated_rows = [
        row_index
        for row_index in range(start_row, worksheet.max_row + 1)
        if any(
            _normalize_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, column_count + 1)
        )
    ]
    last_data_row = max(populated_rows, default=template_row)
    target_last_row = max(last_data_row, start_row + len(rows) - 1)

    if target_last_row >= start_row:
        _reset_project_ledger_merges(
            worksheet,
            start_row=start_row,
            end_row=target_last_row,
            column_count=column_count,
        )

    for row_index in range(start_row, target_last_row + 1):
        if row_index > worksheet.max_row:
            _copy_project_ledger_row_style(
                worksheet, template_row, row_index, column_count
            )
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    for row_offset, row_values in enumerate(rows):
        row_index = start_row + row_offset
        for column_index, column in enumerate(definition.columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            # Export rows intentionally use None to keep repeated main-record
            # columns blank on history rows, so never refill from template data.
            cell.value = row_values.get(column.key)

    _apply_project_ledger_vertical_merges(
        worksheet,
        definition,
        rows,
        start_row=start_row,
    )


class ProjectLedgerWorkbookService:
    """Project ledger service with persisted versions."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.repository = RegistrationProjectLedgerRepository(session)

    async def ensure_seeded(self) -> None:
        if await self.repository.count_versions() == 0:
            versions = _load_seed_versions()
            if versions:
                await self.repository.create_versions(versions)
                await self.session.commit()

    async def get_workbook_detail(self) -> ProjectLedgerWorkbookDetail:
        overview = await self.get_overview()
        return ProjectLedgerWorkbookDetail(
            workbook_name=overview.workbook_name,
            updated_at=overview.updated_at,
            total_records=overview.total_records,
            sheets=overview.sheets,
        )

    async def get_sheet_detail(self, sheet_key: str) -> ProjectLedgerSheetDetail:
        await self.ensure_seeded()
        definition = _find_sheet_definition(sheet_key)
        versions = await self.repository.list_versions(sheet_key=sheet_key)
        return _build_sheet_detail(
            definition,
            versions,
            include_columns=True,
            include_records=True,
            include_history_records=False,
        )

    async def get_entry_history(self, record_id: Any) -> ProjectLedgerRecordHistory:
        await self.ensure_seeded()
        versions = await self.repository.list_versions_by_group(record_id)
        if not versions:
            raise NotFoundException("申报台账记录", str(record_id))
        latest = versions[-1]
        return ProjectLedgerRecordHistory(
            record_id=latest.record_group_id,
            sheet_key=latest.sheet_key,
            sequence=latest.source_sequence,
            history_count=len(versions),
            history_records=[_build_history_record(version) for version in versions],
        )

    async def get_overview(self) -> ProjectLedgerWorkbookOverview:
        await self.ensure_seeded()
        definitions, updated_at = _parse_workbook_definitions()
        versions = await self.repository.list_versions()

        grouped_by_sheet: dict[
            str, dict[object, list[RegistrationProjectLedgerVersion]]
        ] = defaultdict(lambda: defaultdict(list))
        for version in versions:
            grouped_by_sheet[version.sheet_key][version.record_group_id].append(version)

        sheets: list[ProjectLedgerSheetDetail] = []
        for definition in definitions:
            sheet_versions = [
                version
                for group_versions in grouped_by_sheet.get(
                    definition.sheet_key, {}
                ).values()
                for version in group_versions
            ]
            sheets.append(
                _build_sheet_detail(
                    definition,
                    sheet_versions,
                    include_columns=False,
                    include_records=False,
                    include_history_records=False,
                )
            )

        return ProjectLedgerWorkbookOverview(
            workbook_name=PROJECT_LEDGER_WORKBOOK_NAME,
            updated_at=updated_at,
            total_records=sum(sheet.summary.total_records for sheet in sheets),
            records_with_history=sum(
                sheet.summary.records_with_history for sheet in sheets
            ),
            total_history_versions=sum(
                sheet.summary.total_history_versions for sheet in sheets
            ),
            sheets=sheets,
        )

    async def import_workbook(
        self, upload_file: UploadFile
    ) -> ProjectLedgerWorkbookImportResult:
        filename, content = await read_upload_secure(
            upload_file,
            max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
            allowed_extensions={".xlsx"},
            what="申报台账工作簿",
        )

        temp_dir = Path(tempfile.mkdtemp(prefix="project-ledger-import-"))
        import_path = temp_dir / (filename or PROJECT_LEDGER_WORKBOOK_NAME)
        import_path.write_bytes(content)

        await self.ensure_seeded()
        versions = _load_versions_from_workbook_path(import_path)

        try:
            await self.repository.replace_all_versions(versions)
            await self.session.commit()
        except Exception:
            await self.session.rollback()
            raise

        sheet_record_counts = Counter(
            version.sheet_key for version in versions if version.version_number == 1
        )
        return ProjectLedgerWorkbookImportResult(
            workbook_name=filename or PROJECT_LEDGER_WORKBOOK_NAME,
            imported_records=sum(sheet_record_counts.values()),
            sheet_record_counts=dict(sheet_record_counts),
        )

    async def export_workbook(self) -> tuple[Path, str]:
        await self.ensure_seeded()
        workbook_path = _get_workbook_path()
        if not workbook_path.exists():
            raise NotFoundException("注册台账文件", str(workbook_path))

        definitions, _ = _parse_workbook_definitions()
        temp_dir = Path(tempfile.mkdtemp(prefix="project-ledger-export-"))
        export_path = temp_dir / PROJECT_LEDGER_WORKBOOK_NAME
        shutil.copyfile(workbook_path, export_path)

        workbook = load_workbook(export_path)
        try:
            for definition in definitions:
                versions = await self.repository.list_active_versions_by_sheet(
                    definition.sheet_key
                )
                rows = _build_export_rows(definition, versions)
                _fill_project_ledger_sheet(
                    workbook[definition.worksheet_title],
                    definition,
                    rows,
                )
            workbook.save(export_path)
        finally:
            workbook.close()

        return export_path, PROJECT_LEDGER_WORKBOOK_NAME

    async def create_entry(
        self,
        data: ProjectLedgerEntryInput,
    ) -> ProjectLedgerEntryResponse:
        await self.ensure_seeded()
        definition = _find_sheet_definition(data.sheet_key)
        values = _normalize_entry_values(definition, data.values)
        sequence_key = definition.columns[0].key
        source_sequence = _extract_sequence_number(values.get(sequence_key))
        if source_sequence is None:
            source_sequence = (
                await self.repository.get_max_source_sequence(data.sheet_key) + 1
            )
            values[sequence_key] = str(source_sequence)

        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)
        version = RegistrationProjectLedgerVersion(
            record_group_id=self.repository.generate_group_id(),
            sheet_key=definition.sheet_key,
            sheet_name=definition.sheet_name,
            sheet_title=definition.sheet_title,
            source_sequence=source_sequence,
            version_number=1,
            source_row_number=None,
            values_data=values,
            project_name=values.get(project_key) if project_key else None,
            product_name=values.get(product_key) if product_key else None,
        )
        created = await self.repository.create_version(version)
        await self.session.commit()
        return _build_entry_response(created)

    async def update_entry(
        self,
        record_id: Any,
        data: ProjectLedgerEntryInput,
    ) -> ProjectLedgerEntryResponse:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报台账记录", str(record_id))
        if latest.sheet_key != data.sheet_key:
            raise AppException(message="更新记录的子表与当前页面不一致")

        definition = _find_sheet_definition(latest.sheet_key)
        normalized_values = _normalize_entry_values(definition, data.values)
        sequence_key = definition.columns[0].key
        normalized_values[sequence_key] = str(latest.source_sequence)
        latest.values_data = normalized_values
        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)
        latest.project_name = (
            normalized_values.get(project_key) if project_key else None
        )
        latest.product_name = (
            normalized_values.get(product_key) if product_key else None
        )
        await self.session.flush()
        result = await self.session.execute(
            select(RegistrationProjectLedgerVersion).where(
                RegistrationProjectLedgerVersion.id == latest.id
            )
        )
        latest = result.scalar_one()
        await self.session.commit()
        return _build_entry_response(latest)

    async def create_sub_record(
        self,
        record_id: Any,
        data: ProjectLedgerEntryInput,
    ) -> ProjectLedgerEntryResponse:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报台账记录", str(record_id))
        if latest.sheet_key != data.sheet_key:
            raise AppException(message="新增子记录的子表与当前页面不一致")

        definition = _find_sheet_definition(latest.sheet_key)
        merged_values = latest.values_data.copy()
        merged_values.update(_normalize_entry_values(definition, data.values))

        sequence_key = definition.columns[0].key
        merged_values[sequence_key] = str(latest.source_sequence)
        version_number = await self.repository.get_next_version_number(record_id)
        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)

        created = await self.repository.create_version(
            RegistrationProjectLedgerVersion(
                record_group_id=latest.record_group_id,
                sheet_key=latest.sheet_key,
                sheet_name=latest.sheet_name,
                sheet_title=latest.sheet_title,
                source_sequence=latest.source_sequence,
                version_number=version_number,
                source_row_number=None,
                values_data=merged_values,
                project_name=merged_values.get(project_key) if project_key else None,
                product_name=merged_values.get(product_key) if product_key else None,
            )
        )
        await self.session.commit()
        return _build_entry_response(created)

    async def delete_entry(self, record_id: Any) -> None:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报台账记录", str(record_id))
        await self.repository.soft_delete_group(record_id)
        await self.session.commit()
