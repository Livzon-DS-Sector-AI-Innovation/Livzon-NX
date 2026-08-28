"""Declaration progress workbook service."""

from __future__ import annotations

import hashlib
import logging
import re
import shutil
import tempfile
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import MergedCell  # type: ignore[import-untyped]
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.exceptions import AppException, NotFoundException
from app.core.upload_security import read_upload_secure
from app.modules.registration.models import (
    RegistrationDeclarationProgressWorkbookVersion,
)
from app.modules.registration.repository import (
    RegistrationDeclarationProgressWorkbookRepository,
)
from app.modules.registration.schemas.declaration_progress import (
    DeclarationProgressColumn,
    DeclarationProgressEntryInput,
    DeclarationProgressEntryResponse,
    DeclarationProgressHistoryRecord,
    DeclarationProgressRecord,
    DeclarationProgressRecordHistory,
    DeclarationProgressSheetDetail,
    DeclarationProgressSheetSummary,
    DeclarationProgressWorkbookDetail,
    DeclarationProgressWorkbookImportResult,
    DeclarationProgressWorkbookOverview,
)

logger = logging.getLogger(__name__)

DECLARATION_PROGRESS_WORKBOOK_NAME = "宁夏-注册项目信息统计表-2026.06.25.xlsx"
DECLARATION_PROGRESS_EXPORT_NAME = "宁夏-注册项目信息统计表.xlsx"
DECLARATION_PROGRESS_HEADER_ROW = 4
DECLARATION_PROGRESS_DATA_START_ROW = 5
DECLARATION_PROGRESS_EXCLUDED_SHEETS = {"Sheet1"}
STYLE_MARK_NEW = "new"
STYLE_MARK_UPDATED = "updated"
STYLE_COLOR_MAP = {
    STYLE_MARK_NEW: "FF0000FE",
    STYLE_MARK_UPDATED: "FFFF0000",
}


@dataclass(slots=True)
class DeclarationProgressSheetTemplate:
    worksheet_title: str
    sheet_key: str
    sheet_name: str
    sheet_title: str
    column_labels: list[str]
    main_end_label: str | None
    supports_sub_records: bool


@dataclass(slots=True)
class DeclarationProgressSheetDefinition:
    worksheet_title: str
    sheet_key: str
    sheet_name: str
    sheet_title: str
    supports_sub_records: bool
    columns: list[DeclarationProgressColumn]


@dataclass(slots=True)
class DeclarationProgressExportRow:
    values: dict[str, str | None]
    style_marks: dict[str, str | None]


DECLARATION_PROGRESS_SHEET_TEMPLATES: tuple[DeclarationProgressSheetTemplate, ...] = (
    DeclarationProgressSheetTemplate(
        worksheet_title="1.2026国际注册（计划和进行中）项目",
        sheet_key="international-planned-in-progress",
        sheet_name="2026国际注册（计划和进行中）项目",
        sheet_title="2026国际注册（计划和进行中）项目",
        main_end_label="我司文件计划递交时间",
        supports_sub_records=True,
        column_labels=[
            "序号",
            "项目名称",
            "产品",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "质量标准",
            "批量/包装规格",
            "国家/受理机构",
            "MF内部编号",
            "MF官方登记号",
            "与制剂关联审评历史，被官方批准历史",
            "交费/时间",
            "药政活动类型（首次递交/缺陷信回复/变更/年度报告/再注册/委托生产/撤销）",
            "代理机构",
            "制剂公司",
            "制剂剂型/规格/官方登记号",
            "是否为第一供应商，是否涉及首仿",
            "制剂文件计划递交时间",
            "我司文件计划递交时间",
            "项目进展情况以及项目存在问题(预估该项目何时可以被批准)",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title="2.2026年国内注册（计划和进行中）项目",
        sheet_key="domestic-planned-in-progress",
        sheet_name="2026年国内注册（计划和进行中）项目",
        sheet_title="2026年国内注册（计划和进行中）项目",
        main_end_label="我司文件计划递交时间",
        supports_sub_records=True,
        column_labels=[
            "序号",
            "项目名称",
            "产品",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "质量标准",
            "批量/包装规格",
            "受理机构",
            "MF内部编号",
            "官方登记号/批准文号",
            "与制剂关联审评历史，被官方批准历史",
            "交费/时间",
            "药政活动类型（首次递交/缺陷信回复/变更/年度报告/再注册/委托生产/撤销）",
            "制剂公司",
            "制剂剂型/规格/官方登记号",
            "是否为第一供应商，是否涉及首仿",
            "制剂文件计划递交时间",
            "我司文件计划递交时间",
            "项目进展情况以及项目存在问题(预估该项目何时可以被批准)",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title=" 3.2026年新产品项目",
        sheet_key="new-product-projects",
        sheet_name="2026年新产品项目",
        sheet_title="2026年新产品项目",
        main_end_label=None,
        supports_sub_records=False,
        column_labels=[
            "序号",
            "产品名称",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "生产线建成时间",
            "质量标准建立时间",
            "分析方法验证时间",
            "菌种鉴定时间",
            "中试时间",
            "工艺验证时间",
            "稳定性实验时间",
            "结构确证时间（主成分和杂质）",
            "晶型研究时间",
            "参比制剂购买时间",
            "预申报时间",
            "申报机构",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title="4.2026年国际注册（已完成）",
        sheet_key="international-completed",
        sheet_name="2026年国际注册（已完成）",
        sheet_title="2026年国际注册（已完成）",
        main_end_label="制剂文件递交时间",
        supports_sub_records=True,
        column_labels=[
            "序号",
            "项目名称",
            "产品",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "质量标准",
            "批量/包装规格",
            "国家/受理机构",
            "MF内部编号",
            "MF官方登记号",
            "与制剂关联审评历史，被官方批准历史",
            "交费/时间",
            "药政活动类型（首次递交/缺陷信回复/变更/年度报告/再注册/委托生产/撤销）",
            "代理机构",
            "制剂公司",
            "制剂剂型/规格/官方登记号",
            "是否为第一供应商，是否涉及首仿",
            "制剂文件递交时间",
            "我司文件递交时间",
            "LOA签署日期和我司递交官方日期",
            "（该项目）审评结果/批准时间/正式批准信函或证书情况",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title="5.2026年国内注册（已完成）",
        sheet_key="domestic-completed",
        sheet_name="2026年国内注册（已完成）",
        sheet_title="2026年国内注册（已完成）",
        main_end_label="制剂文件递交时间",
        supports_sub_records=True,
        column_labels=[
            "序号",
            "项目名称",
            "产品",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "质量标准",
            "批量/包装规格",
            "受理机构",
            "MF内部编号",
            "官方登记号/批准文号",
            "与制剂关联审评历史，被官方批准历史",
            "交费/时间",
            "药政活动类型（首次递交/缺陷信回复/变更/年度报告/再注册/委托生产/撤销）",
            "制剂公司",
            "制剂剂型/规格/官方登记号",
            "是否为第一供应商，是否涉及首仿",
            "制剂文件递交时间",
            "我司文件递交时间",
            "（该项目）审评结果/批准时间",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title="6.2026年GMP项目",
        sheet_key="gmp-projects",
        sheet_name="2026年GMP项目",
        sheet_title="2026年GMP项目",
        main_end_label=None,
        supports_sub_records=False,
        column_labels=[
            "序号",
            "项目名称",
            "官方机构/国家",
            "涉及产品",
            "质量标准",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "需注册人员准备的资料",
            "审计官人数/名字",
            "审计日期",
            "审计类型（批准前检查、变更批准检查、日常监督检查、有因检查、换证前检查）",
            "是否有审计报告",
            "审计结果",
        ],
    ),
    DeclarationProgressSheetTemplate(
        worksheet_title="7.美国FDA注册进展",
        sheet_key="us-fda-progress",
        sheet_name="美国FDA注册进展",
        sheet_title="美国FDA注册进展",
        main_end_label="制剂文件递交时间",
        supports_sub_records=True,
        column_labels=[
            "序号",
            "产品名称",
            "药物类型（无菌/非无菌，原料药/中间体，人用药/兽用药）",
            "质量标准",
            "批量/包装规格",
            "DMF/VMF号",
            "制剂公司",
            "是否提供LOA",
            "LOA是否向FDA递交/日期",
            "是否提供注册资料/日期",
            "制剂文件递交时间",
            "我司文件递交时间",
            "是否现场审计",
            "是否发货/时间",
            "备注",
        ],
    ),
)

DECLARATION_PROGRESS_SHEET_CONFIG: dict[str, tuple[str, str]] = {
    template.worksheet_title: (template.sheet_key, template.sheet_name)
    for template in DECLARATION_PROGRESS_SHEET_TEMPLATES
}


def _get_workbook_path() -> Path:
    """获取申报进度工作簿路径，优先使用配置路径，回退到默认位置。"""
    settings = get_settings()
    workbook_dir = Path(settings.REGISTRATION_WORKBOOK_DIR)
    workbook_path = (
        workbook_dir / settings.REGISTRATION_DECLARATION_PROGRESS_WORKBOOK_NAME
    )

    if not workbook_path.exists():
        desktop = Path.home() / "Desktop"
        legacy_path = desktop / "注册相关文件" / DECLARATION_PROGRESS_WORKBOOK_NAME
        if legacy_path.exists():
            logger.warning(
                "使用桌面路径: %s，建议配置 REGISTRATION_WORKBOOK_DIR", legacy_path
            )
            return legacy_path

    return workbook_path


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def _normalize_optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    text = value.strip()
    return text or None


def _slugify(text: str) -> str:
    ascii_text = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    if ascii_text:
        return ascii_text
    return hashlib.md5(text.encode("utf-8")).hexdigest()[:12]


def _build_column_key(label: str, used_keys: set[str]) -> str:
    base_key = _slugify(label)
    candidate = base_key
    suffix = 2
    while candidate in used_keys:
        candidate = f"{base_key}_{suffix}"
        suffix += 1
    used_keys.add(candidate)
    return candidate


def _extract_sequence_number(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"\d+", value)
    return int(match.group()) if match else None


def _get_font_color_key(cell: Any) -> str | None:
    color = getattr(cell.font, "color", None)
    if color is None:
        return None
    try:
        if color.type == "rgb":
            return f"rgb:{color.rgb}"
        if color.type == "theme":
            return f"theme:{color.theme}:tint:{color.tint}"
        if color.type == "indexed":
            return f"indexed:{color.indexed}"
    except Exception:
        logger.exception("Failed to parse font color")
        return None
    return None


def _classify_style_mark(cell: Any) -> str | None:
    color_key = _get_font_color_key(cell)
    if not color_key:
        return None
    if color_key in {"rgb:FF000000", "theme:1:tint:0.0", "indexed:8"}:
        return None
    if color_key in {"rgb:FF0000FE", "rgb:FF0000FF"}:
        return STYLE_MARK_NEW
    if color_key in {"rgb:FFFF0000", "rgb:FFC0504D"}:
        return STYLE_MARK_UPDATED
    if color_key.startswith("theme:"):
        if color_key.startswith("theme:4:"):
            return STYLE_MARK_NEW
        if color_key.startswith("theme:5:"):
            return STYLE_MARK_UPDATED
    if color_key.startswith("rgb:") and len(color_key) == 12:
        rgb = color_key[4:]
        red = int(rgb[2:4], 16)
        green = int(rgb[4:6], 16)
        blue = int(rgb[6:8], 16)
        if blue >= red + 40 and blue >= green + 40:
            return STYLE_MARK_NEW
        if red >= blue + 40 and red >= green + 20:
            return STYLE_MARK_UPDATED
    return None


def _build_columns(
    template: DeclarationProgressSheetTemplate,
) -> list[DeclarationProgressColumn]:
    used_keys: set[str] = set()
    main_end_index = (
        template.column_labels.index(template.main_end_label)
        if template.main_end_label in template.column_labels
        else len(template.column_labels) - 1
    )
    columns: list[DeclarationProgressColumn] = []
    for index, label in enumerate(template.column_labels):
        columns.append(
            DeclarationProgressColumn(
                key=_build_column_key(label, used_keys),
                label=label,
                is_main=(not template.supports_sub_records) or index <= main_end_index,
            )
        )
    return columns


def _header_matches(
    actual_headers: list[str],
    template: DeclarationProgressSheetTemplate,
) -> bool:
    if len(actual_headers) != len(template.column_labels):
        return False
    for actual_label, expected_label in zip(actual_headers, template.column_labels):
        if actual_label == expected_label:
            continue
        if template.sheet_key == "domestic-completed" and actual_label.startswith(
            "='[1]5.2019"
        ):
            continue
        return False
    return True


def _parse_workbook_definitions() -> tuple[
    list[DeclarationProgressSheetDefinition], datetime
]:
    workbook_path = _get_workbook_path()
    if not workbook_path.exists():
        raise NotFoundException("申报进度统计表", str(workbook_path))

    definitions: list[DeclarationProgressSheetDefinition] = []
    for template in DECLARATION_PROGRESS_SHEET_TEMPLATES:
        definitions.append(
            DeclarationProgressSheetDefinition(
                worksheet_title=template.worksheet_title,
                sheet_key=template.sheet_key,
                sheet_name=template.sheet_name,
                sheet_title=template.sheet_title,
                supports_sub_records=template.supports_sub_records,
                columns=_build_columns(template),
            )
        )

    return definitions, datetime.fromtimestamp(workbook_path.stat().st_mtime)


def _find_sheet_definition(sheet_key: str) -> DeclarationProgressSheetDefinition:
    definitions, _ = _parse_workbook_definitions()
    for definition in definitions:
        if definition.sheet_key == sheet_key:
            return definition
    raise NotFoundException("申报进度子表", sheet_key)


def _get_project_name_key(definition: DeclarationProgressSheetDefinition) -> str | None:
    return next(
        (
            column.key
            for column in definition.columns
            if column.label in {"项目名称", "项目名称 "}
        ),
        None,
    )


def _get_product_name_key(definition: DeclarationProgressSheetDefinition) -> str | None:
    return next(
        (
            column.key
            for column in definition.columns
            if column.label in {"产品", "产品名称", "涉及产品"}
        ),
        None,
    )


def _load_seed_versions() -> list[RegistrationDeclarationProgressWorkbookVersion]:
    return _load_versions_from_workbook_path(_get_workbook_path())


def _load_versions_from_worksheet(
    definition: DeclarationProgressSheetDefinition,
    worksheet: Any,
) -> list[RegistrationDeclarationProgressWorkbookVersion]:
    grouped_rows: list[
        list[tuple[int, dict[str, str | None], dict[str, str | None]]]
    ] = []
    current_group: list[tuple[int, dict[str, str | None], dict[str, str | None]]] = []
    sequence_key = definition.columns[0].key
    project_key = _get_project_name_key(definition)
    product_key = _get_product_name_key(definition)

    for source_row_number in range(
        DECLARATION_PROGRESS_DATA_START_ROW, worksheet.max_row + 1
    ):
        values: dict[str, str | None] = {}
        style_marks: dict[str, str | None] = {}
        for index, column in enumerate(definition.columns, start=1):
            cell = worksheet.cell(row=source_row_number, column=index)
            cell_value = _normalize_text(cell.value)
            values[column.key] = _normalize_optional_text(cell_value)
            style_mark = _classify_style_mark(cell)
            if values[column.key] is not None and style_mark:
                style_marks[column.key] = style_mark

        if not any(value is not None for value in values.values()):
            continue

        non_sequence_values = [
            value
            for key, value in values.items()
            if key != sequence_key and value is not None
        ]
        if values.get(sequence_key) is not None and not non_sequence_values:
            continue

        if not definition.supports_sub_records:
            grouped_rows.append([(source_row_number, values, style_marks)])
            continue

        if values.get(sequence_key) is not None:
            if current_group:
                grouped_rows.append(current_group)
            current_group = [(source_row_number, values, style_marks)]
            continue

        if current_group:
            current_group.append((source_row_number, values, style_marks))

    if current_group:
        grouped_rows.append(current_group)

    versions: list[RegistrationDeclarationProgressWorkbookVersion] = []
    for fallback_sequence, group_rows in enumerate(grouped_rows, start=1):
        record_group_id = (
            RegistrationDeclarationProgressWorkbookRepository.generate_group_id()
        )
        snapshot_values: dict[str, str | None] = {}
        snapshot_style_marks: dict[str, str | None] = {}
        first_row_values = group_rows[0][1]
        source_sequence = (
            _extract_sequence_number(first_row_values.get(sequence_key))
            or fallback_sequence
        )

        for version_index, (
            source_row_number,
            row_values,
            row_style_marks,
        ) in enumerate(
            group_rows,
            start=1,
        ):
            current_snapshot = snapshot_values.copy()
            current_style_marks = snapshot_style_marks.copy()
            for column in definition.columns:
                next_value = row_values.get(column.key)
                if next_value is None:
                    continue
                current_snapshot[column.key] = next_value
                if row_style_marks.get(column.key):
                    current_style_marks[column.key] = row_style_marks[column.key]
                else:
                    current_style_marks.pop(column.key, None)

            current_snapshot[sequence_key] = str(source_sequence)
            versions.append(
                RegistrationDeclarationProgressWorkbookVersion(
                    record_group_id=record_group_id,
                    sheet_key=definition.sheet_key,
                    sheet_name=definition.sheet_name,
                    sheet_title=definition.sheet_title,
                    source_sequence=source_sequence,
                    version_number=version_index,
                    source_row_number=source_row_number,
                    values_data=current_snapshot.copy(),
                    style_marks=current_style_marks.copy(),
                    project_name=current_snapshot.get(project_key)
                    if project_key
                    else None,
                    product_name=current_snapshot.get(product_key)
                    if product_key
                    else None,
                )
            )
            snapshot_values = current_snapshot
            snapshot_style_marks = current_style_marks

    return versions


def _load_versions_from_workbook_path(
    workbook_path: Path,
) -> list[RegistrationDeclarationProgressWorkbookVersion]:
    definitions, _ = _parse_workbook_definitions()
    definition_map = {
        definition.worksheet_title: definition for definition in definitions
    }

    workbook = load_workbook(workbook_path, data_only=False)
    try:
        missing_sheet_titles = [
            definition.worksheet_title
            for definition in definitions
            if definition.worksheet_title not in workbook.sheetnames
        ]
        if missing_sheet_titles:
            raise AppException(
                message=f"导入失败：缺少子表《{missing_sheet_titles[0]}》"
            )

        unexpected_sheet_titles = [
            worksheet.title
            for worksheet in workbook.worksheets
            if worksheet.title not in DECLARATION_PROGRESS_EXCLUDED_SHEETS
            and worksheet.title not in definition_map
        ]
        if unexpected_sheet_titles:
            raise AppException(
                message=f"导入失败：存在未定义子表《{unexpected_sheet_titles[0]}》"
            )

        for template in DECLARATION_PROGRESS_SHEET_TEMPLATES:
            worksheet = workbook[template.worksheet_title]
            actual_headers = [
                _normalize_text(
                    worksheet.cell(DECLARATION_PROGRESS_HEADER_ROW, column_index).value
                )
                for column_index in range(1, len(template.column_labels) + 1)
            ]
            if not _header_matches(actual_headers, template):
                raise AppException(
                    message=f"导入失败：子表《{template.sheet_name}》表头不匹配"
                )

        versions: list[RegistrationDeclarationProgressWorkbookVersion] = []
        for definition in definitions:
            versions.extend(
                _load_versions_from_worksheet(
                    definition,
                    workbook[definition.worksheet_title],
                )
            )
        return versions
    finally:
        workbook.close()


def _build_history_record(
    version: RegistrationDeclarationProgressWorkbookVersion,
) -> DeclarationProgressHistoryRecord:
    return DeclarationProgressHistoryRecord(
        entry_id=version.id,
        version=version.version_number,
        source_row_number=version.source_row_number,
        values=version.values_data,
        style_marks=version.style_marks,
    )


def _build_record(
    versions: list[RegistrationDeclarationProgressWorkbookVersion],
    *,
    include_history_records: bool = True,
) -> DeclarationProgressRecord:
    latest = versions[-1]
    return DeclarationProgressRecord(
        record_id=latest.record_group_id,
        record_key=str(latest.record_group_id),
        sequence=latest.source_sequence,
        latest_values=latest.values_data,
        latest_style_marks=latest.style_marks,
        history_count=len(versions),
        history_records=(
            [_build_history_record(version) for version in versions]
            if include_history_records
            else []
        ),
    )


def _build_sheet_summary(
    definition: DeclarationProgressSheetDefinition,
    records: list[DeclarationProgressRecord],
) -> DeclarationProgressSheetSummary:
    return DeclarationProgressSheetSummary(
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        title=definition.sheet_title,
        total_records=len(records),
        records_with_history=sum(1 for item in records if item.history_count > 1),
        total_history_versions=sum(max(item.history_count - 1, 0) for item in records),
        main_column_count=sum(1 for column in definition.columns if column.is_main),
        child_column_count=sum(
            1 for column in definition.columns if not column.is_main
        ),
    )


def _build_sheet_detail(
    definition: DeclarationProgressSheetDefinition,
    versions: list[RegistrationDeclarationProgressWorkbookVersion],
    *,
    include_columns: bool = True,
    include_records: bool = True,
    include_history_records: bool = True,
) -> DeclarationProgressSheetDetail:
    grouped_versions: dict[
        object, list[RegistrationDeclarationProgressWorkbookVersion]
    ] = defaultdict(list)
    for version in versions:
        grouped_versions[version.record_group_id].append(version)

    records = (
        [
            _build_record(
                group,
                include_history_records=include_history_records,
            )
            for group in grouped_versions.values()
        ]
        if include_records
        else []
    )
    records.sort(key=lambda item: item.sequence)
    summary = _build_sheet_summary(
        definition,
        records
        if include_records
        else [
            DeclarationProgressRecord(
                record_id=group[-1].record_group_id,
                record_key=str(group[-1].record_group_id),
                sequence=group[-1].source_sequence,
                latest_values={},
                latest_style_marks={},
                history_count=len(group),
                history_records=[],
            )
            for group in grouped_versions.values()
        ],
    )
    return DeclarationProgressSheetDetail(
        sheet_key=definition.sheet_key,
        sheet_name=definition.sheet_name,
        title=definition.sheet_title,
        supports_sub_records=definition.supports_sub_records,
        columns=definition.columns if include_columns else [],
        records=records,
        summary=summary,
    )


def _normalize_entry_values(
    definition: DeclarationProgressSheetDefinition,
    values: dict[str, str | None],
) -> dict[str, str | None]:
    normalized: dict[str, str | None] = {}
    for column in definition.columns:
        normalized[column.key] = _normalize_optional_text(values.get(column.key))
    return normalized


def _build_entry_response(
    version: RegistrationDeclarationProgressWorkbookVersion,
) -> DeclarationProgressEntryResponse:
    return DeclarationProgressEntryResponse(
        record_id=version.record_group_id,
        sheet_key=version.sheet_key,
        version_number=version.version_number,
        sequence=version.source_sequence,
        values=version.values_data,
        style_marks=version.style_marks,
    )


def _copy_declaration_progress_row_style(
    worksheet: Any,
    source_row: int,
    target_row: int,
    column_count: int,
) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[
        source_row
    ].height
    for column_index in range(1, column_count + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)


def _reset_declaration_progress_merges(
    worksheet: Any,
    *,
    start_row: int,
    end_row: int,
    column_count: int,
) -> None:
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        if merged_range.max_col < 1 or merged_range.min_col > column_count:
            continue
        worksheet.unmerge_cells(str(merged_range))


def _apply_declaration_progress_vertical_merges(
    worksheet: Any,
    definition: DeclarationProgressSheetDefinition,
    rows: list[DeclarationProgressExportRow],
    *,
    start_row: int,
) -> None:
    if not rows or not definition.supports_sub_records:
        return

    sequence_key = definition.columns[0].key
    group_start = 0
    total_rows = len(rows)

    for row_index in range(1, total_rows + 1):
        next_is_group_start = (
            row_index == total_rows
            or rows[row_index].values.get(sequence_key) is not None
        )
        if not next_is_group_start:
            continue

        group_end = row_index - 1
        if group_end > group_start:
            sheet_start_row = start_row + group_start
            sheet_end_row = start_row + group_end
            for column_index, column in enumerate(definition.columns, start=1):
                if not column.is_main:
                    continue
                first_value = rows[group_start].values.get(column.key)
                if first_value in (None, ""):
                    continue
                if all(
                    rows[item_index].values.get(column.key) in (None, "")
                    for item_index in range(group_start + 1, group_end + 1)
                ):
                    worksheet.merge_cells(
                        start_row=sheet_start_row,
                        start_column=column_index,
                        end_row=sheet_end_row,
                        end_column=column_index,
                    )

        group_start = row_index


def _apply_style_mark(
    worksheet: Any,
    *,
    template_row: int,
    row_index: int,
    column_index: int,
    style_mark: str | None,
) -> None:
    template_font = copy(worksheet.cell(row=template_row, column=column_index).font)
    if style_mark:
        template_font.color = STYLE_COLOR_MAP[style_mark]
    worksheet.cell(row=row_index, column=column_index).font = template_font


def _build_export_rows(
    definition: DeclarationProgressSheetDefinition,
    versions: list[RegistrationDeclarationProgressWorkbookVersion],
) -> list[DeclarationProgressExportRow]:
    grouped_versions: dict[
        object, list[RegistrationDeclarationProgressWorkbookVersion]
    ] = defaultdict(list)
    for version in versions:
        grouped_versions[version.record_group_id].append(version)

    sequence_key = definition.columns[0].key
    rows: list[DeclarationProgressExportRow] = []
    for group_versions in grouped_versions.values():
        previous_values: dict[str, str | None] = {}
        for index, version in enumerate(group_versions):
            row_values: dict[str, str | None] = {}
            row_style_marks: dict[str, str | None] = {}
            for column in definition.columns:
                current_value = version.values_data.get(column.key)
                current_style_mark = version.style_marks.get(column.key)

                if index == 0 or not definition.supports_sub_records:
                    row_values[column.key] = current_value
                    if current_value is not None and current_style_mark:
                        row_style_marks[column.key] = current_style_mark
                    continue

                if column.key == sequence_key or column.is_main:
                    row_values[column.key] = None
                    continue

                previous_value = previous_values.get(column.key)
                if current_value != previous_value:
                    row_values[column.key] = current_value
                    if current_value is not None and current_style_mark:
                        row_style_marks[column.key] = current_style_mark
                else:
                    row_values[column.key] = None

            row_values[sequence_key] = (
                str(version.source_sequence) if index == 0 else None
            )
            rows.append(
                DeclarationProgressExportRow(
                    values=row_values,
                    style_marks=row_style_marks,
                )
            )
            previous_values = version.values_data

    return rows


def _fill_declaration_progress_sheet(
    worksheet: Any,
    definition: DeclarationProgressSheetDefinition,
    rows: list[DeclarationProgressExportRow],
    *,
    start_row: int = DECLARATION_PROGRESS_DATA_START_ROW,
    template_row: int = DECLARATION_PROGRESS_DATA_START_ROW,
) -> None:
    column_count = len(definition.columns)
    populated_rows = [
        row_index
        for row_index in range(start_row, worksheet.max_row + 1)
        if any(
            _normalize_text(worksheet.cell(row=row_index, column=column_index).value)
            for column_index in range(1, column_count + 1)
        )
    ]
    last_data_row = max(populated_rows, default=template_row)
    target_last_row = max(last_data_row, start_row + len(rows) - 1)

    if target_last_row >= start_row:
        _reset_declaration_progress_merges(
            worksheet,
            start_row=start_row,
            end_row=target_last_row,
            column_count=column_count,
        )

    for row_index in range(start_row, target_last_row + 1):
        _copy_declaration_progress_row_style(
            worksheet, template_row, row_index, column_count
        )
        for column_index in range(1, column_count + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    for row_offset, export_row in enumerate(rows):
        row_index = start_row + row_offset
        for column_index, column in enumerate(definition.columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            cell.value = export_row.values.get(column.key)
            _apply_style_mark(
                worksheet,
                template_row=template_row,
                row_index=row_index,
                column_index=column_index,
                style_mark=export_row.style_marks.get(column.key),
            )

    _apply_declaration_progress_vertical_merges(
        worksheet,
        definition,
        rows,
        start_row=start_row,
    )


class DeclarationProgressWorkbookService:
    """Declaration progress service with persisted versions."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.repository = RegistrationDeclarationProgressWorkbookRepository(session)

    async def ensure_seeded(self) -> None:
        if await self.repository.count_versions() == 0:
            versions = _load_seed_versions()
            if versions:
                await self.repository.create_versions(versions)
                await self.session.commit()

    async def get_workbook_detail(self) -> DeclarationProgressWorkbookDetail:
        overview = await self.get_overview()
        return DeclarationProgressWorkbookDetail(
            workbook_name=overview.workbook_name,
            updated_at=overview.updated_at,
            total_records=overview.total_records,
            sheets=overview.sheets,
        )

    async def get_sheet_detail(self, sheet_key: str) -> DeclarationProgressSheetDetail:
        await self.ensure_seeded()
        definition = _find_sheet_definition(sheet_key)
        versions = await self.repository.list_versions(sheet_key=sheet_key)
        return _build_sheet_detail(
            definition,
            versions,
            include_columns=True,
            include_records=True,
            include_history_records=False,
        )

    async def get_entry_history(
        self, record_id: Any
    ) -> DeclarationProgressRecordHistory:
        await self.ensure_seeded()
        versions = await self.repository.list_versions_by_group(record_id)
        if not versions:
            raise NotFoundException("申报进度记录", str(record_id))
        latest = versions[-1]
        return DeclarationProgressRecordHistory(
            record_id=latest.record_group_id,
            sheet_key=latest.sheet_key,
            sequence=latest.source_sequence,
            history_count=len(versions),
            history_records=[_build_history_record(version) for version in versions],
        )

    async def get_overview(self) -> DeclarationProgressWorkbookOverview:
        await self.ensure_seeded()
        definitions, updated_at = _parse_workbook_definitions()
        versions = await self.repository.list_versions()

        grouped_by_sheet: dict[
            str, dict[object, list[RegistrationDeclarationProgressWorkbookVersion]]
        ] = defaultdict(lambda: defaultdict(list))
        for version in versions:
            grouped_by_sheet[version.sheet_key][version.record_group_id].append(version)

        sheets: list[DeclarationProgressSheetDetail] = []
        for definition in definitions:
            sheet_versions = [
                version
                for group_versions in grouped_by_sheet.get(
                    definition.sheet_key, {}
                ).values()
                for version in group_versions
            ]
            sheets.append(
                _build_sheet_detail(
                    definition,
                    sheet_versions,
                    include_columns=False,
                    include_records=False,
                    include_history_records=False,
                )
            )

        return DeclarationProgressWorkbookOverview(
            workbook_name=DECLARATION_PROGRESS_WORKBOOK_NAME,
            updated_at=updated_at,
            total_records=sum(sheet.summary.total_records for sheet in sheets),
            records_with_history=sum(
                sheet.summary.records_with_history for sheet in sheets
            ),
            total_history_versions=sum(
                sheet.summary.total_history_versions for sheet in sheets
            ),
            sheets=sheets,
        )

    async def import_workbook(
        self,
        upload_file: UploadFile,
    ) -> DeclarationProgressWorkbookImportResult:
        filename, content = await read_upload_secure(
            upload_file,
            max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
            allowed_extensions={".xlsx"},
            what="申报进度工作簿",
        )

        temp_dir = Path(tempfile.mkdtemp(prefix="declaration-progress-import-"))
        import_path = temp_dir / (filename or DECLARATION_PROGRESS_WORKBOOK_NAME)
        import_path.write_bytes(content)

        await self.ensure_seeded()
        versions = _load_versions_from_workbook_path(import_path)

        try:
            await self.repository.replace_all_versions(versions)
            await self.session.commit()
        except Exception:
            await self.session.rollback()
            raise

        sheet_record_counts = Counter(
            version.sheet_key for version in versions if version.version_number == 1
        )
        return DeclarationProgressWorkbookImportResult(
            workbook_name=filename or DECLARATION_PROGRESS_WORKBOOK_NAME,
            imported_records=sum(sheet_record_counts.values()),
            sheet_record_counts=dict(sheet_record_counts),
        )

    async def export_workbook(self) -> tuple[Path, str]:
        await self.ensure_seeded()
        workbook_path = _get_workbook_path()
        if not workbook_path.exists():
            raise NotFoundException("申报进度统计表", str(workbook_path))

        definitions, _ = _parse_workbook_definitions()
        temp_dir = Path(tempfile.mkdtemp(prefix="declaration-progress-export-"))
        export_path = temp_dir / DECLARATION_PROGRESS_EXPORT_NAME
        shutil.copyfile(workbook_path, export_path)

        workbook = load_workbook(export_path)
        try:
            for excluded_sheet in DECLARATION_PROGRESS_EXCLUDED_SHEETS:
                if excluded_sheet in workbook.sheetnames:
                    workbook.remove(workbook[excluded_sheet])

            for definition in definitions:
                versions = await self.repository.list_active_versions_by_sheet(
                    definition.sheet_key
                )
                rows = _build_export_rows(definition, versions)
                _fill_declaration_progress_sheet(
                    workbook[definition.worksheet_title],
                    definition,
                    rows,
                )
            workbook.save(export_path)
        finally:
            workbook.close()

        return export_path, DECLARATION_PROGRESS_EXPORT_NAME

    async def create_entry(
        self,
        data: DeclarationProgressEntryInput,
    ) -> DeclarationProgressEntryResponse:
        await self.ensure_seeded()
        definition = _find_sheet_definition(data.sheet_key)
        values = _normalize_entry_values(definition, data.values)
        sequence_key = definition.columns[0].key
        source_sequence = _extract_sequence_number(values.get(sequence_key))
        if source_sequence is None:
            source_sequence = (
                await self.repository.get_max_source_sequence(data.sheet_key) + 1
            )
            values[sequence_key] = str(source_sequence)

        style_marks = {
            column.key: STYLE_MARK_NEW
            for column in definition.columns
            if values.get(column.key) is not None
        }
        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)
        version = RegistrationDeclarationProgressWorkbookVersion(
            record_group_id=self.repository.generate_group_id(),
            sheet_key=definition.sheet_key,
            sheet_name=definition.sheet_name,
            sheet_title=definition.sheet_title,
            source_sequence=source_sequence,
            version_number=1,
            source_row_number=None,
            values_data=values,
            style_marks=style_marks,
            project_name=values.get(project_key) if project_key else None,
            product_name=values.get(product_key) if product_key else None,
        )
        created = await self.repository.create_version(version)
        await self.session.commit()
        return _build_entry_response(created)

    async def update_entry(
        self,
        record_id: Any,
        data: DeclarationProgressEntryInput,
    ) -> DeclarationProgressEntryResponse:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报进度记录", str(record_id))
        if latest.sheet_key != data.sheet_key:
            raise AppException(message="更新记录的子表与当前页面不一致")

        definition = _find_sheet_definition(latest.sheet_key)
        normalized_values = _normalize_entry_values(definition, data.values)
        sequence_key = definition.columns[0].key
        normalized_values[sequence_key] = str(latest.source_sequence)

        updated_style_marks = latest.style_marks.copy()
        for column in definition.columns:
            current_value = latest.values_data.get(column.key)
            next_value = normalized_values.get(column.key)
            if current_value == next_value:
                continue
            if next_value is None:
                updated_style_marks.pop(column.key, None)
            else:
                updated_style_marks[column.key] = STYLE_MARK_UPDATED

        latest.values_data = normalized_values
        latest.style_marks = updated_style_marks
        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)
        latest.project_name = (
            normalized_values.get(project_key) if project_key else None
        )
        latest.product_name = (
            normalized_values.get(product_key) if product_key else None
        )
        await self.session.flush()
        result = await self.session.execute(
            select(RegistrationDeclarationProgressWorkbookVersion).where(
                RegistrationDeclarationProgressWorkbookVersion.id == latest.id
            )
        )
        latest = result.scalar_one()
        await self.session.commit()
        return _build_entry_response(latest)

    async def create_sub_record(
        self,
        record_id: Any,
        data: DeclarationProgressEntryInput,
    ) -> DeclarationProgressEntryResponse:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报进度记录", str(record_id))
        if latest.sheet_key != data.sheet_key:
            raise AppException(message="新增子记录的子表与当前页面不一致")

        definition = _find_sheet_definition(latest.sheet_key)
        if not definition.supports_sub_records:
            raise AppException(message="当前子表不支持新增子记录")

        merged_values = latest.values_data.copy()
        merged_style_marks = latest.style_marks.copy()
        normalized_values = _normalize_entry_values(definition, data.values)
        changed = False

        for column in definition.columns:
            if column.is_main:
                continue
            next_value = normalized_values.get(column.key)
            if next_value is None or next_value == latest.values_data.get(column.key):
                continue
            merged_values[column.key] = next_value
            merged_style_marks[column.key] = STYLE_MARK_UPDATED
            changed = True

        if not changed:
            raise AppException(message="请至少修改一个子记录字段")

        sequence_key = definition.columns[0].key
        merged_values[sequence_key] = str(latest.source_sequence)
        version_number = await self.repository.get_next_version_number(record_id)
        project_key = _get_project_name_key(definition)
        product_key = _get_product_name_key(definition)

        created = await self.repository.create_version(
            RegistrationDeclarationProgressWorkbookVersion(
                record_group_id=latest.record_group_id,
                sheet_key=latest.sheet_key,
                sheet_name=latest.sheet_name,
                sheet_title=latest.sheet_title,
                source_sequence=latest.source_sequence,
                version_number=version_number,
                source_row_number=None,
                values_data=merged_values,
                style_marks=merged_style_marks,
                project_name=merged_values.get(project_key) if project_key else None,
                product_name=merged_values.get(product_key) if product_key else None,
            )
        )
        await self.session.commit()
        return _build_entry_response(created)

    async def delete_entry(self, record_id: Any) -> None:
        await self.ensure_seeded()
        latest = await self.repository.get_latest_version_by_group(record_id)
        if latest is None:
            raise NotFoundException("申报进度记录", str(record_id))
        await self.repository.soft_delete_group(record_id)
        await self.session.commit()
