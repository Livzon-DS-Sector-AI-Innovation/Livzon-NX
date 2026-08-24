"""Generate pre-job training plan documents from templates."""

import logging
from io import BytesIO
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]

from app.modules.hr.date_format import fmt_date_obj
from app.modules.hr.models import Employee

logger = logging.getLogger(__name__)

OLD_TEMPLATE_NAME = "7.4岗前培训计划.xlsx"


def _find_old_template() -> Path:
    """Locate the old factory xlsx template."""
    candidates = [
        Path("员工培训教育管理规程") / OLD_TEMPLATE_NAME,
        Path("../员工培训教育管理规程") / OLD_TEMPLATE_NAME,
        Path(__file__).resolve().parent.parent.parent.parent
        / "员工培训教育管理规程"
        / OLD_TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {OLD_TEMPLATE_NAME}")


DEPT_CONTENT_MAP: dict[str, list[str]] = {
    "人事行政部": [
        "公司级公用文件(详见附件一)",
        "部门级公用文件(详见附件二)",
        "人事行政部人事行政专员岗位文件(详见附件三)",
        "人事行政专员岗位职责(QP.PM.053)",
        "生产安全知识",
        "岗前培训计划",
    ],
}


def _generate_old(employee: Employee) -> BytesIO:
    """Fill the old factory pre-job training plan xlsx template."""
    template_path = _find_old_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # Part 1: Employee overview
    ws["C5"] = employee.name or ""
    ws["I5"] = employee.department or ""
    ws["C6"] = employee.employee_number or ""
    ws["I6"] = fmt_date_obj(employee.hire_date)
    ws["C7"] = employee.position or ""

    # Part 2: Training content (auto-fill by department)
    content_list = DEPT_CONTENT_MAP.get(employee.department or "", [])
    for i, content in enumerate(content_list):
        row = 11 + i
        if row <= 20:
            ws[f"B{row}"] = content

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_prejob_training_plan(employee: Employee) -> BytesIO:
    """Fill the pre-job training plan template with employee data.

    Returns a BytesIO buffer containing the generated document.
    """
    return _generate_old(employee)
