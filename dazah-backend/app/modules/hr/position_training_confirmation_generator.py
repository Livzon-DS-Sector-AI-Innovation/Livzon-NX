"""岗位培训确认表 文档生成器（APP14模板）.

严格按照 APP14-SMP-HR-002-14岗位培训确认表.xlsx 模板格式导出：
- 加载模板文件，保留原始格式（合并单元格、字体、边框、行高列宽）
- 填充基本信息（R5-R6）：姓名、部门、岗位、入职时间、员工类别
- 填充培训记录（R9-R11+）：按序号罗列所有培训教材名称、培训日期、考核结果
- 保留任职资格确认区域（R13-R16）和备注行（R17）的模板格式
"""

import logging
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

from app.modules.hr.date_format import fmt_date_str

logger = logging.getLogger(__name__)

TEMPLATE_NAME = "APP14岗位培训确认表.xlsx"


def _find_template() -> Path:
    """定位 APP14 模板文件。"""
    candidates = [
        Path("员工培训教育管理规程") / TEMPLATE_NAME,
        Path("../员工培训教育管理规程") / TEMPLATE_NAME,
        Path(__file__).resolve().parent.parent.parent.parent
        / "员工培训教育管理规程"
        / TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {TEMPLATE_NAME}")


def generate_position_training_confirmation(
    employee_name: str,
    employee_number: str | None,
    department: str,
    position: str,
    hire_date: str,
    employee_category: str,
    training_items: list[dict[str, Any]],
) -> BytesIO:
    """根据员工信息和培训记录生成岗位培训确认表 Excel 文档。

    Args:
        employee_name: 员工姓名
        employee_number: 工号
        department: 部门
        position: 岗位
        hire_date: 入职/转岗/复岗时间（字符串格式）
        employee_category: 员工类别（入职/转岗/复岗）
        training_items: 培训记录列表，每项包含:
            - textbook_name: 培训教材名称
            - textbook_code: 文件编码
            - training_date: 培训日期
            - assessment_result: 考核结果
    """
    template_path = _find_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # ─ 基本信息（R5-R6）──
    # R5: 基本信息 | 姓名 | [值] | 部门 | [值] | 岗位 |
    # [值] | 入职/转岗/复岗时间 | [值] | 员工类别 | [勾选]
    # 模板结构：A5:B6=基本信息, C5=姓名, E5=部门, G5=岗位,
    # I5=入职时间, L5=员工类别, M5=勾选
    ws["D5"] = employee_name or ""
    ws["F5"] = department or ""
    ws["H5"] = position or ""
    ws["K5"] = fmt_date_str(hire_date)
    # 员工类别勾选（M5 已包含 £入职£转岗£复岗）
    # 根据 employee_category 标记对应选项
    category_map = {
        "入职": "☑入职£转岗£复岗",
        "转岗": "£入职转岗£复岗",
        "复岗": "£入职£转岗☑复岗",
    }
    ws["M5"] = category_map.get(employee_category, "£入职£转岗£复岗")

    # ── 培训记录（R9-R11+）──
    # R7: 表头（培训课程或项目内容 | 培训日期 | 考核结果 | 员工签名 | 备注）
    # R8: 子表头（序号 | 培训教材名称及文件编码 | ...）
    # R9-R11: 数据行（模板提供3行示例，需动态扩展）

    # 清除模板中的示例数据（R9-R11 和 R12 的省略号）
    for row in range(9, 13):
        for col in range(1, 15):  # A-N
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).strip() in ("1", "2", "3", "……"):
                cell.value = None

    # 填充培训记录
    start_row = 9
    for idx, item in enumerate(training_items):
        row = start_row + idx
        # 如果超出模板行数，需要插入新行
        if row > 12:
            # 在第12行之前插入新行（保持 R13 任职资格确认区域位置）
            ws.insert_rows(12)
            # 复制 R11 的格式到新行
            for col in range(1, 15):
                src_cell = ws.cell(row=11, column=col)
                dst_cell = ws.cell(row=row, column=col)
                if src_cell.has_style:
                    dst_cell.font = (
                        openpyxl.styles.Font(
                            **{
                                "name": src_cell.font.name,
                                "size": src_cell.font.size,
                                "bold": src_cell.font.bold,
                            }
                        )
                        if src_cell.font.name
                        else src_cell.font
                    )
                    dst_cell.alignment = src_cell.alignment
                    dst_cell.border = src_cell.border

        # 序号（A列）
        ws.cell(row=row, column=1, value=idx + 1)
        # 培训教材名称及文件编码（B列，合并 B-C）
        textbook = item.get("textbook_name", "")
        code = item.get("textbook_code", "")
        ws.cell(row=row, column=2, value=f"{textbook}（{code}）" if code else textbook)
        # 培训日期（H-I列）
        ws.cell(row=row, column=8, value=fmt_date_str(item.get("training_date", "")))
        # 考核结果（J-K列）
        ws.cell(row=row, column=10, value=item.get("assessment_result", ""))
        # 员工签名（L-M列）- 留空供手写
        # 备注（N列）- 留空

    # ── 保存 ──
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
