import logging
import re
import zipfile
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, cast
from urllib.parse import quote
from uuid import UUID, uuid4

from fastapi import Depends, File, Form, Query, Request, UploadFile
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from pydantic import BaseModel as _PydanticBaseModel
from pydantic import Field
from sqlalchemy import inspect as sa_inspect
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import get_db
from app.core.deps import CurrentUser
from app.core.exceptions import (
    AppException,
    NotFoundException,
    RecruitmentNotConfigured,
)
from app.core.jobs import get_job_status, submit_job
from app.core.response import paginated_response, success_response
from app.core.upload_security import MAX_UPLOAD_FILES, read_upload_secure
from app.modules.hr.ai_chat_api import router as ai_chat_router
from app.modules.hr.ai_exam_api import router as ai_exam_router
from app.modules.hr.analysis_api import router as analysis_router
from app.modules.hr.constants import DEPT_APPROVAL_EXCLUDE, DEPT_APPROVAL_NO_EXPAND
from app.modules.hr.contract_api import router as contract_router
from app.modules.hr.contract_settings_api import router as contract_settings_router
from app.modules.hr.date_format import fmt_date_str
from app.modules.hr.document_generator import generate_onboarding_training_record
from app.modules.hr.esg_api import router as esg_router
from app.modules.hr.feishu_settings_api import router as feishu_settings_router
from app.modules.hr.models import (
    HrDepartment,
    HrDeptApprovalConfig,
    TrainingContentUsed,
    TrainingDocument,
    TrainingSession,
)
from app.modules.hr.new_employee_training_api import (
    router as new_employee_training_router,
)
from app.modules.hr.notification_document_generator import (
    generate_training_notification,
)
from app.modules.hr.onboarding_evaluation_document_generator import (
    generate_onboarding_evaluation,
)
from app.modules.hr.oral_exam_document_generator import generate_oral_exam_result
from app.modules.hr.plan_tracking_api import router as plan_tracking_router
from app.modules.hr.position_training_api import router as position_training_router
from app.modules.hr.position_training_mapping_api import (
    router as position_training_mapping_router,
)
from app.modules.hr.practical_exam_document_generator import (
    generate_practical_exam_zip,
    parse_practical_exam_questions,
)
from app.modules.hr.prejob_document_generator import generate_prejob_training_plan
from app.modules.hr.push_settings_schemas import SendNoticeRequest, SendNoticeResult
from app.modules.hr.push_settings_service import PushSettingsService
from app.modules.hr.recruitment_service import (
    OnboardingService as RecruitmentOnboardingService,
)
from app.modules.hr.recruitment_service import RecruitmentService
from app.modules.hr.schemas import (
    AnnualTrainingPlanCreate,
    AnnualTrainingPlanItemBatchUpdate,
    AnnualTrainingPlanItemResponse,
    AnnualTrainingPlanResponse,
    AnnualTrainingPlanUpdate,
    AttachmentPreview,
    AttachmentPreviewEnvelope,
    CandidateResponse,
    CandidateUpdate,
    CustomTrainingDepartmentCreate,
    DepartmentCreate,
    DepartmentResponse,
    DepartmentUpdate,
    DepartureRecordCreate,
    DepartureRecordUpdate,
    DeptApprovalConfigCreate,
    DeptApprovalConfigUpdate,
    EmailConfigUpdate,
    EmployeeCreate,
    EmployeePublicCreate,
    EmployeeResponse,
    EmployeeTrainingListMemberCreate,
    EmployeeTrainingListMemberOut,
    EmployeeTrainingListMemberUpdate,
    EmployeeTrainingRecordOut,
    EmployeeTrainingSummaryOut,
    EmployeeUpdate,
    ExamScoreConfirmRequest,
    ExamScoreConfirmResponse,
    ExamScoreImportResponse,
    ExamScoreItem,
    FeishuApprovalWebhookPayload,
    ImportConfirmRequest,
    ImportConfirmResponse,
    ImportFeishuMembersRequest,
    ImportFeishuMembersResult,
    ImportPreviewData,
    ImportPreviewResponse,
    ImportSheetPreview,
    JobPostingCreate,
    JobPostingResponse,
    JobPostingUpdate,
    MarkLedgerImportedRequest,
    MarkTrainingContentUsedRequest,
    OffboardingRecordCreate,
    OffboardingRecordResponse,
    OffboardingRecordUpdate,
    OnboardingCreate,
    OnboardingEvaluationInput,
    OnboardingResponse,
    OralExamExportRequest,
    PlanAttachmentListEnvelope,
    PlanAttachmentResponse,
    PlanAttachmentSectionListEnvelope,
    PlanAttachmentSectionResponse,
    PositionTransferApproveNodeRequest,
    PositionTransferRecordCreate,
    PositionTransferRecordResponse,
    PositionTransferRecordUpdate,
    PositionTransferRejectNodeRequest,
    PositionTransferSubmitRequest,
    PracticalExamExportRequest,
    SendOfferEmailPayload,
    TeamCreate,
    TeamResponse,
    TeamUpdate,
    TrainingAttachmentExportRequest,
    TrainingConflictCheckRequest,
    TrainingConflictCheckResponse,
    TrainingContentUsedOut,
    TrainingDeptMappingCreate,
    TrainingDeptMappingUpdate,
    TrainingDocumentOut,
    TrainingDocumentUpsert,
    TrainingEvaluationInput,
    TrainingLedgerCreate,
    TrainingLedgerPageCreate,
    TrainingLedgerPageResponse,
    TrainingLedgerResponse,
    TrainingLedgerUpdate,
    TrainingNotificationInput,
    TrainingNotifyInput,
    TrainingPersonnelConfigCreate,
    TrainingPersonnelConfigOut,
    TrainingSessionOut,
    TrainingSessionUpsert,
    TrainingSignInSheetInput,
)
from app.modules.hr.service import (
    AnnualTrainingPlanItemService,
    AnnualTrainingPlanService,
    DepartmentService,
    DepartureRecordService,
    EmployeeService,
    EmployeeTrainingListService,
    OffboardingRecordService,
    OnboardingRecordService,
    PlanAttachmentService,
    PositionTransferRecordService,
    TeamService,
    TrainingLedgerPageService,
    TrainingLedgerService,
    TrainingPersonnelConfigService,
)
from app.modules.hr.signin_document_generator import generate_training_sign_in_sheet
from app.modules.hr.trainer_api import router as trainer_router
from app.modules.hr.training_attachment_document_generator import (
    generate_training_attachment,
)
from app.modules.hr.training_evaluation_api import router as training_evaluation_router
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE
from app.shared.schemas import ApiResponseEnvelope, PageParams

logger = logging.getLogger(__name__)

router = create_module_router(MODULES_BY_CODE["hr"])

router.include_router(feishu_settings_router)

router.include_router(contract_settings_router)

router.include_router(contract_router)

router.include_router(trainer_router)

router.include_router(training_evaluation_router)

router.include_router(position_training_router)

router.include_router(position_training_mapping_router)

router.include_router(plan_tracking_router)

router.include_router(new_employee_training_router)

router.include_router(esg_router)

router.include_router(ai_chat_router)

router.include_router(ai_exam_router)

router.include_router(analysis_router)


def _require_user(current_user: CurrentUser) -> None:
    """验证用户已登录，未登录则抛出 401 异常"""
    if current_user is None:
        raise AppException(status_code=401, message="请先登录")


async def _assert_employee_page_access(
    db: AsyncSession,
    current_user: CurrentUser,
) -> None:
    """员工档案页面（员工管理/员工档案）访问校验。

    通过条件：通配（super_admin）/ hr:write（人力资源部）/ hr:employee:read。
    其他部门人员即使有 hr:read（人事查看员）也不可访问员工档案接口。
    """
    _require_user(current_user)
    assert current_user is not None
    from app.platform.identity.rbac import resolve_user_permissions

    perms = await resolve_user_permissions(db, current_user.id)
    if "*" in perms or "hr:write" in perms or "hr:employee:read" in perms:
        return
    raise AppException(
        status_code=403,
        message="无权限访问员工档案，仅人力资源部可用",
    )


async def _assert_hr_write(db: AsyncSession, current_user: CurrentUser) -> None:
    """Guard legacy HR write endpoints when called outside the middleware."""

    _require_user(current_user)
    assert current_user is not None
    from app.platform.identity.rbac import resolve_user_permissions

    permissions = await resolve_user_permissions(db, current_user.id)
    if "*" in permissions or "hr:write" in permissions:
        return
    raise AppException(status_code=403, message="无权修改人事模块数据")


def _legacy_record_payload(record: Any) -> dict[str, Any]:
    """Serialize a legacy ORM row without dropping columns unknown to new code."""

    mapper = sa_inspect(record).mapper
    return cast(
        dict[str, Any],
        jsonable_encoder(
            {column.key: getattr(record, column.key) for column in mapper.column_attrs}
        ),
    )


# ─── 部门级数据隔离工具（人事/培训模块可见范围）────────────────────────


async def _resolve_visible_scope(
    db: AsyncSession, current_user: CurrentUser
) -> set[str] | None:
    """解析当前用户可见部门的档案别名集合；None = 全部可见（管理员 hr:write）。"""
    from app.modules.hr.training_dept_resolver import resolve_visible_dept_alias_set

    _require_user(current_user)
    return await resolve_visible_dept_alias_set(db, current_user)


async def _assert_dept_in_scope(
    db: AsyncSession, current_user: CurrentUser, department: str | None
) -> set[str] | None:
    """校验 department 参数在用户可见范围内；返回可见别名集合（None=全部）。

    - department 为空：直接返回可见范围（调用方据此做默认过滤）
    - department 非空且不在可见范围：抛 403
    """
    alias_set = await _resolve_visible_scope(db, current_user)
    if alias_set is None:
        return None
    if department and department not in alias_set:
        raise AppException(
            status_code=403,
            message=f"无权访问部门「{department}」的数据",
            detail="如需访问其他部门数据，请联系管理员在 HR 设置中配置可见部门",
        )
    return alias_set


async def _resolve_visible_norms(
    db: AsyncSession, current_user: CurrentUser
) -> set[str] | None:
    """解析当前用户可见的培训规范部门名集合；None = 全部可见（管理员）。"""
    from app.modules.hr.training_dept_resolver import visible_training_dept_names

    _require_user(current_user)
    return await visible_training_dept_names(db, current_user)


def get_employee_service(session: AsyncSession = Depends(get_db)) -> EmployeeService:
    return EmployeeService(session)


def get_department_service(
    session: AsyncSession = Depends(get_db),
) -> DepartmentService:
    return DepartmentService(session)


def get_offboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OffboardingRecordService:
    return OffboardingRecordService(session)


def get_onboarding_service(
    session: AsyncSession = Depends(get_db),
) -> OnboardingRecordService:
    return OnboardingRecordService(session)


def get_departure_service(
    session: AsyncSession = Depends(get_db),
) -> DepartureRecordService:
    return DepartureRecordService(session)


def get_team_service(
    session: AsyncSession = Depends(get_db),
) -> TeamService:
    return TeamService(session)


def get_training_ledger_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerService:
    return TrainingLedgerService(session)


def get_training_ledger_page_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingLedgerPageService:
    return TrainingLedgerPageService(session)


def get_employee_training_list_service(
    session: AsyncSession = Depends(get_db),
) -> EmployeeTrainingListService:
    return EmployeeTrainingListService(session)


def get_annual_training_plan_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanService:
    return AnnualTrainingPlanService(session)


def get_annual_training_plan_item_service(
    session: AsyncSession = Depends(get_db),
) -> AnnualTrainingPlanItemService:
    return AnnualTrainingPlanItemService(session)


def get_plan_attachment_service(
    session: AsyncSession = Depends(get_db),
) -> PlanAttachmentService:
    return PlanAttachmentService(session)


def get_position_transfer_service(
    session: AsyncSession = Depends(get_db),
) -> PositionTransferRecordService:
    return PositionTransferRecordService(session)


def get_training_personnel_config_service(
    session: AsyncSession = Depends(get_db),
) -> TrainingPersonnelConfigService:
    return TrainingPersonnelConfigService(session)


# ─── Employee Routes ───


@router.get("/employees", summary="员工列表")
async def list_employees(
    department: str | None = Query(None, description="部门筛选"),
    status: str | None = Query(None, description="状态筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_employee_page_access(db, current_user)
    alias_set = await _assert_dept_in_scope(db, current_user, department)
    employees, total = await service.list_employees(
        department=department,
        status=status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=alias_set,
    )
    data = [
        EmployeeResponse.model_validate(e).model_dump(mode="json") for e in employees
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/employees/stats", summary="员工统计")
async def get_employee_stats(
    db: AsyncSession = Depends(get_db),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_employee_page_access(db, current_user)
    alias_set = await _resolve_visible_scope(db, current_user)
    stats = await service.get_employee_stats(dept_alias_set=alias_set)
    return success_response(data=stats)


@router.get("/employees/max-seq", summary="获取最大序号")
async def get_max_seq_number(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(db, current_user)
    from app.modules.hr.repository import EmployeeRepository

    repo = EmployeeRepository(db)
    max_seq = await repo.get_max_seq_number()
    return success_response(data={"max_seq": max_seq, "next_seq": max_seq + 1})


@router.get("/employees/new-hires", summary="新员工列表（最近首次签合同）")
async def list_new_hires(
    days: int = Query(7, ge=0, description="最近天数（默认7天）"),
    db: AsyncSession = Depends(get_db),
    service: TrainingPersonnelConfigService = Depends(
        get_training_personnel_config_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _resolve_visible_scope(db, current_user)
    new_hires = await service.list_new_hires(days=days, dept_alias_set=alias_set)
    data = [item.model_dump(mode="json") for item in new_hires]
    return success_response(data=data)


@router.get("/training-personnel-configs", summary="培训人员配置列表")
async def list_training_personnel_configs(
    level: str | None = Query(None, description="培训级别: 公司级/部门级"),
    department: str | None = Query(None, description="部门(公司级为空)"),
    db: AsyncSession = Depends(get_db),
    service: TrainingPersonnelConfigService = Depends(
        get_training_personnel_config_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_dept_in_scope(db, current_user, department)
    configs = await service.list_configs(level=level, department=department)
    data = [
        TrainingPersonnelConfigOut.model_validate(c).model_dump(mode="json")
        for c in configs
    ]
    return success_response(data=data)


@router.post(
    "/training-personnel-configs",
    summary="保存培训人员配置（按级别+部门+配置名 upsert）",
)
async def upsert_training_personnel_config(
    payload: TrainingPersonnelConfigCreate,
    service: TrainingPersonnelConfigService = Depends(
        get_training_personnel_config_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    config = await service.upsert_config(payload)
    return success_response(
        data=TrainingPersonnelConfigOut.model_validate(config).model_dump(mode="json"),
        message="培训人员配置保存成功",
    )


@router.delete(
    "/training-personnel-configs/{config_id}",
    summary="删除一条培训人员配置（软删除）",
)
async def delete_training_personnel_config(
    config_id: UUID,
    service: TrainingPersonnelConfigService = Depends(
        get_training_personnel_config_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_config(config_id)
    return success_response(message="培训人员配置已删除")


@router.get("/employees/contract-expiring", summary="合同到期人员列表")
async def list_contract_expiring(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    department: str | None = Query(None, description="部门筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _assert_dept_in_scope(db, current_user, department)
    employees, total = await service.list_contract_expiring(
        start_date=start_date,
        end_date=end_date,
        department=department,
        page=page,
        page_size=page_size,
        dept_alias_set=alias_set,
    )
    # list_contract_expiring 返回的是 dict 列表，不是 Employee 对象
    return paginated_response(
        data=employees, total=total, page=page, page_size=page_size
    )


async def _resolve_contract_approvers(
    session: Any, dept_name: str
) -> tuple[str | None, str | None, str | None, str | None]:
    """解析部门合同审批人: (部门负责人名, 部门负责人open_id,
    分管领导名, 分管领导open_id)。

    合同到期审批：部门负责人（第一级）= 部门经理
    manager（无经理回退直属领导，再回退部门表 leader_name）；
    分管领导（第二级）= 部门总监 director（无总监自动跳过第二级）；不再使用主管领导 vp。
    """
    from app.modules.hr.models import HrDepartment, HrDeptApprovalConfig, HrFeishuMember

    dept_leader_name: str | None = None
    dept_leader_open_id: str | None = None
    supervisor_name: str | None = None
    supervisor_open_id: str | None = None

    # 1. 部门级审批人配置（岗位调动同款表）：经理优先，回退直属领导；分管领导=总监
    cfg_result = await session.execute(
        select(HrDeptApprovalConfig)
        .where(
            HrDeptApprovalConfig.department_name == dept_name,
            HrDeptApprovalConfig.is_deleted.is_(False),
        )
        .limit(1)
    )
    cfg = cfg_result.scalars().first()
    if cfg:
        dept_leader_name = cfg.manager_name or cfg.direct_leader_name
        dept_leader_open_id = cfg.manager_open_id or cfg.direct_leader_open_id
        supervisor_name = cfg.director_name
        supervisor_open_id = cfg.director_open_id

    # 2. 回退：部门表 leader_name -> 飞书成员缓存
    if not dept_leader_name or not dept_leader_open_id:
        dept_result = await session.execute(
            select(HrDepartment)
            .where(
                HrDepartment.name == dept_name,
                HrDepartment.is_deleted.is_(False),
            )
            .limit(1)
        )
        dept = dept_result.scalars().first()
        if dept and dept.leader_name:
            dept_leader_name = dept.leader_name
            if not dept_leader_open_id:
                member_result = await session.execute(
                    select(HrFeishuMember.open_id)
                    .where(
                        HrFeishuMember.name == dept.leader_name,
                        HrFeishuMember.is_deleted.is_(False),
                        HrFeishuMember.status == "1",
                    )
                    .limit(1)
                )
                dept_leader_open_id = member_result.scalar_one_or_none()

    # 3. 分管领导 open_id 缺失时按姓名补查
    if supervisor_name and not supervisor_open_id:
        member_result = await session.execute(
            select(HrFeishuMember.open_id)
            .where(
                HrFeishuMember.name == supervisor_name,
                HrFeishuMember.is_deleted.is_(False),
                HrFeishuMember.status == "1",
            )
            .limit(1)
        )
        supervisor_open_id = member_result.scalar_one_or_none()

    return dept_leader_name, dept_leader_open_id, supervisor_name, supervisor_open_id


@router.post(
    "/employees/contract-expiring/push-notify", summary="触发合同到期提醒（后台任务）"
)
async def push_contract_expiring_notify(
    payload: dict[str, Any] | None = None,
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    from app.core.database import async_session_factory
    from app.core.jobs import is_job_running, submit_job
    from app.core.redis import cache_get, cache_set

    task_id = "hr:push:contract-expiring-notify"

    # 检查是否已有任务在运行（心跳过期的孤儿状态允许重新提交）
    if await is_job_running(task_id):
        return success_response(
            data={"task_id": task_id, "state": "running"},
            message="推送任务正在执行中，请稍后查看状态",
        )

    body = payload or {}
    start_str = body.get("start_date")
    end_str = body.get("end_date")

    async def _do_push() -> Any:
        from datetime import date as date_type

        if start_str and end_str:
            start_date = date_type.fromisoformat(start_str)
            end_date = date_type.fromisoformat(end_str)
        else:
            from datetime import datetime, timedelta

            now = datetime.now()
            q = (now.month - 1) // 3
            start_date = date_type(now.year, q * 3 + 1, 1)
            end_date = date_type(now.year, q * 3 + 4, 1) - timedelta(days=1)

        async with async_session_factory() as session:
            from app.modules.hr.service import EmployeeService

            service = EmployeeService(session)
            all_employees, total = await service.list_contract_expiring(
                start_date=start_date,
                end_date=end_date,
                page=1,
                page_size=200,
            )
            if not all_employees:
                return {
                    "pushed": 0,
                    "failed": 0,
                    "total_expiring": 0,
                    "skipped_pushed": 0,
                    "skipped_approved": 0,
                }

            # 增量去重：跳过已推送过或已审批过的员工
            employees = []
            skipped_pushed = 0
            skipped_approved = 0
            for emp in all_employees:
                emp_no = emp.get("employee_number", "")
                # 已审批过的跳过
                approved = await cache_get(f"hr:contract_approved:{emp_no}")
                if approved:
                    skipped_approved += 1
                    continue
                # 已推送过的跳过（7天TTL，避免短期重复推送同一人）
                already_pushed = await cache_get(f"hr:contract:pushed:{emp_no}")
                if already_pushed:
                    skipped_pushed += 1
                    continue
                employees.append(emp)

            if not employees:
                return {
                    "pushed": 0,
                    "failed": 0,
                    "total_expiring": total,
                    "skipped_pushed": skipped_pushed,
                    "skipped_approved": skipped_approved,
                }

            # 手动触发：审批卡片自动按「审批流程设置」的部门经理/总监发送
            # （不再受提醒开关控制
            # ）
            pushed, failed = 0, 0

            # 按部门分组到期人员，解析部门负责人（部门经理，回退直属领导/leader_nam
            # e）与分管领导（部门总监）
            dept_employees: dict[str, list[dict[str, Any]]] = {}
            dept_leader_openids: dict[str, str] = {}  # dept_name -> leader_open_id
            dept_leader_names: dict[str, str] = {}  # dept_name -> leader_name
            dept_supervisors: dict[
                str, tuple[str, str]
            ] = {}  # dept_name -> (supervisor_name, supervisor_open_id)
            for emp in employees:
                dept_name = emp.get("department", "")
                if not dept_name:
                    continue
                if dept_name not in dept_employees:
                    dept_employees[dept_name] = []
                    (
                        leader_name,
                        leader_open_id,
                        sup_name,
                        sup_open_id,
                    ) = await _resolve_contract_approvers(session, dept_name)
                    if leader_open_id:
                        dept_leader_openids[dept_name] = leader_open_id
                    if leader_name:
                        dept_leader_names[dept_name] = leader_name
                    if sup_name:
                        dept_supervisors[dept_name] = (sup_name, sup_open_id or "")
                dept_employees[dept_name].append(emp)

            # 先归档到合同管理模块（写入审批阶段与分管领导），再发卡片：
            # 保证用户点击卡片按钮时记录已存在且含分管领导，避免竞态导致单级通过
            synced = 0
            try:
                from app.modules.hr.contract_service import ContractService

                contract_service = ContractService(session)
                for emp in employees:
                    # 检查是否已审批过（按员工维度防重）
                    emp_no = emp.get("employee_number", "")
                    approved = await cache_get(f"hr:contract_approved:{emp_no}")
                    if approved:
                        continue  # 已审批，跳过归档
                    record = await contract_service.sync_from_contract_expiry(emp)
                    if record:
                        # 已最终审批（通过/拒绝）的记录不重置，防止覆盖用户审批结果
                        if getattr(record, "approval_status", "dept_pending") in (
                            "approved",
                            "rejected",
                        ):
                            continue
                        dept_name = emp.get("department", "")
                        sup_name, sup_open_id = dept_supervisors.get(
                            dept_name, (None, "")
                        )
                        record.approval_status = "dept_pending"
                        if sup_name:
                            record.supervisor_name = sup_name
                            record.supervisor_open_id = sup_open_id or None
                        synced += 1
                # 必须 flush+commit：否则任务事务回滚，归档记录（含分管领导字段）丢失
                await session.flush()
                await session.commit()
            except Exception:
                logger.exception("归档合同到期数据到合同管理模块失败")

            # 给部门经理发本部门到期人员的审批卡片（按部门汇总，一张卡片包
            # 含所有人员+审批按钮
            # ）
            for dept_name, dept_emps in dept_employees.items():
                leader_open_id = dept_leader_openids.get(dept_name)
                leader_name = dept_leader_names.get(dept_name, "")
                if not leader_open_id:
                    continue

                # 构建汇总内容：列出该部门所有到期人员
                lines = []
                for emp in dept_emps:
                    emp_name = emp.get("name", "")
                    emp_no = emp.get("employee_number", "")
                    emp_seq = emp.get("contract_sequence", "")
                    emp_end = emp.get("contract_end_date", "")
                    lines.append(
                        f"- **{emp_name}**（工号：{emp_no}），"
                        f"第{emp_seq}次合同于 **{emp_end}** 到期"
                    )

                dept_title = f"【{dept_name}】合同到期提醒与审批"
                dept_content = (
                    "合同到期提醒：贵部门以下人员合同即将到期，请审批是否续签：\n\n"
                    + "\n".join(lines)
                    + "\n\n请逐一审批（同意后将提交分管领导复核）："
                )

                # 构建审批按钮：value 交互回调（卡片内审批，不跳转浏览器；stage
                # 参与后续防重）
                elements = []
                from app.modules.hr.contract_api import build_contract_approval_actions

                for emp in dept_emps:
                    emp_name = emp.get("name", "")
                    emp_no = emp.get("employee_number", "")
                    elements.append(
                        build_contract_approval_actions(
                            emp_no,
                            emp_name,
                            leader_name,
                            "dept",
                            dept_name=dept_name,
                        )
                    )

                try:
                    from app.platform.integrations.feishu.notification import (
                        send_user_card_with_message_id,
                    )

                    message_id = await send_user_card_with_message_id(
                        open_id=leader_open_id,
                        title=dept_title,
                        content=dept_content,
                        elements=elements,
                    )
                    if message_id:
                        pushed += 1
                        # 保存 message_id 与员工按钮快照，供审批后置灰卡片
                        import json as _json

                        from app.core.redis import cache_set as _cache_set

                        snapshot = {
                            "title": dept_title,
                            "content": dept_content,
                            "emps": [
                                {
                                    "employee_number": e.get("employee_number", ""),
                                    "employee_name": e.get("name", ""),
                                    "leader_name": leader_name,
                                }
                                for e in dept_emps
                            ],
                        }
                        await _cache_set(
                            f"hr:contract:card:{dept_name}:msgid",
                            message_id,
                            ex=86400 * 14,
                        )
                        await _cache_set(
                            f"hr:contract:card:{dept_name}:emps",
                            _json.dumps(snapshot, ensure_ascii=False),
                            ex=86400 * 14,
                        )
                    else:
                        failed += 1
                except Exception:
                    failed += 1

            # 标记本次推送的员工（7天TTL，防止短期重复推送同一人）
            for emp in employees:
                emp_no = emp.get("employee_number", "")
                if emp_no:
                    await cache_set(f"hr:contract:pushed:{emp_no}", "1", ex=86400 * 7)

            return {
                "pushed": pushed,
                "failed": failed,
                "total_expiring": total,
                "skipped_pushed": skipped_pushed,
                "skipped_approved": skipped_approved,
                "auto_created_renewals": synced,
            }

    await submit_job(_do_push, task_id=task_id, ttl=600)
    return success_response(
        data={"task_id": task_id, "state": "running"},
        message="推送任务已启动，正在后台发送飞书消息...",
    )


@router.get("/employees/contract-expiring/push-status", summary="查询推送任务状态")
async def get_push_notify_status(
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)

    status = await get_job_status("hr:push:contract-expiring-notify")
    if not status:
        return success_response(
            data={"state": "idle", "progress": "无推送任务", "result": None}
        )
    return success_response(data=status)


# ─── 合同到期提醒模板配置（内存存储，重启后丢失，后续可迁移到数据库）───

_template_config: dict[str, Any] | None = None


@router.post("/employees/contract-expiring/template", summary="保存导入的模板配置")
async def save_template_config(
    payload: dict[str, Any],
    current_user: CurrentUser = None,
) -> Any:
    """保存模板格式配置，导出时使用"""
    _require_user(current_user)
    global _template_config
    _template_config = payload
    logger.info("模板配置已更新: %s", payload.get("title_text", ""))
    return success_response(message="模板配置已保存")


@router.get("/employees/contract-expiring/export", summary="导出合同到期提醒Excel")
async def export_contract_expiring(
    start_date: date = Query(..., description="开始日期"),
    end_date: date = Query(..., description="结束日期"),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    import io
    import logging
    from urllib.parse import quote

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side

    logger = logging.getLogger(__name__)

    try:
        employees, total = await service.list_contract_expiring(
            start_date=start_date,
            end_date=end_date,
            page=1,
            page_size=100,
        )

        sorted_emps = sorted(employees, key=lambda x: x.get("department", ""))

        # 读取模板配置
        template = _template_config or {}
        header_row = template.get("header_row", 3)
        template_headers = template.get(
            "headers",
            ["序号", "姓名", "一级部门", "二级部门", "合同到期日期", "车间领导审批"],
        )
        data_start_row = header_row + 1
        total_rows = template.get("total_rows", 35)
        title_text = template.get("title_text", "")

        wb = Workbook()
        ws = wb.active
        ws.title = template.get("sheet_name", "Sheet1")

        # 样式定义
        title_font = Font(name="宋体", size=18, bold=True)
        header_font = Font(name="宋体", size=11, bold=True)
        data_font = Font(name="宋体", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 标题行
        ws.merge_cells("A1:F2")
        title_cell = ws["A1"]
        quarter_num = (start_date.month - 1) // 3 + 1
        title_cell.value = (
            title_text if title_text else f"第{quarter_num}季度合同到期审批"
        )
        title_cell.font = title_font
        title_cell.alignment = center_align

        # 表头（6 列：序号/姓名/一级部门/二级部门/合同到期日期/车间领导审批）
        for col_idx, header in enumerate(template_headers[:6], 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 数据行（从模板配置的行开始）
        current_dept = ""
        dept_start_row = data_start_row

        for idx, emp in enumerate(sorted_emps):
            row = data_start_row + idx
            dept = emp.get("department", "") or ""
            sub_dept = emp.get("sub_department", "") or ""

            ws.cell(row=row, column=1, value=idx + 1).alignment = center_align
            ws.cell(
                row=row, column=2, value=str(emp.get("name", "") or "")
            ).alignment = center_align
            ws.cell(row=row, column=3, value=str(dept)).alignment = center_align
            ws.cell(row=row, column=4, value=str(sub_dept)).alignment = center_align

            end_dt = emp.get("contract_end_date", "")
            end_dt_str = (
                end_dt.isoformat()
                if hasattr(end_dt, "isoformat")
                else str(end_dt)
                if end_dt
                else ""
            )
            ws.cell(row=row, column=5, value=end_dt_str).alignment = center_align
            ws.cell(row=row, column=6, value="").alignment = center_align

            for col in range(1, 7):
                ws.cell(row=row, column=col).font = data_font
                ws.cell(row=row, column=col).border = thin_border

            if dept != current_dept:
                if current_dept != "" and dept_start_row < row - 1:
                    ws.merge_cells(
                        start_row=dept_start_row,
                        start_column=6,
                        end_row=row - 1,
                        end_column=6,
                    )
                current_dept = dept
                dept_start_row = row

        if dept_start_row < data_start_row + len(sorted_emps) - 1:
            ws.merge_cells(
                start_row=dept_start_row,
                start_column=6,
                end_row=data_start_row - 1 + len(sorted_emps),
                end_column=6,
            )

        # 填充空数据行（与导入模板一致）
        actual_total = max(data_start_row + len(sorted_emps), total_rows)
        for row in range(data_start_row + len(sorted_emps), actual_total):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col, value=None)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

        # 列宽 - 严格匹配模板
        ws.column_dimensions["A"].width = 12.78
        ws.column_dimensions["B"].width = 15.33
        ws.column_dimensions["C"].width = 13.56
        ws.column_dimensions["D"].width = 13.56
        ws.column_dimensions["E"].width = 21.44
        ws.column_dimensions["F"].width = 21.22

        # 行高 - 全部设为25磅
        for row in range(1, actual_total + 1):
            ws.row_dimensions[row].height = 25

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = (
            f"{start_date.year}年Q{(start_date.month - 1) // 3 + 1}合同到期提醒.xlsx"
        )
        encoded_filename = quote(filename)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                (
                    "Content-Disposition"
                ): f"attachment; filename*=UTF-8''{encoded_filename}"
            },
        )
    except Exception as e:
        logger.exception("Export error: %s", e)
        from app.core.exceptions import AppException

        raise AppException(status_code=500, message=f"导出失败: {str(e)}")


@router.post("/employees", summary="创建员工")
async def create_employee(
    payload: EmployeeCreate,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    created = await service.create_employee(payload)
    if isinstance(created, tuple):
        employee, sync_status = created
    else:
        employee, sync_status = created, "success"
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工创建成功",
        meta={"feishu_sync_status": sync_status},
        status_code=201,
    )


@router.post("/employees/public-create", summary="扫码公开创建员工")
async def create_employee_public(
    payload: EmployeePublicCreate,
    service: EmployeeService = Depends(get_employee_service),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """扫码填写表单后公开创建员工档案（不需要登录）。

    工号留空，状态默认"在职"。按姓名去重，已存在则返回提示。
    """
    from app.modules.hr.models import Employee
    from app.modules.hr.repository import EmployeeRepository

    emp_repo = EmployeeRepository(db)
    existing = await emp_repo.get_by_name(payload.name)
    if existing:
        raise AppException(status_code=409, message=f"{payload.name} 已存在员工档案")

    employee = Employee(**payload.model_dump(exclude_none=True))
    employee.status = "在职"
    result = await emp_repo.create(employee)

    # 同步到飞书
    sync_status = "success"
    try:
        bitable = await service._get_bitable()
        rid = await bitable.create(service._to_bitable_fields(result))
        if rid:
            result.feishu_record_id = rid
            result.feishu_synced_at = date.today()
            await emp_repo.update(result)
    except Exception as e:
        sync_status = f"failed: {str(e)}"

    return success_response(
        data=EmployeeResponse.model_validate(result).model_dump(mode="json"),
        message="员工档案创建成功",
        meta={"feishu_sync_status": sync_status},
        status_code=201,
    )


@router.post("/employees/sync-from-feishu", summary="从飞书多维表格同步员工数据")
async def sync_employees_from_feishu(
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """手动触发：从飞书多维表格拉取全部员工数据并 upsert 到本地 PG。"""
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，"
        f"删除 {stats.get('deleted', 0)} 条，"
        f"跳过 {stats.get('skipped', 0)} 条，"
        f"失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


@router.get("/employees/sync-status", summary="飞书同步状态")
async def get_employee_sync_status(
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """查看本地与飞书的数据同步统计。"""
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    status = await service.get_sync_status()
    return success_response(
        data=status.model_dump(mode="json"),
    )


@router.get("/employees/by-number/{employee_number}", summary="根据工号查询员工")
async def get_employee_by_number(
    employee_number: str,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    employee = await service.get_employee_by_number(employee_number)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.get("/employees/{employee_id}", summary="员工详情")
async def get_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    employee = await service.get_employee(employee_id)
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
    )


@router.put("/employees/{employee_id}", summary="更新员工")
async def update_employee(
    employee_id: UUID,
    payload: EmployeeUpdate,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    updated = await service.update_employee(employee_id, payload)
    if isinstance(updated, tuple):
        employee, sync_status = updated
    else:
        employee, sync_status = updated, "success"
    return success_response(
        data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
        message="员工更新成功",
        meta={"feishu_sync_status": sync_status},
    )


@router.delete("/employees/{employee_id}", summary="删除员工")
async def delete_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    sync_status = await service.delete_employee(employee_id)
    return success_response(
        message="员工删除成功", meta={"feishu_sync_status": sync_status}
    )


@router.post("/employees/{employee_id}/sync-to-feishu", summary="同步单个员工到飞书")
async def sync_employee_to_feishu(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """将本地单个员工强制同步到飞书多维表格。"""
    _require_user(current_user)
    await _assert_employee_page_access(service.session, current_user)
    record_id = await service.sync_to_feishu(employee_id)
    return success_response(
        data={"feishu_record_id": record_id},
        message="员工已同步到飞书",
    )


@router.post("/webhook/feishu-approval", summary="飞书审批完成回调")
async def feishu_approval_webhook(  # public endpoint - webhook from Feishu
    request: Request,
    payload: FeishuApprovalWebhookPayload,
    service: EmployeeService = Depends(get_employee_service),
) -> Any:
    """接收飞书审批完成通知，更新员工状态为在职。"""
    from app.core.security import verify_feishu_signature

    raw_body = await request.body()
    if not verify_feishu_signature(
        timestamp=request.headers.get("X-Lark-Request-Timestamp"),
        nonce=request.headers.get("X-Lark-Request-Nonce"),
        body=raw_body.decode("utf-8", errors="replace"),
        signature=request.headers.get("X-Lark-Signature"),
    ):
        raise AppException(status_code=401, message="回调签名校验失败")

    employee_number = payload.employee_number
    if not employee_number:
        return success_response(message="缺少工号")

    try:
        employee = await service.approve_employee(employee_number)
        return success_response(
            data=EmployeeResponse.model_validate(employee).model_dump(mode="json"),
            message="员工审批通过，状态已更新为在职",
        )
    except Exception:
        logger.exception("审批处理失败")
        raise AppException(status_code=500, message="审批处理失败，请稍后重试")


@router.get(
    "/employees/{employee_id}/onboarding-training-record",
    summary="导出员工入职培训记录",
)
async def export_onboarding_training_record(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """根据员工数据自动生成并下载入职培训记录 Word 文档。"""
    _require_user(current_user)
    employee = await service.get_employee(employee_id)
    try:
        buffer: BytesIO = generate_onboarding_training_record(employee)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    filename = f"onboarding_training_record_{employee.employee_number}.docx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/prejob-training-plan",
    summary="导出员工岗前培训计划",
)
async def export_prejob_training_plan(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """根据员工数据自动生成并下载岗前培训计划 Excel 文档。"""
    _require_user(current_user)
    employee = await service.get_employee(employee_id)
    try:
        buffer: BytesIO = generate_prejob_training_plan(employee)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    filename = f"prejob_training_plan_{employee.employee_number}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/employees/{employee_id}/onboarding-evaluation",
    summary="导出员工上岗评估表",
)
async def export_onboarding_evaluation_by_employee(
    employee_id: UUID,
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """根据员工档案预填基本信息并导出上岗评估表 Excel 文档。"""
    _require_user(current_user)
    employee = await service.get_employee(employee_id)

    payload = OnboardingEvaluationInput(
        employee_name=employee.name or "",
        employee_number=employee.employee_number or None,
        gender=employee.gender or None,
        department_position=f"{employee.department or ''}/{employee.position or ''}",
        hire_date=employee.hire_date,
    )
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    safe_date = (
        str(employee.hire_date).replace("-", "") if employee.hire_date else "nodate"
    )
    filename = f"onboarding_evaluation_{employee.employee_number}_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/training-sign-in-sheet", summary="生成培训签到表")
async def export_training_sign_in_sheet(
    payload: TrainingSignInSheetInput,
    current_user: CurrentUser = None,
) -> Any:
    """根据填写的培训信息自动生成培训签到表 Word 文档（APP3-SMP-HR-002-14）。

    签到表始终为一份 Word 文档：每 42 人一页，超出自动追加续页。
    """
    _require_user(current_user)
    safe_date = str(payload.training_date).replace("-", "")

    try:
        buffer: BytesIO = generate_training_sign_in_sheet(payload)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    filename = quote(f"7.5培训签到表_{safe_date}.docx")
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/training-notification", summary="生成培训通知")
async def export_training_notification(
    payload: TrainingNotificationInput,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """根据填写的培训信息自动生成培训通知 Word 文档。

    若应出席受训人员包含李健文(110000673)或黄丽耘(110001372)，
    自动为其创建培训台账记录。
    """
    _require_user(current_user)
    try:
        buffer: BytesIO = generate_training_notification(payload)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    # 自动关联创建培训台账记录
    employee_service = EmployeeService(service.repo.session)
    for name in payload.trainee_names:
        emp = await employee_service.repo.list_employees(
            keyword=name, page=1, page_size=1
        )
        if emp[0] and emp[0][0]:
            employee = emp[0][0]
            if employee.employee_number in {"110000673", "110001372"}:
                await service.create_from_notification(
                    employee_number=employee.employee_number,
                    training_date=payload.training_date,
                    training_subject=payload.subject,
                    training_method=None,
                    trainer=payload.trainer,
                    source_id=f"notification_{payload.training_date}_{payload.subject}",
                )

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    # 命名规范：日期-培训通知.docx（如 2026-08-06-培训通知.docx）
    d = payload.training_date
    filename = quote(f"{d:%Y-%m-%d}-培训通知.docx")
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/training-evaluation", summary="生成培训效果评估表")
async def export_training_evaluation(
    payload: TrainingEvaluationInput,
    current_user: CurrentUser = None,
) -> Any:
    "根据填写的培训信息自动生成培训评估表 Word 文档（APP4-SMP-HR-002-14 模板保真）。"
    _require_user(current_user)
    from app.modules.hr.training_evaluation_document_generator import (
        generate_training_evaluation_doc,
    )

    buffer: BytesIO = generate_training_evaluation_doc(payload.model_dump())

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    safe_date = (
        str(payload.training_date).replace("-", "")
        if payload.training_date
        else "nodate"
    )
    filename = quote(f"APP4培训评估表_{safe_date}.docx")
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/onboarding-evaluation", summary="生成员工上岗评估表")
async def export_onboarding_evaluation(
    payload: OnboardingEvaluationInput,
    current_user: CurrentUser = None,
) -> Any:
    """根据填写的评估信息自动生成员工上岗评估表 Excel 文档。"""
    _require_user(current_user)
    buffer: BytesIO = generate_onboarding_evaluation(payload)

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    safe_date = (
        str(payload.signature_date).replace("-", "")
        if payload.signature_date
        else "nodate"
    )
    filename = f"onboarding_evaluation_{safe_date}.xlsx"
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Department Routes ───


@router.get("/departments", summary="部门列表")
async def list_departments(
    keyword: str | None = Query(None, description="部门名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    departments, total = await service.list_departments(
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        DepartmentResponse.model_validate(d).model_dump(mode="json")
        for d in departments
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departments", summary="创建部门")
async def create_department(
    payload: DepartmentCreate,
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    department = await service.create_department(payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门创建成功",
        status_code=201,
    )


@router.get("/departments/tree", summary="部门树形结构")
async def get_department_tree(
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    tree = await service.get_department_tree()
    return success_response(data=tree)


@router.get("/departments/org-tree", summary="组织架构树（含人员）")
async def get_org_tree(
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    tree = await service.get_org_tree()
    return success_response(data=tree)


@router.post("/departments/sync-from-feishu", summary="从飞书同步部门信息（后台任务）")
async def sync_departments_from_feishu(
    force_refresh: bool = Query(False, description="强制刷新缓存，重新从飞书获取"),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    from app.core.database import async_session_factory
    from app.core.jobs import is_job_running, submit_job

    task_id = "hr:sync:departments"

    # 检查是否已有同步任务在运行（心跳过期的孤儿状态允许重新提交）
    if await is_job_running(task_id):
        return success_response(
            data={"task_id": task_id, "state": "running"},
            message="同步任务正在执行中，请稍后查询状态",
        )

    # 强制刷新时清除缓存
    if force_refresh:
        from app.core.redis import cache_delete

        await cache_delete("hr:feishu:departments")

    async def _do_sync() -> Any:
        async with async_session_factory() as session:
            from app.modules.hr.service import DepartmentService

            # DepartmentService 接收 session 并自建 repository，不能传 repo 进去
            service = DepartmentService(session)
            stats = await service.sync_departments_from_feishu()
            await session.commit()
            return stats

    await submit_job(_do_sync, task_id=task_id, ttl=600)
    return success_response(
        data={"task_id": task_id, "state": "running"},
        message="同步任务已启动",
    )


@router.get("/departments/sync-status", summary="查询飞书同步任务状态")
async def get_sync_status(
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)

    status = await get_job_status("hr:sync:departments")
    if not status:
        return success_response(
            data={"state": "idle", "progress": "无同步任务", "result": None}
        )
    return success_response(data=status)


@router.get("/departments/{department_id}", summary="部门详情")
async def get_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    department = await service.get_department(department_id)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
    )


@router.put("/departments/{department_id}", summary="更新部门")
async def update_department(
    department_id: UUID,
    payload: DepartmentUpdate,
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    department = await service.update_department(department_id, payload)
    return success_response(
        data=DepartmentResponse.model_validate(department).model_dump(mode="json"),
        message="部门更新成功",
    )


@router.delete("/departments/{department_id}", summary="删除部门")
async def delete_department(
    department_id: UUID,
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_department(department_id)
    return success_response(message="部门删除成功")


# ─── Team Routes ───


@router.get("/teams", summary="班组列表")
async def list_teams(
    department_id: UUID | None = Query(None, description="部门筛选"),
    keyword: str | None = Query(None, description="班组名称或编码关键词"),
    page_params: PageParams = Depends(),
    service: TeamService = Depends(get_team_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    teams, total = await service.list_teams(
        department_id=department_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [TeamResponse.model_validate(t).model_dump(mode="json") for t in teams]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/teams", summary="创建班组")
async def create_team(
    payload: TeamCreate,
    service: TeamService = Depends(get_team_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    team = await service.create_team(payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组创建成功",
        status_code=201,
    )


@router.get("/teams/{team_id}", summary="班组详情")
async def get_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    team = await service.get_team(team_id)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
    )


@router.put("/teams/{team_id}", summary="更新班组")
async def update_team(
    team_id: UUID,
    payload: TeamUpdate,
    service: TeamService = Depends(get_team_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    team = await service.update_team(team_id, payload)
    return success_response(
        data=TeamResponse.model_validate(team).model_dump(mode="json"),
        message="班组更新成功",
    )


@router.delete("/teams/{team_id}", summary="删除班组")
async def delete_team(
    team_id: UUID,
    service: TeamService = Depends(get_team_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_team(team_id)
    return success_response(message="班组删除成功")


# ─── OffboardingRecord Routes ───


@router.get("/offboarding-records", summary="离职记录列表")
async def list_offboarding_records(
    employee_id: UUID | None = Query(None, description="员工ID筛选"),
    keyword: str | None = Query(None, description="姓名或工号关键词"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _resolve_visible_scope(db, current_user)
    records, total = await service.list_records(
        employee_id=employee_id,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=alias_set,
    )
    data = [
        OffboardingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/offboarding-records", summary="创建离职记录")
async def create_offboarding_record(
    payload: OffboardingRecordCreate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.create_record(payload)
    # 手动构建响应，避免触发未加载的 relationship
    data = {
        "id": str(record.id),
        "employee_id": str(record.employee_id),
        "offboarding_date": (
            record.offboarding_date.isoformat() if record.offboarding_date else None
        ),
        "offboarding_type": record.offboarding_type,
        "reason": record.reason,
        "handover_status": record.handover_status,
        "notes": record.notes,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }
    return success_response(
        data=data,
        message="离职记录创建成功，员工状态已更新为离职",
        status_code=201,
    )


@router.get("/offboarding-records/{record_id}", summary="离职记录详情")
async def get_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/offboarding-records/{record_id}", summary="更新离职记录")
async def update_offboarding_record(
    record_id: UUID,
    payload: OffboardingRecordUpdate,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=OffboardingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="离职记录更新成功",
    )


@router.delete("/offboarding-records/{record_id}", summary="删除离职记录")
async def delete_offboarding_record(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_record(record_id)
    return success_response(message="离职记录删除成功")


@router.post("/offboarding-records/{record_id}/certificate", summary="生成离职证明")
async def generate_offboarding_certificate(
    record_id: UUID,
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    doc_buffer, filename, _record = await service.generate_termination_certificate(
        record_id
    )
    doc_buffer.seek(0)
    encoded_filename = quote(filename)
    return StreamingResponse(
        doc_buffer,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.post("/offboarding-records/sync-from-feishu", summary="从飞书同步离职管理数据")
async def sync_offboarding_from_feishu(
    service: OffboardingRecordService = Depends(get_offboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """手动触发：从飞书多维表格拉取全部离职记录并 upsert 到本地 PG。
    飞书中不存在的本地记录会被软删除（严格同步）。"""
    _require_user(current_user)
    stats = await service.sync_from_feishu()
    msg = (
        f"同步完成：新增 {stats['created']} 条，"
        f"更新 {stats['updated']} 条，"
        f"删除 {stats['deleted']} 条，"
        f"失败 {stats['failed']} 条"
    )
    return success_response(
        data=stats,
        message=msg,
    )


# ─── PositionTransferRecord Routes ───


@router.get("/position-transfers", summary="岗位调动列表")
async def list_position_transfers(
    employee_id: UUID | None = Query(None, description="员工ID筛选"),
    approval_status: str | None = Query(None, description="审批状态筛选"),
    keyword: str | None = Query(None, description="姓名/工号关键词"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _resolve_visible_scope(db, current_user)
    records, total = await service.list_records(
        employee_id=employee_id,
        approval_status=approval_status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=alias_set,
    )
    data = [
        PositionTransferRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/position-transfers", summary="创建岗位调动记录")
async def create_position_transfer(
    payload: PositionTransferRecordCreate,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.create_record(payload)
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
        message="岗位调动记录创建成功",
        status_code=201,
    )


@router.get("/position-transfers/approvals", summary="审批列表")
async def list_position_transfer_approvals(
    tab: str = Query(
        "my_applications", description="my_applications / pending_approval / approved"
    ),
    page_params: PageParams = Depends(),
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    records, total = await service.list_approvals(
        current_user=current_user,
        tab=tab,
        page=page_params.page,
        page_size=page_params.page_size,
    )
    data = [
        PositionTransferRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/position-transfers/sync-from-feishu", summary="从飞书同步岗位调动数据")
async def sync_position_transfers_from_feishu(
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    stats = await service.sync_from_feishu()
    return success_response(data=stats, message="岗位调动数据同步完成")


@router.get("/position-transfers/{record_id}", summary="岗位调动记录详情")
async def get_position_transfer(
    record_id: UUID,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
    )


@router.put("/position-transfers/{record_id}", summary="更新岗位调动记录")
async def update_position_transfer(
    record_id: UUID,
    payload: PositionTransferRecordUpdate,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
        message="岗位调动记录更新成功",
    )


@router.delete("/position-transfers/{record_id}", summary="删除岗位调动记录")
async def delete_position_transfer(
    record_id: UUID,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_record(record_id)
    return success_response(message="岗位调动记录删除成功")


@router.post("/position-transfers/{record_id}/submit", summary="提交审批")
async def submit_position_transfer_approval(
    record_id: UUID,
    payload: PositionTransferSubmitRequest,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    logger.info(
        "[SUBMIT-v3] record=%s, supervisor=%s, custom=%s",
        record_id,
        payload.is_supervisor_level,
        payload.custom_approvers,
    )
    record = await service.submit_for_approval(record_id, payload)
    approval_flow = record.approval_flow or {}
    logger.info(
        "[SUBMIT-v3] 完成 record=%s, steps=%d, current=%d",
        record_id,
        len(approval_flow.get("steps", [])),
        approval_flow.get("current_step", -1),
    )
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
        message="审批已提交",
    )


@router.post("/position-transfers/{record_id}/approve-node", summary="审批通过当前节点")
async def approve_position_transfer_node(
    record_id: UUID,
    payload: PositionTransferApproveNodeRequest,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.approve_current_node(record_id, payload)
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
        message="审批通过",
    )


@router.post("/position-transfers/{record_id}/reject-node", summary="审批拒绝当前节点")
async def reject_position_transfer_node(
    record_id: UUID,
    payload: PositionTransferRejectNodeRequest,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.reject_current_node(record_id, payload)
    return success_response(
        data=PositionTransferRecordResponse.model_validate(record).model_dump(
            mode="json"
        ),
        message="审批已拒绝",
    )


@router.get("/position-transfers/{record_id}/export", summary="导出内调申请表PDF")
async def export_position_transfer_pdf(
    record_id: UUID,
    service: PositionTransferRecordService = Depends(get_position_transfer_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    pdf_bytes, filename = await service.export_approval_pdf(record_id)
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
        },
    )


# ─── Department-level Approval Config Routes ───


@router.get("/dept-approval-configs", summary="部门级审批人配置列表（自动合并部门表）")
async def list_dept_approval_configs(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    # 查所有部门
    all_result = await db.execute(
        select(HrDepartment)
        .where(HrDepartment.is_deleted.is_(False))
        .order_by(HrDepartment.sort_order, HrDepartment.name)
    )
    all_depts = all_result.scalars().all()

    # 构建子部门映射
    children_map: dict[str, list[Any]] = {}
    for d in all_depts:
        if d.parent_id:
            pid = str(d.parent_id)
            children_map.setdefault(pid, []).append(d)

    # 排除 + 展开规则
    exclude_depts = DEPT_APPROVAL_EXCLUDE
    no_expand_depts = DEPT_APPROVAL_NO_EXPAND

    top_depts = [d for d in all_depts if not d.parent_id]
    display_depts = []
    for dept in top_depts:
        if dept.name in exclude_depts:
            continue
        children = children_map.get(str(dept.id), [])
        if children and dept.name not in no_expand_depts:
            display_depts.extend(children)
        else:
            display_depts.append(dept)

    # 查已有配置，按 department_id 索引
    config_result = await db.execute(
        select(HrDeptApprovalConfig).where(HrDeptApprovalConfig.is_deleted.is_(False))
    )
    config_map = {c.department_id: c for c in config_result.scalars().all()}

    data = []
    for dept in display_depts:
        cfg = config_map.get(dept.id)
        data.append(
            {
                "id": str(cfg.id) if cfg else None,
                "department_id": str(dept.id),
                "department_name": dept.name,
                "direct_leader_name": cfg.direct_leader_name
                if cfg
                else dept.leader_name,
                "direct_leader_open_id": cfg.direct_leader_open_id if cfg else None,
                "manager_name": cfg.manager_name if cfg else None,
                "manager_open_id": cfg.manager_open_id if cfg else None,
                "director_name": cfg.director_name if cfg else None,
                "director_open_id": cfg.director_open_id if cfg else None,
                "vp_name": cfg.vp_name if cfg else None,
                "vp_open_id": cfg.vp_open_id if cfg else None,
                "sort_order": cfg.sort_order if cfg else dept.sort_order,
            }
        )
    return success_response(data=data)


@router.post("/dept-approval-configs", summary="新增部门级审批人配置")
async def create_dept_approval_config(
    payload: DeptApprovalConfigCreate,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    config = HrDeptApprovalConfig(**payload.model_dump())
    db.add(config)
    await db.flush()
    await db.commit()
    await db.refresh(config)
    return success_response(
        data={"id": str(config.id), "department_name": config.department_name},
        message="部门审批人配置创建成功",
        status_code=201,
    )


@router.put("/dept-approval-configs/{config_id}", summary="更新部门级审批人配置")
async def update_dept_approval_config(
    config_id: UUID,
    payload: DeptApprovalConfigUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await db.execute(
        select(HrDeptApprovalConfig).where(
            HrDeptApprovalConfig.id == config_id,
            HrDeptApprovalConfig.is_deleted.is_(False),
        )
    )
    config = result.scalar_one_or_none()
    if not config:
        raise NotFoundException("部门审批配置", str(config_id))
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(config, field, value)
    await db.flush()
    await db.commit()
    return success_response(message="部门审批人配置更新成功")


@router.delete("/dept-approval-configs/{config_id}", summary="删除部门级审批人配置")
async def delete_dept_approval_config(
    config_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await db.execute(
        select(HrDeptApprovalConfig).where(
            HrDeptApprovalConfig.id == config_id,
            HrDeptApprovalConfig.is_deleted.is_(False),
        )
    )
    config = result.scalar_one_or_none()
    if not config:
        raise NotFoundException("部门审批配置", str(config_id))
    config.is_deleted = True
    await db.flush()
    await db.commit()
    return success_response(message="部门审批人配置删除成功")


@router.post(
    "/dept-approval-configs/init-from-departments", summary="从部门列表初始化审批人配置"
)
async def init_dept_approval_configs_from_departments(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    # 查所有顶级部门
    result = await db.execute(
        select(HrDepartment)
        .where(
            HrDepartment.is_deleted.is_(False),
            HrDepartment.parent_id.is_(None),
        )
        .order_by(HrDepartment.sort_order, HrDepartment.name)
    )
    departments = result.scalars().all()

    created = 0
    skipped = 0
    for dept in departments:
        # 检查是否已有配置
        existing = await db.execute(
            select(HrDeptApprovalConfig).where(
                HrDeptApprovalConfig.department_id == dept.id,
                HrDeptApprovalConfig.is_deleted.is_(False),
            )
        )
        if existing.scalar_one_or_none():
            skipped += 1
            continue
        config = HrDeptApprovalConfig(
            department_id=dept.id,
            department_name=dept.name,
            direct_leader_name=dept.leader_name,
            sort_order=dept.sort_order,
        )
        db.add(config)
        created += 1

    await db.flush()
    await db.commit()
    return success_response(
        data={"created": created, "skipped": skipped, "total": len(departments)},
        message=f"初始化完成：新增 {created} 条，跳过 {skipped} 条",
    )


# ─── TrainingLedger Routes ───


@router.get("/training-ledgers", summary="培训台账列表")
async def list_training_ledgers(
    employee_number: str | None = Query(None, description="工号筛选"),
    department: str | None = Query(None, description="按授课部门筛选"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    session_id: UUID | None = Query(None, description="培训会话筛选（防重复入台账）"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _assert_dept_in_scope(db, current_user, department)
    if session_id:
        records, total = await service.list_records(
            session_id=session_id,
            page=page_params.page,
            page_size=page_params.page_size,
        )
        return success_response(
            data=[
                TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
                for r in records
            ],
            meta={
                "page": page_params.page,
                "page_size": page_params.page_size,
                "total": total,
            },
        )
    if department:
        records, total = await service.list_by_department(
            department=department,
            page=page_params.page,
            page_size=page_params.page_size,
        )
    else:
        records, total = await service.list_records(
            employee_number=employee_number,
            date_from=date_from,
            date_to=date_to,
            page=page_params.page,
            page_size=page_params.page_size,
            sort_by="training_date",
            sort_order="asc",
            dept_alias_set=alias_set,
        )
    data = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-ledgers", summary="创建培训台账记录")
async def create_training_ledger(
    payload: TrainingLedgerCreate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.create_record(payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录创建成功",
        status_code=201,
    )


@router.post(
    "/training-ledgers/check-conflict",
    summary="检测培训时间冲突",
    response_model=ApiResponseEnvelope[TrainingConflictCheckResponse],
)
async def check_training_conflict(
    payload: TrainingConflictCheckRequest,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await service.check_conflict(
        training_date=payload.training_date,
        time_start=payload.time_start,
        time_end=payload.time_end,
        instructor=payload.instructor,
        trainees=payload.trainees,
        exclude_session_id=payload.exclude_session_id,
    )
    return success_response(data=result)


# ─── TrainingLedgerPage Routes (must be before /{record_id}) ───


@router.get("/training/departments", summary="获取培训模块所有有数据的部门")
async def list_training_departments(
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """台账/ESG/年度计划/岗位清单/培训师/培训会话
    中有数据的部门并集，供培训各页面部门Tab/筛选使用.

    部门级数据隔离：非管理员仅返回可见范围内的部门（前端 Tab/下拉据此自动收窄）。
    """
    alias_set = await _resolve_visible_scope(db, current_user)
    departments = await service.list_training_departments()
    if alias_set is not None:
        departments = [d for d in departments if d in alias_set]
    return success_response(data=departments)


@router.get("/training/departments/custom", summary="获取手动添加的自定义培训部门列表")
async def list_custom_training_departments(
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """手动添加的自定义部门列表，供前端判断哪些部门可删除."""
    alias_set = await _resolve_visible_scope(db, current_user)
    departments = await service.list_custom_training_departments()
    if alias_set is not None:
        departments = [d for d in departments if d in alias_set]
    return success_response(data=departments)


@router.post("/training/departments", summary="添加自定义培训部门")
async def add_custom_training_department(
    body: CustomTrainingDepartmentCreate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """手动添加一个自定义培训部门，补充数据驱动部门."""
    _require_user(current_user)
    result = await service.add_custom_training_department(body.name)
    return success_response(data=result)


@router.delete("/training/departments/{name}", summary="删除自定义培训部门")
async def delete_custom_training_department(
    name: str,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """软删除一个自定义培训部门（不能删除有培训数据的数据驱动部门）."""
    _require_user(current_user)
    deleted = await service.delete_custom_training_department(name)
    if not deleted:
        raise AppException(status_code=404, message=f"自定义部门「{name}」不存在")
    return success_response(data={"deleted": True})


# ─── 培训部门映射配置（HR 设置维护，替代硬编码字典）───


@router.get("/training/dept-mappings", summary="培训部门映射配置列表")
async def list_training_dept_mappings(
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """全部映射配置（特殊映射/别名归并/201归一/候选人来源/台账拆分/打印统一/弹窗规则）。

    前端与后端解析共用同一份数据源。
    """
    _require_user(current_user)
    data = await service.list_dept_mappings()
    return success_response(data=data)


@router.post("/training/dept-mappings", summary="新增培训部门映射")
async def create_training_dept_mapping(
    body: TrainingDeptMappingCreate,
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """新增一条映射（source+match_level+mapping_type+target 查重）。"""
    _require_user(current_user)
    assert current_user is not None
    from app.platform.identity.rbac import resolve_user_permissions

    perms = await resolve_user_permissions(db, current_user.id)
    if "*" not in perms and "hr:write" not in perms:
        raise AppException(status_code=403, message="需要 hr:write 权限")
    data = await service.create_dept_mapping(body, user_id=current_user.id)
    return success_response(data=data)


@router.put("/training/dept-mappings/{mapping_id}", summary="更新培训部门映射")
async def update_training_dept_mapping(
    mapping_id: UUID,
    body: TrainingDeptMappingUpdate,
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """更新一条映射（仅非 None 字段生效）。"""
    _require_user(current_user)
    assert current_user is not None
    from app.platform.identity.rbac import resolve_user_permissions

    perms = await resolve_user_permissions(db, current_user.id)
    if "*" not in perms and "hr:write" not in perms:
        raise AppException(status_code=403, message="需要 hr:write 权限")
    data = await service.update_dept_mapping(mapping_id, body, user_id=current_user.id)
    return success_response(data=data)


@router.delete("/training/dept-mappings/{mapping_id}", summary="删除培训部门映射")
async def delete_training_dept_mapping(
    mapping_id: UUID,
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """软删除一条映射。"""
    _require_user(current_user)
    assert current_user is not None
    from app.platform.identity.rbac import resolve_user_permissions

    perms = await resolve_user_permissions(db, current_user.id)
    if "*" not in perms and "hr:write" not in perms:
        raise AppException(status_code=403, message="需要 hr:write 权限")
    deleted = await service.delete_dept_mapping(mapping_id)
    if not deleted:
        raise AppException(status_code=404, message="映射不存在")
    return success_response(data={"deleted": True})


# ─── 用户可见部门配置（部门级数据隔离管理，仅 hr:write 管理员可用）───


class DeptScopeUpdate(_PydanticBaseModel):
    """更新用户可见培训部门配置。"""

    visible_depts: list[str] = Field(default_factory=list)
    user_name: str | None = Field(
        default=None, description="联系人姓名（用户未登录过时预创建用）"
    )
    user_department: str | None = Field(
        default=None, description="联系人部门（用户未登录过时预创建用）"
    )


@router.get("/dept-scopes", summary="可见部门配置列表（含用户名）")
async def list_dept_scopes(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """返回全部可见部门配置（join 用户名），供管理员在 HR 设置页维护。"""
    _require_user(current_user)
    from app.modules.hr.models import HrUserDeptScope
    from app.platform.identity.models import User

    result = await db.execute(
        select(HrUserDeptScope, User.name, User.department)
        .outerjoin(User, HrUserDeptScope.user_id == User.id)
        .where(HrUserDeptScope.is_deleted.is_(False))
        .order_by(User.name.asc())
    )
    data = [
        {
            "user_id": str(scope.user_id),
            "user_name": user_name or "",
            "user_department": user_dept or "",
            "visible_depts": scope.visible_depts or [],
            "updated_at": scope.updated_at.isoformat() if scope.updated_at else None,
        }
        for scope, user_name, user_dept in result.all()
    ]
    return success_response(data=data)


async def _resolve_scope_user_uuid(db: AsyncSession, user_id: str) -> UUID | None:
    """将 user_id 参数解析为系统用户 UUID。

    前端添加用户来自飞书联系人搜索，传的是飞书 open_id（如 ou_xxx）；
    已保存配置的行传的是系统 UUID。两者都接受：UUID 直接用，
    非 UUID 按 feishu_open_id 查 identity.users 解析。
    """
    try:
        return UUID(user_id)
    except ValueError:
        pass
    from app.platform.identity.models import User

    result = await db.execute(
        select(User.id).where(
            User.feishu_open_id == user_id,
            User.is_deleted.is_(False),
        )
    )
    return result.scalar_one_or_none()


@router.get("/dept-scopes/{user_id}", summary="查询单用户可见部门配置")
async def get_dept_scope(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    from app.modules.hr.models import HrUserDeptScope

    user_uuid = await _resolve_scope_user_uuid(db, user_id)
    if user_uuid is None:
        # 用户尚未配置过（未登录/未预创建）→ 返回空配置
        return success_response(data={"user_id": user_id, "visible_depts": []})
    result = await db.execute(
        select(HrUserDeptScope).where(
            HrUserDeptScope.user_id == user_uuid,
            HrUserDeptScope.is_deleted.is_(False),
        )
    )
    scope = result.scalar_one_or_none()
    return success_response(
        data={
            "user_id": str(user_uuid),
            "visible_depts": scope.visible_depts or [] if scope else [],
        }
    )


@router.put("/dept-scopes/{user_id}", summary="设置用户可见培训部门（upsert）")
async def upsert_dept_scope(
    user_id: str,
    body: DeptScopeUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """设置用户可见培训部门（培训规范名数组）；空数组 = 清除配置，回退白名单（不可见）。

    支持提前配置：用户未登录过系统时，用飞书联系人信息预创建用户记录，
    该用户首次飞书登录后配置立即生效。
    """
    _require_user(current_user)
    from app.modules.hr.models import HrUserDeptScope
    from app.platform.identity.models import User

    user_uuid = await _resolve_scope_user_uuid(db, user_id)
    if user_uuid is None:
        # 预创建用户（提前配置）：联系人姓名/部门来自前端，未登录过也能配置
        new_user = User(
            name=body.user_name or f"待登录用户({user_id[:12]})",
            department=body.user_department,
            feishu_open_id=user_id,
        )
        db.add(new_user)
        await db.flush()
        user_uuid = new_user.id
    # 不过滤 is_deleted：软删除的记录复用恢复，避免 user_id 唯一约束冲突
    result = await db.execute(
        select(HrUserDeptScope).where(HrUserDeptScope.user_id == user_uuid)
    )
    scope = result.scalar_one_or_none()
    if scope is None:
        scope = HrUserDeptScope(user_id=user_uuid, visible_depts=body.visible_depts)
        db.add(scope)
    else:
        scope.visible_depts = body.visible_depts
        scope.is_deleted = False
    await db.commit()
    await db.refresh(scope)
    return success_response(
        data={"user_id": str(user_uuid), "visible_depts": scope.visible_depts or []},
        message="可见部门配置已保存",
    )


@router.delete("/dept-scopes/{user_id}", summary="清除用户可见部门配置")
async def delete_dept_scope(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """清除配置（软删除），该用户回退白名单（不可见任何部门）。"""
    _require_user(current_user)
    from app.modules.hr.models import HrUserDeptScope

    user_uuid = await _resolve_scope_user_uuid(db, user_id)
    if user_uuid is None:
        # 用户不存在（从未配置过）→ 幂等返回成功
        return success_response(message="可见部门配置已清除")
    result = await db.execute(
        select(HrUserDeptScope).where(
            HrUserDeptScope.user_id == user_uuid,
            HrUserDeptScope.is_deleted.is_(False),
        )
    )
    scope = result.scalar_one_or_none()
    if scope is not None:
        scope.is_deleted = True
        await db.commit()
    return success_response(message="可见部门配置已清除")


@router.get("/training-ledgers/pages", summary="已创建的培训台账页面列表")
async def list_training_ledger_pages(
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    pages_with_dept = await service.list_pages_with_department()
    data = [
        {
            "id": str(page.id),
            "employee_number": page.employee_number,
            "employee_name": page.employee_name,
            "department": dept or "未知部门",
            "created_at": page.created_at.isoformat() if page.created_at else None,
            "updated_at": page.updated_at.isoformat() if page.updated_at else None,
        }
        for page, dept in pages_with_dept
    ]
    return success_response(data=data)


@router.post("/training-ledgers/pages", summary="创建培训台账页面")
async def create_training_ledger_page(
    payload: TrainingLedgerPageCreate,
    service: TrainingLedgerPageService = Depends(get_training_ledger_page_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    page = await service.create_page(payload)
    return success_response(
        data=TrainingLedgerPageResponse(
            id=page.id,
            employee_number=page.employee_number,
            employee_name=page.employee_name,
            department=None,
            created_at=page.created_at,
            updated_at=page.updated_at,
        ).model_dump(mode="json"),
        message="培训台账页面创建成功",
        status_code=201,
    )


def _generate_training_ledger_excel(
    employee: dict[str, Any], records: list[dict[str, Any]]
) -> BytesIO:
    """Generate training ledger Excel based on employee training ledger format."""
    records = list(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "员工培训台账"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    ws.merge_cells("A1:G1")
    ws["A1"] = "丽珠集团新北江制药股份有限公司"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "员工培训台账"
    ws["A2"].font = bold_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 24

    ws["A3"] = "姓名"
    ws["A3"].font = bold_font
    ws["A3"].alignment = center_align
    ws["A3"].border = thin_border
    ws["B3"] = employee.get("name", "")
    ws["B3"].border = thin_border
    ws["C3"] = "性别"
    ws["C3"].font = bold_font
    ws["C3"].alignment = center_align
    ws["C3"].border = thin_border
    ws["D3"] = employee.get("gender", "")
    ws["D3"].border = thin_border
    ws["E3"] = "工作卡号"
    ws["E3"].font = bold_font
    ws["E3"].alignment = center_align
    ws["E3"].border = thin_border
    ws.merge_cells("F3:G3")
    ws["F3"] = employee.get("employee_number", "")
    ws["F3"].border = thin_border
    ws["G3"].border = thin_border

    ws["A4"] = "部门"
    ws["A4"].font = bold_font
    ws["A4"].alignment = center_align
    ws["A4"].border = thin_border
    ws["B4"] = employee.get("department", "")
    ws["B4"].border = thin_border
    ws["C4"] = "岗位/职务"
    ws["C4"].font = bold_font
    ws["C4"].alignment = center_align
    ws["C4"].border = thin_border
    ws["D4"] = employee.get("position", "")
    ws["D4"].border = thin_border
    ws["E4"] = "入厂时间"
    ws["E4"].font = bold_font
    ws["E4"].alignment = center_align
    ws["E4"].border = thin_border
    ws.merge_cells("F4:G4")
    ws["F4"] = employee.get("factory_entry_date") or employee.get("hire_date", "")
    ws["F4"].border = thin_border
    ws["G4"].border = thin_border

    ws["A5"] = "岗位变动"
    ws["A5"].font = bold_font
    ws["A5"].alignment = center_align
    ws["A5"].border = thin_border
    ws.merge_cells("B5:G5")
    ws["B5"] = employee.get("transfer_history", "无")
    ws["B5"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=5, column=c).border = thin_border

    ws["A6"] = "记录"
    ws["A6"].font = bold_font
    ws["A6"].alignment = center_align
    ws["A6"].border = thin_border
    ws.merge_cells("B6:G6")
    ws["B6"] = ""
    ws["B6"].border = thin_border
    for c in range(3, 8):
        ws.cell(row=6, column=c).border = thin_border

    headers = [
        "年月日",
        "培训课程",
        "培训方式",
        "课时",
        "培训单位/培训师",
        "考核成绩",
        "备注",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[7].height = 24

    for idx, record in enumerate(records, 8):
        values = [
            record.get("training_date", ""),
            record.get("training_subject", ""),
            record.get("training_method", ""),
            record.get("duration_hours", ""),
            record.get("trainer", ""),
            record.get("assessment_result", ""),
            record.get("remarks", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 3, 4, 6, 7) else left_align

    while len(records) < 12:
        row = 8 + len(records)
        for col in range(1, 8):
            ws.cell(row=row, column=col, value="").border = thin_border
        records.append({})

    footer_row = 8 + len(records)
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    ws.cell(
        row=footer_row,
        column=1,
        value="备注：笔试考核设置为满分100分，考试合格线为80分。",
    )
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 8):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/training-ledgers/export", summary="导出培训台账Excel")
async def export_training_ledger(
    employee_number: str = Query(..., description="员工工号"),
    ledger_service: TrainingLedgerService = Depends(get_training_ledger_service),
    employee_service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    """根据员工数据生成并导出培训台账 Excel 文件。"""
    _require_user(current_user)
    employee = await employee_service.get_employee_by_number(employee_number)
    if not employee:
        raise NotFoundException(resource="员工")

    records, _ = await ledger_service.list_records(
        employee_number=employee_number,
        page=1,
        page_size=1000,
        sort_by="training_date",
        sort_order="asc",
    )

    employee_dict = EmployeeResponse.model_validate(employee).model_dump(mode="json")
    record_dicts = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]

    buffer = _generate_training_ledger_excel(employee_dict, record_dicts)
    buffer.seek(0)

    safe_name = employee.name or "unknown"
    filename = f"{safe_name}培训台账.xlsx"
    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )


# ── 部门级培训台账 导入/导出（年度培训统计表 SMP-HR-002-14）──

_DEPT_LEDGER_HEADERS = [
    "培训时间",
    "培训日期",
    "培训时长（h）",
    "培训内容",
    "授课部门",
    "授课人",
    "一级/二级",
    "涉及部门",
    "培训对象",
    "培训类型",
    "考核方式",
    "部门/公司计划",
    "人药/兽药",
    "成绩汇总",
]
_DEPT_LEDGER_FIELDS = [
    "training_datetime",
    "training_date",
    "duration_hours",
    "training_content",
    "teaching_dept",
    "instructor",
    "level_category",
    "involved_depts",
    "trainees",
    "training_type",
    "ledger_assessment_method",
    "plan_source",
    "drug_category",
    "score_summary",
]

# 表头别名 → 字段名映射（兼容系统导出格式与各部门实际统计格式）
_DEPT_LEDGER_HEADER_ALIASES = {
    "培训时间": "training_datetime",
    "日期/时间": "training_datetime",
    "培训日期": "training_date",
    "培训时长（h）": "duration_hours",
    "时长（h）": "duration_hours",
    "时长/h": "duration_hours",
    "培训内容": "training_content",
    "授课部门": "teaching_dept",
    "授课人": "instructor",
    "培训师": "instructor",
    "一级/二级": "level_category",
    "培训级别": "level_category",
    "涉及部门": "involved_depts",
    "受训部门": "involved_depts",
    "培训对象": "trainees",
    "受训人员名单": "trainees",
    "受训人员": "trainees",
    "培训类型": "training_type",
    "考核方式": "ledger_assessment_method",
    "考核结果": "score_summary",
    "培训结果": "score_summary",
    "成绩汇总": "score_summary",
    "部门/公司计划": "plan_source",
    "人药/兽药": "drug_category",
    "备注": "remarks",
}


def _parse_excel_date(v: Any) -> date | None:
    """把 Excel 单元格值转为 date，支持日期对象与常见文本格式.

    兼容部门统计表中的 "2026.01.06\\n09:00~10:00" 这类带时间的文本，
    仅取其中的日期部分。
    """
    from datetime import datetime as _dt

    if v is None or v == "":
        return None
    if isinstance(v, _dt):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # 取第一行（日期/时间混合文本中的日期部分）
    first = s.split("\n")[0].strip()
    first = first.split()[0] if first else ""
    if not first:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return _dt.strptime(first, fmt).date()
        except ValueError:
            continue
    return None


_TIME_RANGE_RE = re.compile(r"(\d{1,2}):(\d{2})\s*[~～\-—至到]\s*(\d{1,2}):(\d{2})")


def _calc_duration_from_text(v: Any) -> float | None:
    """从培训时间文本中提取时间段（如 09:00~10:00）计算时长，单位小时.

    仅在时长缺失时使用；值来自 Excel 单元格原文的确定性推导，
    解析不出返回 None，绝不猜测。
    """
    if v is None:
        return None
    m = _TIME_RANGE_RE.search(str(v))
    if not m:
        return None
    h1, m1, h2, m2 = (int(g) for g in m.groups())
    if h1 > 23 or h2 > 23 or m1 > 59 or m2 > 59:
        return None
    start = h1 * 60 + m1
    end = h2 * 60 + m2
    if end <= start:
        return None
    return round((end - start) / 60, 2)


def _locate_header_row(ws: Any) -> tuple[int, list[str]]:
    '在前 5 行内定位严格表头行（含"培训内容" + 日期类列），返回 (行号, 表头文本列表).'
    for ri, row in enumerate(
        ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1
    ):
        cells = [str(c).strip() if c is not None else "" for c in row]
        has_content = "培训内容" in cells
        has_date = any(h in cells for h in ("培训日期", "培训时间", "日期/时间"))
        if has_content and has_date:
            return ri, cells
    return 0, []


def _map_headers_by_alias(headers: list[str]) -> dict[str, str]:
    """按别名表把表头映射为系统字段，返回 {列索引str: 字段名}."""
    mapping: dict[str, str] = {}
    for ci, text_val in enumerate(headers):
        if text_val in _DEPT_LEDGER_HEADER_ALIASES:
            mapping[str(ci)] = _DEPT_LEDGER_HEADER_ALIASES[text_val]
    return mapping


def _find_candidate_header_row(ws: Any) -> tuple[int, list[str]]:
    """宽松候选表头：前 5 行中非空单元格最多且 >=3 的行（供 AI 兜底分析）."""
    best_row: int = 0
    best_cells: list[str] = []
    for ri, row in enumerate(
        ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1
    ):
        cells = [str(c).strip() if c is not None else "" for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) >= 3 and len(non_empty) > len([c for c in best_cells if c]):
            best_row, best_cells = ri, cells
    return best_row, best_cells


def _read_excel_header_map(ws: Any) -> tuple[int, dict[int, str]]:
    """在前 5 行内定位表头行，返回 (表头行索引, 列索引->字段名).

    兼容两种格式：
    - 系统导出格式（培训时间/培训日期/授课人/涉及部门/培训对象...）
    - 部门实际统计格式（日期/时间/培训师/受训部门/受训人员名单...）
    """
    header_row, headers = _locate_header_row(ws)
    if not header_row:
        return 0, {}
    return header_row, {int(k): v for k, v in _map_headers_by_alias(headers).items()}


def _generate_dept_ledger_excel(
    department: str, records: list[dict[str, Any]]
) -> BytesIO:
    """生成部门年度培训统计表 Excel（与页面统计表列一致）."""
    wb = Workbook()
    ws = wb.active
    ws.title = "年度培训统计表"

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_DEPT_LEDGER_HEADERS)
    )
    ws["A1"] = f"{department} 年度培训统计表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    for ci, h in enumerate(_DEPT_LEDGER_HEADERS, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.alignment = center
        cell.border = thin

    for ri, rec in enumerate(records, start=3):
        for ci, field in enumerate(_DEPT_LEDGER_FIELDS, start=1):
            val = rec.get(field)
            if isinstance(val, date):
                val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = (
                left
                if field in ("training_content", "involved_depts", "trainees")
                else center
            )

    return _wb_to_buffer(wb)


def _wb_to_buffer(wb: Any) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/training-ledgers/export-by-dept", summary="按部门导出年度培训统计表Excel")
async def export_training_ledger_by_dept(
    department: str = Query(..., description="部门"),
    date_from: date | None = Query(None, description="培训日期起（筛选全年/月份）"),
    date_to: date | None = Query(None, description="培训日期止"),
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_dept_in_scope(db, current_user, department)
    records, _ = await service.list_by_department(
        department=department, page=1, page_size=10000
    )
    if date_from:
        records = [
            r for r in records if r.training_date and r.training_date >= date_from
        ]
    if date_to:
        records = [r for r in records if r.training_date and r.training_date <= date_to]
    record_dicts = [
        TrainingLedgerResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    buffer = _generate_dept_ledger_excel(department, record_dicts)
    encoded = quote(f"{department}年度培训统计表.xlsx", safe="")
    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded}"},
    )


# ─── 员工培训清单（配置表方案）───


@router.get(
    "/training/employee-training-lists",
    summary="员工培训清单-员工汇总列表",
    response_model=ApiResponseEnvelope[list[EmployeeTrainingSummaryOut]],
)
async def list_employee_training_lists(
    department: str = Query(..., description="部门"),
    name: str | None = Query(None, description="姓名筛选"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    db: AsyncSession = Depends(get_db),
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_dept_in_scope(db, current_user, department)
    data = await service.list_employee_training_summary(
        department=department, name=name, date_from=date_from, date_to=date_to
    )
    return success_response(data=data)


@router.get(
    "/training/employee-training-list/members",
    summary="员工培训清单-部门配置人员",
    response_model=ApiResponseEnvelope[list[EmployeeTrainingListMemberOut]],
)
async def list_employee_training_members(
    department: str = Query(..., description="部门"),
    db: AsyncSession = Depends(get_db),
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_dept_in_scope(db, current_user, department)
    data = await service.list_employee_members(department)
    return success_response(data=data)


@router.get(
    "/training/employee-training-list/records",
    summary="员工培训清单-个人培训记录",
    response_model=ApiResponseEnvelope[list[EmployeeTrainingRecordOut]],
)
async def get_employee_training_records(
    name: str = Query(..., description="员工姓名"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    """某员工全部培训台账记录（含个人考核结果，右面板清单表格数据源）."""
    _require_user(current_user)
    data = await service.get_employee_training_records(
        name=name, date_from=date_from, date_to=date_to
    )
    return success_response(data=data)


@router.post(
    "/training/employee-training-list/members/import-feishu",
    summary="员工培训清单-一键导入飞书联系人",
    response_model=ApiResponseEnvelope[ImportFeishuMembersResult],
)
async def import_feishu_members(
    payload: ImportFeishuMembersRequest,
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await service.import_feishu_members(department=payload.department)
    return success_response(
        data=ImportFeishuMembersResult(**result).model_dump(mode="json"),
        message=f"导入完成：共 {result['total']} 人",
    )


@router.post(
    "/training/employee-training-list/members",
    summary="员工培训清单-手动添加人员",
    response_model=ApiResponseEnvelope[EmployeeTrainingListMemberOut],
)
async def add_employee_training_member(
    payload: EmployeeTrainingListMemberCreate,
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    member = await service.add_member(
        department=payload.department,
        name=payload.name,
        employee_number=payload.employee_number,
    )
    return success_response(
        data=EmployeeTrainingListMemberOut(**member).model_dump(mode="json"),
        message="人员已添加",
    )


@router.delete(
    "/training/employee-training-list/members/{member_id}",
    summary="员工培训清单-移除人员",
)
async def remove_employee_training_member(
    member_id: UUID,
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.remove_member(member_id)
    return success_response(message="人员已移除")


@router.put(
    "/training/employee-training-list/members/{member_id}",
    summary="员工培训清单-编辑人员姓名",
    response_model=ApiResponseEnvelope[EmployeeTrainingListMemberOut],
)
async def update_employee_training_member(
    member_id: UUID,
    payload: EmployeeTrainingListMemberUpdate,
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    member = await service.update_member_name(member_id, payload.name)
    return success_response(
        data=EmployeeTrainingListMemberOut(**member).model_dump(mode="json"),
        message="姓名已更新",
    )


@router.get("/training/employee-training-list/export", summary="员工培训清单-导出")
async def export_employee_training_list(
    department: str = Query(..., description="部门"),
    name: str | None = Query(None, description="姓名（空=导出整个部门 zip）"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    db: AsyncSession = Depends(get_db),
    service: EmployeeTrainingListService = Depends(get_employee_training_list_service),
    current_user: CurrentUser = None,
) -> Any:
    """单员工导出单 xlsx；部门导出 zip（每人一份 xlsx）."""
    await _assert_dept_in_scope(db, current_user, department)
    from app.modules.hr.employee_training_list_document_generator import (
        generate_employee_training_list,
    )
    from app.modules.hr.practical_exam_document_generator import _safe_filename

    def _export_year(records: list[dict[str, Any]]) -> int:
        """年份取记录中最近一次培训的年份，无记录用当前年."""
        years: list[date] = []
        for record in records:
            raw_date = record.get("training_date")
            if isinstance(raw_date, datetime):
                years.append(raw_date.date())
            elif isinstance(raw_date, date):
                years.append(raw_date)
            elif isinstance(raw_date, str):
                try:
                    years.append(date.fromisoformat(raw_date[:10]))
                except ValueError:
                    continue
        if years:
            return max(years).year
        return date.today().year

    if name:
        records = await service.get_employee_training_records(
            name=name, date_from=date_from, date_to=date_to
        )
        if not records:
            raise AppException(
                status_code=400, message=f"{name} 在筛选条件下暂无培训记录，无法导出"
            )
        year = _export_year(records)
        buffer = generate_employee_training_list(department, name, records, year=year)
        encoded = quote(f"{year}年-{department}-{name}-员工培训清单.xlsx", safe="")
        return StreamingResponse(
            iter([buffer.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded}"},
        )

    # 整部门 zip
    members = await service.list_employee_members(department)
    zip_buffer = BytesIO()
    file_count = 0
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for member in members:
            records = await service.get_employee_training_records(
                name=member["name"], date_from=date_from, date_to=date_to
            )
            if not records:
                continue
            year = _export_year(records)
            buffer = generate_employee_training_list(
                department, member["name"], records, year=year
            )
            zf.writestr(
                f"{year}年-{_safe_filename(member['name'])}-员工培训清单.xlsx",
                buffer.read(),
            )
            file_count += 1
    if file_count == 0:
        raise AppException(
            status_code=400,
            message=f"{department} 在筛选条件下暂无培训记录，无法导出",
        )
    zip_buffer.seek(0)
    encoded_zip = quote(f"{department}-员工培训清单.zip", safe="")
    return StreamingResponse(
        iter([zip_buffer.read()]),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded_zip}"},
    )


@router.post(
    "/training-ledgers/import-by-dept", summary="按部门导入年度培训统计表Excel"
)
async def import_training_ledger_by_dept(
    file: UploadFile = File(..., description="Excel文件(.xlsx)"),
    department: str = Query(..., description="导入到哪个部门"),
    sheet_name: str | None = Query(
        None, description="要导入的工作表名（默认第一个可识别的工作表）"
    ),
    all_sheets: bool = Query(False, description="合并导入所有可识别的工作表"),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    """解析部门年度培训统计表 Excel，批量创建台账记录，授课部门记为所选部门.

    支持各部门不同表头格式（别名映射），自动跳过空行/月份分隔行。
    多工作表文件默认只导入第一个可识别的工作表，可通过 sheet_name 指定
    或 all_sheets=true 合并导入全部可识别工作表。
    """
    from openpyxl import load_workbook

    _require_user(current_user)
    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".xlsx"},
        what="培训统计表",
    )
    wb = load_workbook(BytesIO(content), data_only=True)

    # 扫描所有工作表，找出可识别表头的
    recognized: list[tuple[str, int, dict[int, str]]] = []
    for sname in wb.sheetnames:
        hr, cm = _read_excel_header_map(wb[sname])
        if cm:
            recognized.append((sname, hr, cm))

    if not recognized:
        # 诊断辅助：返回第一个工作表前 5 行实际出现的文本，便于定位新格式
        ws0 = wb[wb.sheetnames[0]]
        detected: list[str] = []
        for row in ws0.iter_rows(min_row=1, max_row=5, values_only=True):
            for c in row:
                text = str(c).strip() if c is not None else ""
                if text and len(text) <= 20 and text not in detected:
                    detected.append(text)
        preview = "、".join(detected[:15]) or "（空）"
        raise AppException(
            status_code=400,
            message=f"未识别到表头。文件前几行内容为：{preview}。请联系管理员将该表头加入导入别名映射",
        )

    if all_sheets:
        targets = recognized
    elif sheet_name:
        targets = [r for r in recognized if r[0] == sheet_name]
        if not targets:
            available = "、".join(r[0] for r in recognized)
            raise AppException(
                status_code=400,
                message=f"工作表「{sheet_name}」无法识别表头。可导入的工作表：{available}",
            )
    else:
        targets = recognized[:1]

    created = 0
    per_sheet: list[str] = []
    for sname, header_row, col_map in targets:
        sheet_created = await _import_rows_with_mapping(
            wb[sname], header_row, col_map, department, service
        )
        created += sheet_created
        per_sheet.append(f"[{sname}]{sheet_created}条")

    message = f"成功导入{created}条台账记录到{department}：{'、'.join(per_sheet)}"
    if not all_sheets and len(recognized) > len(targets):
        others = "、".join(r[0] for r in recognized if r not in targets)
        message += f"。该文件还有其他可识别工作表：{others}，如需导入请指定工作表名"
    return success_response(data={"created": created}, message=message)


def _cell_text(v: Any) -> str | None:
    """Excel 单元格转文本：数字/日期等非字符串值统一转 str，空白返回 None.

    部门统计表常见数字单元格（如级别=1、备注=数字），直接传给 str 字段
    会触发 Pydantic string_type 校验错误。
    """
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _clip(s: str | None, n: int) -> str | None:
    """截断到字段最大长度，避免脏数据触发 string_too_long 校验错误."""
    return s[:n] if s else s


async def _import_rows_with_mapping(
    ws: Any,
    header_row: int,
    col_map: dict[int, str],
    department: str,
    service: TrainingLedgerService,
) -> int:
    """按列映射导入工作表数据行，返回创建条数.

    AI 只做列映射，导入的每个值都来自 Excel 单元格原文。
    """
    created = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = {col_map[ci]: (row[ci] if ci < len(row) else None) for ci in col_map}
        content_text = _cell_text(vals.get("training_content")) or ""
        training_date = _parse_excel_date(
            vals.get("training_date") or vals.get("training_datetime")
        )
        if not content_text and training_date is None:
            continue  # 空行/月份分隔行跳过
        # 课时兼容数字与文本（如 "1"、"3.5"）
        duration_raw = vals.get("duration_hours")
        duration_val = None
        if isinstance(duration_raw, (int, float)):
            duration_val = float(duration_raw)
        elif isinstance(duration_raw, str) and duration_raw.strip():
            try:
                duration_val = float(duration_raw.strip())
            except ValueError:
                duration_val = None
        # 时长缺失时，从培训时间原文的时间段（如 09:00~10:00）自动计算
        if duration_val is None:
            duration_val = _calc_duration_from_text(
                vals.get("training_datetime") or vals.get("training_date")
            )
        data = TrainingLedgerCreate(
            employee_number="",
            training_date=training_date or date.today(),
            training_subject=_clip(content_text or "部门培训", 256),
            duration_hours=duration_val,
            training_datetime=_clip(
                (_cell_text(vals.get("training_datetime")) or "")
                .replace("\n", " ")
                .strip()
                or None,
                64,
            ),
            training_content=_clip(content_text or None, 512),
            teaching_dept=_clip(department, 128),
            instructor=_clip(_cell_text(vals.get("instructor")), 128),
            level_category=_clip(_cell_text(vals.get("level_category")), 16),
            involved_depts=_cell_text(vals.get("involved_depts")),
            trainees=_cell_text(vals.get("trainees")),
            training_type=_clip(_cell_text(vals.get("training_type")), 32),
            ledger_assessment_method=_clip(
                _cell_text(vals.get("ledger_assessment_method")), 32
            ),
            plan_source=_clip(_cell_text(vals.get("plan_source")), 32),
            drug_category=_clip(_cell_text(vals.get("drug_category")), 32),
            score_summary=_cell_text(vals.get("score_summary")),
            remarks=_clip(_cell_text(vals.get("remarks")), 512),
            source_type="manual",
        )
        await service.create_record(data)
        created += 1
    return created


def _count_data_rows(
    ws: Any, header_row: int, col_map: dict[int, str]
) -> tuple[int, list[list[str]]]:
    """统计可导入行数并取前 3 行样例（预览用）."""
    count = 0
    samples: list[list[str]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = {col_map[ci]: (row[ci] if ci < len(row) else None) for ci in col_map}
        content_text = (
            (vals.get("training_content") or "").strip()
            if isinstance(vals.get("training_content"), str)
            else ""
        )
        training_date = _parse_excel_date(
            vals.get("training_date") or vals.get("training_datetime")
        )
        if not content_text and training_date is None:
            continue
        count += 1
        if len(samples) < 3:
            samples.append(
                [
                    str(row[ci])[:100].replace("\n", " ")
                    if ci < len(row) and row[ci] is not None
                    else ""
                    for ci in sorted(col_map)
                ]
            )
    return count, samples


# ── AI 智能导入（预览分析 + 确认导入 + 格式记忆）──


def _get_import_mapping_repo(session: AsyncSession) -> Any:
    from app.modules.hr.repository import TrainingImportMappingRepository

    return TrainingImportMappingRepository(session)


@router.post(
    "/training-ledgers/import-preview",
    summary="AI识别导入预览",
    response_model=ImportPreviewResponse,
)
async def preview_training_import(
    file: UploadFile = File(..., description="Excel文件(.xlsx)"),
    department: str = Query(..., description="导入到哪个部门"),
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    "分析 Excel 所有工作表：记忆表 → 规则别名 → AI 兜底，返回列映射与样例供人工确认."
    from openpyxl import load_workbook

    from app.modules.hr.training_import_ai import (
        analyze_headers_by_llm,
        field_catalog_payload,
        header_fingerprint,
    )

    _require_user(current_user)
    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".xlsx"},
        what="培训导入预览文件",
    )
    wb = load_workbook(BytesIO(content), data_only=True)
    mapping_repo = _get_import_mapping_repo(session)

    sheets_payload: list[ImportSheetPreview] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        preview = ImportSheetPreview(name=sname)

        # ① 严格定位表头
        header_row, headers = _locate_header_row(ws)
        mapping: dict[str, str] = {}
        source = "none"
        judgment: str | None = None

        if header_row:
            fingerprint = header_fingerprint(headers)
            # 记忆表优先
            memory = await mapping_repo.get_by_dept_fingerprint(department, fingerprint)
            if memory:
                import json as _json

                try:
                    mapping = _json.loads(memory.mapping_json)
                except ValueError:
                    mapping = {}
                if mapping:
                    source = "memory"
                    memory.used_count += 1
                    await mapping_repo.update(memory)
            if not mapping:
                # 规则别名表
                rule_mapping = _map_headers_by_alias(headers)
                if len(rule_mapping) >= 4:
                    mapping = rule_mapping
                    source = "rule"
        else:
            # 严格条件未满足：找候选表头行供 AI 分析
            cand_row, cand_headers = _find_candidate_header_row(ws)
            if cand_row:
                header_row, headers = cand_row, cand_headers

        # ② AI 兜底（规则未命中时）
        if source == "none" and headers:
            ai_result = await analyze_headers_by_llm(sname, headers)
            if ai_result["mapping"]:
                mapping = ai_result["mapping"]
                source = "ai"
            judgment = ai_result["judgment"] or None

        if header_row and mapping:
            col_map = {int(k): v for k, v in mapping.items()}
            data_count, samples = _count_data_rows(ws, header_row, col_map)
            preview.header_row = header_row
            preview.mapping = mapping
            preview.data_row_count = data_count
            preview.sample_rows = samples

        preview.source = source
        preview.headers = [h for h in headers][:30]
        preview.ai_judgment = judgment
        sheets_payload.append(preview)

    data = ImportPreviewData(
        sheets=sheets_payload,
        field_catalog=field_catalog_payload(),
    )
    return success_response(data=data.model_dump(mode="json"))


@router.post(
    "/training-ledgers/import-confirm",
    summary="确认导入并记住格式",
    response_model=ImportConfirmResponse,
)
async def confirm_training_import(
    file: UploadFile = File(..., description="Excel文件(.xlsx)"),
    department: str = Form(..., description="导入到哪个部门"),
    sheets: str = Form(..., description="JSON: [{name, header_row, mapping}]"),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """按用户确认的列映射执行导入，并把映射存入记忆表供下次复用."""
    import json as _json

    from openpyxl import load_workbook

    from app.modules.hr.models import TrainingImportMapping
    from app.modules.hr.training_import_ai import header_fingerprint

    _require_user(current_user)
    try:
        payload = ImportConfirmRequest(
            department=department, sheets=_json.loads(sheets)
        )
    except ValueError:
        raise AppException(status_code=400, message="sheets 参数格式错误")

    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".xlsx"},
        what="培训导入确认文件",
    )
    wb = load_workbook(BytesIO(content), data_only=True)
    mapping_repo = _get_import_mapping_repo(session)

    created = 0
    per_sheet: list[str] = []
    for sheet_cfg in payload.sheets:
        if sheet_cfg.name not in wb.sheetnames:
            raise AppException(
                status_code=400, message=f"工作表「{sheet_cfg.name}」不存在"
            )
        ws = wb[sheet_cfg.name]
        col_map = {int(k): v for k, v in sheet_cfg.mapping.items() if v}
        if not col_map:
            continue
        sheet_created = await _import_rows_with_mapping(
            ws, sheet_cfg.header_row, col_map, department, service
        )
        created += sheet_created
        per_sheet.append(f"[{sheet_cfg.name}]{sheet_created}条")

        # 存入记忆表（表头指纹 → 映射）
        header_cells = [
            str(c.value).strip() if c.value is not None else ""
            for c in ws[sheet_cfg.header_row]
        ]
        fingerprint = header_fingerprint(header_cells)
        existing = await mapping_repo.get_by_dept_fingerprint(department, fingerprint)
        if existing:
            existing.mapping_json = _json.dumps(sheet_cfg.mapping, ensure_ascii=False)
            existing.header_row = sheet_cfg.header_row
            existing.confirmed_at = datetime.now(UTC)
            existing.used_count += 1
            await mapping_repo.update(existing)
        else:
            await mapping_repo.create(
                TrainingImportMapping(
                    department=department,
                    header_fingerprint=fingerprint,
                    header_row=sheet_cfg.header_row,
                    mapping_json=_json.dumps(sheet_cfg.mapping, ensure_ascii=False),
                    confirmed_at=datetime.now(UTC),
                    used_count=1,
                )
            )

    if created == 0:
        raise AppException(
            status_code=400, message="未导入任何数据，请检查列映射与数据行"
        )
    return success_response(
        data={
            "created": created,
            "echo_sheets": [s.model_dump() for s in payload.sheets],
        },
        message=f"成功导入{created}条台账记录到{department}：{'、'.join(per_sheet)}。格式已记住，下次同格式文件将自动识别",
    )


@router.delete("/training-ledgers/by-dept", summary="清空部门全部培训台账（软删除）")
async def clear_training_ledgers_by_dept(
    department: str = Query(..., description="部门"),
    db: AsyncSession = Depends(get_db),
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    deleted = await service.delete_by_department(department)
    if deleted == 0:
        raise AppException(
            status_code=400, message=f"{department} 没有可清除的台账记录"
        )

    # 台账清空后同步复位"已培训"标记，允许在培训资料中重新勾选附件
    used_result = await db.execute(
        select(TrainingContentUsed).where(TrainingContentUsed.is_deleted.is_(False))
    )
    now = datetime.now(UTC)
    reset = 0
    for row in used_result.scalars().all():
        row.is_deleted = True
        row.updated_at = now
        reset += 1
    await db.flush()

    return success_response(
        data={"deleted": deleted, "reset_content_used": reset},
        message=f"已清空{department}的{deleted}条培训台账记录",
    )


# ── 笔试成绩导入 ──────────────────────────────────────────


@router.post(
    "/training-ledgers/import-exam-scores",
    summary="解析笔试成绩文件（.docx/.xlsx）",
    response_model=ExamScoreImportResponse,
)
async def import_exam_scores(
    file: UploadFile = File(..., description="成绩文件（.docx 或 .xlsx）"),
    record_id: UUID = Form(..., description="目标台账记录 ID"),
) -> Any:
    from app.modules.hr.exam_score_parser import parse_exam_scores

    try:
        scores = await parse_exam_scores(file)
    except ValueError as e:
        raise AppException(status_code=400, message=str(e))

    if not scores:
        raise AppException(
            status_code=400, message="未从文件中解析出任何成绩数据，请检查文件格式"
        )

    return ExamScoreImportResponse(
        code=0,
        message=f"解析出 {len(scores)} 条成绩",
        data=[ExamScoreItem(name=s.name, score=s.score) for s in scores],
    )


@router.post(
    "/training-ledgers/confirm-exam-scores",
    summary="确认导入笔试成绩并跨部门同步",
    response_model=ExamScoreConfirmResponse,
)
async def confirm_exam_scores(
    payload: ExamScoreConfirmRequest,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
) -> Any:
    from app.modules.hr.exam_score_parser import ExamScore, format_score_summary

    record = await service.get_record(payload.record_id)

    # 格式化成绩汇总
    score_items = [ExamScore(name=s.name, score=s.score) for s in payload.scores]
    summary_text = format_score_summary(score_items)

    # 更新当前记录的 score_summary
    record.score_summary = summary_text
    # 如果考核方式为空，自动设为"笔试"
    if not record.ledger_assessment_method:
        record.ledger_assessment_method = "笔试"
    await service.repo.update(record)

    # 通过 session_id 同步到其他部门记录
    synced = 0
    if record.session_id:
        sync_data = {"score_summary": summary_text}
        if not record.ledger_assessment_method:
            sync_data["ledger_assessment_method"] = "笔试"
        synced = await service.repo.sync_by_session_id(
            session_id=record.session_id,
            exclude_id=record.id,
            update_data=sync_data,
        )

    return ExamScoreConfirmResponse(
        code=0,
        message=f"已导入 {len(score_items)} 条成绩，同步更新 {synced} 条部门记录",
        data={"synced_count": synced, "score_summary": summary_text},
    )


@router.get("/training-ledgers/{record_id}", summary="培训台账记录详情")
async def get_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
    )


@router.put("/training-ledgers/{record_id}", summary="更新培训台账记录")
async def update_training_ledger(
    record_id: UUID,
    payload: TrainingLedgerUpdate,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=TrainingLedgerResponse.model_validate(record).model_dump(mode="json"),
        message="培训台账记录更新成功",
    )


@router.delete("/training-ledgers/{record_id}", summary="删除培训台账记录")
async def delete_training_ledger(
    record_id: UUID,
    service: TrainingLedgerService = Depends(get_training_ledger_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_record(record_id)
    return success_response(message="培训台账记录删除成功")


# ─── AnnualTrainingPlan Routes ───


@router.get("/annual-training-plans", summary="年度培训计划列表")
async def list_annual_training_plans(
    year: int | None = Query(None, description="年度筛选"),
    department: str | None = Query(None, description="部门筛选"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    current_user: CurrentUser = None,
) -> Any:
    alias_set = await _assert_dept_in_scope(db, current_user, department)
    plans, total = await service.list_plans(
        year=year,
        department=department,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=alias_set,
    )
    data = [
        AnnualTrainingPlanResponse.model_validate(p).model_dump(mode="json")
        for p in plans
    ]
    return paginated_response(
        data=data,
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/annual-training-plans", summary="创建年度培训计划")
async def create_annual_training_plan(
    payload: AnnualTrainingPlanCreate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    plan = await service.create_plan(payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划创建成功",
    )


@router.get("/annual-training-plans/{plan_id}", summary="年度培训计划详情")
async def get_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    plan = await service.get_plan(plan_id)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
    )


@router.put("/annual-training-plans/{plan_id}", summary="更新年度培训计划")
async def update_annual_training_plan(
    plan_id: UUID,
    payload: AnnualTrainingPlanUpdate,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    plan = await service.update_plan(plan_id, payload)
    return success_response(
        data=AnnualTrainingPlanResponse.model_validate(plan).model_dump(mode="json"),
        message="年度培训计划更新成功",
    )


@router.delete("/annual-training-plans/{plan_id}", summary="删除年度培训计划")
async def delete_annual_training_plan(
    plan_id: UUID,
    service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_plan(plan_id)
    return success_response(message="年度培训计划删除成功")


@router.get("/annual-training-plans/{plan_id}/items", summary="年度计划明细列表")
async def list_annual_training_plan_items(
    plan_id: UUID,
    service: AnnualTrainingPlanItemService = Depends(
        get_annual_training_plan_item_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    items = await service.list_items(plan_id)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(data=data)


@router.put(
    "/annual-training-plans/{plan_id}/items/batch", summary="批量更新年度计划明细"
)
async def batch_update_annual_training_plan_items(
    plan_id: UUID,
    payload: AnnualTrainingPlanItemBatchUpdate,
    service: AnnualTrainingPlanItemService = Depends(
        get_annual_training_plan_item_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    items = await service.batch_update_items(plan_id, payload)
    data = [
        AnnualTrainingPlanItemResponse.model_validate(i).model_dump(mode="json")
        for i in items
    ]
    return success_response(
        data=data,
        message="年度计划明细更新成功",
    )


def _generate_annual_plan_excel(
    plan: dict[str, Any], items: list[dict[str, Any]]
) -> BytesIO:
    """Generate annual training plan Excel based on 7.7 template format."""
    items = list(items)
    wb = Workbook()
    ws = wb.active
    ws.title = "年度培训计划"

    # Styles
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{plan['year']} 年培训计划"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Department row
    ws.merge_cells("A2:I2")
    ws["A2"] = f"部门：{plan['department']}"
    ws["A2"].font = bold_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 22

    # Header row
    headers = [
        "序号",
        "培训季度及课时",
        "培训内容及使用教材",
        "培训对象",
        "授课单位及授课人",
        "考核方式",
        "培训跟踪",
        "确认人/日期",
        "备注",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 28

    # Data rows
    for idx, item in enumerate(items, 1):
        row = 3 + idx
        quarter = item.get("month") or ""
        hours = item.get("duration_hours")
        quarter_hours = f"{quarter}\n{hours}课时" if hours else quarter
        confirmer = item.get("confirmer") or ""
        confirm_date = (
            f" / {fmt_date_str(str(item.get('confirm_date')))}"
            if item.get("confirm_date")
            else ""
        )

        values = [
            idx,
            quarter_hours,
            item.get("content_and_textbook") or "",
            item.get("target_audience") or "",
            item.get("position_and_count") or "",
            item.get("training_method") or "",
            item.get("tracking_status") or "",
            f"{confirmer}{confirm_date}",
            item.get("remarks") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 2, 6, 7) else left_align
        ws.row_dimensions[row].height = 36

    # Pad to at least 12 rows
    while len(items) < 12:
        row = 3 + len(items) + 1
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col, value="")
            cell.border = thin_border
        ws.row_dimensions[row].height = 36
        items.append({})

    # Footer row
    footer_row = 4 + len(items) + 1
    ws.merge_cells(f"A{footer_row}:E{footer_row}")
    ws.cell(row=footer_row, column=1, value="制表人/日期：")
    ws.cell(row=footer_row, column=1).alignment = left_align
    ws.cell(row=footer_row, column=1).border = thin_border
    for c in range(2, 6):
        ws.cell(row=footer_row, column=c).border = thin_border

    ws.merge_cells(f"F{footer_row}:I{footer_row}")
    ws.cell(row=footer_row, column=6, value="部门负责人/日期：")
    ws.cell(row=footer_row, column=6).alignment = left_align
    ws.cell(row=footer_row, column=6).border = thin_border
    for c in range(7, 10):
        ws.cell(row=footer_row, column=c).border = thin_border

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get(
    "/annual-training-plans/{plan_id}/export", summary="导出年度培训计划Word文档"
)
async def export_annual_training_plan(
    plan_id: UUID,
    plan_service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    item_service: AnnualTrainingPlanItemService = Depends(
        get_annual_training_plan_item_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    """根据年度计划数据生成并导出 Word 文档（APP1/APP2模板格式）。

    公司级计划使用 APP2 模板，部门级计划使用 APP1 模板。
    """
    _require_user(current_user)
    plan = await plan_service.get_plan(plan_id)
    items = await item_service.list_items(plan_id)

    from app.modules.hr.annual_plan_document_generator import generate_annual_plan_doc

    buffer = generate_annual_plan_doc(plan, items)
    buffer.seek(0)

    safe_dept = plan.department.replace(" ", "_")
    level = plan.plan_level or "公司级"
    if level == "公司级":
        filename = f"APP2-SMP-HR-002-14年度公司培训计划表_{plan.year}_{safe_dept}.docx"
    else:
        filename = f"APP1-SMP-HR-002-14年度部门培训计划表_{plan.year}_{safe_dept}.docx"

    encoded_filename = quote(filename, safe="")

    return StreamingResponse(
        iter([buffer.read()]),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename*=utf-8''{encoded_filename}"
        },
    )


@router.post("/annual-training-plans/import", summary="导入年度培训计划Word文档")
async def import_annual_training_plan(
    file: UploadFile = File(..., description="Word文档(.docx)"),
    year: int = Query(..., description="年度"),
    plan_level: str = Query(
        None, description="计划级别: 公司级, 部门级（留空时从文档自动识别）"
    ),
    department: str = Query(None, description="部门（留空时从文档自动识别）"),
    plan_service: AnnualTrainingPlanService = Depends(get_annual_training_plan_service),
    item_service: AnnualTrainingPlanItemService = Depends(
        get_annual_training_plan_item_service
    ),
    current_user: CurrentUser = None,
) -> Any:
    """从 Word 文档导入年度培训计划明细。

    支持 APP1/APP2 格式的 .docx 文档：
    - 自动识别表格中的序号、培训类型、培训时间、培训内容、培训对象、授课单位、考核方式
    - 自动从文档标题识别计划级别（公司级/部门级），从"部门：XXX"段落识别部门
    - 跳过审批栏行（制表人、签名/日期等）
    - 自动创建或复用已有计划，全量替换明细
    """
    _require_user(current_user)

    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".docx"},
        what="年度培训计划",
    )
    result = await plan_service.import_from_docx(
        content,
        year=year,
        plan_level=plan_level,
        department=department,
        item_service=item_service,
    )

    return success_response(
        data=result,
        message=f"成功导入{result['imported_count']}条明细",
    )


@router.get(
    "/annual-training-plans/{plan_id}/attachments",
    summary="计划附件清单",
    responses={200: {"model": PlanAttachmentListEnvelope}},
)
async def list_plan_attachments(
    plan_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    attachments = await service.list_by_plan(plan_id)
    data = [
        PlanAttachmentResponse.model_validate(a).model_dump(mode="json")
        for a in attachments
    ]
    return success_response(data=data)


@router.post("/annual-training-plans/{plan_id}/attachments", summary="批量上传计划附件")
async def upload_plan_attachments(
    plan_id: UUID,
    files: list[UploadFile] = File(..., description="附件文件（可多选）"),
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    """批量上传附件；文件名含"附件X"时自动归一化编号，否则编号为空可后续展示."""
    _require_user(current_user)
    if len(files) > MAX_UPLOAD_FILES:
        raise AppException(
            status_code=400, message=f"单次最多上传 {MAX_UPLOAD_FILES} 个附件"
        )
    created = []
    for f in files:
        file_name, data = await read_upload_secure(
            f,
            max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
            allowed_extensions={
                ".doc",
                ".docx",
                ".pdf",
                ".xls",
                ".xlsx",
                ".png",
                ".jpg",
                ".jpeg",
            },
            what="培训计划附件",
        )
        attachment = await service.upload(
            plan_id=plan_id, file_name=file_name, data=data
        )
        created.append(
            PlanAttachmentResponse.model_validate(attachment).model_dump(mode="json")
        )
    return success_response(data=created, message=f"成功上传{len(created)}个附件")


@router.get(
    "/annual-training-plan-attachments/{attachment_id}/download", summary="下载计划附件"
)
async def download_plan_attachment(
    attachment_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    attachment = await service.get(attachment_id)
    encoded = quote(attachment.file_name, safe="")
    return StreamingResponse(
        iter([service.read_data(attachment)]),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded}"},
    )


@router.delete(
    "/annual-training-plan-attachments/{attachment_id}", summary="删除计划附件"
)
async def delete_plan_attachment(
    attachment_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete(attachment_id)
    return success_response(message="附件删除成功")


@router.post(
    "/annual-training-plan-attachments/mark-ledger-imported",
    summary="标记计划附件已导入培训台账（置灰不可再选）",
)
async def mark_attachments_ledger_imported(
    body: MarkLedgerImportedRequest,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    count = await service.mark_ledger_imported(body.ids)
    return success_response(message=f"已标记 {count} 个附件为已导入台账")


@router.get("/training-content-used", summary="已培训附件文件清单条目列表")
async def list_training_content_used(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await db.execute(
        select(TrainingContentUsed).where(TrainingContentUsed.is_deleted.is_(False))
    )
    data = [
        TrainingContentUsedOut.model_validate(r).model_dump(mode="json")
        for r in result.scalars().all()
    ]
    return success_response(data=data)


@router.post(
    "/training-content-used", summary="标记附件文件清单条目已培训（置灰不可再选）"
)
async def mark_training_content_used(
    body: MarkTrainingContentUsedRequest,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    assert current_user is not None
    now = datetime.now(UTC)
    count = 0
    for item in body.items:
        name = re.sub(r"\s+", "", item.name or "")
        if not name:
            continue
        existing = (
            await db.execute(
                select(TrainingContentUsed).where(
                    TrainingContentUsed.entry_name == name,
                    TrainingContentUsed.is_deleted.is_(False),
                )
            )
        ).scalar_one_or_none()
        if existing:
            continue
        db.add(
            TrainingContentUsed(
                entry_name=name,
                entry_code=item.code,
                attachment_id=item.attachment_id,
                used_at=now,
                created_by=current_user.id,
                updated_by=current_user.id,
            )
        )
        count += 1
    await db.flush()
    return success_response(message=f"已标记 {count} 个文件条目为已培训")


@router.post("/training-sessions/upsert", summary="保存/更新培训会话")
async def upsert_training_session(
    body: TrainingSessionUpsert,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    assert current_user is not None
    sess = None
    if body.id:
        sess = await db.get(TrainingSession, body.id)
    data = body.model_dump(exclude={"id"}, exclude_unset=True)
    if sess is None:
        sess = TrainingSession(
            id=body.id or uuid4(),
            created_by=current_user.id,
            updated_by=current_user.id,
            **data,
        )
        db.add(sess)
    else:
        for key, value in data.items():
            setattr(sess, key, value)
        sess.updated_by = current_user.id
    await db.flush()
    return success_response(data={"id": str(sess.id)})


@router.get("/training-sessions/{session_id}", summary="培训会话详情")
async def get_training_session(
    session_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    sess = await db.get(TrainingSession, session_id)
    if sess is None or sess.is_deleted:
        raise NotFoundException("培训会话不存在")
    return success_response(
        data=TrainingSessionOut.model_validate(sess).model_dump(mode="json")
    )


@router.post("/training-documents/upsert", summary="保存会话资料（同会话同类覆盖更新）")
async def upsert_training_document(
    body: TrainingDocumentUpsert,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    assert current_user is not None
    existing = (
        await db.execute(
            select(TrainingDocument).where(
                TrainingDocument.session_id == body.session_id,
                TrainingDocument.doc_type == body.doc_type,
                TrainingDocument.is_deleted.is_(False),
            )
        )
    ).scalar_one_or_none()
    if existing:
        existing.payload = body.payload
        existing.title = body.title
        existing.updated_by = current_user.id
        doc = existing
    else:
        doc = TrainingDocument(
            session_id=body.session_id,
            doc_type=body.doc_type,
            title=body.title,
            payload=body.payload,
            created_by=current_user.id,
            updated_by=current_user.id,
        )
        db.add(doc)
    await db.flush()
    return success_response(data={"id": str(doc.id)})


@router.get("/training-sessions/{session_id}/documents", summary="会话已有资料列表")
async def list_session_documents(
    session_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    rows = (
        (
            await db.execute(
                select(TrainingDocument)
                .where(
                    TrainingDocument.session_id == session_id,
                    TrainingDocument.is_deleted.is_(False),
                )
                .order_by(TrainingDocument.doc_type)
            )
        )
        .scalars()
        .all()
    )
    data = [TrainingDocumentOut.model_validate(r).model_dump(mode="json") for r in rows]
    return success_response(data=data)


@router.get("/training-documents/{doc_id}", summary="资料详情（payload 恢复编辑）")
async def get_training_document(
    doc_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    doc = await db.get(TrainingDocument, doc_id)
    if doc is None or doc.is_deleted:
        raise NotFoundException("培训资料不存在")
    return success_response(
        data=TrainingDocumentOut.model_validate(doc).model_dump(mode="json")
    )


@router.post("/training-oral-exam/export", summary="生成口试培训考核结果表（APP10）")
async def export_oral_exam(
    payload: OralExamExportRequest,
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    try:
        buffer = generate_oral_exam_result(payload)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    safe = (payload.training_date or "").replace("-", "") or "nodate"
    filename = quote(f"口试培训考核结果表_{safe}.docx")
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post(
    "/training-practical-exam/export",
    summary="生成实操培训考核结果表（APP13，每人一份，打包zip）",
)
async def export_practical_exam(
    payload: PracticalExamExportRequest,
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    try:
        buffer = generate_practical_exam_zip(payload)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    from app.modules.hr.practical_exam_document_generator import _safe_filename

    safe_content = _safe_filename(payload.training_content or "")
    safe_date = payload.training_date or "nodate"
    filename = quote(f"{safe_date}-{safe_content}-实操.zip")
    return StreamingResponse(
        _iterfile(),
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/training-attachment", summary="生成培训附件")
async def export_training_attachment(
    payload: TrainingAttachmentExportRequest,
    current_user: CurrentUser = None,
) -> Any:
    "根据填写的附件文件清单自动生成培训附件 Word（附件： +"
    " 序号/文件名称/文件编号表格）。"
    _require_user(current_user)
    try:
        buffer = generate_training_attachment(payload.items)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))

    def _iterfile() -> Any:
        buffer.seek(0)
        yield buffer.read()

    filename = quote("培训附件.docx")
    return StreamingResponse(
        _iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post(
    "/training-practical-exam/import",
    summary="导入实操试题（APP13格式docx，提取描述与培训日期）",
)
async def import_practical_exam_questions(
    file: UploadFile = File(..., description="APP13格式的实操试题docx"),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".docx", ".doc"},
        what="实操试题",
    )
    try:
        data = parse_practical_exam_questions(content)
    except FileNotFoundError as e:
        raise AppException(status_code=400, message=str(e))
    except Exception as e:
        raise AppException(status_code=400, message=f"解析实操试题失败：{e}")
    # 保存导入的 docx 作为导出基底（保留试题段落结构/方框，导出时只填部门/姓名/日期）
    from app.modules.hr.practical_exam_document_generator import _imported_template_path

    try:
        _imported_template_path().write_bytes(content)
    except Exception:
        logger.exception("保存实操试题导入文件失败")
    return success_response(data=data, message="导入成功")


@router.get(
    "/annual-training-plans/{plan_id}/attachment-sections",
    summary="计划附件条目列表",
    responses={200: {"model": PlanAttachmentSectionListEnvelope}},
)
async def list_plan_attachment_sections(
    plan_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    """返回该计划下所有附件条目（附件X），供明细行引用匹配与跨模块索引."""
    _require_user(current_user)
    sections = await service.list_sections(plan_id)
    data = [
        PlanAttachmentSectionResponse.model_validate(s).model_dump(mode="json")
        for s in sections
    ]
    return success_response(data=data)


@router.get(
    "/plan-attachment-sections/{section_id}/preview",
    summary="附件条目预览",
    responses={200: {"model": AttachmentPreviewEnvelope}},
)
async def preview_plan_attachment_section(
    section_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    preview = await service.preview_section(section_id)
    return success_response(data=preview)


@router.get(
    "/annual-training-plan-attachments/{attachment_id}/preview",
    summary="整文件附件预览",
    responses={200: {"model": AttachmentPreviewEnvelope}},
)
async def preview_plan_attachment(
    attachment_id: UUID,
    service: PlanAttachmentService = Depends(get_plan_attachment_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    preview = await service.preview_attachment(attachment_id)
    return success_response(
        data=AttachmentPreview.model_validate(preview).model_dump(mode="json")
    )


@router.get("/email/config", summary="查询邮箱配置状态")
async def get_email_config(
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """查询邮箱 IMAP/SMTP 配置状态和上次扫描结果。密码不返回。"""
    _require_user(current_user)
    from app.shared.config_reader import get_module_setting

    imap_host = await get_module_setting("hr", "HR_MAIL_IMAP_HOST", "")
    imap_user = await get_module_setting("hr", "HR_MAIL_IMAP_USER", "")
    fetch_enabled = await get_module_setting("hr", "HR_MAIL_FETCH_ENABLED", "false")
    smtp_host = await get_module_setting("hr", "HR_MAIL_SMTP_HOST", "")
    smtp_user = await get_module_setting("hr", "HR_MAIL_SMTP_USER", "")
    from_addr = await get_module_setting("hr", "HR_MAIL_FROM", "")

    # 读取抓取状态
    last_scan_at = await get_module_setting("hr", "HR_MAIL_LAST_SCAN_AT", "")
    last_fetched_count = await get_module_setting(
        "hr", "HR_MAIL_LAST_FETCHED_COUNT", "0"
    )
    last_fetch_status = await get_module_setting("hr", "HR_MAIL_LAST_FETCH_STATUS", "")

    import json

    schedule_hours_raw = await get_module_setting(
        "hr", "HR_MAIL_FETCH_SCHEDULE_HOURS", "[]"
    )
    try:
        fetch_schedule_hours = json.loads(schedule_hours_raw)
    except Exception:
        fetch_schedule_hours = []

    return success_response(
        data={
            "imap_host": imap_host,
            "imap_user": imap_user,
            "imap_port": await get_module_setting("hr", "HR_MAIL_IMAP_PORT", "993"),
            "smtp_host": smtp_host,
            "smtp_user": smtp_user,
            "smtp_port": await get_module_setting("hr", "HR_MAIL_SMTP_PORT", "465"),
            "from_addr": from_addr,
            "fetch_enabled": fetch_enabled.lower() == "true",
            "fetch_interval_hours": int(
                await get_module_setting("hr", "HR_MAIL_FETCH_INTERVAL_HOURS", "1")
            ),
            "fetch_schedule_hours": fetch_schedule_hours,
            "watch_dir": await get_module_setting("hr", "HR_RESUME_WATCH_DIR", ""),
            "offer_subject": await get_module_setting(
                "hr", "HR_MAIL_OFFER_SUBJECT", "录用通知 - {name}"
            ),
            "offer_body": await get_module_setting("hr", "HR_MAIL_OFFER_BODY", ""),
            "reject_subject": await get_module_setting(
                "hr", "HR_MAIL_REJECT_SUBJECT", "面试结果通知 - {name}"
            ),
            "reject_body": await get_module_setting("hr", "HR_MAIL_REJECT_BODY", ""),
            "last_scan_at": last_scan_at or None,
            "last_fetched_count": int(last_fetched_count) if last_fetched_count else 0,
            "last_fetch_status": last_fetch_status or None,
        }
    )


@router.put("/email/config", summary="更新邮箱配置")
async def update_email_config(
    payload: EmailConfigUpdate,
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """更新 IMAP/SMTP 邮箱配置。密码加密存储。"""
    _require_user(current_user)
    from app.core.llm import encrypt_api_key
    from app.shared.config_reader import set_module_setting

    key_map = {
        "imap_host": "HR_MAIL_IMAP_HOST",
        "imap_port": "HR_MAIL_IMAP_PORT",
        "imap_user": "HR_MAIL_IMAP_USER",
        "imap_pass": "HR_MAIL_IMAP_PASS",
        "smtp_host": "HR_MAIL_SMTP_HOST",
        "smtp_port": "HR_MAIL_SMTP_PORT",
        "smtp_user": "HR_MAIL_SMTP_USER",
        "smtp_pass": "HR_MAIL_SMTP_PASS",
        "from_addr": "HR_MAIL_FROM",
        "fetch_enabled": "HR_MAIL_FETCH_ENABLED",
        "fetch_interval_hours": "HR_MAIL_FETCH_INTERVAL_HOURS",
        "fetch_schedule_hours": "HR_MAIL_FETCH_SCHEDULE_HOURS",
        "watch_dir": "HR_RESUME_WATCH_DIR",
        "offer_subject": "HR_MAIL_OFFER_SUBJECT",
        "offer_body": "HR_MAIL_OFFER_BODY",
        "reject_subject": "HR_MAIL_REJECT_SUBJECT",
        "reject_body": "HR_MAIL_REJECT_BODY",
    }

    import json

    payload_dict = payload.model_dump(exclude_none=True)
    for form_key, db_key in key_map.items():
        value = payload_dict.get(form_key)
        if value is None:
            continue
        if db_key in ("HR_MAIL_IMAP_PASS", "HR_MAIL_SMTP_PASS"):
            # 留空时保持原密码不变
            if not value:
                continue
            value = encrypt_api_key(value)
        if db_key == "HR_MAIL_FETCH_ENABLED":
            value = "true" if value else "false"
        if db_key == "HR_MAIL_FETCH_SCHEDULE_HOURS":
            # 列表序列化为 JSON 字符串
            value = json.dumps(value)
        await set_module_setting(session, "hr", db_key, str(value))

    return success_response(message="邮箱配置更新成功")


@router.post("/email/config/test", summary="测试邮箱连接")
async def test_email_config(
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """测试 IMAP 和 SMTP 连接是否正常。"""
    _require_user(current_user)
    import imaplib as _imaplib
    import smtplib as _smtplib

    from app.core.llm import decrypt_api_key
    from app.shared.config_reader import get_module_setting

    results = {"imap": "not_tested", "smtp": "not_tested"}
    # Test IMAP
    try:
        host = await get_module_setting("hr", "HR_MAIL_IMAP_HOST")
        port = int(await get_module_setting("hr", "HR_MAIL_IMAP_PORT", "993"))
        user = await get_module_setting("hr", "HR_MAIL_IMAP_USER")
        pwd = await get_module_setting("hr", "HR_MAIL_IMAP_PASS")
        if host and user and pwd:
            pwd = decrypt_api_key(pwd)
            mail = _imaplib.IMAP4_SSL(host, port, timeout=10)
            mail.login(user, pwd)
            mail.logout()
            results["imap"] = "success"
    except Exception as e:
        results["imap"] = str(e)
    # Test SMTP
    try:
        host = await get_module_setting("hr", "HR_MAIL_SMTP_HOST")
        port = int(await get_module_setting("hr", "HR_MAIL_SMTP_PORT", "465"))
        user = await get_module_setting("hr", "HR_MAIL_SMTP_USER")
        pwd = await get_module_setting("hr", "HR_MAIL_SMTP_PASS")
        if host and user and pwd:
            pwd = decrypt_api_key(pwd)
            server = _smtplib.SMTP_SSL(host, port, timeout=10)
            server.login(user, pwd)
            server.quit()
            results["smtp"] = "success"
    except Exception as e:
        results["smtp"] = str(e)
    return success_response(data=results)


@router.post("/email/send-offer", summary="发送录用通知邮件")
async def send_offer_email(
    payload: SendOfferEmailPayload,
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """发送录用通知邮件给候选人。自动填写Offer模板并作为附件发送。"""
    _require_user(current_user)
    from app.modules.hr.mail_sender import send_email_with_template
    from app.shared.config_reader import get_module_setting

    candidate_id = payload.candidate_id
    to_email = payload.to_email
    subject = payload.subject
    body = payload.body

    if not all([candidate_id, to_email, subject, body]):
        raise AppException(status_code=400, message="缺少必要参数")

    # 获取 PDF 模板路径（仅 Offer 邮件附带，面试通知不带附件）
    from pathlib import Path as _Path

    raw_template = await get_module_setting("hr", "HR_MAIL_OFFER_TEMPLATE_PATH", "")
    template_path = str(_Path(raw_template).resolve()) if raw_template else ""
    is_offer = subject and "面试结果" not in subject

    # Offer 邮件附带模板 PDF（重命名为候选人姓名），面试通知不带附件
    attachment_path: str | None = None
    if is_offer and template_path and _Path(template_path).exists():
        import shutil
        import tempfile

        try:
            from app.modules.hr.recruitment_service import RecruitmentService

            svc = RecruitmentService()
            candidate = await svc.get_candidate(candidate_id)
            cand_name = candidate.get("name", "候选人")
        except Exception:
            cand_name = "候选人"
        tmp = _Path(tempfile.gettempdir()) / f"录用通知书_{cand_name}.pdf"
        shutil.copy2(template_path, str(tmp))
        attachment_path = str(tmp)

    # 发送邮件
    success = await send_email_with_template(
        to_email=to_email,
        subject=subject,
        html_body=body,
        attachment_path=attachment_path,
    )

    if not success:
        raise AppException(status_code=500, message="邮件发送失败")

    return success_response(message="邮件发送成功")


@router.post("/email/upload-offer-template", summary="上传 Offer PDF 模板")
async def upload_offer_template(
    file: UploadFile = File(...),
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """上传录用通知书 PDF 模板文件。"""
    _require_user(current_user)
    from pathlib import Path

    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".pdf"},
        what="Offer 模板",
    )

    # 保存到 templates 目录
    templates_dir = Path("templates") / "hr" / "offer"
    templates_dir.mkdir(parents=True, exist_ok=True)

    file_path = templates_dir / "offer_template.pdf"
    file_path.write_bytes(content)

    # 保存路径到数据库
    from app.shared.config_reader import set_module_setting

    await set_module_setting(
        session, "hr", "HR_MAIL_OFFER_TEMPLATE_PATH", str(file_path)
    )

    return success_response(data={"path": str(file_path)}, message="模板上传成功")


@router.get("/email/offer-template", summary="获取 Offer PDF 模板信息")
async def get_offer_template(
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """获取 Offer PDF 模板文件信息。"""
    _require_user(current_user)
    from app.shared.config_reader import get_module_setting

    template_path = await get_module_setting("hr", "HR_MAIL_OFFER_TEMPLATE_PATH", "")

    if not template_path:
        return success_response(data={"has_template": False})

    file_path = Path(template_path)
    if not file_path.exists():
        return success_response(data={"has_template": False})

    return success_response(
        data={
            "has_template": True,
            "filename": file_path.name,
            "size": file_path.stat().st_size,
        }
    )


@router.post("/email/fetch-now", summary="手动触发邮箱抓取")
async def trigger_mail_fetch(
    payload: dict[str, Any] = {"scan_all": True},
    current_user: CurrentUser = None,
) -> Any:
    """手动触发一次 IMAP 邮箱扫描（异步任务，通过 job_id 轮询状态）。

    默认扫描所有邮件（包括已读），强制重新下载。
    """
    _require_user(current_user)
    from app.modules.hr.mail_fetcher import fetch_resumes_from_mail

    scan_all = payload.get("scan_all", True)
    force_redownload = payload.get("force_redownload", True)

    async def _run_fetch(**kwargs: Any) -> Any:
        return await fetch_resumes_from_mail(**kwargs)

    job_id = await submit_job(
        _run_fetch, scan_all=scan_all, force_redownload=force_redownload
    )
    return success_response(data={"job_id": job_id}, message="邮箱抓取已提交")


@router.post("/email/browse-folder", summary="打开原生文件夹选择对话框")
async def browse_folder(
    current_user: CurrentUser = None,
) -> Any:
    """弹出 Windows 原生文件夹选择对话框，返回用户选择的路径。"""
    _require_user(current_user)
    import subprocess
    import sys
    from pathlib import Path

    # 用 Python tkinter 弹窗（同步阻塞，但通过 subprocess 隔离）
    py_script = r"""
import tkinter as tk
from tkinter import filedialog
root = tk.Tk()
root.withdraw()
path = filedialog.askdirectory(title="选择简历下载文件夹")
if path:
    print(path)
"""

    try:
        # 使用当前 Python 解释器（确保 tkinter 可用）
        result = subprocess.run(
            [sys.executable, "-c", py_script],
            capture_output=True,
            text=True,
            timeout=60,
        )
        path = result.stdout.strip()
        if path:
            # 路径安全校验：只允许本地绝对路径
            path_obj = Path(path)
            if not path_obj.is_absolute():
                raise AppException(status_code=403, message="只能选择本地绝对路径")
            return success_response(data={"path": path})
        return success_response(data={"path": None})
    except subprocess.TimeoutExpired:
        raise AppException(status_code=408, message="对话框超时（60秒）")
    except AppException:
        raise
    except Exception as e:
        raise AppException(status_code=500, message=f"打开文件夹对话框失败：{str(e)}")


# ─── 招聘管理（Recruitment）Routes ───
# 数据通过飞书多维表格操作，Repository
# 层对未配置场景做了兜底（读操作返回空、写操作抛 RuntimeError）


def get_recruitment_service() -> RecruitmentService:
    return RecruitmentService()


def get_recruitment_onboarding_service() -> RecruitmentOnboardingService:
    return RecruitmentOnboardingService()


def _handle_write_unconfigured(func: Any) -> Any:
    """装饰器：将 Repository 层的 RecruitmentNotConfigured 转为 503 响应。"""
    from functools import wraps

    @wraps(func)
    async def wrapper(*args: Any, **kwargs: Any) -> Any:
        try:
            return await func(*args, **kwargs)
        except RecruitmentNotConfigured as e:
            logger.warning("Recruitment write operation blocked: %s", e.message)
            raise

    return wrapper


# ─── Job Posting Routes ───


@router.get("/jobs", summary="职位列表")
async def list_jobs(
    keyword: str | None = Query(None, description="关键词搜索"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(100, ge=1, le=500, description="每页条数"),
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取招聘职位列表。飞书多维表格未配置时返回空列表。"""
    _require_user(current_user)
    items, total = await service.list_jobs(
        keyword=keyword, page=page, page_size=page_size
    )
    return paginated_response(
        data=[JobPostingResponse(**item).model_dump(mode="json") for item in items],
        page=page,
        page_size=page_size,
        total=total,
    )


@router.post("/jobs", summary="创建职位")
@_handle_write_unconfigured
async def create_job(
    payload: JobPostingCreate,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """创建招聘职位，同步到飞书多维表格。"""
    _require_user(current_user)
    result = await service.create_job(payload.model_dump(exclude_none=True))
    return success_response(
        data=JobPostingResponse(**result).model_dump(mode="json"),
        message="职位创建成功",
        status_code=201,
    )


@router.get("/jobs/{record_id}", summary="职位详情")
@_handle_write_unconfigured
async def get_job(
    record_id: str,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取招聘职位详情。"""
    _require_user(current_user)
    result = await service.get_job(record_id)
    return success_response(
        data=JobPostingResponse(**result).model_dump(mode="json"),
    )


@router.put("/jobs/{record_id}", summary="更新职位")
@_handle_write_unconfigured
async def update_job(
    record_id: str,
    payload: JobPostingUpdate,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """更新招聘职位信息，同步到飞书多维表格。"""
    _require_user(current_user)
    result = await service.update_job(record_id, payload.model_dump(exclude_none=True))
    return success_response(
        data=JobPostingResponse(**result).model_dump(mode="json"),
        message="职位更新成功",
    )


# ─── Candidate Routes ───


@router.get("/candidates", summary="候选人列表")
async def list_candidates(
    keyword: str | None = Query(None, description="关键词搜索"),
    fit_level: str | None = Query(None, description="符合程度筛选"),
    interview_status: str | None = Query(None, description="面试状态筛选"),
    job_id: str | None = Query(None, description="应聘职位筛选"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=500, description="每页条数"),
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取候选人列表。飞书多维表格未配置时返回空列表。"""
    _require_user(current_user)
    items, total = await service.list_candidates(
        keyword=keyword,
        fit_level=fit_level,
        interview_status=interview_status,
        job_id=job_id,
        page=page,
        page_size=page_size,
    )
    return paginated_response(
        data=[CandidateResponse(**item).model_dump(mode="json") for item in items],
        page=page,
        page_size=page_size,
        total=total,
    )


@router.get("/candidates/sync-from-feishu", summary="从飞书同步候选人")
async def sync_candidates_from_feishu(
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """从飞书多维表格同步候选人数据。"""
    _require_user(current_user)
    items, total = await service.list_candidates(page=1, page_size=500)
    return success_response(
        data={"total": total, "items": items},
        message="同步成功",
    )


@router.post("/candidates/ai-analyze-batch", summary="AI批量分析简历")
@_handle_write_unconfigured
async def batch_ai_analyze_candidates(
    payload: dict[str, Any],
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """对指定候选人进行 AI 简历批量分析（支持多维度评分）。"""
    _require_user(current_user)
    candidate_ids = payload.get("candidate_ids", [])
    job_id = payload.get("job_id")
    result = await service.batch_analyze(candidate_ids=candidate_ids, job_id=job_id)
    return success_response(
        data=result,
        message=(
            f"分析完成：共{result.get('total', 0)}人，成功{result.get('success', 0)}人"
        ),
    )


@router.get("/candidates/{record_id}", summary="候选人详情")
@_handle_write_unconfigured
async def get_candidate(
    record_id: str,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取候选人详情，含简历附件信息。"""
    _require_user(current_user)
    result = await service.get_candidate(record_id)
    return success_response(
        data=CandidateResponse(**result).model_dump(mode="json"),
    )


@router.get("/candidates/{record_id}/resume-file", summary="下载候选人简历文件")
async def get_candidate_resume_file(
    record_id: str,
    current_user: CurrentUser = None,
) -> Any:
    """从飞书下载候选人简历文件（PDF/DOCX）。"""
    _require_user(current_user)
    from app.modules.hr.feishu.client import FeishuClient
    from app.modules.hr.recruitment_repository import (
        TBL_CANDIDATE,
        RecruitmentBitableRepo,
    )

    repo = RecruitmentBitableRepo()
    client = await repo._get_client()
    if not client:
        raise AppException(status_code=404, message="飞书多维表格未配置")

    # 获取候选人记录
    records = await client.search_records(TBL_CANDIDATE, page_size=500)
    candidate_record = None
    for r in records:
        if r.get("record_id") == record_id:
            candidate_record = r
            break

    if not candidate_record:
        raise NotFoundException("候选人", record_id)

    # 获取简历附件信息
    raw_fields = candidate_record.get("fields", {})
    attachments = raw_fields.get("简历附件")
    if not attachments or not isinstance(attachments, list) or len(attachments) == 0:
        raise AppException(status_code=404, message="该候选人没有简历附件")

    file_token = attachments[0].get("file_token", "")
    file_name = attachments[0].get("name", "resume.pdf")

    if not file_token:
        raise AppException(status_code=404, message="简历文件 token 无效")

    # 从飞书下载文件
    try:
        feishu_client = FeishuClient()
        file_bytes = await feishu_client.download_file(file_token)
    except Exception as e:
        logger.exception(
            "failed to download resume from feishu", extra={"record_id": record_id}
        )
        raise AppException(status_code=500, message=f"下载简历失败：{str(e)}")

    # 返回文件
    import io

    from fastapi.responses import StreamingResponse

    content_type = "application/pdf"
    if file_name.lower().endswith(".docx"):
        content_type = (
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
    elif file_name.lower().endswith(".doc"):
        content_type = "application/msword"

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type=content_type,
        headers={
            "Content-Disposition": f"inline; filename*=utf-8''{quote(file_name)}",
        },
    )


@router.put("/candidates/{record_id}", summary="更新候选人")
@_handle_write_unconfigured
async def update_candidate(
    record_id: str,
    payload: CandidateUpdate,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """更新候选人信息（包括面试状态、匹配度等），面试状态变更时自动发送邮件通知。"""
    _require_user(current_user)
    result = await service.update_candidate(
        record_id, payload.model_dump(exclude_none=True)
    )
    return success_response(
        data=CandidateResponse(**result).model_dump(mode="json"),
        message="候选人更新成功",
    )


@router.delete("/candidates/{record_id}", summary="删除候选人")
@_handle_write_unconfigured
async def delete_candidate(
    record_id: str,
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    """删除（软删除）候选人记录。"""
    _require_user(current_user)
    await service.delete_candidate(record_id)
    return success_response(message="候选人已删除")


# ─── Recruitment Onboarding Routes ───


@router.get("/onboarding/names", summary="获取入职姓名列表（公开）")
async def get_onboarding_names(
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
) -> Any:
    """获取飞书入职信息表中的所有姓名，供扫码填写时选择。"""
    items, _ = await service.list_onboarding(page=1, page_size=500)
    names = [item.get("name", "") for item in items if item.get("name")]
    return success_response(data=names)


@router.get("/onboarding", summary="入职列表")
async def list_onboarding_records(
    keyword: str | None = Query(None, description="关键词搜索"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=500, description="每页条数"),
    db: AsyncSession = Depends(get_db),
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取招聘入职管理列表。飞书多维表格未配置时返回空列表。"""
    alias_set = await _resolve_visible_scope(db, current_user)
    items, total = await service.list_onboarding(
        keyword=keyword, page=page, page_size=page_size, dept_alias_set=alias_set
    )
    return paginated_response(
        data=[OnboardingResponse(**item).model_dump(mode="json") for item in items],
        page=page,
        page_size=page_size,
        total=total,
    )


@router.post("/onboarding/from-interview", summary="从面试通过转入职")
@_handle_write_unconfigured
async def create_onboarding_from_interview(
    payload: dict[str, Any],
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """将从面试通过的候选人转入入职流程。"""
    _require_user(current_user)
    candidate_id = payload.get("candidate_id")
    if not candidate_id:
        raise AppException(status_code=400, message="缺少 candidate_id 参数")
    result = await service.create_from_interview(candidate_id)
    return success_response(data=result, message="入职记录创建成功", status_code=201)


@router.get("/onboarding/dashboard", summary="入职看板统计")
async def get_onboarding_dashboard(
    db: AsyncSession = Depends(get_db),
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取入职管理看板统计数据。飞书多维表格未配置时返回空统计。"""
    alias_set = await _resolve_visible_scope(db, current_user)
    result = await service.get_dashboard(dept_alias_set=alias_set)
    return success_response(data=result)


@router.get("/onboarding/{record_id}", summary="入职详情")
@_handle_write_unconfigured
async def get_onboarding_record_recruitment(
    record_id: str,
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """获取招聘入职记录详情。"""
    _require_user(current_user)
    result = await service.get_onboarding(record_id)
    return success_response(data=result)


@router.put("/onboarding/{record_id}", summary="更新入职信息")
@_handle_write_unconfigured
async def update_onboarding_record_recruitment(
    record_id: str,
    payload: OnboardingCreate,
    service: RecruitmentOnboardingService = Depends(get_recruitment_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    """更新招聘入职记录信息。"""
    _require_user(current_user)
    result = await service.update_onboarding(
        record_id, payload.model_dump(exclude_none=True)
    )
    return success_response(data=result, message="入职信息更新成功")


@router.post(
    "/onboarding/{record_id}/sync-to-employee",
    summary="入职完成同步到员工档案和合同管理",
)
@_handle_write_unconfigured
async def sync_onboarding_to_employee(
    record_id: str,
    onboarding_service: RecruitmentOnboardingService = Depends(
        get_recruitment_onboarding_service
    ),
    employee_service: EmployeeService = Depends(get_employee_service),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """入职状态完成后，将入职信息同步到员工档案和合同管理。

    按姓名从飞书员工档案表自动匹配工号，无需手动输入。
    - 员工档案：按工号去重，已存在则返回提示
    - 合同管理：按工号+姓名去重，已存在则跳过
    """
    _require_user(current_user)

    # 1. 从飞书入职表获取入职记录
    onboarding = await onboarding_service.get_onboarding(record_id)
    name = onboarding.get("name", "")
    department = onboarding.get("department", "")
    level = onboarding.get("level", "")
    onboard_date = onboarding.get("onboard_date", "")

    if not name:
        raise AppException(status_code=400, message="入职记录缺少姓名，无法同步")

    # 2. 按姓名从飞书员工档案表查找工号
    from app.modules.hr.recruitment_repository import (
        TBL_EMPLOYEE,
        RecruitmentBitableRepo,
    )

    repo = RecruitmentBitableRepo()
    client = await repo._get_client()
    if not client:
        raise AppException(status_code=500, message="飞书连接未配置")

    # 搜索员工档案表，按姓名匹配
    records = await client.search_records(TBL_EMPLOYEE, page_size=500)
    employee_number = None
    for r in records:
        fields = r.get("fields", {})
        name_field = fields.get("姓名", "")
        emp_name = (
            name_field[0].get("text", "")
            if isinstance(name_field, list) and name_field
            else str(name_field)
        )
        if emp_name == name:
            # 找到匹配的员工，获取工号
            emp_no_field = fields.get("工号", "")
            if isinstance(emp_no_field, (int, float)):
                employee_number = str(int(emp_no_field))
            elif emp_no_field:
                employee_number = str(emp_no_field)
            break

    # 3. 按姓名去重（本地 PG 表）
    from app.modules.hr.repository import EmployeeRepository

    emp_repo = EmployeeRepository(db)
    existing_emp = await emp_repo.get_by_name(name)
    if existing_emp:
        raise AppException(
            status_code=409,
            message=f"{name} 已存在员工档案，无需重复创建",
        )

    # 4. 创建员工档案（工号可空）
    from datetime import date as date_type

    try:
        hire_date = (
            date_type.fromisoformat(onboard_date) if onboard_date else date_type.today()
        )
    except ValueError:
        hire_date = date_type.today()

    from app.modules.hr.schemas import EmployeeCreate

    emp_data = EmployeeCreate(
        employee_number=employee_number,
        name=name,
        department=department or "未分配",
        position=level or "未定",
        hire_date=hire_date,
    )
    created = await employee_service.create_employee(emp_data)
    if isinstance(created, tuple):
        employee, sync_status = created
    else:
        employee, sync_status = created, "success"

    # 5. 同时创建合同管理记录（去重）
    contract_synced = False
    try:
        from app.modules.hr.contract_service import ContractService

        contract_service = ContractService(db)
        contract_data = {
            "employee_number": employee_number,
            "name": name,
            "department": department,
            "level": level,
        }
        contract_record = await contract_service.sync_from_onboarding(contract_data)
        contract_synced = contract_record is not None
    except Exception:
        logger.warning("合同管理同步失败，员工档案已创建", exc_info=True)

    return success_response(
        data={
            "employee_id": str(employee.id),
            "employee_number": employee_number,
            "name": name,
            "feishu_sync_status": sync_status,
            "contract_synced": contract_synced,
        },
        message=(
            f"员工档案创建成功（{name}），合同管理"
            f"{'同步成功' if contract_synced else '已存在或同步失败'}"
        ),
        status_code=201,
    )


# ─── Legacy turnover and new-factory compatibility routes ───
#
# These routes intentionally delegate to the current repositories/services.
# They keep the former URLs alive for migrated pages without introducing clone
# tables or a second permission/audit path.


@router.get("/onboarding-records", summary="老厂入职台账列表")
async def list_onboarding_records_compat(
    department: str | None = Query(None),
    position: str | None = Query(None),
    is_employed: str | None = Query(None),
    keyword: str | None = Query(None),
    sort_by: str = Query("hire_date"),
    sort_order: str = Query("desc"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: OnboardingRecordService = Depends(get_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_records(
        department=department,
        position=position,
        is_employed=is_employed,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[_legacy_record_payload(record) for record in records],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/onboarding-records/sync-from-feishu", summary="从飞书同步老厂入职台账")
async def sync_onboarding_records_compat(
    db: AsyncSession = Depends(get_db),
    service: OnboardingRecordService = Depends(get_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    stats = await service.sync_from_feishu()
    await db.commit()
    return success_response(data=stats, message="老厂入职台账同步完成")


@router.get("/onboarding-records/sync-status", summary="老厂入职台账同步状态")
async def onboarding_sync_status_compat(
    service: OnboardingRecordService = Depends(get_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    status = await service.get_sync_status()
    return success_response(data=status.model_dump(mode="json"))


@router.get("/onboarding-records/{record_id}", summary="入职记录详情")
async def get_onboarding_record_compat(
    record_id: UUID,
    service: OnboardingRecordService = Depends(get_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(data=_legacy_record_payload(record))


@router.get("/departure-records", summary="老厂离职台账列表")
async def list_departure_records_compat(
    department: str | None = Query(None),
    offboarding_type: str | None = Query(None),
    keyword: str | None = Query(None),
    sort_by: str = Query("offboarding_date"),
    sort_order: str = Query("desc"),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_records(
        department=department,
        offboarding_type=offboarding_type,
        keyword=keyword,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[_legacy_record_payload(record) for record in records],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/departure-records", summary="创建离职台账记录")
async def create_departure_record_compat(
    payload: DepartureRecordCreate,
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    await _assert_dept_in_scope(db, current_user, payload.department)
    record = await service.create_record(payload)
    await db.commit()
    return success_response(
        data=_legacy_record_payload(record),
        message="离职台账记录创建成功",
        status_code=201,
    )


@router.get("/departure-records/{record_id}", summary="离职台账记录详情")
async def get_departure_record_compat(
    record_id: UUID,
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(data=_legacy_record_payload(record))


@router.put("/departure-records/{record_id}", summary="更新离职台账记录")
async def update_departure_record_compat(
    record_id: UUID,
    payload: DepartureRecordUpdate,
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    if payload.department:
        await _assert_dept_in_scope(db, current_user, payload.department)
    record = await service.update_record(record_id, payload)
    await db.commit()
    return success_response(
        data=_legacy_record_payload(record),
        message="离职台账记录更新成功",
    )


@router.delete("/departure-records/{record_id}", summary="删除离职台账记录")
async def delete_departure_record_compat(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    record = await service.get_record(record_id)
    await _assert_dept_in_scope(db, current_user, getattr(record, "department", None))
    await service.delete_record(record_id)
    await db.commit()
    return success_response(message="离职台账记录删除成功")


@router.post("/departure-records/sync-from-feishu", summary="从飞书同步老厂离职台账")
async def sync_departure_records_compat(
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    stats = await service.sync_from_feishu()
    await db.commit()
    return success_response(data=stats, message="老厂离职台账同步完成")


@router.get("/departure-records/sync-status", summary="老厂离职台账同步状态")
async def departure_sync_status_compat(
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    status = await service.get_sync_status()
    return success_response(data=status.model_dump(mode="json"))


@router.get("/new/employees", summary="新厂员工列表")
async def list_new_employees_compat(
    department: str | None = Query(None),
    status: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_employee_page_access(db, current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_employees(
        department=department,
        status=status,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[
            EmployeeResponse.model_validate(record).model_dump(mode="json")
            for record in records
        ],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/new/onboarding-records", summary="新厂入职台账列表")
async def list_new_onboarding_records_compat(
    department: str | None = Query(None),
    position: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: OnboardingRecordService = Depends(get_onboarding_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_records(
        department=department,
        position=position,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[_legacy_record_payload(record) for record in records],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/new/departure-records", summary="新厂离职台账列表")
async def list_new_departure_records_compat(
    department: str | None = Query(None),
    offboarding_type: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_records(
        department=department,
        offboarding_type=offboarding_type,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[_legacy_record_payload(record) for record in records],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/new/offboarding-records", summary="新厂离职管理列表")
async def list_new_offboarding_records_compat(
    department: str | None = Query(None),
    offboarding_type: str | None = Query(None),
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: DepartureRecordService = Depends(get_departure_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _assert_dept_in_scope(db, current_user, department)
    records, total = await service.list_records(
        department=department,
        offboarding_type=offboarding_type,
        keyword=keyword,
        page=page_params.page,
        page_size=page_params.page_size,
        dept_alias_set=scope,
    )
    return paginated_response(
        data=[_legacy_record_payload(record) for record in records],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.get("/new/departments", summary="新厂部门列表")
async def list_new_departments_compat(
    keyword: str | None = Query(None),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: DepartmentService = Depends(get_department_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    scope = await _resolve_visible_scope(db, current_user)
    query_page_size = 5000 if scope is not None else page_params.page_size
    records, total = await service.list_departments(
        keyword=keyword,
        page=1 if scope is not None else page_params.page,
        page_size=query_page_size,
    )
    if scope is not None:
        records = [record for record in records if record.name in scope]
        total = len(records)
        start = (page_params.page - 1) * page_params.page_size
        records = records[start : start + page_params.page_size]
    return paginated_response(
        data=[
            DepartmentResponse.model_validate(record).model_dump(mode="json")
            for record in records
        ],
        page=page_params.page,
        page_size=page_params.page_size,
        total=total,
    )


@router.post("/training-notifications/send", summary="发送培训通知到飞书")
async def send_training_notification_compat(
    payload: TrainingNotifyInput,
    db: AsyncSession = Depends(get_db),
    service: EmployeeService = Depends(get_employee_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    await _assert_dept_in_scope(db, current_user, payload.department)
    result = await service.notify_training(payload)
    message = (
        f"发送完成：成功 {result.get('sent', 0)} 人，失败 {result.get('failed', 0)} 人"
    )
    return success_response(
        data=result,
        message=message,
    )


@router.post("/candidates/{candidate_id}/send-notice", summary="发送候选人通知")
async def send_candidate_notice_compat(
    candidate_id: str,
    payload: SendNoticeRequest,
    db: AsyncSession = Depends(get_db),
    service: RecruitmentService = Depends(get_recruitment_service),
    current_user: CurrentUser = None,
) -> Any:
    await _assert_hr_write(db, current_user)
    assert current_user is not None
    if payload.scene_code not in {"interview_notice", "offer_notice"}:
        raise AppException(status_code=422, message="不支持的通知场景")

    candidate = await service.get_candidate(candidate_id)
    expected_status = "已安排" if payload.scene_code == "interview_notice" else "通过"
    if candidate.get("interview_status") != expected_status:
        notice_label = "面试通知" if expected_status == "已安排" else "Offer"
        raise AppException(
            status_code=422,
            message=f"候选人当前状态不允许发送{notice_label}",
        )

    contact = candidate.get("contact")
    email = candidate.get("email")
    if not email and isinstance(contact, str) and "@" in contact:
        email = contact
    variables = {
        "name": candidate.get("name", ""),
        "email": email or "",
        "department": candidate.get("department", ""),
        "position": candidate.get("job_position") or candidate.get("job_id", ""),
        "interview_time": candidate.get("interview_time", ""),
        "interviewer": candidate.get("interviewer", ""),
        "location": candidate.get("location", ""),
        "onboard_date": candidate.get("onboard_date", ""),
    }
    try:
        result = await PushSettingsService(db).send_notice_for_candidate(
            candidate_id=candidate_id,
            candidate_name=str(candidate.get("name", "")),
            candidate_email=str(email or ""),
            scene_code=payload.scene_code,
            variables=variables,
            triggered_by=str(current_user.id),
        )
        await db.commit()
    except Exception:
        await db.rollback()
        raise

    # Do not expose SMTP/Feishu exception text or provider responses to the UI.
    safe_result = {
        "scene_code": result.get("scene_code", payload.scene_code),
        "scene_label": result.get("scene_label", payload.scene_code),
        "email_sent": bool(result.get("email_sent")),
        "email_recipient": result.get("email_recipient"),
        "email_error": "邮件发送失败" if result.get("email_error") else None,
        "feishu_sent": bool(result.get("feishu_sent")),
        "feishu_recipients": result.get("feishu_recipients", []),
        "feishu_errors": ["飞书发送失败"] * len(result.get("feishu_errors", [])),
    }
    return success_response(
        data=SendNoticeResult.model_validate(safe_result).model_dump(mode="json")
    )
