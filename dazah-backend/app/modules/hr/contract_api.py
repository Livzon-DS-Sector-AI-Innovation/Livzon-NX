"""合同管理 API"""

import hashlib
import hmac
import io
import logging
from datetime import UTC, date, datetime
from typing import Any
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import get_db
from app.core.deps import CurrentUser
from app.core.exceptions import AppException, NotFoundException
from app.core.response import paginated_response, success_response
from app.modules.hr.contract_schemas import (
    ContractApprovalResultItem,
    ContractManagementCreate,
    ContractManagementResponse,
    ContractManagementUpdate,
    ContractRenewRequest,
    ContractSignStatusRequest,
)
from app.modules.hr.contract_service import ContractService
from app.modules.hr.schemas import OnboardingSyncRequest

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/contracts", tags=["人事-合同管理"])


# ─── 合同审批回调签名（HR 模块内实现，stage 参与 HMAC）───
# 全局层 app/core/security.py 由架构负责人维护不可修改，
# 此处复用 SECRET_KEY 实现含 stage 的签名，防止伪造 stage 越级审批。


def _sign_contract_callback_token(
    *,
    employee_number: str,
    employee_name: str,
    action: str,
    leader_name: str = "",
    stage: str = "dept",
    secret_key: str | None = None,
) -> str:
    """为合同审批卡片回调参数生成 HMAC 签名 token（含审批阶段）。"""
    settings = get_settings()
    key = secret_key or settings.SECRET_KEY
    payload = f"{employee_number}|{employee_name}|{action}|{leader_name}|{stage}"
    digest = hmac.new(
        key.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256
    ).hexdigest()
    return digest


def _verify_contract_callback_token(
    *,
    employee_number: str,
    employee_name: str,
    action: str,
    leader_name: str = "",
    stage: str = "dept",
    token: str | None = None,
    secret_key: str | None = None,
) -> bool:
    """校验合同审批卡片回调签名 token。"""
    if not token:
        return False
    expected = _sign_contract_callback_token(
        employee_number=employee_number,
        employee_name=employee_name,
        action=action,
        leader_name=leader_name,
        stage=stage,
        secret_key=secret_key,
    )
    return hmac.compare_digest(expected, token)


def _require_user(current_user: CurrentUser) -> Any:
    if current_user is None:
        from app.core.exceptions import AppException

        raise AppException(status_code=401, message="请先登录")


@router.get("", summary="合同管理列表")
async def list_contracts(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str | None = Query(None, description="姓名/工号搜索"),
    department: str | None = Query(None, description="一级部门筛选"),
    contract_sequence: str | None = Query(None, description="第几次合同筛选"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    alias_set = await _assert_dept_in_scope(db, current_user, department)
    service = ContractService(db)
    # 台账只显示审批通过 + 飞书同步历史记录（审批中/拒绝的仅在审批结果页可见）
    records, total = await service.repo.list(
        page=page,
        page_size=page_size,
        keyword=keyword,
        department=department,
        contract_sequence=contract_sequence,
        approval_statuses=["approved", "synced"],
        dept_alias_set=alias_set,
    )
    result = {
        "data": [
            ContractManagementResponse.model_validate(r).model_dump(mode="json")
            for r in records
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }
    return success_response(data=result)


@router.post("", summary="创建合同记录")
async def create_contract(
    data: ContractManagementCreate,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    service = ContractService(db)
    record = await service.create(data)
    return success_response(data=record.model_dump(), message="创建成功")


class ApprovalCallbackRequest(BaseModel):
    """合同审批回调请求体（POST）。"""

    employee_number: str
    employee_name: str
    action: str
    leader_name: str = ""
    stage: str = "dept"  # dept=部门负责人 / supervisor=分管领导
    token: str = ""


@router.get("/approval-callback", summary="合同审批回调（GET，飞书卡片按钮导航）")
async def approval_callback_get(
    employee_number: str = Query("", description="工号"),
    employee_name: str = Query("", description="姓名"),
    action: str = Query("", description="approve/reject"),
    leader_name: str = Query("", description="审批人姓名"),
    stage: str = Query("dept", description="审批阶段: dept/supervisor"),
    token: str = Query("", description="HMAC 签名 token"),
    db: AsyncSession = Depends(get_db),
) -> Any:
    """飞书 URL 按钮回调（GET 导航）。"""
    return await _approval_callback_core(
        employee_number=employee_number,
        employee_name=employee_name,
        action=action,
        leader_name=leader_name,
        stage=stage,
        token=token,
        db=db,
    )


@router.post("/approval-callback", summary="合同审批回调（POST 兼容）")
async def approval_callback_post(
    payload: ApprovalCallbackRequest,
    db: AsyncSession = Depends(get_db),
) -> Any:
    """飞书卡片按钮回调 POST 兼容入口。"""
    return await _approval_callback_core(
        employee_number=payload.employee_number,
        employee_name=payload.employee_name,
        action=payload.action,
        leader_name=payload.leader_name or "",
        stage=payload.stage or "dept",
        token=payload.token or "",
        db=db,
    )


async def process_contract_approval(
    *,
    employee_number: str,
    employee_name: str,
    action: str,
    leader_name: str,
    stage: str,
    db: AsyncSession,
) -> str:
    """合同审批核心逻辑（两级审批），HTTP 回调与飞书卡片回调共用。

    - stage=dept 通过：进入分管领导审批（supervisor_pending），异步推送分管领导卡片
    - stage=dept 拒绝：直接终止（不同意续签，转离职）
    - stage=supervisor 通过：最终同意续签
    - stage=supervisor 拒绝：最终不同意续签（转离职）

    返回状态文本（含防重/失败信息）；调用方负责展示与提交事务。
    """
    if stage not in ("dept", "supervisor"):
        return "回调参数错误，请重新发起审批"

    # 防重复：部门负责人阶段与最终阶段分别去重
    from app.core.redis import cache_get, cache_set

    dedup_key = (
        f"hr:contract:dept:{employee_number}"
        if stage == "dept"
        else f"hr:contract_approved:{employee_number}"
    )
    if await cache_get(dedup_key):
        return "已审批过，请勿重复操作"
    await cache_set(dedup_key, action, ex=86400 * 7)

    try:
        if stage == "dept":
            if action == "approve":
                status_text = await _on_dept_approved(
                    db, employee_number, employee_name, leader_name
                )
            else:
                status_text = await _on_final_rejected(
                    db,
                    employee_number,
                    employee_name,
                    leader_name,
                    reject_by="部门负责人",
                )
        else:
            if action == "approve":
                status_text = await _on_supervisor_approved(
                    db, employee_number, employee_name, leader_name
                )
            else:
                status_text = await _on_final_rejected(
                    db,
                    employee_number,
                    employee_name,
                    leader_name,
                    reject_by="分管领导",
                )
        await db.commit()
    except Exception:
        await db.rollback()
        logger.exception(
            "合同审批处理失败",
            extra={"employee_number": employee_number, "hr_module": "hr"},
        )
        status_text = "处理失败，请联系 HR 处理"
    return status_text


async def _approval_callback_core(
    *,
    employee_number: str,
    employee_name: str,
    action: str,
    leader_name: str,
    stage: str,
    token: str,
    db: AsyncSession,
) -> Any:
    """合同审批 HTTP 回调（GET/POST 兼容旧 URL 按钮卡片）：验签后委托公共逻辑。"""
    from fastapi.responses import HTMLResponse

    # 验签：stage 参与 HMAC，防止伪造 stage 越级审批
    if not _verify_contract_callback_token(
        employee_number=employee_number,
        employee_name=employee_name,
        action=action,
        leader_name=leader_name,
        stage=stage,
        token=token,
    ):
        return HTMLResponse(
            content=_build_callback_html(
                "⚠️ 回调校验失败，请重新发起审批", employee_name, employee_number, True
            )
        )

    status_text = await process_contract_approval(
        employee_number=employee_number,
        employee_name=employee_name,
        action=action,
        leader_name=leader_name,
        stage=stage,
        db=db,
    )
    return HTMLResponse(
        content=_build_callback_html(status_text, employee_name, employee_number, True)
    )


# ─── 两级审批处理辅助函数 ───


def _fmt_contract_seq(value: str | None) -> str:
    """合同次数展示：存储值已含"次"（首次/第二次…），避免重复加"第…次"包装。"""
    if not value:
        return ""
    if isinstance(value, int):
        return f"第{value}次"
    if value == "首次" or value.endswith("次"):
        return value
    return f"第{value}次"


def _compute_renew_seq(emp: Any) -> int:
    """从 Employee 表的 6 组合同日期中找出最晚非空到期日，确定当前合同次数。"""
    from datetime import date as date_type
    from datetime import datetime as dt

    current_seq = 0
    if not emp:
        return 0
    contract_pairs = [
        (1, emp.contract_end_date, emp.contract_start_date),
        (2, emp.contract_end_2, emp.contract_start_2),
        (3, emp.contract_end_3, emp.contract_start_3),
        (4, emp.contract_end_4, emp.contract_start_4),
    ]
    max_date = None
    for seq, end_dt, _ in contract_pairs:
        if isinstance(end_dt, date_type) and (max_date is None or end_dt > max_date):
            max_date = end_dt
            current_seq = seq
    for seq, attr_name in [(5, "contract_end_5"), (6, "contract_end_6")]:
        val = getattr(emp, attr_name, None)
        if isinstance(val, str) and val.strip():
            for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
                try:
                    parsed = dt.strptime(val.strip(), fmt).date()
                    if max_date is None or parsed > max_date:
                        max_date = parsed
                        current_seq = seq
                    break
                except ValueError:
                    continue
    return current_seq


async def _find_contract_record(db: AsyncSession, employee_number: str) -> Any:
    from app.modules.hr.models import ContractManagement as ContractRecord

    result = await db.execute(
        select(ContractRecord)
        .where(
            ContractRecord.employee_number == employee_number,
            ContractRecord.is_deleted.is_(False),
        )
        .limit(1)
    )
    return result.scalars().first()


async def _find_employee(db: AsyncSession, employee_number: str) -> Any:
    from app.modules.hr.models import Employee

    result = await db.execute(
        select(Employee)
        .where(
            Employee.employee_number == employee_number, Employee.is_deleted.is_(False)
        )
        .limit(1)
    )
    return result.scalars().first()


async def _create_contract_record(
    db: AsyncSession,
    emp: Any,
    employee_number: str,
    employee_name: str,
    leader_name: str,
    renew: bool = True,
) -> Any:
    """从 Employee 表补全信息创建合同记录。

    renew=True 时合同次数为当前次数+1（同意续签场景）；
    renew=False 时保持当前次数（不同意续签场景）。
    """
    from app.modules.hr.models import ContractManagement as ContractRecord

    def _date_to_str(d: Any) -> Any:
        if d is None:
            return None
        if hasattr(d, "isoformat"):
            return d.isoformat()
        return str(d) if d else None

    new_record = ContractRecord(
        employee_number=employee_number,
        name=employee_name,
        dept_leader_name=leader_name,
    )
    if emp:
        new_record.gender = emp.gender
        new_record.dept_level1 = emp.department
        new_record.dept_level2 = emp.sub_department
        new_record.position = emp.position
        new_record.job_level = emp.level
        new_record.domain_account = emp.domain_account
        new_record.id_card = emp.id_card
        new_record.id_card_expiry = emp.id_card_expiry
        new_record.archive_number = emp.archive_number
        new_record.contract_start_1 = emp.contract_start_date
        new_record.contract_end_1 = emp.contract_end_date
        new_record.contract_start_2 = emp.contract_start_2
        new_record.contract_end_2 = _date_to_str(emp.contract_end_2)
        new_record.contract_start_3 = emp.contract_start_3
        new_record.contract_end_3 = _date_to_str(emp.contract_end_3)
        new_record.contract_start_4 = emp.contract_start_4
        new_record.contract_end_4 = _date_to_str(emp.contract_end_4)
        new_record.contract_start_5 = emp.contract_start_5
        new_record.contract_end_5 = emp.contract_end_5
        new_record.contract_start_6 = emp.contract_start_6
        new_record.contract_end_6 = emp.contract_end_6

    current_seq = _compute_renew_seq(emp)
    seq_labels = {
        1: "首次",
        2: "第二次",
        3: "第三次",
        4: "第四次",
        5: "第五次",
        6: "第六次",
    }
    if current_seq > 0:
        seq = min(current_seq + 1, 6) if renew else current_seq
        new_record.contract_sequence = seq_labels.get(seq)
    db.add(new_record)
    await db.flush()
    return new_record


async def _on_dept_approved(
    db: AsyncSession, employee_number: str, employee_name: str, leader_name: str
) -> str:
    """部门负责人同意 → 进入分管领导审批；无分管领导则直接最终通过。"""
    record = await _find_contract_record(db, employee_number)
    emp = await _find_employee(db, employee_number)
    if record is None:
        record = await _create_contract_record(
            db, emp, employee_number, employee_name, leader_name
        )

    record.dept_leader_name = leader_name
    record.dept_approved_at = datetime.now(UTC)
    record.approval_status = "supervisor_pending"
    await db.flush()

    if not record.supervisor_open_id:
        # 未配置分管领导：跳过第二级，直接最终通过
        return await _finalize_approved(db, record, emp, supervisor_name=None)

    # 异步推送分管领导审批卡片（HTTP 回调内不执行飞书调用）
    from app.core.jobs import submit_job

    await submit_job(
        _send_supervisor_card_task,
        task_id=f"hr:contract:supervisor_card:{employee_number}",
        ttl=600,
        employee_number=employee_number,
    )
    return "✅ 已同意续签（等待分管领导审批）"


async def _on_supervisor_approved(
    db: AsyncSession, employee_number: str, employee_name: str, leader_name: str
) -> str:
    """分管领导同意 → 最终通过。"""
    record = await _find_contract_record(db, employee_number)
    emp = await _find_employee(db, employee_number)
    if record is None:
        raise AppException(message="合同记录不存在，请重新发起审批")
    return await _finalize_approved(
        db, record, emp, supervisor_name=leader_name or record.supervisor_name
    )


async def _finalize_approved(
    db: AsyncSession, record: Any, emp: Any, supervisor_name: str | None
) -> str:
    """最终通过：写意见、回写员工、同步飞书、异步通知 HR 与办事员。"""
    from app.core.redis import cache_set
    from app.modules.hr.contract_service import _SEQ_LABELS

    if supervisor_name:
        record.supervisor_name = supervisor_name
        record.supervisor_approved_at = datetime.now(UTC)
    record.contract_opinion = "同意续签"
    record.approval_status = "approved"
    # 审批通过：合同期次自动
    # +1（首次→第二次...），台账显示"需要签署的第几次合同"；新日期由 HR
    # 填写
    if record.contract_sequence:
        for seq, label in _SEQ_LABELS.items():
            if label == record.contract_sequence and seq + 1 in _SEQ_LABELS:
                record.contract_sequence = _SEQ_LABELS[seq + 1]
                break
    if emp:
        emp.contract_opinion = "同意续签"
        emp.dept_leader_name = record.dept_leader_name
    await db.flush()

    # 同步飞书合同管理表（仅原有字段，新流程字段白名单外不同步）
    try:
        from app.modules.hr.contract_sync_service import ContractSyncService

        sync_svc = ContractSyncService(db)
        if getattr(record, "feishu_record_id", None):
            await sync_svc.push_update(record)
        else:
            await sync_svc.push_create(record)
    except Exception:
        logger.exception(
            "审批通过同步飞书合同管理表失败",
            extra={"employee_number": record.employee_number, "hr_module": "hr"},
        )

    # 异步通知 HR 结果 + 办事员签署（HTTP 回调内不执行飞书调用）
    from app.core.jobs import submit_job

    await submit_job(
        _send_contract_result_task,
        task_id=f"hr:contract:result:{record.employee_number}",
        ttl=600,
        employee_number=record.employee_number,
    )

    await cache_set(
        f"hr:contract_approved:{record.employee_number}", "approved", ex=86400 * 7
    )
    return "✅ 已同意续签"


async def _on_final_rejected(
    db: AsyncSession,
    employee_number: str,
    employee_name: str,
    leader_name: str,
    reject_by: str,
) -> str:
    """最终拒绝：写意见、创建离职记录、异步通知 HR。"""
    record = await _find_contract_record(db, employee_number)
    emp = await _find_employee(db, employee_number)

    # 记录不存在时创建（保证审批结果页面能看到不通过记录）
    if record is None:
        record = await _create_contract_record(
            db, emp, employee_number, employee_name, leader_name, renew=False
        )

    if record:
        # 拒绝时只记录拒绝人姓名，不写 approved_at（该字段语义是"同意时间"，
        # 拒绝也写入会被审批结果页/导出误判为"该级同意"）
        if reject_by == "部门负责人":
            record.dept_leader_name = leader_name or record.dept_leader_name
        else:
            record.supervisor_name = leader_name or record.supervisor_name
        record.contract_opinion = "不同意续签"
        record.approval_status = "rejected"
        await db.flush()

    if emp:
        emp.contract_opinion = "不同意续签"
        emp.dept_leader_name = record.dept_leader_name if record else leader_name
        await db.flush()

    # 创建离职记录（从员工档案读取全量信息）
    offboarding_error: str | None = None
    try:
        await _create_offboarding_record(
            db, emp, employee_number, employee_name, record, leader_name, reject_by
        )
    except Exception as exc:
        logger.exception(
            "审批拒绝创建离职记录失败",
            extra={"employee_number": employee_number, "hr_module": "hr"},
        )
        offboarding_error = str(exc)

    # 异步通知 HR 结果（HTTP 回调内不执行飞书调用）
    from app.core.jobs import submit_job

    await submit_job(
        _send_contract_result_task,
        task_id=f"hr:contract:result:{employee_number}",
        ttl=600,
        employee_number=employee_number,
    )

    # 标记最终拒绝（防止再次推送/重复审批，7天TTL）
    from app.core.redis import cache_set

    await cache_set(f"hr:contract_approved:{employee_number}", "rejected", ex=86400 * 7)

    status_text = "❌ 已标记为不续签（转离职流程）"
    if offboarding_error:
        status_text += "\n⚠️ 离职记录创建失败，请联系 HR 处理"
    return status_text


async def _create_offboarding_record(
    db: AsyncSession,
    emp: Any,
    employee_number: str,
    employee_name: str,
    record: Any,
    leader_name: str,
    reject_by: str,
) -> None:
    """创建离职记录（审批不同意续签）。"""
    from datetime import date as date_type
    from datetime import datetime as dt

    from app.modules.hr.schemas import OffboardingRecordCreate
    from app.modules.hr.service import OffboardingRecordService

    offboarding_service = OffboardingRecordService(db)

    def _date_to_str(d: date_type | None) -> str | None:
        return d.isoformat() if d is not None else None

    def _int_to_str(v: int | None) -> str | None:
        return str(v) if v is not None else None

    # 取员工最近一次合同截止日期作为最后工作日
    contract_dates = []
    for attr in [
        "contract_end_date",
        "contract_end_2",
        "contract_end_3",
        "contract_end_4",
    ]:
        val = getattr(emp, attr, None)
        if isinstance(val, date_type):
            contract_dates.append(val)
    for attr in ["contract_end_5", "contract_end_6"]:
        val = getattr(emp, attr, None)
        if isinstance(val, str) and val.strip():
            try:
                contract_dates.append(dt.strptime(val.strip(), "%Y-%m-%d").date())
            except ValueError:
                try:
                    contract_dates.append(dt.strptime(val.strip(), "%Y/%m/%d").date())
                except ValueError:
                    pass

    last_working_day = max(contract_dates) if contract_dates else date_type.today()

    off_data = OffboardingRecordCreate(
        employee_id=emp.id if emp else None,
        employee_number=employee_number,
        name=employee_name,
        department=record.dept_level1 if record else (emp.department if emp else ""),
        sub_department=record.dept_level2
        if record
        else (emp.sub_department if emp else ""),
        position=record.position if record else (emp.position if emp else ""),
        level=record.job_level if record else (emp.level if emp else ""),
        offboarding_type="合同到期",
        offboarding_date=last_working_day,
        reason=f"{reject_by}{leader_name}审批不同意续签",
        **(
            {}
            if not emp
            else {
                "seq_number": emp.seq_number,
                "gender": emp.gender,
                "domain_account": emp.domain_account,
                "id_card": emp.id_card,
                "id_card_expiry": emp.id_card_expiry,
                "phone": emp.phone,
                "email": emp.email,
                "emergency_contact_name": emp.emergency_contact_name,
                "emergency_contact_phone": emp.emergency_contact_phone,
                "emergency_contact_relation": emp.emergency_contact_relation,
                "ethnic_group": emp.ethnic_group,
                "native_place": emp.native_place,
                "political_status": emp.political_status,
                "marital_status": emp.marital_status,
                "health_status": emp.health_status,
                "household_type": emp.household_type,
                "status_category": emp.status_category,
                "current_address": emp.current_address,
                "employment_type": emp.employment_type,
                "probation_status": emp.probation_status,
                "probation_effective_date": emp.probation_effective_date,
                "education": emp.education,
                "degree": emp.degree,
                "major": emp.major,
                "school": emp.school,
                "graduation_date": emp.graduation_date,
                "qualification_type": emp.qualification_type,
                "qualifications": emp.qualifications,
                "certificate_number": emp.certificate_number,
                "certificate_review_date": emp.certificate_review_date,
                "hire_date": emp.hire_date,
                "work_start_date": emp.work_start_date,
                "factory_entry_date": emp.factory_entry_date,
                "livo_entry_date": emp.livo_entry_date,
                "work_years": _int_to_str(emp.work_years),
                "work_experience_1": emp.work_experience_1,
                "work_experience_2": emp.work_experience_2,
                "work_experience_3": emp.work_experience_3,
                "work_experience_4": emp.work_experience_4,
                "archive_number": emp.archive_number,
                "birth_year": emp.birth_year,
                "birth_month": emp.birth_month,
                "birth_day": emp.birth_day,
                "age": emp.age,
                "contract_start_date": emp.contract_start_date,
                "contract_end_date": emp.contract_end_date,
                "contract_start_2": emp.contract_start_2,
                "contract_end_2": _date_to_str(emp.contract_end_2),
                "contract_start_3": _date_to_str(emp.contract_start_3),
                "contract_end_3": _date_to_str(emp.contract_end_3),
                "contract_start_4": _date_to_str(emp.contract_start_4),
                "contract_end_4": _date_to_str(emp.contract_end_4),
                "contract_start_5": _date_to_str(emp.contract_start_5),
                "contract_end_5": emp.contract_end_5,
                "contract_start_6": emp.contract_start_6,
                "contract_end_6": emp.contract_end_6,
            }
        ),
    )
    await offboarding_service.create_record(off_data)


# ─── 异步飞书卡片任务（submit_job 后台执行，HTTP 回调内不调用飞书）───


async def _resolve_contract_clerk_ids(
    session: Any,
    configs: Any,
    department: str,
    global_clerk_ids: list[str],
    hr_ids: list[str],
) -> list[str]:
    """解析合同签署办事员：优先该部门在「按部门配置接收人」中显式指定的人；
    未配置该部门则回退全局 sign_clerk_open_ids，再回退 HR 接收人。"""
    from app.modules.hr.models import HrReminderDeptRecipient

    if department:
        for c in configs:
            if getattr(c, "is_deleted", False):
                continue
            result = await session.execute(
                select(HrReminderDeptRecipient)
                .where(
                    HrReminderDeptRecipient.reminder_config_id == c.id,
                    HrReminderDeptRecipient.department == department,
                    HrReminderDeptRecipient.is_deleted.is_(False),
                )
                .limit(1)
            )
            row = result.scalar_one_or_none()
            if row and row.recipient_open_ids:
                return list(dict.fromkeys(row.recipient_open_ids))
    return list(dict.fromkeys(global_clerk_ids)) or list(dict.fromkeys(hr_ids))


def build_contract_approval_actions(
    emp_no: str, emp_name: str, leader_name: str, stage: str, dept_name: str = ""
) -> dict[str, Any]:
    """构建飞书卡片审批按钮组（value 交互回调，卡片内审批不跳转浏览器）。"""
    return {
        "tag": "action",
        "actions": [
            {
                "tag": "button",
                "text": {"tag": "plain_text", "content": f"✅ {emp_name} 同意续签"},
                "type": "primary",
                "value": {
                    "module": "hr_contract_approval",
                    "action": "approve",
                    "employee_number": emp_no,
                    "employee_name": emp_name,
                    "stage": stage,
                    "leader_name": leader_name,
                    "dept_name": dept_name,
                },
            },
            {
                "tag": "button",
                "text": {"tag": "plain_text", "content": f"❌ {emp_name} 不同意"},
                "type": "danger",
                "value": {
                    "module": "hr_contract_approval",
                    "action": "reject",
                    "employee_number": emp_no,
                    "employee_name": emp_name,
                    "stage": stage,
                    "leader_name": leader_name,
                    "dept_name": dept_name,
                },
            },
        ],
    }


async def update_contract_approval_card(
    emp_no: str,
    emp_name: str,
    action: str,
    stage: str,
    dept_name: str,
) -> None:
    "审批后更新飞书审批卡片：已处理员工按钮置灰（部门汇总卡重建元"
    "素 / 分管领导卡整卡替换）。"
    import json

    from app.core.database import async_session_factory
    from app.core.redis import cache_get
    from app.modules.hr.feishu_settings_service import (
        get_hr_feishu_app_credentials,
    )
    from app.platform.integrations.feishu.notification import update_card

    # 卡片由人事应用发送，必须由人事应用更新（卡片归属应用一致）
    # 未配置时跳过置灰（回调仍返回 toast，不阻断审批结果提示）
    try:
        async with async_session_factory() as session:
            app_id, app_secret = await get_hr_feishu_app_credentials(session)
    except Exception as exc:
        logger.warning("合同审批卡片置灰跳过: 人事飞书凭证未配置: %s", exc)
        return

    result_text = "✅ 已同意续签" if action == "approve" else "❌ 已不同意（转离职）"
    try:
        if stage == "dept":
            # 部门汇总卡片：重建元素（已处理员工 -> 灰色文本，未处理保留按钮）
            message_id = await cache_get(f"hr:contract:card:{dept_name}:msgid")
            emps_raw = await cache_get(f"hr:contract:card:{dept_name}:emps")
            if not message_id or not emps_raw:
                logger.warning("合同卡片置灰跳过: 部门卡片快照缺失 dept=%s", dept_name)
                return
            snapshot = json.loads(emps_raw)
            elements: list[dict[str, Any]] = []
            if snapshot.get("content"):
                elements.append(
                    {
                        "tag": "div",
                        "text": {"tag": "lark_md", "content": snapshot["content"]},
                    }
                )
                elements.append({"tag": "hr"})
            for e in snapshot.get("emps", []):
                if e.get("employee_number") == emp_no:
                    elements.append(
                        {
                            "tag": "div",
                            "text": {
                                "tag": "lark_md",
                                "content": (
                                    f"**{e.get('employee_name', emp_name)}**"
                                    f"（{emp_no}）：{result_text}"
                                ),
                            },
                        }
                    )
                else:
                    elements.append(
                        build_contract_approval_actions(
                            e.get("employee_number", ""),
                            e.get("employee_name", ""),
                            e.get("leader_name", ""),
                            "dept",
                        )
                    )
            card = {
                "config": {"wide_screen_mode": True},
                "header": {
                    "title": {
                        "tag": "plain_text",
                        "content": snapshot.get("title", "合同到期提醒与审批"),
                    },
                    "template": "blue",
                },
                "elements": elements,
            }
            await update_card(message_id, card, app_id=app_id, app_secret=app_secret)
        else:
            # 分管领导单员工卡片：整卡替换为结果
            message_id = await cache_get(f"hr:contract:supervisor_card:{emp_no}:msgid")
            if not message_id:
                logger.warning(
                    "合同卡片置灰跳过: 分管领导卡片 message_id 缺失 emp=%s", emp_no
                )
                return
            card = {
                "config": {"wide_screen_mode": True},
                "header": {
                    "title": {
                        "tag": "plain_text",
                        "content": f"合同续签审批 - {emp_name}",
                    },
                    "template": "green" if action == "approve" else "red",
                },
                "elements": [
                    {
                        "tag": "div",
                        "text": {
                            "tag": "lark_md",
                            "content": f"**{emp_name}**（{emp_no}）：{result_text}",
                        },
                    },
                ],
            }
            await update_card(message_id, card, app_id=app_id, app_secret=app_secret)
        logger.info(
            "合同审批卡片已置灰: emp=%s stage=%s action=%s", emp_no, stage, action
        )
    except Exception:
        logger.warning("合同审批卡片置灰失败", exc_info=True)


async def _send_supervisor_card_task(employee_number: str) -> None:
    """给分管领导发单员工审批卡片（部门负责人通过后触发）。"""
    from app.core.database import async_session_factory
    from app.modules.hr.models import ContractManagement as ContractRecord

    async with async_session_factory() as session:
        result = await session.execute(
            select(ContractRecord)
            .where(
                ContractRecord.employee_number == employee_number,
                ContractRecord.is_deleted.is_(False),
            )
            .limit(1)
        )
        record = result.scalars().first()
        if not record or not record.supervisor_open_id:
            logger.warning(
                "分管领导卡片跳过: %s 无记录或未配置分管领导", employee_number
            )
            return
        if record.approval_status != "supervisor_pending":
            logger.info(
                "分管领导卡片跳过: %s 当前状态 %s",
                employee_number,
                record.approval_status,
            )
            return

        supervisor_name = record.supervisor_name or ""

        content = (
            f"部门负责人 **{record.dept_leader_name or ''}** 已同意续签，请您审批：\n\n"
            f"- **{record.name}**（工号：{record.employee_number}）\n"
            f"- 部门：{record.dept_level1 or ''} / {record.dept_level2 or ''}\n"
            f"- {_fmt_contract_seq(record.contract_sequence)}合同，到期日见合同管理台账"
        )
        elements = [
            build_contract_approval_actions(
                record.employee_number,
                record.name,
                supervisor_name,
                "supervisor",
                dept_name=record.dept_level1 or "",
            ),
        ]
        try:
            # 卡片由人事应用发送，必须由人事应用发送（卡片归属应用一致）
            from app.modules.hr.feishu_settings_service import (
                get_hr_feishu_app_credentials,
            )
            from app.platform.integrations.feishu.notification import (
                send_user_card_with_message_id,
            )

            app_id, app_secret = await get_hr_feishu_app_credentials(session)
            message_id = await send_user_card_with_message_id(
                open_id=record.supervisor_open_id,
                title=f"【{record.dept_level1 or ''}】合同续签审批（分管领导）",
                content=content,
                elements=elements,
                app_id=app_id,
                app_secret=app_secret,
            )
            if message_id:
                # 保存 message_id 供审批后置灰卡片
                from app.core.redis import cache_set

                await cache_set(
                    f"hr:contract:supervisor_card:{record.employee_number}:msgid",
                    message_id,
                    ex=86400 * 14,
                )
            logger.info(
                "分管领导审批卡片发送: %s (%s) -> %s message_id=%s",
                record.name,
                record.employee_number,
                supervisor_name,
                message_id,
            )
        except Exception:
            logger.exception(
                "分管领导审批卡片发送失败",
                extra={"employee_number": record.employee_number, "hr_module": "hr"},
            )


async def _send_contract_result_task(employee_number: str) -> None:
    """审批最终落地后：通知 HR 结果卡片；（通过时）通知办事员签署。"""
    from app.core.database import async_session_factory
    from app.modules.hr.models import ContractManagement as ContractRecord
    from app.modules.hr.models import HrReminderConfig

    async with async_session_factory() as session:
        result = await session.execute(
            select(ContractRecord)
            .where(
                ContractRecord.employee_number == employee_number,
                ContractRecord.is_deleted.is_(False),
            )
            .limit(1)
        )
        record = result.scalars().first()
        if not record or record.approval_status not in ("approved", "rejected"):
            logger.warning("审批结果通知跳过: %s 无最终审批记录", employee_number)
            return

        configs = (
            (
                await session.execute(
                    select(HrReminderConfig).where(
                        HrReminderConfig.entity_code == "contract_renewal",
                        # 审批结果/签署通知是审批流程必需环节，不受提醒开关(is_e
                        # nabled)控制
                        HrReminderConfig.is_deleted.is_(False),
                    )
                )
            )
            .scalars()
            .all()
        )

        hr_open_ids: list[str] = []
        clerk_open_ids: list[str] = []
        for c in configs:
            if c.recipient_open_ids:
                hr_open_ids.extend(c.recipient_open_ids)
            if c.sign_clerk_open_ids:
                clerk_open_ids.extend(c.sign_clerk_open_ids)
        hr_open_ids = list(dict.fromkeys(hr_open_ids))
        # 全局办事员（按部门配置的办事员在签署通知时再解析）
        global_clerk_ids = list(dict.fromkeys(clerk_open_ids))

        approved = record.approval_status == "approved"
        result_line = "✅ 同意续签" if approved else "❌ 不同意续签（转离职）"
        content = (
            f"合同到期审批结果：\n\n"
            f"- **{record.name}**（工号：{record.employee_number}）\n"
            f"- 部门：{record.dept_level1 or ''} / {record.dept_level2 or ''}\n"
            f"- 审批结果：{result_line}\n"
            f"- 部门负责人：{record.dept_leader_name or '-'}\n"
            f"- 分管领导：{record.supervisor_name or '-'}"
        )

        from app.modules.hr.feishu_settings_service import (
            get_hr_feishu_app_credentials,
        )
        from app.platform.integrations.feishu.notification import send_user_card

        app_id, app_secret = await get_hr_feishu_app_credentials(session)

        # 1. HR 结果卡片
        for open_id in hr_open_ids:
            try:
                await send_user_card(
                    open_id=open_id,
                    title=f"合同审批结果 - {record.name}",
                    content=content,
                    app_id=app_id,
                    app_secret=app_secret,
                )
            except Exception:
                logger.exception(
                    "审批结果通知失败",
                    extra={
                        "employee_number": record.employee_number,
                        "open_id": open_id,
                        "hr_module": "hr",
                    },
                )

        # 2. 通过时：办事员签署通知（按员工部门解析办事员：部门配置优先，回退全局/HR）
        if approved:
            clerk_open_ids = await _resolve_contract_clerk_ids(
                session,
                configs,
                record.dept_level1 or "",
                global_clerk_ids,
                hr_open_ids,
            )
            if not clerk_open_ids:
                logger.warning("签署通知跳过: %s 未配置办事员", record.employee_number)
                return
            sign_content = (
                f"合同审批已通过，请通知员工到人事签署合同：\n\n"
                f"- **{record.name}**（工号：{record.employee_number}）\n"
                f"- 部门：{record.dept_level1 or ''} / {record.dept_level2 or ''}\n"
                f"- 合同次数：{_fmt_contract_seq(record.contract_sequence)}"
            )
            for open_id in clerk_open_ids:
                try:
                    await send_user_card(
                        open_id=open_id,
                        title=f"合同签署通知 - {record.name}",
                        content=sign_content,
                        app_id=app_id,
                        app_secret=app_secret,
                    )
                except Exception:
                    logger.exception(
                        "签署通知失败",
                        extra={
                            "employee_number": record.employee_number,
                            "open_id": open_id,
                            "hr_module": "hr",
                        },
                    )


def _build_callback_html(
    status_text: str, emp_name: str, emp_no: str, auto_close: bool = True
) -> str:
    """构建审批回调页面 HTML（自动关闭）"""
    close_script = (
        "<script>setTimeout(()=>{window.close()},1500)</script>" if auto_close else ""
    )
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>审批完成</title>
<style>body{{font-family:sans-serif;display:flex;justify-content:center;align-items:center;height:100vh;margin:0;background:#f5f5f5}}
.card{{background:#fff;padding:40px;border-radius:12px;
box-shadow:0 2px 12px rgba(0,0,0,.1);text-align:center;max-width:90%}}
h2{{margin:0 0 12px;font-size:20px}}
p{{color:#666;margin:0;font-size:14px}}</style></head>
<body><div class="card"><h2>{status_text}</h2>
<p>{emp_name}（{emp_no}）</p></div>{close_script}</body></html>"""


@router.post("/{record_id}/renew", summary="填写续签日期（快捷操作）")
async def renew_contract(
    record_id: UUID,
    data: ContractRenewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """快捷填写续签合同日期，自动写入对应的合同字段并回写员工档案表。

    请求体：{ "start_date": "2026-08-16", "end_date": "2029-08-15" }
    根据 contract_sequence 确定写入第几组合同日期字段。
    同时回写 Employee 表对应字段。
    """
    _require_user(current_user)
    from datetime import date as date_type

    from sqlalchemy import select

    from app.modules.hr.models import ContractManagement as ContractRecord
    from app.modules.hr.models import Employee

    result = await db.execute(
        select(ContractRecord).where(
            ContractRecord.id == record_id,
            ContractRecord.is_deleted.is_(False),
        )
    )
    record = result.scalars().first()
    if not record:
        raise NotFoundException("合同记录不存在")

    start_str = data.start_date
    end_str = data.end_date
    if not start_str or not end_str:
        from app.core.exceptions import AppException

        raise AppException(status_code=400, message="请填写续签开始日期和截止日期")

    try:
        renew_start = date_type.fromisoformat(start_str)
        renew_end = date_type.fromisoformat(end_str)
    except ValueError:
        from app.core.exceptions import AppException

        raise AppException(
            status_code=400, message="日期格式错误，请使用 YYYY-MM-DD 格式"
        )

    # 根据 contract_sequence 确定写入哪组字段
    seq_map = {
        "首次": (1, "contract_start_1", "contract_end_1"),
        "第二次": (2, "contract_start_2", "contract_end_2"),
        "第三次": (3, "contract_start_3", "contract_end_3"),
        "第四次": (4, "contract_start_4", "contract_end_4"),
        "第五次": (5, "contract_start_5", "contract_end_5"),
        "第六次": (6, "contract_start_6", "contract_end_6"),
    }
    seq_label = record.contract_sequence or "首次"
    if seq_label not in seq_map:
        from app.core.exceptions import AppException

        raise AppException(status_code=400, message=f"无法识别合同次数: {seq_label}")

    seq_num, cm_start_field, cm_end_field = seq_map[seq_label]

    # 写入合同管理表
    setattr(record, cm_start_field, renew_start)
    # contract_end_2/3/4/5/6 是 String 类型，需要转字符串
    if cm_end_field in (
        "contract_end_2",
        "contract_end_3",
        "contract_end_4",
        "contract_end_5",
        "contract_end_6",
    ):
        setattr(record, cm_end_field, renew_end.isoformat())
    else:
        setattr(record, cm_end_field, renew_end)
    await db.flush()

    # 回写员工档案表
    emp_result = await db.execute(
        select(Employee)
        .where(
            Employee.employee_number == record.employee_number,
            Employee.is_deleted.is_(False),
        )
        .limit(1)
    )
    emp = emp_result.scalars().first()
    if emp:
        # Employee 表的字段名与 CM 表不同（首次用
        # contract_start_date/contract_end_date）
        emp_field_map = {
            1: ("contract_start_date", "contract_end_date"),
            2: ("contract_start_2", "contract_end_2"),
            3: ("contract_start_3", "contract_end_3"),
            4: ("contract_start_4", "contract_end_4"),
            5: ("contract_start_5", "contract_end_5"),
            6: ("contract_start_6", "contract_end_6"),
        }
        if seq_num in emp_field_map:
            emp_start_field, emp_end_field = emp_field_map[seq_num]
            setattr(emp, emp_start_field, renew_start)
            # Employee 表 contract_end_2/3/4 是 Date，5/6 是 String
            if emp_end_field in ("contract_end_5", "contract_end_6"):
                setattr(emp, emp_end_field, renew_end.isoformat())
            else:
                setattr(emp, emp_end_field, renew_end)
            await db.flush()

    # 同步到飞书多维表格
    # 1. 同步合同管理表到飞书
    try:
        from app.modules.hr.contract_sync_service import ContractSyncService

        sync_svc = ContractSyncService(db)
        if record.feishu_record_id:
            await sync_svc.push_update(record)
        else:
            await sync_svc.push_create(record)
    except Exception:
        import logging

        logging.getLogger(__name__).exception(
            "续签日期同步飞书合同管理表失败",
            extra={"employee_number": record.employee_number, "hr_module": "hr"},
        )

    # 2. 同步员工档案表到飞书
    if emp:
        try:
            from app.modules.hr.service import EmployeeService

            emp_service = EmployeeService(db)
            await emp_service._sync_single_to_feishu(emp)
        except Exception:
            import logging

            logging.getLogger(__name__).exception(
                "续签日期同步飞书员工档案表失败",
                extra={"employee_number": record.employee_number, "hr_module": "hr"},
            )

    await db.commit()
    await db.refresh(record)
    return success_response(
        data=ContractManagementResponse.model_validate(record).model_dump(mode="json"),
        message=f"续签日期已保存（{seq_label}）并同步至员工档案",
    )


@router.post("/sync-from-onboarding", summary="从入职管理同步合同信息")
async def sync_contract_from_onboarding(
    data: OnboardingSyncRequest,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """入职完成时同步合同信息到合同管理模块。

    支持两种模式：
    1. 提供 employee_number：直接使用该工号
    2. 不提供 employee_number：按姓名从飞书员工档案表自动查找工号
    """
    _require_user(current_user)
    service = ContractService(db)

    # 如果没有工号，按姓名查找
    if not data.employee_number and data.name:
        from app.modules.hr.recruitment_repository import (
            TBL_EMPLOYEE,
            RecruitmentBitableRepo,
        )

        repo = RecruitmentBitableRepo()
        client = await repo._get_client()
        if client:
            records = await client.search_records(TBL_EMPLOYEE, page_size=500)
            for r in records:
                fields = r.get("fields", {})
                name_field = fields.get("姓名", "")
                emp_name = (
                    name_field[0].get("text", "")
                    if isinstance(name_field, list) and name_field
                    else str(name_field)
                )
                if emp_name == data.name:
                    emp_no_field = fields.get("工号", "")
                    if isinstance(emp_no_field, (int, float)):
                        data.employee_number = str(int(emp_no_field))
                    elif emp_no_field:
                        data.employee_number = str(emp_no_field)
                    break

    record = await service.sync_from_onboarding(data.model_dump(exclude_none=True))
    if record:
        return success_response(data={"synced": True}, message="合同信息已同步")
    return success_response(data={"synced": False}, message="该员工已有合同记录，跳过")


# ─── 审批结果列表 / 导出 / 签署状态 ───


def _latest_contract_end(record: Any) -> str:
    """取合同记录 6 组截止日期中最晚的一个，返回字符串。"""
    from datetime import date as date_type

    candidates: list[str] = []
    for attr in (
        "contract_end_1",
        "contract_end_2",
        "contract_end_3",
        "contract_end_4",
    ):
        val = getattr(record, attr, None)
        if isinstance(val, date_type):
            candidates.append(val.isoformat())
        elif isinstance(val, str) and val.strip():
            candidates.append(val.strip())
    for attr in ("contract_end_5", "contract_end_6"):
        val = getattr(record, attr, None)
        if isinstance(val, str) and val.strip():
            candidates.append(val.strip())
    return max(candidates) if candidates else ""


@router.get("/approval-results", summary="合同审批结果列表（按季度/部门/结果筛选）")
async def list_contract_approval_results(
    start_date: date | None = Query(None, description="开始日期"),
    end_date: date | None = Query(None, description="结束日期"),
    department: str | None = Query(None, description="一级部门筛选"),
    result: str | None = Query(None, description="approved/rejected"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    alias_set = await _assert_dept_in_scope(db, current_user, department)
    from app.modules.hr.models import ContractManagement as ContractRecord

    stmt = select(ContractRecord).where(
        ContractRecord.is_deleted.is_(False),
        ContractRecord.approval_status.in_(["approved", "rejected"]),
    )
    if result == "approved":
        stmt = stmt.where(ContractRecord.approval_status == "approved")
    elif result == "rejected":
        stmt = stmt.where(ContractRecord.approval_status == "rejected")
    if alias_set is not None:
        # 部门级数据隔离：可见部门别名集合（一级/二级部门任一命中）
        stmt = stmt.where(
            or_(
                ContractRecord.dept_level1.in_(alias_set),
                ContractRecord.dept_level2.in_(alias_set),
            )
        )
    elif department:
        stmt = stmt.where(ContractRecord.dept_level1.ilike(f"{department}%"))
    # 审批完成时间 = 最后一级处理时间（拒绝记录不再写 approved_at，用 updated_at 兜底）
    completed_at = func.coalesce(
        ContractRecord.supervisor_approved_at,
        ContractRecord.dept_approved_at,
        ContractRecord.updated_at,
    )
    if start_date:
        stmt = stmt.where(
            completed_at
            >= datetime.combine(start_date, datetime.min.time(), tzinfo=UTC)
        )
    if end_date:
        stmt = stmt.where(
            completed_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=UTC)
        )

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar() or 0
    stmt = stmt.order_by(
        completed_at.desc().nullslast(), ContractRecord.employee_number.asc()
    )
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    records = (await db.execute(stmt)).scalars().all()

    data = []
    for r in records:
        item = ContractApprovalResultItem.model_validate(r).model_dump(mode="json")
        item["contract_end_date"] = _latest_contract_end(r)
        completed = r.supervisor_approved_at or r.dept_approved_at
        item["completed_at"] = completed.isoformat() if completed else None
        data.append(item)
    return paginated_response(data=data, total=total, page=page, page_size=page_size)


@router.get("/approval-results/export", summary="导出合同审批结果 Excel")
async def export_contract_approval_results(
    start_date: date | None = Query(None, description="开始日期"),
    end_date: date | None = Query(None, description="结束日期"),
    department: str | None = Query(None, description="一级部门筛选"),
    result: str | None = Query(None, description="approved/rejected"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    alias_set = await _assert_dept_in_scope(db, current_user, department)
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import (  # type: ignore[import-untyped]
        Alignment,
        Border,
        Font,
        Side,
    )

    from app.modules.hr.models import ContractManagement as ContractRecord

    stmt = select(ContractRecord).where(
        ContractRecord.is_deleted.is_(False),
        ContractRecord.approval_status.in_(["approved", "rejected"]),
    )
    if result == "approved":
        stmt = stmt.where(ContractRecord.approval_status == "approved")
    elif result == "rejected":
        stmt = stmt.where(ContractRecord.approval_status == "rejected")
    if alias_set is not None:
        # 部门级数据隔离：可见部门别名集合（一级/二级部门任一命中）
        stmt = stmt.where(
            or_(
                ContractRecord.dept_level1.in_(alias_set),
                ContractRecord.dept_level2.in_(alias_set),
            )
        )
    elif department:
        stmt = stmt.where(ContractRecord.dept_level1.ilike(f"{department}%"))
    # 审批完成时间 = 最后一级处理时间（拒绝记录不再写 approved_at，用 updated_at 兜底）
    completed_at = func.coalesce(
        ContractRecord.supervisor_approved_at,
        ContractRecord.dept_approved_at,
        ContractRecord.updated_at,
    )
    if start_date:
        stmt = stmt.where(
            completed_at
            >= datetime.combine(start_date, datetime.min.time(), tzinfo=UTC)
        )
    if end_date:
        stmt = stmt.where(
            completed_at <= datetime.combine(end_date, datetime.max.time(), tzinfo=UTC)
        )
    stmt = stmt.order_by(completed_at.desc().nullslast())
    records = (await db.execute(stmt)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "合同审批结果"

    title_font = Font(name="宋体", size=14, bold=True)
    header_font = Font(name="宋体", size=11, bold=True)
    data_font = Font(name="宋体", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "序号",
        "工号",
        "姓名",
        "一级部门",
        "二级部门",
        "合同到期日期",
        "部门经理审批",
        "分管领导审批",
        "审批结果",
    ]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="合同到期审批结果汇总")
    title_cell.font = title_font
    title_cell.alignment = center_align

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for idx, r in enumerate(records):
        row = 4 + idx
        opinion = "同意续签" if r.approval_status == "approved" else "不同意续签"

        # 两级审批各自结果：部门经理/分管领导 同意（MM-DD）/不同意
        def _fmt_result(name: str | None, approved_at: Any, dept_ok: bool) -> str:
            if approved_at:
                return f"{name or '-'}：同意（{approved_at.strftime('%m-%d')}）"
            if not dept_ok:
                return f"{name or '-'}：不同意"
            return "-"

        dept_result = _fmt_result(r.dept_leader_name, r.dept_approved_at, False)
        supervisor_result = _fmt_result(
            r.supervisor_name, r.supervisor_approved_at, not bool(r.dept_approved_at)
        )
        values = [
            idx + 1,
            r.employee_number,
            r.name,
            r.dept_level1 or "",
            r.dept_level2 or "",
            _latest_contract_end(r),
            dept_result,
            supervisor_result,
            opinion,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 16
    ws.column_dimensions["C"].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"合同审批结果_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@router.put("/{record_id}/sign-status", summary="标记合同签署状态（已签署/拒签）")
async def update_contract_sign_status(
    record_id: UUID,
    data: ContractSignStatusRequest,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    if data.signed_status not in ("已签署", "拒签"):
        raise AppException(status_code=400, message="签署状态必须为 已签署 或 拒签")
    from app.modules.hr.models import ContractManagement as ContractRecord

    result = await db.execute(
        select(ContractRecord).where(
            ContractRecord.id == record_id,
            ContractRecord.is_deleted.is_(False),
        )
    )
    record = result.scalar_one_or_none()
    if not record:
        raise NotFoundException("合同记录不存在")

    record.signed_status = data.signed_status
    record.signed_at = datetime.now(UTC) if data.signed_status == "已签署" else None
    await db.flush()

    # 同步飞书合同管理表（仅原有字段）
    try:
        from app.modules.hr.contract_sync_service import ContractSyncService

        sync_svc = ContractSyncService(db)
        if record.feishu_record_id:
            await sync_svc.push_update(record)
        else:
            await sync_svc.push_create(record)
    except Exception:
        logger.exception(
            "签署状态同步飞书合同管理表失败",
            extra={"record_id": str(record_id), "hr_module": "hr"},
        )

    await db.commit()
    await db.refresh(record)
    return success_response(
        data=ContractManagementResponse.model_validate(record).model_dump(mode="json"),
        message=f"已标记为{data.signed_status}",
    )


@router.get("/{record_id}", summary="合同记录详情")
async def get_contract(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    service = ContractService(db)
    try:
        record = await service.get(record_id)
        return success_response(data=record.model_dump())
    except ValueError:
        raise NotFoundException("合同记录不存在")


@router.post("/sync-from-feishu", summary="从飞书多维表格同步合同数据")
async def sync_contracts_from_feishu(
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """方向 B：从飞书多维表格全量拉取合同数据"""
    _require_user(current_user)
    service = ContractService(db)
    try:
        result = await service.sync_from_feishu()
        await db.commit()
        if "error" in result:
            return success_response(data=result, message=f"同步失败: {result['error']}")
        return success_response(
            data=result,
            message=(
                f"同步完成: 新增 {result['created']} 条, 更新 "
                f"{result['updated']} 条, 删除 {result.get('deleted', 0)} 条"
            ),
        )
    except Exception:
        # 外部依赖（飞书）失败：502 + 统一错误响应，不伪造成功
        await db.rollback()
        import logging

        logging.getLogger(__name__).exception("合同同步飞书失败")
        raise AppException(
            status_code=502,
            message="合同同步飞书失败（飞书服务不可用），请稍后重试",
        )


@router.put("/{record_id}", summary="更新合同记录")
async def update_contract(
    record_id: UUID,
    data: ContractManagementUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    service = ContractService(db)
    try:
        record = await service.update(record_id, data)
        return success_response(data=record.model_dump(), message="更新成功")
    except ValueError:
        raise NotFoundException("合同记录不存在")


@router.delete("/{record_id}", summary="删除合同记录")
async def delete_contract(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    service = ContractService(db)
    try:
        await service.delete(record_id)
        return success_response(message="删除成功")
    except ValueError:
        raise NotFoundException("合同记录不存在")
