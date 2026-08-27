"""HR-QD-01 年度员工培训清单 Excel 文档生成器（清单锁定版模板保真填充）.

模板结构（Sheet1，A1:M56）：
- 锁定区：A2/C2/D2 程序头、A6:E6 表头（序号|培训时间|培训内容|考核结果|备注）、
  A56 QA确认人
- 可编辑区：A4 标题、A5 "部门：___ 姓名：___"、A7:E55 数据行（49行，细边框、居中、wrap）
- 工作表保护开启（加载-保存天然保留）

保真原则：仅填充可编辑区，保留锁定区文本与全部单元格格式；
数据行内容居中、5号字(10.5pt)、中文宋体、英文数字 Times New Roman
（继承模板样式并统一字号）。
"""

from __future__ import annotations

import logging
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

from app.modules.hr.date_format import fmt_date_str

logger = logging.getLogger(__name__)

TEMPLATE_NAME = "HR-QD-01年度员工培训清单(清单锁定版)2026.01.20.xlsx"

# 数据行区间（模板预置 49 行；行 56 为 QA 确认合并行，动态扩展时保持其位置）
DATA_START_ROW = 7
LAST_TEMPLATE_ROW = 55
QA_ROW_INDEX = 56  # 1-based 行号（insert_rows 后自动下移）


def _find_template() -> Path:
    candidates = [
        Path("员工培训教育管理规程") / TEMPLATE_NAME,
        Path("../员工培训教育管理规程") / TEMPLATE_NAME,
        Path(__file__).resolve().parent.parent.parent
        / "员工培训教育管理规程"
        / TEMPLATE_NAME,
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"模板文件未找到: {TEMPLATE_NAME}")


def _apply_data_style(ws: Any, row: int, col: int) -> None:
    """复制模板数据行样式并统一为 5 号字（10.5pt）、居中、自动换行."""
    dst = ws.cell(row=row, column=col)
    src = ws.cell(row=LAST_TEMPLATE_ROW, column=col)
    if src.has_style:
        f = copy(src.font)
        f.size = 10.5  # 5号
        dst.font = f
        # 用户要求：可编辑区域填写内容一律居中
        dst.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        dst.border = copy(src.border)


def generate_employee_training_list(
    department: str,
    name: str,
    records: list[dict[str, Any]],
    year: int | None = None,
) -> BytesIO:
    """按 HR-QD-01 模板生成单个员工的年度员工培训清单 xlsx.

    Args:
        department: 培训部门（写入 A5 "部门："）
        name: 员工姓名（写入 A5 "姓名："）
        records: 培训记录列表，每项含 training_datetime/training_date/
            training_content/personal_score/remarks
        year: 标题年份（写入 A4 "{year}年度员工培训清单"），缺省取记录最新年份或当前年
    """
    if year is None:
        from datetime import date as _date
        from datetime import datetime as _datetime

        years: list[int] = []
        for record in records:
            value = record.get("training_date")
            if isinstance(value, (_date, _datetime)):
                years.append(value.year)
        year = max(years) if years else _date.today().year

    template_path = _find_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb["Sheet1"]

    # A4 标题（保留模板尾部占位空格）
    ws["A4"] = f"{year}年度员工培训清单{' ' * 20}"

    # A5 部门/姓名（合并单元格 A5:E5，保留单元格样式 Times New Roman 16pt bold）
    # 模板原始占位：" 部门：" 后 74 个半角空格（≈37 个全角字符位）再 "姓名："
    # 部门值后补全角空格对齐，使"姓名："保持在模板原始位置附近
    dept_pad = max(1, 37 - len(department))
    ws["A5"] = f" 部门：{department}{'　' * dept_pad}姓名：{name}"

    # 超过模板预置行数时先批量插入行（行56 QA 合并区自动下移），
    # 并从 LAST_TEMPLATE_ROW 复制样式
    extra = len(records) - (LAST_TEMPLATE_ROW - DATA_START_ROW + 1)
    if extra > 0:
        ws.insert_rows(QA_ROW_INDEX, extra)
        for offset in range(extra):
            new_row = QA_ROW_INDEX + offset
            for col in range(1, 6):
                _apply_data_style(ws, new_row, col)

    for idx, record in enumerate(records):
        row = DATA_START_ROW + idx
        # A 序号
        ws.cell(row=row, column=1, value=idx + 1)
        _apply_data_style(ws, row, 1)
        # B 培训时间（training_datetime 优先，空则 training_date）
        time_text = record.get("training_datetime") or (
            fmt_date_str(record["training_date"]) if record.get("training_date") else ""
        )
        ws.cell(row=row, column=2, value=time_text)
        _apply_data_style(ws, row, 2)
        # C 培训内容
        ws.cell(row=row, column=3, value=record.get("training_content") or "")
        _apply_data_style(ws, row, 3)
        # D 考核结果（个人成绩）
        ws.cell(row=row, column=4, value=record.get("personal_score") or "")
        _apply_data_style(ws, row, 4)
        # E 备注（空值统一填 -）
        ws.cell(row=row, column=5, value=record.get("remarks") or "-")
        _apply_data_style(ws, row, 5)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
