"""ESG 培训报表 API 端点."""

import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.datavalidation import (  # type: ignore[import-untyped]
    DataValidation,
)
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.core.database import get_db
from app.core.deps import CurrentUser
from app.core.exceptions import AppException
from app.core.response import paginated_response, success_response
from app.core.upload_security import read_upload_secure
from app.modules.hr.esg_service import EsgTrainingRecordService
from app.modules.hr.schemas import (
    BatchDeleteRequest,
    EsgListFilters,
    EsgTrainingRecordCreate,
    EsgTrainingRecordResponse,
    EsgTrainingRecordUpdate,
)
from app.shared.schemas import PageParams

logger = logging.getLogger(__name__)

router = APIRouter()

# 导出格式严格按桌面模板《ESG培训报表模板（更新时间2026.05.29）.xls》：
# 元信息三行 + 17 列表头（CODE<中文名>(*)）+ 培训类型下拉选项行（D 列）+ 数据行
_ESG_HEADERS = [
    "CULTIVATE_YEAR<培训日期>(*)",
    "CULTIVATE_NAME<培训名称>(*)",
    "CULTIVATE_WAY<培训方式>(*)",
    "CALIBER<口径>(*)",
    "CULTIVATE_TYPE<培训类型>(*)",
    "NAME<姓名>(*)",
    "USERID<员工账号>(*)",
    "ADDRESS<身份所属地>(*)",
    "DEPARTMENT<部门>(*)",
    "EMPLOYEE_LEVEL<层级>(*)",
    "GENDER<性别>",
    "AGE<年龄>",
    "CULTIVATE_DURATION<培训时长(h)>(*)",
    "IS_INSIDE<是否通过本次培训成功实现晋升>(*)",
    "REMARK<备注>",
    "APPLYCOM<单位名称>",
    "APPLYCOMNO<单位编码>",
]
# 导出数据列映射（与 _ESG_HEADERS 一一对应；IS_INSIDE 系统无数据，导出留空）
_ESG_FIELDS = [
    "training_date",
    "training_name",
    "training_method",
    "caliber",
    "training_type",
    "employee_name",
    "employee_account",
    "location_address",
    "department",
    "employee_level",
    "gender",
    "age",
    "duration",
    None,
    "remarks",
    "apply_company",
    "apply_company_no",
]
# 模板 D 列培训类型下拉选项（模板 9 项）
_ESG_TRAINING_TYPES = [
    "EHS类",
    "质量类",
    "商业道德反贪腐",
    "负责任营销",
    "数据安全、隐私保护",
    "领导力",
    "管理类",
    "多元化",
    "女性领导力发展计划",
]
# 模板列宽（xlrd 单位 1/256 字符宽 → openpyxl 字符数）
_ESG_TEMPLATE_COL_WIDTHS = [
    37.6,
    37.6,
    36.3,
    22.3,
    37.6,
    18.1,
    26.5,
    30.7,
    26.5,
    32.1,
    20.9,
    16.7,
    47.4,
    58.6,
    20.9,
    29.3,
    32.1,
]
# 模板行高（xlrd 单位 1/20 pt → pt）：标题/创建日期/制表信息/表头/选项行
_ESG_TEMPLATE_ROW_HEIGHTS = [19.2, 13.2, 13.2, 15.6, 13.2]
# 导入兜底口径：Excel 中缺省时按集团填报规则补默认值
# （口径=部门组织、身份所属地=中国大陆、培训方式=线下；导出不做覆盖，按实际数据导出）
_ESG_IMPORT_DEFAULTS = {
    "training_method": "线下",
    "caliber": "部门组织",
    "location_address": "中国大陆",
}


def get_esg_service(
    session: AsyncSession = Depends(get_db),
) -> EsgTrainingRecordService:
    return EsgTrainingRecordService(session)


# 单次导入行数上限（防超大文件拖垮导入与数据库）
ESG_IMPORT_MAX_ROWS = 2000


def _parse_esg_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# 模板英文表头 CODE 前缀 → 字段名（用于导入时识别集团模板格式）
_ESG_HEADER_CODE_MAP: dict[str, str] = {
    "CULTIVATE_YEAR": "training_date",
    "CULTIVATE_NAME": "training_name",
    "CULTIVATE_WAY": "training_method",
    "CALIBER": "caliber",
    "CULTIVATE_TYPE": "training_type",
    "NAME": "employee_name",
    "USERID": "employee_account",
    "ADDRESS": "location_address",
    "DEPARTMENT": "department",
    "EMPLOYEE_LEVEL": "employee_level",
    "GENDER": "gender",
    "AGE": "age",
    "CULTIVATE_DURATION": "duration",
    "REMARK": "remarks",
    "APPLYCOM": "apply_company",
    "APPLYCOMNO": "apply_company_no",
}
# 中文表头名 → 字段名（兼容历史导出文件）
_ESG_HEADER_CN_MAP: dict[str, str] = {
    "培训日期": "training_date",
    "培训名称": "training_name",
    "培训方式": "training_method",
    "口径": "caliber",
    "培训类型": "training_type",
    "姓名": "employee_name",
    "员工账号": "employee_account",
    "身份所属地": "location_address",
    "部门": "department",
    "层级": "employee_level",
    "性别": "gender",
    "年龄": "age",
    "培训时长": "duration",
    "备注": "remarks",
    "单位名称": "apply_company",
    "单位编码": "apply_company_no",
}


def _esg_header_to_field(text: str) -> str | None:
    """识别 ESG 表头单元格，返回字段名（兼容模板英文 CODE 与中文表头）.

    模板表头形如 CULTIVATE_YEAR<培训日期>(*)：取 < 前的 CODE 前缀匹配；
    中文表头直接匹配中文名。
    """
    if not text:
        return None
    code = text.split("<", 1)[0].strip().upper()
    if code in _ESG_HEADER_CODE_MAP:
        return _ESG_HEADER_CODE_MAP[code]
    return _ESG_HEADER_CN_MAP.get(text.strip())


@router.get("/esg-training-records", summary="ESG培训报表列表（支持各列筛选）")
async def list_esg_records(
    department: str | None = None,
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    filters: EsgListFilters = Depends(),
    page_params: PageParams = Depends(),
    db: AsyncSession = Depends(get_db),
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    await _assert_dept_in_scope(db, current_user, department)
    if not department:
        raise AppException(status_code=400, message="请提供部门参数")
    records, total = await service.list_by_department(
        department=department,
        page=page_params.page,
        page_size=page_params.page_size,
        date_from=date_from,
        date_to=date_to,
        filters=filters,
    )
    data = [
        EsgTrainingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]
    return paginated_response(
        data=data, page=page_params.page, page_size=page_params.page_size, total=total
    )


@router.get(
    "/esg-training-records/filter-options",
    summary="ESG培训报表各枚举列筛选选项（部门+日期范围内去重）",
)
async def list_esg_filter_options(
    department: str = Query(..., description="部门"),
    date_from: date | None = Query(None, description="培训日期起"),
    date_to: date | None = Query(None, description="培训日期止"),
    db: AsyncSession = Depends(get_db),
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    await _assert_dept_in_scope(db, current_user, department)
    options = await service.filter_options(
        department=department, date_from=date_from, date_to=date_to
    )
    return success_response(data=options)


@router.post("/esg-training-records", summary="创建ESG培训记录")
async def create_esg_record(
    payload: EsgTrainingRecordCreate,
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.create_record(payload)
    return success_response(
        data=EsgTrainingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="ESG培训记录创建成功",
        status_code=201,
    )


@router.get("/esg-training-records/export", summary="按部门导出ESG培训报表Excel")
async def export_esg_records(
    department: str = Query(..., description="部门"),
    date_from: date | None = Query(None, description="培训日期起（筛选全年/月份）"),
    date_to: date | None = Query(None, description="培训日期止"),
    db: AsyncSession = Depends(get_db),
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    from app.modules.hr.api import _assert_dept_in_scope

    await _assert_dept_in_scope(db, current_user, department)
    records, _ = await service.list_by_department(
        department=department, page=1, page_size=10000
    )
    if date_from:
        records = [
            r for r in records if r.training_date and r.training_date >= date_from
        ]
    if date_to:
        records = [r for r in records if r.training_date and r.training_date <= date_to]
    dicts = [
        EsgTrainingRecordResponse.model_validate(r).model_dump(mode="json")
        for r in records
    ]

    # 布局按模板《ESG培训报表模板（更新时间2026.05.29）.xls》（不含选项行）：
    # 行1 标题 / 行2 创建日期 / 行3 制表信息 / 行4 表头 / 行5+ 数据
    wb = Workbook()
    ws = wb.active
    ws.title = "集团培训明细表"
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 元信息三行（模板样式：标题宋体15，其余宋体10，垂直居中；填写内容均加边框）
    today = date.today()
    meta_rows = [
        "员工培训信息收集:集团培训明细表",
        f"创建日期:{today.strftime('%Y-%m-%d')} ",
        "制表信息:人力资源部/人力资源部//",
    ]
    for ri, text in enumerate(meta_rows, start=1):
        cell = ws.cell(row=ri, column=1, value=text)
        cell.font = Font(name="宋体", size=15 if ri == 1 else 10)
        cell.alignment = Alignment(vertical="center")
        cell.border = thin

    # 表头行（模板样式：Arial 12 居中、细边框、浅灰 C0C0C0 填充）
    for ci, h in enumerate(_ESG_HEADERS, start=1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = Font(name="Arial", size=12)
        cell.alignment = center
        cell.border = thin
        cell.fill = PatternFill(fill_type="solid", fgColor="C0C0C0")

    # 数据行（从第 5 行起；全部单元格加边框+居中，Arial 10）
    for ri, rec in enumerate(dicts, start=5):
        for ci, field in enumerate(_ESG_FIELDS, start=1):
            if field is None:
                # 空占位列（IS_INSIDE）也写边框，保持网格完整
                cell = ws.cell(row=ri, column=ci, value="")
                cell.border = thin
                cell.alignment = center
                continue
            val = rec.get(field)
            if isinstance(val, date):
                val = val.strftime("%Y-%m-%d")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.alignment = center
            cell.font = Font(name="Arial", size=10)

    # 列宽/行高对齐模板
    for ci, width in enumerate(_ESG_TEMPLATE_COL_WIDTHS, start=1):
        ws.column_dimensions[chr(64 + ci)].width = width
    for ri, height in enumerate(_ESG_TEMPLATE_ROW_HEIGHTS, start=1):
        ws.row_dimensions[ri].height = height

    # E 列（培训类型）数据区下拉（模板同款 9 项）
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(_ESG_TRAINING_TYPES) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"E5:E{4 + max(len(dicts), 1)}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    # 文件名：导出月份-导出部门-ESG培训报表
    filename = f"{today.year}年{today.month}月-{department}-ESG培训报表.xlsx"
    encoded = quote(filename, safe="")
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=utf-8''{encoded}"},
    )


@router.post("/esg-training-records/import", summary="按部门导入ESG培训报表Excel")
async def import_esg_records(
    file: UploadFile = File(..., description="Excel文件(.xlsx)"),
    department: str = Query(..., description="导入到哪个部门"),
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    """解析 ESG 培训报表 Excel，批量创建记录，部门记为所选部门."""
    from openpyxl import load_workbook

    _require_user(current_user)
    _, content = await read_upload_secure(
        file,
        max_bytes=get_settings().MAX_UPLOAD_SIZE_MB * 1024 * 1024,
        allowed_extensions={".xlsx"},
        what="ESG 培训报表",
    )
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    header_row, col_map = 0, {}
    for ri, row in enumerate(
        ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1
    ):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # 兼容模板英文表头（CULTIVATE_YEAR<培训日期>(*)）与中文表头（培训日期）
        if "培训名称" in cells and "姓名" in cells:
            for ci, text in enumerate(cells):
                field = _esg_header_to_field(text)
                if field:
                    col_map[ci] = field
            header_row = ri
            break
    if not col_map:
        raise AppException(
            status_code=400, message="未识别到表头，请使用导出的ESG培训报表格式"
        )

    created = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = {col_map[ci]: (row[ci] if ci < len(row) else None) for ci in col_map}
        name = (
            (vals.get("employee_name") or "").strip()
            if isinstance(vals.get("employee_name"), str)
            else ""
        )
        tname = (
            (vals.get("training_name") or "").strip()
            if isinstance(vals.get("training_name"), str)
            else ""
        )
        if not name and not tname:
            continue
        age = vals.get("age")
        duration = vals.get("duration")
        data = EsgTrainingRecordCreate(
            training_date=_parse_esg_date(vals.get("training_date")) or date.today(),
            training_name=tname or "培训",
            # 导入兜底：Excel 缺省时按集团填报口径补默认值（线下/部门组织/中国大陆）
            training_method=(vals.get("training_method") or "").strip()
            or _ESG_IMPORT_DEFAULTS["training_method"],
            caliber=(vals.get("caliber") or "").strip()
            or _ESG_IMPORT_DEFAULTS["caliber"],
            training_type=(vals.get("training_type") or "").strip() or None,
            employee_name=name or "未知",
            employee_account=vals.get("employee_account") or None,
            location_address=(vals.get("location_address") or "").strip()
            or _ESG_IMPORT_DEFAULTS["location_address"],
            department=department,
            employee_level=vals.get("employee_level") or None,
            gender=vals.get("gender") or None,
            age=int(age) if isinstance(age, (int, float)) else None,
            duration=float(duration) if isinstance(duration, (int, float)) else None,
            remarks=vals.get("remarks") or None,
            apply_company=vals.get("apply_company") or None,
            apply_company_no=vals.get("apply_company_no") or None,
        )
        if created >= ESG_IMPORT_MAX_ROWS:
            raise AppException(
                status_code=400,
                message=(
                    "单次导入不得超过 "
                    f"{ESG_IMPORT_MAX_ROWS} 条记录，请拆分文件后分批导入"
                ),
            )
        await service.create_record(data)
        created += 1

    return success_response(
        data={"created": created}, message=f"成功导入{created}条ESG记录到{department}"
    )


@router.post(
    "/esg-training-records/sync-from-ledger", summary="从培训台账同步生成ESG记录"
)
async def sync_esg_from_ledger(
    department: str = Query(
        ..., description="选中部门（前端部门Tab），仅同步该部门员工"
    ),
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    result = await service.sync_from_ledger(department=department)
    return success_response(
        data=result,
        message=(
            f"同步完成：新建 {result['created']} 条，"
            f"跳过 {result['skipped_existing']} 条已存在记录，"
            f"{result['skipped_unmatched']} 条因员工档案未匹配跳过，"
            f"{result['skipped_other_dept']} 条非本部门人员跳过"
        ),
    )


@router.get("/esg-training-records/{record_id}", summary="ESG培训记录详情")
async def get_esg_record(
    record_id: UUID,
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.get_record(record_id)
    return success_response(
        data=EsgTrainingRecordResponse.model_validate(record).model_dump(mode="json")
    )


@router.put("/esg-training-records/{record_id}", summary="更新ESG培训记录")
async def update_esg_record(
    record_id: UUID,
    payload: EsgTrainingRecordUpdate,
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    record = await service.update_record(record_id, payload)
    return success_response(
        data=EsgTrainingRecordResponse.model_validate(record).model_dump(mode="json"),
        message="ESG培训记录更新成功",
    )


@router.delete("/esg-training-records/{record_id}", summary="删除ESG培训记录")
async def delete_esg_record(
    record_id: UUID,
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    await service.delete_record(record_id)
    return success_response(message="ESG培训记录删除成功")


@router.post(
    "/esg-training-records/batch-delete", summary="批量删除ESG培训记录（软删除）"
)
async def batch_delete_esg_records(
    payload: BatchDeleteRequest,
    service: EsgTrainingRecordService = Depends(get_esg_service),
    current_user: CurrentUser = None,
) -> Any:
    _require_user(current_user)
    if not payload.ids:
        raise AppException(status_code=400, message="请先选择要删除的记录")
    result = await service.batch_delete_records(payload.ids)
    message = f"已删除{result['deleted']}条ESG培训记录"
    if result["failed"]:
        message += f"，{len(result['failed'])}条不存在或已删除"
    return success_response(data=result, message=message)


def _require_user(current_user: CurrentUser) -> None:
    if current_user is None:
        raise AppException(status_code=401, message="请先登录")
