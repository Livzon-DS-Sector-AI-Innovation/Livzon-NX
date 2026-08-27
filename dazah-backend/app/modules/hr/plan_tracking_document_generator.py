"""培训计划跟踪表 Excel 文档生成器（APP11模板）."""

import logging
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]

from app.modules.hr.date_format import fmt_date_obj

logger = logging.getLogger(__name__)


def _find_template() -> Path:
    """Locate the xlsx template, trying several path candidates."""
    candidates = [
        Path("员工培训教育管理规程/APP11培训计划跟踪表.xlsx"),
        Path("../员工培训教育管理规程/APP11培训计划跟踪表.xlsx"),
        Path(__file__).resolve().parent.parent.parent.parent
        / "员工培训教育管理规程"
        / "APP11培训计划跟踪表.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("模板文件未找到: APP11培训计划跟踪表.xlsx")


def generate_plan_tracking_excel(
    records: Any,
    year: int | None = None,
    month: int | None = None,
    plan_level: str | None = None,
) -> BytesIO:
    """Fill the plan tracking template with tracking data.

    Args:
        records: List of PlanTrackingRecord ORM objects
        year/month/plan_level: 期间信息，提供时填充标题行（A4），
            单元格原有字体格式（宋体18粗居中）自动保留。

    Returns:
        BytesIO buffer containing the generated xlsx
    """
    template_path = _find_template()
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb.active

    # 模板结构: 行2 程序标题, 行4 年度月份标题, 行5
    # 培训内容/培训跟踪工作, 行6 表头, 行7-13 数据行, 行14 底部合并行
    # 列: A=序号 B=培训内容或使用教材 C=实际培训时间 D=培训对象
    # E=培训类型 F=考核方式 G=是否按照计划完成 H=跟踪人 I=跟踪日期
    # J=备注

    # 填充标题（锁定区以外的可编辑空位）
    if year and month and plan_level:
        ws["A4"] = f"{year}年度{month}月（{plan_level}）培训计划跟踪表"

    data_start_row = 7  # 数据从行7开始（行6是表头）
    template_data_rows = 7  # 模板预置数据行 7-13

    # 超过模板行数时先批量插入行（行14底部合并区自动下移），并从行13复制样式
    extra = len(records) - template_data_rows
    if extra > 0:
        style_src_row = 13
        ws.insert_rows(14, extra)
        for offset in range(extra):
            new_row = 14 + offset
            for col in range(1, 11):
                src = ws.cell(row=style_src_row, column=col)
                dst = ws.cell(row=new_row, column=col)
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)

    for idx, record in enumerate(records):
        row = data_start_row + idx

        # Fill cells
        ws.cell(row=row, column=1, value=idx + 1)  # 序号
        ws.cell(row=row, column=2, value=record.training_content or "")  # 培训内容
        ws.cell(row=row, column=3, value=record.actual_time or "")  # 实际培训时间
        ws.cell(row=row, column=4, value=record.target_audience or "")  # 培训对象
        ws.cell(row=row, column=5, value=record.training_type or "")  # 培训类型
        ws.cell(
            row=row, column=6, value=record.tracking_assessment_method or ""
        )  # 考核方式

        # 是否按照计划完成 (checkbox style)
        if record.is_completed is True:
            ws.cell(row=row, column=7, value="☑是 □否")
        elif record.is_completed is False:
            ws.cell(row=row, column=7, value="□是 ☑否")
        else:
            ws.cell(row=row, column=7, value="□是 □否")

        ws.cell(row=row, column=8, value=record.tracker or "")  # 跟踪人
        ws.cell(
            row=row,
            column=9,
            value=fmt_date_obj(record.track_date),
        )  # 跟踪日期
        ws.cell(row=row, column=10, value=record.remarks or "")  # 备注

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
