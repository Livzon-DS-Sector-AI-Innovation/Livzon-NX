"""DR 多拉菌素 — 排产放罐计划 API（接罐计划）

数据源：102 发酵车间的 DR 排产 Excel（计划员维护，多个 sheet 为不同版本快照）。
解析「放罐」行（工段行 × 日期列甘特表）：每批哪天放罐、用哪个发酵罐。
放罐 = 发酵液产出 → 201三车间（提炼）据此接罐，即「接罐计划」。

关联 dr_fermentation_batches 标注该批号 DB 是否已有记录（是否已实际投产/接罐），
帮助提炼车间区分「历史已完成」与「未来待接罐」。
"""
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl  # type: ignore[import-untyped]
from fastapi import Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.response import success_response
from app.platform.identity.deps import CurrentUser
from app.shared.module_api import create_module_router
from app.shared.module_registry import MODULES_BY_CODE

logger = logging.getLogger(__name__)
router = create_module_router(MODULES_BY_CODE["production"])

# 排产 Excel 固定存放目录（随代码挂载进容器，替换文件即更新排产）
_SCHEDULE_DIR = Path(__file__).resolve().parent / "schedule_data"

_MONTH_RE = re.compile(r"(\d{4})年(\d{1,2})月")
_DUMP_LABELS = ("放罐",)


class DumpPlanItem(BaseModel):
    batch_no: str  # 批号（DR-xxxxx 正式批 / 中试xxxx / ZS-xxx）
    tank_no: str  # 发酵罐号（B401/B402/B304 等）
    product_type: str  # 正式批 / 中试批
    dump_date: str  # 放罐日期 YYYY-MM-DD
    year: int
    month: int
    day: int
    in_db: bool  # dr_fermentation_batches 是否已有该批号
    is_past: bool  # 放罐日是否已过（今天之前）
    status: str  # completed=已放罐 / upcoming=待放罐
    # ── 接罐任务（production.receiving_task）状态 ──
    task_status: str | None = (
        None  # pending待接罐 / confirmed已确认 / delayed已填延期 / pending_approval待审批 / cancelled已取消  # noqa: E501
    )
    actual_time: str | None = None  # 实际接罐确认时间
    confirmed_by: str | None = None  # 确认人（姓名）
    actual_tank_no: str | None = None  # 实测罐号
    delay_reason: str | None = None  # 延期原因


# 资质硬校验开关：默认关（staff_tank_qualification 尚未录人员数据，不阻塞主流程），
# 待资质表有数据后再打开。开启后：无该罐资质的人员确认 → 自动转 pending_approval 需班组长审批。  # noqa: E501
_QUALIFICATION_ENFORCED = False

_UNDONE_STATUS = ("pending", "pending_approval")


class ConfirmBody(BaseModel):
    actual_tank_no: str | None = None  # 实测罐号（默认取计划罐号）
    note: str | None = None  # 备注
    operator: str | None = None  # 未登录时手动填确认人


class DelayBody(BaseModel):
    delay_reason: str  # 延期原因：等料/等水/等蒸汽/人员不足/设备问题/其他
    note: str | None = None
    operator: str | None = None


class ApproveBody(BaseModel):
    approve: bool = True  # True=批准 / False=驳回
    operator: str | None = None  # 审批人（默认取登录用户）


async def _sync_receiving_tasks(session: AsyncSession, plans: Any) -> Any:
    """排产解析后按批号生成/更新待接罐任务（幂等，GET 内调用可安全重复）。

    - 批号不存在 → 新建 pending（计划罐号/日期取排产）
    - 已存在且仍 pending → 跟随最新排产更新罐号/计划日期
    - 已确认/已延期/待审批 → 不动（执行记录不被排产覆盖）
    """
    for y, mo, d, batch, tank in plans:
        try:
            pd = date(y, mo, d)
        except ValueError:
            continue
        row = (
            await session.execute(
                text(
                    "SELECT id, status FROM production.receiving_task "
                    "WHERE batch_no = :b AND is_deleted = false LIMIT 1"
                ),
                {"b": batch},
            )
        ).fetchone()
        if row is None:
            await session.execute(
                text(
                    "INSERT INTO production.receiving_task (batch_no, tank_no, plan_date, status) "  # noqa: E501
                    "VALUES (:b, :t, :d, 'pending')"
                ),
                {"b": batch, "t": tank, "d": pd},
            )
        elif row[1] == "pending":
            await session.execute(
                text(
                    "UPDATE production.receiving_task SET tank_no = :t, plan_date = :d, updated_at = now() "  # noqa: E501
                    "WHERE id = :id"
                ),
                {"t": tank, "d": pd, "id": row[0]},
            )
    await session.commit()


async def _receiving_task_map(session: AsyncSession) -> dict[str, dict[str, Any]]:
    """批号 → 任务状态映射（供 dump-plans 每批附加任务状态）。

    actual_time 存为 timestamptz（UTC），读出须转本地时区，避免差 8 小时。
    """
    rows = (
        await session.execute(
            text(
                "SELECT batch_no, status, "
                "to_char(actual_time AT TIME ZONE 'Asia/Shanghai', 'YYYY-MM-DD HH24:MI:SS') AS actual_time, "  # noqa: E501
                "confirmed_by, actual_tank_no, delay_reason, approval_status "
                "FROM production.receiving_task WHERE is_deleted = false"
            )
        )
    ).fetchall()
    m: dict[str, dict[str, Any]] = {}
    for bn, st, at, cb, atk, dr, aps in rows:
        m[bn] = {
            "task_status": st,
            "actual_time": at,
            "confirmed_by": cb,
            "actual_tank_no": atk,
            "delay_reason": dr,
            "approval_status": aps,
        }
    return m


def _pick_latest_sheet(wb: Any) -> Any:
    """取 sheet 名中日期最大的版本（当前排产）。无日期可解析则返回第一个 sheet。"""
    best, best_d = None, None
    for ws in wb.worksheets:
        m = re.search(r"(\d{4})\.(\d{2})\.(\d{2})", ws.title)
        if not m:
            continue
        d = (int(m.group(1)), int(m.group(2)), int(m.group(3)))
        if best_d is None or d > best_d:
            best, best_d = ws, d
    return best or (wb.worksheets[0] if wb.worksheets else None)


def _parse_dump_plans(ws: Any) -> Any:
    """解析一个 sheet 的全部「放罐」行 → [(year, month, day, batch_no, tank)]"""
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]
    plans = []
    cur_ym = None
    for i, row in enumerate(rows, start=1):
        a = str(row[0]).strip() if row[0] is not None else ""
        m = _MONTH_RE.search(a)
        if m and "多拉计划" in a:
            cur_ym = (int(m.group(1)), int(m.group(2)))
        if a in _DUMP_LABELS and cur_ym:
            for c in range(1, len(row)):
                v = row[c]
                if v is None:
                    continue
                day = c  # 列 B.. → 1..31
                batch = str(v).strip()
                sub = rows[i][c] if i < len(rows) else None
                tank = str(sub).strip() if sub is not None else ""
                if re.fullmatch(r"\d+", tank):  # 304 → B304
                    tank = "B" + tank
                plans.append((cur_ym[0], cur_ym[1], day, batch, tank))
    return plans


@router.get("/dr/schedule/dump-plans", summary="DR 排产放罐计划（接罐计划）")
async def dr_dump_plans(
    session: AsyncSession = Depends(get_db),
    from_date: str = Query("", description="起始日期 YYYY-MM-DD，空则全部"),
    to_date: str = Query("", description="截止日期 YYYY-MM-DD，空则全部"),
) -> Any:
    files = sorted(_SCHEDULE_DIR.glob("*.xlsx"))
    if not files:
        return success_response(data={"version": None, "items": [], "summary": {}})

    wb = openpyxl.load_workbook(files[-1], data_only=True)
    ws = _pick_latest_sheet(wb)
    plans = _parse_dump_plans(ws)

    # 解析后自动生成/更新接罐任务（幂等：新批号建 pending，已确认不动）
    await _sync_receiving_tasks(session, plans)

    latest = files[-1]
    # 上传时间 = 当前生效排产文件的修改时间（上传/替换即更新）
    upload_time = datetime.fromtimestamp(latest.stat().st_mtime).strftime(
        "%Y-%m-%d %H:%M:%S"
    )
    version = {
        "file": latest.name,
        "sheet": ws.title if ws else "",
        "upload_time": upload_time,
    }
    if not plans:
        return success_response(data={"version": version, "items": [], "summary": {}})

    # DB 已有批号集合（实际投产/已接罐）
    db_rows = (
        await session.execute(
            text(
                "SELECT DISTINCT batch_no FROM production.dr_fermentation_batches WHERE is_deleted = false"  # noqa: E501
            )
        )
    ).fetchall()
    db_set = {r[0] for r in db_rows}

    # 接罐任务状态映射（批号 → 状态）
    task_map = await _receiving_task_map(session)

    today = date.today()
    fd = date.fromisoformat(from_date) if from_date else None
    td = date.fromisoformat(to_date) if to_date else None

    items = []
    for y, mo, d, batch, tank in plans:
        try:
            dd = date(y, mo, d)
        except ValueError:
            continue
        if fd and dd < fd:
            continue
        if td and dd > td:
            continue
        in_db = batch in db_set
        is_past = dd < today
        product_type = (
            "中试批" if ("中试" in batch or batch.startswith("ZS-")) else "正式批"
        )
        t = task_map.get(batch, {})
        items.append(
            DumpPlanItem(
                batch_no=batch,
                tank_no=tank,
                product_type=product_type,
                dump_date=dd.isoformat(),
                year=y,
                month=mo,
                day=d,
                in_db=in_db,
                is_past=is_past,
                status="completed" if is_past else "upcoming",
                task_status=t.get("task_status"),
                actual_time=t.get("actual_time"),
                confirmed_by=t.get("confirmed_by"),
                actual_tank_no=t.get("actual_tank_no"),
                delay_reason=t.get("delay_reason"),
            )
        )
    items.sort(key=lambda it: it.dump_date)

    future = [it for it in items if not it.is_past]
    summary = {
        "total": len(items),
        "past": sum(1 for it in items if it.is_past),
        "upcoming": len(future),
    }
    return success_response(
        data={
            "version": version,
            "today": today.isoformat(),
            "items": [it.model_dump() for it in items],
            "summary": summary,
        }
    )


@router.post(
    "/dr/schedule/upload", summary="DR 排产 Excel 上传更新（替代手动替换文件）"
)
async def dr_schedule_upload(
    file: UploadFile = File(..., description="最新排产 Excel（.xlsx）"),
) -> Any:
    """计划员在系统内上传最新排产 Excel → 保存到 schedule_data/。

    直接使用原文件名保存（同名覆盖，与「手动替换文件」语义一致）；
    dump-plans 按文件名取最新（sorted(glob)[-1]）→ 上传即生效。
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="缺少文件名")
    ext = os.path.splitext(file.filename)[1].lower()
    if ext != ".xlsx":
        raise HTTPException(
            status_code=400, detail=f"仅支持 .xlsx 文件，收到 {ext or '(无扩展名)'}"
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件超过 20MB 限制")

    _SCHEDULE_DIR.mkdir(parents=True, exist_ok=True)
    saved_name = os.path.basename(file.filename or "schedule.xlsx")
    saved_path = _SCHEDULE_DIR / saved_name
    saved_path.write_bytes(content)

    # 校验文件有效，并解析放罐条数返回给前端
    try:
        wb = openpyxl.load_workbook(saved_path, data_only=True)
    except Exception as e:
        saved_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=f"文件不是有效的 xlsx：{e}")

    ws = _pick_latest_sheet(wb)
    plans = _parse_dump_plans(ws)
    logger.info(
        "[DR排产] 上传新排产 %s：sheet=%s 放罐 %d 条",
        saved_name,
        ws.title if ws else "-",
        len(plans),
    )
    return success_response(
        data={
            "file": saved_name,
            "sheet": ws.title if ws else "",
            "total": len(plans),
        },
        message=f"排产已更新：{saved_name}（{len(plans)} 条放罐计划）",
    )


@router.get("/dr/schedule/tasks", summary="DR 接罐任务列表（对账/管理）")
async def dr_receiving_tasks(
    session: AsyncSession = Depends(get_db),
    month: str = Query("", description="月份 YYYY-MM，空则全部"),
    status: str = Query(
        "", description="pending/confirmed/delayed/pending_approval/cancelled，空则全部"
    ),
) -> Any:
    """接罐任务明细列表：排产解析自动生成，确认/延期/审批在此更新。"""
    where = ["is_deleted = false"]
    params = {}
    if month:
        where.append("to_char(plan_date, 'YYYY-MM') = :month")
        params["month"] = month
    if status:
        where.append("status = :status")
        params["status"] = status
    rows = (
        await session.execute(
            text(
                "SELECT batch_no, tank_no, plan_date, status, "
                "to_char(actual_time AT TIME ZONE 'Asia/Shanghai', 'YYYY-MM-DD HH24:MI:SS') AS actual_time, "  # noqa: E501
                "confirmed_by, actual_tank_no, delay_reason, approver, approval_status "
                "FROM production.receiving_task WHERE "
                + " AND ".join(where)
                + " ORDER BY plan_date"
            ),
            params,
        )
    ).fetchall()
    items = [
        {
            "batch_no": r[0],
            "tank_no": r[1],
            "plan_date": r[2].isoformat() if r[2] else None,
            "status": r[3],
            "actual_time": r[4],
            "confirmed_by": r[5],
            "actual_tank_no": r[6],
            "delay_reason": r[7],
            "approver": r[8],
            "approval_status": r[9],
        }
        for r in rows
    ]
    return success_response(data={"items": items, "total": len(items)})


def _operator_name(current_user: Any, fallback: str | None) -> str:
    """确认/审批人：优先登录用户姓名，其次手动填写，兜底「未登录」。"""
    if current_user is not None and getattr(current_user, "name", None):
        return str(current_user.name)
    return fallback or "未登录"


@router.post("/dr/schedule/tasks/{batch_no}/confirm", summary="确认接罐")
async def confirm_receiving(
    batch_no: str,
    body: ConfirmBody,
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """现场接罐确认。

    - 默认 status=confirmed（记 actual_time / confirmed_by / actual_tank_no）
    - 开启资质校验且当前用户无该罐资质 → 转 pending_approval 待班组长审批
    """
    row = (
        await session.execute(
            text(
                "SELECT id, status, tank_no FROM production.receiving_task "
                "WHERE batch_no = :b AND is_deleted = false LIMIT 1"
            ),
            {"b": batch_no},
        )
    ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail=f"未找到批号 {batch_no} 的接罐任务")
    tid, cur_status, plan_tank = row[0], row[1], row[2]
    if cur_status == "confirmed":
        raise HTTPException(
            status_code=400, detail=f"批号 {batch_no} 已确认接罐，不能重复操作"
        )
    if cur_status == "delayed":
        raise HTTPException(
            status_code=400, detail=f"批号 {batch_no} 已填报延期，不能确认"
        )

    operator = _operator_name(current_user, body.operator)
    actual_tank = body.actual_tank_no or plan_tank or ""

    # 资质校验：开关默认关。开启后，若资质表存在该工号记录且未通过 → 需班组长审批
    needs_approval = False
    if _QUALIFICATION_ENFORCED and current_user is not None:
        emp_no = getattr(current_user, "employee_no", None)
        if emp_no:
            q = (
                await session.execute(
                    text(
                        "SELECT qualified FROM production.staff_tank_qualification "
                        "WHERE staff_no = :s AND tank_no = :t AND is_deleted = false LIMIT 1"  # noqa: E501
                    ),
                    {"s": emp_no, "t": actual_tank},
                )
            ).fetchone()
            needs_approval = bool(q is not None and not q[0])

    if needs_approval:
        await session.execute(
            text(
                "UPDATE production.receiving_task SET status = 'pending_approval', confirmed_by = :op, "  # noqa: E501
                "actual_tank_no = :atk, note = :note, updated_at = now() WHERE id = :id"
            ),
            {"op": operator, "atk": actual_tank, "note": body.note, "id": tid},
        )
        await session.commit()
        return success_response(
            message=f"{batch_no} 已提交接罐确认，需班组长审批",
            data={
                "task_status": "pending_approval",
                "confirmed_by": operator,
                "actual_tank_no": actual_tank,
            },
        )

    now = datetime.now()
    await session.execute(
        text(
            "UPDATE production.receiving_task SET status = 'confirmed', actual_time = :at, confirmed_by = :op, "  # noqa: E501
            "actual_tank_no = :atk, note = :note, updated_at = now() WHERE id = :id"
        ),
        {"at": now, "op": operator, "atk": actual_tank, "note": body.note, "id": tid},
    )
    await session.commit()
    logger.info(
        "[接罐] %s 确认接罐，实测罐 %s，确认人 %s", batch_no, actual_tank, operator
    )
    return success_response(
        message=f"{batch_no} 已确认接罐",
        data={
            "task_status": "confirmed",
            "actual_time": now.strftime("%Y-%m-%d %H:%M:%S"),
            "confirmed_by": operator,
            "actual_tank_no": actual_tank,
        },
    )


@router.post("/dr/schedule/tasks/{batch_no}/delay", summary="填报延期原因")
async def delay_receiving(
    batch_no: str,
    body: DelayBody,
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """未接罐批次填报延期原因（等料/等水/等蒸汽/人员不足/设备问题/其他）。"""
    if not body.delay_reason.strip():
        raise HTTPException(status_code=400, detail="请选择延期原因")

    row = (
        await session.execute(
            text(
                "SELECT id, status FROM production.receiving_task "
                "WHERE batch_no = :b AND is_deleted = false LIMIT 1"
            ),
            {"b": batch_no},
        )
    ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail=f"未找到批号 {batch_no} 的接罐任务")
    tid, cur_status = row[0], row[1]
    if cur_status not in _UNDONE_STATUS:
        raise HTTPException(
            status_code=400, detail=f"批号 {batch_no} 当前状态不允许填延期"
        )

    operator = _operator_name(current_user, body.operator)
    await session.execute(
        text(
            "UPDATE production.receiving_task SET status = 'delayed', delay_reason = :r, note = :note, "  # noqa: E501
            "updated_at = now() WHERE id = :id"
        ),
        {"r": body.delay_reason, "note": body.note, "id": tid},
    )
    await session.commit()
    logger.info("[接罐] %s 填报延期：%s（%s）", batch_no, body.delay_reason, operator)
    return success_response(
        message=f"{batch_no} 已填报延期原因：{body.delay_reason}",
        data={"task_status": "delayed", "delay_reason": body.delay_reason},
    )


@router.post("/dr/schedule/tasks/{batch_no}/approve", summary="班组长审批接罐")
async def approve_receiving(
    batch_no: str,
    body: ApproveBody,
    session: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
) -> Any:
    """班组长对「待审批」的接罐确认进行批准/驳回。"""
    row = (
        await session.execute(
            text(
                "SELECT id, status FROM production.receiving_task "
                "WHERE batch_no = :b AND is_deleted = false LIMIT 1"
            ),
            {"b": batch_no},
        )
    ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail=f"未找到批号 {batch_no} 的接罐任务")
    tid, cur_status = row[0], row[1]
    if cur_status != "pending_approval":
        raise HTTPException(
            status_code=400, detail=f"批号 {batch_no} 当前状态不需要审批"
        )

    operator = _operator_name(current_user, body.operator)
    if body.approve:
        now = datetime.now()
        await session.execute(
            text(
                "UPDATE production.receiving_task SET status = 'confirmed', actual_time = :at, "  # noqa: E501
                "approver = :op, approval_status = 'approved', updated_at = now() WHERE id = :id"  # noqa: E501
            ),
            {"at": now, "op": operator, "id": tid},
        )
        await session.commit()
        logger.info("[接罐] %s 接罐已由 %s 批准", batch_no, operator)
        return success_response(
            message=f"{batch_no} 已批准接罐",
            data={
                "task_status": "confirmed",
                "approver": operator,
                "approval_status": "approved",
            },
        )

    await session.execute(
        text(
            "UPDATE production.receiving_task SET status = 'pending', approver = :op, "
            "approval_status = 'rejected', updated_at = now() WHERE id = :id"
        ),
        {"op": operator, "id": tid},
    )
    await session.commit()
    logger.info("[接罐] %s 接罐被 %s 驳回", batch_no, operator)
    return success_response(
        message=f"{batch_no} 已驳回，恢复待接罐",
        data={
            "task_status": "pending",
            "approver": operator,
            "approval_status": "rejected",
        },
    )
