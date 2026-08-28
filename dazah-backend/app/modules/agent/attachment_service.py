from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import logging
import uuid
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Literal

from fastapi import HTTPException, status
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.storage import delete_object, get_object, is_enabled, upload_object

from .models import AgentAttachment
from .repository import AgentRepository

_STORAGE_MODULE = "agent"
_READ_CHUNK_MAX_CHARS = 20_000
logger = logging.getLogger(__name__)


class AgentAttachmentService:
    def __init__(self, repo: AgentRepository | None = None) -> None:
        self.repo = repo or AgentRepository()

    @staticmethod
    def object_key(
        *, user_id: uuid.UUID, session_id: uuid.UUID, attachment_id: uuid.UUID
    ) -> str:
        return f"sessions/{user_id}/{session_id}/{attachment_id}/source"

    @staticmethod
    async def _readback_matches(key: str, expected: bytes) -> bool:
        stored = await asyncio.to_thread(get_object, _STORAGE_MODULE, key)
        return bool(
            stored is not None
            and stored[0] == expected
            and hashlib.sha256(stored[0]).digest() == hashlib.sha256(expected).digest()
        )

    @staticmethod
    async def _restore_object(
        key: str,
        raw: bytes,
        content_type: str,
    ) -> None:
        await asyncio.to_thread(
            upload_object,
            _STORAGE_MODULE,
            key,
            raw,
            len(raw),
            content_type,
        )
        if not await AgentAttachmentService._readback_matches(key, raw):
            raise RuntimeError("attachment object compensation readback failed")

    async def persist(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID,
        message_id: uuid.UUID | None,
        user_id: uuid.UUID,
        uploads: list[dict[str, Any]],
    ) -> list[AgentAttachment]:
        if not uploads:
            return []
        if not is_enabled():
            raise HTTPException(
                status.HTTP_503_SERVICE_UNAVAILABLE,
                "附件持久存储当前不可用，请稍后重试",
            )

        persisted: list[AgentAttachment] = []
        uploaded_keys: list[str] = []
        try:
            for upload in uploads:
                attachment_id = uuid.UUID(str(upload["attachment_id"]))
                raw = upload["data"]
                if not isinstance(raw, bytes):
                    raise TypeError("attachment data must be bytes")
                key = self.object_key(
                    user_id=user_id,
                    session_id=session_id,
                    attachment_id=attachment_id,
                )
                try:
                    await asyncio.to_thread(
                        upload_object,
                        _STORAGE_MODULE,
                        key,
                        raw,
                        len(raw),
                        str(upload["content_type"]),
                    )
                    uploaded_keys.append(key)
                    if not await self._readback_matches(key, raw):
                        raise RuntimeError("attachment object readback mismatch")
                except Exception as exc:
                    raise HTTPException(
                        status.HTTP_503_SERVICE_UNAVAILABLE,
                        "附件持久存储写入失败，请稍后重试",
                    ) from exc
                persisted.append(
                    await self.repo.create_attachment(
                        db,
                        attachment_id=attachment_id,
                        session_id=session_id,
                        message_id=message_id,
                        user_id=user_id,
                        filename=str(upload["filename"]),
                        content_type=str(upload["content_type"]),
                        size=len(raw),
                        kind=str(upload["kind"]),
                        object_key=key,
                        sha256=hashlib.sha256(raw).hexdigest(),
                        extracted_text=(str(upload.get("text") or "") or None),
                    )
                )
            await db.flush()
        except Exception:
            for key in uploaded_keys:
                try:
                    await asyncio.to_thread(delete_object, _STORAGE_MODULE, key)
                except Exception:
                    logger.exception(
                        "Failed to clean up attachment object after persist failure: "
                        "object_key=%s",
                        key,
                    )
            raise
        return persisted

    async def list_for_session(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID,
        user_id: uuid.UUID,
    ) -> list[AgentAttachment]:
        return await self.repo.list_session_attachments(
            db,
            session_id=session_id,
            user_id=user_id,
        )

    async def materialize_for_context(
        self,
        attachments: list[AgentAttachment],
        *,
        text_limit: int,
    ) -> list[dict[str, Any]]:
        restored: list[dict[str, Any]] = []
        for attachment in attachments[:5]:
            item: dict[str, Any] = {
                "attachment_id": str(attachment.id),
                "filename": attachment.filename,
                "content_type": attachment.content_type,
                "size": attachment.size,
                "kind": attachment.kind,
                "truncated": False,
            }
            if attachment.kind == "image":
                stored = await asyncio.to_thread(
                    get_object, _STORAGE_MODULE, attachment.object_key
                )
                if stored is None or not await self._readback_matches(
                    attachment.object_key, stored[0]
                ):
                    raise HTTPException(
                        status.HTTP_409_CONFLICT,
                        f"附件原文件已丢失或损坏：{attachment.filename}",
                    )
                raw, _ = stored
                if hashlib.sha256(raw).hexdigest() != attachment.sha256:
                    raise HTTPException(
                        status.HTTP_409_CONFLICT,
                        f"附件原文件校验失败：{attachment.filename}",
                    )
                item["data_base64"] = base64.b64encode(raw).decode("ascii")
            else:
                content = attachment.extracted_text or "（未提取到可读文本）"
                item["text"] = content[:text_limit]
                item["truncated"] = len(content) > text_limit
            restored.append(item)
        return restored

    async def require(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
        attachment_ref: str,
    ) -> AgentAttachment:
        if session_id is None or user_id is None:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "附件操作必须在有效会话中执行",
            )
        attachment = await self.repo.get_session_attachment(
            db,
            session_id=session_id,
            user_id=user_id,
            attachment_ref=attachment_ref.strip(),
        )
        if attachment is None:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "会话附件不存在")
        return attachment

    async def read(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
        attachment_ref: str,
        offset: int,
        limit: int,
    ) -> dict[str, Any]:
        attachment = await self.require(
            db,
            session_id=session_id,
            user_id=user_id,
            attachment_ref=attachment_ref,
        )
        content = attachment.extracted_text or ""
        bounded_limit = min(max(limit, 1), _READ_CHUNK_MAX_CHARS)
        start = min(max(offset, 0), len(content))
        end = min(start + bounded_limit, len(content))
        return {
            "attachment_id": str(attachment.id),
            "filename": attachment.filename,
            "content_type": attachment.content_type,
            "version": attachment.version,
            "offset": start,
            "next_offset": end if end < len(content) else None,
            "total_chars": len(content),
            "content": content[start:end],
        }

    async def mutate_tabular(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
        attachment_ref: str,
        action: Literal["append_row", "update_row", "delete_row"],
        sheet_name: str | None,
        row_number: int | None,
        values: list[Any],
    ) -> dict[str, Any]:
        attachment = await self.require(
            db,
            session_id=session_id,
            user_id=user_id,
            attachment_ref=attachment_ref,
        )
        stored = await asyncio.to_thread(
            get_object, _STORAGE_MODULE, attachment.object_key
        )
        if stored is None:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "附件原文件已丢失，无法执行修改",
            )
        raw, _ = stored
        extension = Path(attachment.filename).suffix.lower()
        if extension == ".xlsx":
            updated = self._mutate_xlsx(
                raw,
                action=action,
                sheet_name=sheet_name,
                row_number=row_number,
                values=values,
            )
        elif extension == ".csv":
            updated = self._mutate_csv(
                raw,
                action=action,
                row_number=row_number,
                values=values,
            )
        else:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "当前仅支持修改 XLSX 和 CSV 会话附件",
            )

        from .service import _extract_attachment_text

        extracted = await asyncio.to_thread(
            _extract_attachment_text, extension, updated
        )
        try:
            await asyncio.to_thread(
                upload_object,
                _STORAGE_MODULE,
                attachment.object_key,
                updated,
                len(updated),
                attachment.content_type,
            )
            if not await self._readback_matches(attachment.object_key, updated):
                raise RuntimeError("attachment object readback mismatch")
        except Exception as exc:
            raise HTTPException(
                status.HTTP_503_SERVICE_UNAVAILABLE,
                "附件持久存储更新失败，请稍后重试",
            ) from exc
        original_state = (
            attachment.size,
            attachment.sha256,
            attachment.extracted_text,
            attachment.version,
            attachment.updated_by,
        )
        try:
            attachment.size = len(updated)
            attachment.sha256 = hashlib.sha256(updated).hexdigest()
            attachment.extracted_text = extracted[:100_000]
            attachment.version += 1
            attachment.updated_by = user_id
            await db.flush()
        except Exception:
            (
                attachment.size,
                attachment.sha256,
                attachment.extracted_text,
                attachment.version,
                attachment.updated_by,
            ) = original_state
            try:
                await self._restore_object(
                    attachment.object_key,
                    raw,
                    attachment.content_type,
                )
            except Exception:
                logger.exception(
                    "Failed to restore attachment object after database mutation "
                    "failure: attachment_id=%s",
                    attachment.id,
                )
            raise
        return {
            "attachment_id": str(attachment.id),
            "filename": attachment.filename,
            "action": action,
            "sheet_name": sheet_name,
            "row_number": row_number,
            "version": attachment.version,
            "sha256": attachment.sha256,
        }

    async def delete(
        self,
        db: AsyncSession,
        *,
        session_id: uuid.UUID | None,
        user_id: uuid.UUID | None,
        attachment_ref: str,
    ) -> dict[str, Any]:
        attachment = await self.require(
            db,
            session_id=session_id,
            user_id=user_id,
            attachment_ref=attachment_ref,
        )
        stored = await asyncio.to_thread(
            get_object, _STORAGE_MODULE, attachment.object_key
        )
        if stored is None:
            raise HTTPException(status.HTTP_409_CONFLICT, "附件原文件已丢失，无法删除")
        raw, stored_content_type = stored
        try:
            await asyncio.to_thread(
                delete_object, _STORAGE_MODULE, attachment.object_key
            )
            if (
                await asyncio.to_thread(
                    get_object, _STORAGE_MODULE, attachment.object_key
                )
                is not None
            ):
                raise RuntimeError("attachment object still exists after deletion")
        except Exception as exc:
            raise HTTPException(
                status.HTTP_503_SERVICE_UNAVAILABLE,
                "附件持久存储删除失败，请稍后重试",
            ) from exc
        original_deleted = attachment.is_deleted
        original_updated_by = attachment.updated_by
        try:
            attachment.is_deleted = True
            attachment.updated_by = user_id
            await db.flush()
        except Exception:
            attachment.is_deleted = original_deleted
            attachment.updated_by = original_updated_by
            try:
                await self._restore_object(
                    attachment.object_key,
                    raw,
                    stored_content_type or attachment.content_type,
                )
            except Exception:
                logger.exception(
                    "Failed to restore deleted attachment after database failure: "
                    "attachment_id=%s",
                    attachment.id,
                )
            raise
        return {
            "attachment_id": str(attachment.id),
            "filename": attachment.filename,
            "deleted": True,
        }

    @staticmethod
    def _mutate_xlsx(
        raw: bytes,
        *,
        action: str,
        sheet_name: str | None,
        row_number: int | None,
        values: list[Any],
    ) -> bytes:
        workbook = load_workbook(BytesIO(raw))
        try:
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise HTTPException(
                        status.HTTP_400_BAD_REQUEST,
                        f"工作表不存在：{sheet_name}",
                    )
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            if action == "append_row":
                if not values:
                    raise HTTPException(
                        status.HTTP_400_BAD_REQUEST, "新增行必须提供 values"
                    )
                worksheet.append(values)
                row_number = worksheet.max_row
            else:
                if (
                    row_number is None
                    or row_number < 1
                    or row_number > worksheet.max_row
                ):
                    raise HTTPException(
                        status.HTTP_400_BAD_REQUEST, "row_number 超出工作表范围"
                    )
                if action == "delete_row":
                    worksheet.delete_rows(row_number, 1)
                else:
                    if not values:
                        raise HTTPException(
                            status.HTTP_400_BAD_REQUEST, "修改行必须提供 values"
                        )
                    for column, value in enumerate(values, start=1):
                        worksheet.cell(row=row_number, column=column, value=value)
            output = BytesIO()
            workbook.save(output)
            return output.getvalue()
        finally:
            workbook.close()

    @staticmethod
    def _mutate_csv(
        raw: bytes,
        *,
        action: str,
        row_number: int | None,
        values: list[Any],
    ) -> bytes:
        decoded: str | None = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                decoded = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "CSV 编码不受支持，请使用 UTF-8 或 GB18030",
            )
        rows = list(csv.reader(StringIO(decoded)))
        if action == "append_row":
            if not values:
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST, "新增行必须提供 values"
                )
            rows.append([str(value) if value is not None else "" for value in values])
        else:
            if row_number is None or row_number < 1 or row_number > len(rows):
                raise HTTPException(
                    status.HTTP_400_BAD_REQUEST, "row_number 超出 CSV 范围"
                )
            index = row_number - 1
            if action == "delete_row":
                rows.pop(index)
            else:
                if not values:
                    raise HTTPException(
                        status.HTTP_400_BAD_REQUEST, "修改行必须提供 values"
                    )
                rows[index] = [
                    str(value) if value is not None else "" for value in values
                ]
        output = StringIO(newline="")
        writer = csv.writer(output)
        writer.writerows(rows)
        return output.getvalue().encode("utf-8-sig")
